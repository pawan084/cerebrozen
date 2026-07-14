"""Prompt registry — loads editable prompts from the workbook (cell B7,
continuing into B8, B9, ... if the prompt exceeds Excel's 32,767-char cell limit).

Graph *structure* is code; prompt *text* stays in agent_prompts.xlsx so
non-technical users keep editing it (constitution: prompts editable). Each
graph stage maps to a workbook sheet; `reload()` re-reads the file so a prompt
edit takes effect with no redeploy.
"""

from __future__ import annotations

import hashlib
import logging
import os
import re
import threading
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from app import config
from app.llm.prompt_store import resolve_workbook

logger = logging.getLogger("cerebrozen.prompts")

# Prompt text starts at B7. Excel caps a single cell at 32,767 characters, so a
# prompt that hits that ceiling continues downward in the same column — B8, B9,
# ... — with no other content allowed in those cells. Column B rows 8+ on a
# stage sheet are reserved exclusively for this continuation; the reader below
# concatenates them in order and stops at the first blank cell.
PROMPT_CELL = "B7"
PROMPT_START_ROW = 7
PROMPT_COL = 2  # column B

# The workbook tab that drives agent enable/disable. Control lives here (with the
# prompt engineers), NOT in code/env flags: an agent is active iff its row's
# `enabled` column is truthy. Columns: agent_name | role | enabled | model |
# sheet_name | description. We key the enabled map by the (stripped) `sheet_name`
# so it lines up with STAGE_SHEET. An agent missing from the catalog → disabled.
CATALOG_SHEET = "Catalog"
CATALOG_HEADER_ROW = 1

# Always-on regardless of the catalog: the guardrail/identity layer and the
# closing layer. `feedback_mood_capture_agent` is the SOLE legitimate path to a
# terminal close (the deck's "End of chat" agent), so it must never be gated off
# — if it's disabled the session simply can't end naturally, which would strand
# every conversation. `environment` is the always-on safety layer.
ALWAYS_ENABLED = {"environment", "feedback_mood_capture_agent"}

# Graph stage -> workbook sheet name. Stage key == sheet name throughout.
# CBT + Capability are the Phase 3 paths: registered here so their prompts load
# the instant they exist in the workbook. A missing/empty sheet logs
# `prompt.missing_sheet` and returns "" — the node treats that as "not authored
# yet" and falls back to CIM (see nodes._run_path_stage).
STAGE_SHEET: Dict[str, str] = {
    "environment": "environment_system_agent",  # the always-on guardrail layer
    "repeat_user_checkin_agent": "repeat_user_checkin_agent",
    "coaching_intake_agent": "coaching_intake_agent",
    "challenge_context_agent": "challenge_context_agent",
    "core_coaching_agent": "core_coaching_agent",  # unified CIM + CBT engine
    "CH_coaching_agent": "CH_coaching_agent",
    # Post-coaching simulation gate — decides/offers role_play vs SJT vs skip. Loads the
    # instant its sheet is authored + enabled; a missing/disabled entry falls back to the
    # deterministic simulation gate (nodes._post_actions_stage).
    "simulation_decision_agent": "simulation_decision_agent",
    # Phase 6 simulation (CIM/CBT only) — reached when enabled in the Catalog tab.
    "role_play_agent": "role_play_agent",
    "SJT_simulation_agent": "SJT_simulation_agent",
    # Learning-aid support node (CIM/CBT/CH) — after coaching/simulation, before the
    # closing layer; reached when enabled in the Catalog tab. A missing/empty sheet
    # skips cleanly to the closing layer, like simulation.
    "learning_aid_agent": "learning_aid_agent",
    # Closing layer (mood + feedback capture → EndOfConversation). The workbook
    # sheet name has a trailing space; the .strip() lookup above resolves it.
    "feedback_mood_capture_agent": "feedback_mood_capture_agent",
    # Background builders (off the request path) — loaded here so they're editable
    # in the workbook like every other prompt.
    "dynamic_actions_insights_agent": "dynamic_actions_insights_agent",
    "user_context_builder_agent": "user_context_builder_agent",
    "pattern_agent": "pattern_agent",
    # Standalone per-action check-in (UI-triggered, outside the main flow). Loads the
    # instant its `action_checkin_agent` sheet exists in the workbook; a missing/empty
    # sheet makes the node return a graceful "unavailable" reply (see nodes.action_checkin_node).
    "action_checkin_agent": "action_checkin_agent",
}

# Advisory size budget per prompt. Excel's hard 32,767-char cell cap forces a
# continuation spill (B8, B9, …), but a prompt anywhere near that size is itself
# a latency/cost/quality smell (~6K tokens) — the validation report flags it
# well before the spill so authors trim or modularize instead of growing blobs.
PROMPT_SIZE_WARN_CHARS = 24_000

# How many rows past the first blank cell the loader scans for stray content.
# Content BELOW a blank cell is silently dropped by the continuation reader
# (it stops at the first blank), which is exactly how an accidental blank row
# corrupts a spilled prompt — so the loader looks ahead and reports orphans.
_ORPHAN_SCAN_ROWS = 20

# Same token grammar as app.rag.placeholders.PLACEHOLDER_RE (kept local to avoid
# an import cycle through the RAG extraction stack).
_PLACEHOLDER_RE = re.compile(r"\{[A-Za-z0-9_.&\-]+\}")

# Context tokens the runtime injects into the resolver's context dict (profile
# read, engine metadata, guardrails aliases, node bridges). Used ONLY by the
# advisory placeholder audit in the validation report: a token that is neither
# a registered RAG extraction, nor a variable-capture-registry variable, nor
# listed here will be blanked at runtime by the resolver — flagging it at
# load time turns a silent blank into a reviewable warning.
KNOWN_CONTEXT_TOKENS = frozenset({
    # identity / request metadata
    "userName", "user_name", "name", "username", "user_id", "session_id",
    "language", "Time", "timezone", "country", "conversation_mode",
    "session_continued", "invoking_agent",
    # org / role
    "org_id", "orgId", "organizationName", "organizationValues",
    "user_level", "level", "user_role", "userPosition",
    "userRoleContext", "user_role_context",
    # routing / path
    "coaching_path", "coachingPath", "currentPhase", "userRepeatFresh",
    # live turn signals
    "user_message", "conversation", "conversation_history", "history_text",
    "user_goal_challenge", "user_challenge", "currentChallenge",
    "session_goal", "skill_to_develop", "confirmed_competency",
    "coaching_shift_summary", "timeAvailable", "presenting_issue_summary",
    # intake / profile
    "coachingHistory", "coachingNeeds", "userMotivations", "user_motivations",
    "userStrengths", "userGaps", "userWorkEnvironment",
    "userThinkingPreference", "user_thinking_preference",
    "userBehavioralPreference", "userBehavioralChoices",
    "coachability_score", "coachabilityScore",
    "coachability_detail", "coachabilityDetail",
    "coaching_style_preference", "coaching_style_context",
    "coaching_style_context_selected_style",
    "behavioral_intake_responses", "ic_profile", "nbi_scores", "pattern",
    "prev_pattern_table", "session_count", "sessions_completed", "last_session_at",
    # repeat-user / memory
    "previousUserContext", "previousUserActions", "previousUserInsights",
    "pastConversation", "checkinDue", "checkinEligibleActions", "checkinSessionIds",
    "repeatingUserCustomPrompt", "customCoachingStylePrompt",
    "customBehavioralQuestionPrompt",
    # CH / actions
    "idp_competencies", "deep_link_skill", "competency_source",
    "committed_action", "committed_by_when", "action_item", "action_outcome",
    "learning_aid", "LearningAid",
    # CH carry-over fields bridged from coaching_progress into user_context each
    # turn (nodes._run_stage context bridge) — resolvable once the CH arc runs.
    "user_blueprint", "mastery_rubric", "long_term_goal", "short_term_goal",
    "user_career_aspirations", "user_strengths", "userStrengths",
    "user_gaps", "userGaps", "user_work_environment", "userWorkEnvironment",
    "current_phase", "sessionContinued",
    # Session-captured contract fields (state notepad / CH + builder contracts):
    # emitted by agents in context_update / variables_set during the arc and
    # re-injected into the resolver context on later turns.
    "stated_outcome", "synthesised_goal", "synthesised_strengths",
    "confirmed_competency_and_behaviours", "accountability_plan",
    "basic_behavior", "intermediate_behavior", "mastery_description",
    "criteria", "impact_across_levels", "personal_and_org_relevance",
    "challenges_framed_constructively", "skill_1_name", "skill_2_name",
    "support_needed", "timeline", "user_concerns", "goal_priority_order",
    "ch_coaching_shift_summary", "ch_committed_action", "ch_committed_by_when",
    "ch_thinking_preference_used", "committed_actions",
    "attempts_so_far", "early_pattern_signals", "emerging_insight",
    "real_issue_hypothesis", "selected_concept_name", "selected_module",
    "session_transcript", "DeepLinkSkill", "IDPCompetencies",
})

# Crisis-path scripted reply (no LLM). Kept in CODE, not the workbook: it is a safety
# mechanism, not coaching craft — a prompt author must not be able to edit or disable it.
#
# THE HELPLINE IS PER-REGION AND MUST BE SET PER CLIENT.
# It shipped hardcoded to 988 (US) — which is worse than useless to a user in India or the
# UK: it hands someone in crisis a number that does not answer. `CEREBROZEN_CRISIS_LINE`
# overrides it; the default names no country and points at an international directory that
# resolves to the caller's own region.
#
#   CEREBROZEN_CRISIS_LINE="988 (US)"
#   CEREBROZEN_CRISIS_LINE="the Tele-MANAS helpline on 14416 (India)"
#   CEREBROZEN_CRISIS_LINE="Samaritans on 116 123 (UK)"
_DEFAULT_CRISIS_LINE = (
    "your local emergency number, or find a helpline for your country at findahelpline.com"
)
CRISIS_LINE = os.environ.get("CEREBROZEN_CRISIS_LINE", "").strip() or _DEFAULT_CRISIS_LINE

# The crisis reply itself lives in app/graph/crisis.py, with the safety screen that
# triggers it and its per-language variants. It is NOT duplicated here: a second copy that
# nothing serves is a trap — the white-label tests once asserted on exactly such a string,
# so a client could "fix" the crisis line, watch the tests pass, and still ship the reply
# the node actually sends. CRISIS_LINE above is the one piece both need.


def _read_prompt_cell(ws) -> Tuple[str, List[int]]:
    """Read a stage sheet's prompt starting at B7, continuing into B8, B9, ...
    while a hard 32,767-char-per-cell Excel limit forces a prompt to be split
    across cells. Fragments are joined with no separator (the split can land
    mid-word) and `.strip()` is applied once to the combined result — stripping
    each fragment individually would eat a legitimate space at a split boundary.
    Stops at the first blank cell in the column.

    Also returns the rows of any ORPHANED content found within _ORPHAN_SCAN_ROWS
    below the first blank cell: that content is not part of the loaded prompt
    (the reader stopped above it), which is how an accidental blank row silently
    truncates a spilled prompt — the validation report surfaces it."""
    parts = []
    row = PROMPT_START_ROW
    while True:
        value = ws.cell(row=row, column=PROMPT_COL).value
        if value is None or str(value) == "":
            break
        parts.append(str(value))
        row += 1
    orphans = [
        r for r in range(row + 1, row + 1 + _ORPHAN_SCAN_ROWS)
        if ws.cell(row=r, column=PROMPT_COL).value not in (None, "")
    ]
    return "".join(parts).strip(), orphans


def _parse_enabled(value: object) -> bool:
    """Coerce a Catalog `enabled` cell to bool. openpyxl returns a real bool for
    cells Excel stored as TRUE/FALSE, but a string for text-typed cells — handle
    both. Anything unrecognised (blank, None) → disabled."""
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    return str(value).strip().lower() in {"true", "1", "yes", "y", "enabled", "on"}


class PromptRegistry:
    def __init__(self, path: Optional[str] = None) -> None:
        # path=None -> resolve per PROMPT_SOURCE on each reload (local or S3);
        # an explicit path (tests) pins the file and skips resolution.
        self.path = path
        self._prompts: Dict[str, str] = {}
        # sheet_name (stripped) -> enabled, read from the Catalog tab each reload.
        self._enabled: Dict[str, bool] = {}
        # sheet_name (stripped) -> model id, read from the Catalog `model` column.
        self._models: Dict[str, str] = {}
        # Content hash of the loaded prompts+catalog. Changes exactly when the
        # loaded content changes, so it doubles as the LLM prompt-cache buster
        # (see responses_client._cache_header) and as the version admins compare
        # across environments.
        self.version: str = ""
        self.loaded_at: str = ""
        self.source: str = ""       # codebase | s3 | s3-fallback
        self.source_path: str = ""
        # True when the registry is NOT serving what it was configured to serve:
        # S3 fell back to the bundled file, or a reload failed and the previous
        # in-memory prompts were kept. Surfaced by the admin API.
        self.degraded: bool = False
        self.degraded_reason: str = ""
        # Structured issue report from the last load — see _validate().
        self.validation: Dict[str, Any] = {}
        self._lock = threading.Lock()
        self.reload()

    def reload(self) -> None:
        """Re-read the workbook and atomically swap the in-memory registry.

        Resolution happens each reload so S3 edits propagate (the resolver
        re-downloads in s3 mode); an explicit self.path pins the file (tests).
        A reload that fails AFTER a successful boot keeps the previous prompts
        and marks the registry degraded — a bad upload must never take live
        prompts down; only the very first load is allowed to raise."""
        with self._lock:
            if self.path:
                src = {"path": self.path, "source": "pinned", "fallback": False, "error": None}
            else:
                src = resolve_workbook()
            try:
                prompts, enabled, models, orphans, missing = self._load(src["path"])
            except Exception as exc:  # noqa: BLE001
                if self._prompts:
                    self.degraded = True
                    self.degraded_reason = f"reload failed, serving previous prompts: {exc}"
                    logger.error(
                        "prompts.reload_failed_keeping_previous",
                        extra={"path": src["path"], "error": str(exc)},
                    )
                    return
                raise  # first load: fail fast — there is nothing to serve.
            self._prompts = prompts
            self._enabled = enabled
            self._models = models
            self.version = self._compute_version(prompts, enabled, models)
            self.loaded_at = datetime.now(timezone.utc).isoformat()
            self.source = src["source"]
            self.source_path = src["path"]
            self.degraded = bool(src["fallback"])
            self.degraded_reason = src["error"] or ""
            self.validation = self._validate(prompts, enabled, models, orphans, missing)
            logger.info(
                "prompts.loaded",
                extra={
                    "path": src["path"],
                    "source": self.source,
                    "version": self.version,
                    "degraded": self.degraded,
                    "sizes": {k: len(v) for k, v in prompts.items()},
                    "enabled": enabled,
                    "validation_issues": self.validation.get("issue_count", 0),
                },
            )
            _set_current_version(self.version)

    def _load(self, path: str):
        """Read prompts + catalog from the workbook file. Pure read — no state swap."""
        # data_only=False: prompts are plain text, and data_only reads Excel's
        # cached values which openpyxl drops on save (so an openpyxl-edited
        # workbook would read empty). Read the actual cell values instead.
        wb = load_workbook(path, read_only=True, data_only=False)
        try:
            lookup = {s.strip(): s for s in wb.sheetnames}
            prompts: Dict[str, str] = {}
            orphans: Dict[str, List[int]] = {}
            missing: List[str] = []
            for stage, sheet in STAGE_SHEET.items():
                actual = lookup.get(sheet.strip())
                if actual is None:
                    logger.warning("prompt.missing_sheet", extra={"stage": stage, "sheet": sheet})
                    prompts[stage] = ""
                    missing.append(stage)
                    continue
                prompts[stage], stray = _read_prompt_cell(wb[actual])
                if stray:
                    orphans[stage] = stray
            enabled, models = self._read_catalog(wb, lookup)
        finally:
            wb.close()
        return prompts, enabled, models, orphans, missing

    @staticmethod
    def _compute_version(
        prompts: Dict[str, str], enabled: Dict[str, bool], models: Dict[str, str]
    ) -> str:
        """Deterministic short hash of the loaded content (prompt text + catalog).
        Identical content → identical version, regardless of file bytes/mtime."""
        h = hashlib.sha256()
        for stage in sorted(prompts):
            h.update(stage.encode())
            h.update(prompts[stage].encode())
        for sheet in sorted(enabled):
            h.update(f"{sheet}={enabled[sheet]}|{models.get(sheet, '')}".encode())
        return h.hexdigest()[:12]

    def _validate(
        self,
        prompts: Dict[str, str],
        enabled: Dict[str, bool],
        models: Dict[str, str],
        orphans: Dict[str, List[int]],
        missing: List[str],
    ) -> Dict[str, Any]:
        """Advisory issue report for the loaded workbook (never raises, never gates).

        Catches at load time the classes of defect that otherwise surface as live
        incidents: an enabled agent with no prompt or no model (turn-time
        RuntimeError), stray content below a blank continuation cell (silent
        truncation), a placeholder that no data source can resolve (blanked at
        runtime), and prompts past the size budget."""
        report: Dict[str, Any] = {
            "missing_sheets": list(missing),  # stage sheet absent from the workbook
            "not_in_catalog": [],       # stage has no Catalog row at all
            "enabled_no_prompt": [],    # enabled agent with an empty prompt
            "enabled_no_model": [],     # enabled agent with a blank model cell
            "orphaned_continuation": {},  # stage -> rows with content below a blank cell
            "oversize": {},             # stage -> chars, past PROMPT_SIZE_WARN_CHARS
            "unknown_placeholders": {},  # stage -> tokens nothing can resolve
        }
        known_extra = self._known_tokens_from_registries()
        for stage, sheet in STAGE_SHEET.items():
            key = sheet.strip()
            text = prompts.get(stage, "")
            is_on = stage in ALWAYS_ENABLED or enabled.get(key, False)
            if stage not in ("environment",) and key not in enabled:
                report["not_in_catalog"].append(stage)
            if is_on:
                if not text:
                    report["enabled_no_prompt"].append(stage)
                # `environment` is the wrapper layer — it never makes its own LLM
                # call, so a blank model cell is not an issue for it.
                if not models.get(key) and stage != "environment":
                    report["enabled_no_model"].append(stage)
            if stage in orphans:
                report["orphaned_continuation"][stage] = orphans[stage]
            if len(text) > PROMPT_SIZE_WARN_CHARS:
                report["oversize"][stage] = len(text)
            unknown = sorted(
                tok.strip("{}") for tok in set(_PLACEHOLDER_RE.findall(text))
                if not self._token_resolvable(tok.strip("{}"), known_extra)
            )
            if unknown:
                report["unknown_placeholders"][stage] = unknown
        report["issue_count"] = (
            len(report["missing_sheets"]) + len(report["not_in_catalog"])
            + len(report["enabled_no_prompt"]) + len(report["enabled_no_model"])
            + len(report["orphaned_continuation"]) + len(report["oversize"])
            + len(report["unknown_placeholders"])
        )
        report["ok"] = report["issue_count"] == 0
        if not report["ok"]:
            logger.warning("prompts.validation_issues", extra={"report": report})
        return report

    @staticmethod
    def _known_tokens_from_registries() -> frozenset:
        """Tokens resolvable by the RAG extraction registry or declared in the
        variable-capture registry. Lazy imports + broad except: validation is
        advisory and must never break a prompt load."""
        tokens: set = set()
        try:
            from app.rag.registry import get_registry as _rag_registry
            tokens.update(_rag_registry().binding_tokens())
        except Exception:  # noqa: BLE001
            pass
        try:
            from app.stores.variable_capture_registry import VariableCaptureRegistry
            reg = VariableCaptureRegistry.get()
            for var_name, cfg in reg.all_vars.items():
                tokens.add(var_name)
                if cfg.prompt_placeholder:
                    tokens.add(cfg.prompt_placeholder.strip("{}"))
        except Exception:  # noqa: BLE001
            pass
        try:
            # Every declared state field is a legitimate context token once a node
            # lifts it into user_context (e.g. {specific_person_identified}).
            from app.graph.state import CereBroZenState
            tokens.update(CereBroZenState.__annotations__.keys())
        except Exception:  # noqa: BLE001
            pass
        return frozenset(tokens)

    @staticmethod
    def _token_resolvable(name: str, known_extra: frozenset) -> bool:
        if name in KNOWN_CONTEXT_TOKENS or name in known_extra:
            return True
        # Dot-path tokens resolve against nested context dicts; audit the root.
        root = name.split(".", 1)[0]
        return root in KNOWN_CONTEXT_TOKENS or root in known_extra

    @staticmethod
    def _read_catalog(
        wb, lookup: Dict[str, str]
    ) -> Tuple[Dict[str, bool], Dict[str, str]]:
        """Build {sheet_name: enabled} and {sheet_name: model} from the Catalog tab.
        A missing Catalog tab or unreadable row → that agent is absent from both maps.
        Never raises (a bad catalog must not break prompt loading)."""
        _empty: Tuple[Dict[str, bool], Dict[str, str]] = ({}, {})
        actual = lookup.get(CATALOG_SHEET.strip())
        if actual is None:
            logger.warning("prompt.missing_catalog", extra={"sheet": CATALOG_SHEET})
            return _empty
        ws = wb[actual]
        rows = ws.iter_rows(min_row=CATALOG_HEADER_ROW, values_only=True)
        try:
            header = [str(h).strip().lower() if h is not None else "" for h in next(rows)]
        except StopIteration:
            return _empty
        try:
            sheet_col = header.index("sheet_name")
            enabled_col = header.index("enabled")
        except ValueError:
            logger.warning("prompt.catalog_columns_missing", extra={"header": header})
            return _empty
        model_col = header.index("model") if "model" in header else -1
        enabled_out: Dict[str, bool] = {}
        models_out: Dict[str, str] = {}
        for row in rows:
            if row is None or sheet_col >= len(row):
                continue
            sheet = row[sheet_col]
            if sheet is None or not str(sheet).strip():
                continue
            key = str(sheet).strip()
            flag = row[enabled_col] if enabled_col < len(row) else None
            enabled_out[key] = _parse_enabled(flag)
            if model_col >= 0 and model_col < len(row):
                raw = row[model_col]
                if raw is not None and str(raw).strip():
                    models_out[key] = str(raw).strip()
        return enabled_out, models_out

    def get(self, stage: str) -> str:
        return self._prompts.get(stage, "")

    def is_enabled(self, stage: str) -> bool:
        """Whether a stage's agent is switched ON in the Catalog tab. The single
        source of truth for agent activation (checkin / simulation / learning-aid):
        prompt engineers toggle it in the workbook, no code/env change. A stage
        absent from the catalog is treated as DISABLED. The guardrail layer and the
        closing layer are ALWAYS_ENABLED — the closer can't be turned off, else a
        session could never reach its terminal close."""
        if stage in ALWAYS_ENABLED:
            return True
        sheet = STAGE_SHEET.get(stage, stage)
        return self._enabled.get(sheet.strip(), False)

    def set_enabled(self, stage: str, value: bool) -> None:
        """Test/dev override: force a stage's enablement for this process only (does
        NOT touch the workbook, and a reload() will reset it to the Catalog value).
        Production enablement always comes from the Catalog tab. ALWAYS_ENABLED
        stages (the guardrail + closing layer) ignore this — they can't be turned
        off, by design."""
        sheet = STAGE_SHEET.get(stage, stage)
        self._enabled[sheet.strip()] = bool(value)

    def model_for(self, stage: str) -> Optional[str]:
        """Model id for this stage from the Catalog `model` column.

        Returns None and logs an error when the agent is absent from the Catalog or
        its model cell is blank. Callers must NOT silently fall back to a default —
        the Catalog is the single source of truth for agent model selection."""
        sheet = STAGE_SHEET.get(stage, stage)
        model = self._models.get(sheet.strip())
        if not model:
            logger.error(
                "prompt.model_missing_in_catalog",
                extra={"stage": stage, "sheet": sheet},
            )
        return model or None

    def sizes(self) -> Dict[str, int]:
        """Per-stage prompt lengths — used by the /v1/prompts admin endpoints."""
        return {k: len(v) for k, v in self._prompts.items()}

    def info(self) -> Dict[str, Any]:
        """Provenance + health snapshot for the admin API and /health."""
        return {
            "version": self.version,
            "loaded_at": self.loaded_at,
            "source": self.source,
            "source_path": self.source_path,
            "degraded": self.degraded,
            "degraded_reason": self.degraded_reason,
            "validation": self.validation,
        }

    @property
    def environment(self) -> str:
        return self._prompts.get("environment", "")


# Module-level view of the loaded registry version, so the LLM client can stamp
# the prompt-cache header without importing the runtime singleton (no cycle):
# reload() updates it, responses_client reads it.
_CURRENT_VERSION = ""


def _set_current_version(version: str) -> None:
    global _CURRENT_VERSION
    _CURRENT_VERSION = version


def current_workbook_version() -> str:
    """Short content hash of the prompts currently being served ("" before boot)."""
    return _CURRENT_VERSION


def validate_prompt_text(stage: str, text: str, *, enabled: bool = True) -> Dict[str, Any]:
    """Validate-on-save: check one prompt body BEFORE it is written to the workbook.

    Returns {"errors": [...], "warnings": [...]}. Errors block the save (an enabled
    agent with an empty prompt cannot run a turn); warnings are advisory and are
    returned to the caller so the author sees them at edit time rather than
    discovering them as a live incident."""
    errors: List[str] = []
    warnings: List[str] = []
    body = (text or "").strip()
    if enabled and not body:
        errors.append("prompt is empty but the agent is enabled — it cannot run a turn")
    if len(body) > PROMPT_SIZE_WARN_CHARS:
        warnings.append(
            f"prompt is {len(body)} chars (> {PROMPT_SIZE_WARN_CHARS} budget) — "
            "consider trimming or moving shared blocks into the environment layer"
        )
    known = PromptRegistry._known_tokens_from_registries()
    unknown = sorted(
        tok.strip("{}") for tok in set(_PLACEHOLDER_RE.findall(body))
        if not PromptRegistry._token_resolvable(tok.strip("{}"), known)
    )
    if unknown:
        warnings.append(
            "placeholders no data source can resolve (they will be blanked at "
            f"runtime): {', '.join('{' + t + '}' for t in unknown)}"
        )
    return {"errors": errors, "warnings": warnings, "size": len(body)}
