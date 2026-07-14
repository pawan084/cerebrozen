"""SSKB ingestion — the CereBroZen-global knowledge base.

Real S3 layout is directory-first, one typed subfolder per source, under
s3://<bucket>/sskb/:
  - sskb_concept/        → evidence-based concepts   (Extract1)
  - sskb_microlearning/  → micro-learning bites       (Extract6)
  - sskb_competency/     → master competency framework(Extract8)
  - sskb_curated/        → curated content            (Extract7)

The directory is the authority for what an object is (not filename sniffing):
an unrecognized subfolder is skipped and logged rather than guessed at. Content
under sskb_concept/sskb_microlearning/sskb_competency is real documents
(PDF/DOCX/PPTX/TXT/MD) — chunked like CSKB client docs via `chunk_doc()`, with
the extraction-time LLM pulling structured fields out of the chunk text (see
extractors.py). sskb_curated/ keeps the structured xlsx content-library parser
as a special case (falls back to chunking for any other file type there).
"""

from __future__ import annotations

import logging
from typing import List

from app import config
from app.rag import store
from app.rag.ingest import (
    chunk_doc,
    detect_content_format,
    download_to_temp,
    embed_and_upsert,
    iter_s3_objects,
    make_id,
)

logger = logging.getLogger("cerebrozen.rag")

CURATED_SHEET = "content library"

# subdir name -> (source tag, item_type tag). Both are registry-filterable
# scalar columns; `source` is what Extract1/6/7/8 filter on (source=_const:...).
_SSKB_DIR = {
    "sskb_concept": ("concept", "concept"),
    "sskb_microlearning": ("micro_learning", "micro_learning"),
    "sskb_competency": ("competency", "competency"),
    "sskb_curated": ("curated", "curated_content"),
}


# --- curated content (xlsx) --------------------------------------------------


def parse_curated(path: str) -> List[dict]:
    """Parse the 'content library' tab into curated-content records.

    Columns: A Name | B Author Name | C Author Title | D Logo | E Synopsis |
             F Link (cell hyperlink = URL) | G Main Skill | H Level 1 |
             I Level 2 - Final | J Status. Other tabs are ignored.
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)  # not read_only: need cell hyperlinks
    lookup = {s.strip().lower(): s for s in wb.sheetnames}
    actual = lookup.get(CURATED_SHEET)
    if actual is None:
        logger.warning("rag.curated_tab_missing", extra={"sheets": wb.sheetnames})
        wb.close()
        return []
    ws = wb[actual]

    header = {}
    for col, cell in enumerate(ws[1], start=1):
        if cell.value:
            header[str(cell.value).strip().lower()] = col

    def col(name: str) -> int:
        return header.get(name.lower(), 0)

    c_name, c_author, c_syn = col("name"), col("author name"), col("synopsis")
    c_link, c_skill = col("link"), col("main skill")
    c_l1, c_l2 = col("level 1"), col("level 2 - final")

    records: List[dict] = []
    for row in ws.iter_rows(min_row=2):
        def val(c: int) -> str:
            return str(row[c - 1].value).strip() if c and row[c - 1].value is not None else ""

        name = val(c_name)
        if not name:
            continue
        # Real URL is the cell hyperlink target; display text is just a label.
        link = ""
        if c_link:
            cell = row[c_link - 1]
            link = (cell.hyperlink.target if cell.hyperlink else None) or val(c_link)
        synopsis = val(c_syn)
        skill, level1, level2 = val(c_skill), val(c_l1), val(c_l2)
        text = f"{name}. {synopsis}".strip()
        records.append(
            {
                "id": make_id("sskb", "curated", name, link),
                "text": text,
                "kb": "sskb",
                "source": "curated",
                "item_type": "curated_content",
                "title": name,
                "author": val(c_author),
                "source_link": link,
                "content_format": detect_content_format(link),
                # skill/topic/cluster hold the content taxonomy; `level` stays
                # reserved for user seniority (consistent with CSKB), so it is "".
                "skill": skill,
                "topic": level1,
                "cluster": level2,
                "doc_group": "content_library",
                # extras → meta blob (sub_heading for deterministic Extract7 mapping)
                "synopsis": synopsis,
            }
        )
    wb.close()
    logger.info("rag.curated_parsed", extra={"records": len(records)})
    return records


# --- orchestration -----------------------------------------------------------


def _subdir(key: str) -> str:
    """First path segment under sskb/, e.g. 'sskb/sskb_concept/x.pdf' -> 'sskb_concept'."""
    rel = key[len(config.RAG_SSKB_PREFIX) + 1:] if key.startswith(config.RAG_SSKB_PREFIX) else key
    parts = [p for p in rel.split("/") if p]
    return parts[0] if len(parts) > 1 else ""  # >1: needs a subdir AND a filename


def _records_for_key(key: str) -> List[dict]:
    """Directory-first dispatch: the subfolder decides source/item_type. `.xlsx`
    under sskb_curated/ uses the structured content-library parser; everything
    else (PDF/DOCX/PPTX/TXT/MD, in any recognized subfolder) is chunked."""
    subdir = _subdir(key)
    mapping = _SSKB_DIR.get(subdir)
    if mapping is None:
        logger.info("rag.sskb_skip_unknown_dir", extra={"key": key, "subdir": subdir})
        return []
    source, item_type = mapping

    if subdir == "sskb_curated" and key.lower().endswith(".xlsx"):
        return parse_curated(download_to_temp(key))

    return chunk_doc(
        key,
        id_prefix=f"sskb:{source}",
        extra_fields={"kb": "sskb", "source": source, "item_type": item_type},
    )


def ingest_sskb_local(curated_path: str = "", micro_path: str = "", concepts_path: str = "", competency_path: str = "") -> dict:
    """Dev path: ingest from local files (no S3). `micro_path`/`concepts_path`/
    `competency_path` are chunked generically (same as the real PDF sources)."""
    written = {"curated": 0, "micro_learning": 0, "concepts": 0, "competency": 0}
    if curated_path:
        written["curated"] = embed_and_upsert("sskb", parse_curated(curated_path))
    if micro_path:
        recs = chunk_doc(micro_path, id_prefix="sskb:micro_learning", local=True,
                          extra_fields={"kb": "sskb", "source": "micro_learning", "item_type": "micro_learning"})
        written["micro_learning"] = embed_and_upsert("sskb", recs)
    if concepts_path:
        recs = chunk_doc(concepts_path, id_prefix="sskb:concept", local=True,
                          extra_fields={"kb": "sskb", "source": "concept", "item_type": "concept"})
        written["concepts"] = embed_and_upsert("sskb", recs)
    if competency_path:
        recs = chunk_doc(competency_path, id_prefix="sskb:competency", local=True,
                          extra_fields={"kb": "sskb", "source": "competency", "item_type": "competency"})
        written["competency"] = embed_and_upsert("sskb", recs)
    return written


def ingest_sskb() -> dict:
    """Incrementally ingest sskb/ from S3: embed only docs not already indexed
    (dedup by S3 key+ETag); re-embed a doc whose ETag changed; skip the rest.
    Also prunes: a doc_key that was previously indexed but no longer shows up in
    S3 (deleted upstream) has its chunks removed too. Dispatch is directory-first
    (see `_SSKB_DIR`) — any number of files per subfolder are picked up
    automatically, and a subfolder with no content yet (e.g. sskb_curated/ until
    something is uploaded there) just yields nothing."""
    indexed = store.indexed_docs("sskb")  # {doc_key: etag}
    seen: set = set()
    written = {s: 0 for s in _SSKB_DIR}
    written.update({"skipped_existing": 0, "skipped_type": 0, "pruned": 0})
    for key, etag in iter_s3_objects(config.RAG_SSKB_PREFIX + "/"):
        seen.add(key)
        if key.lower().rsplit(".", 1)[-1] not in ("pdf", "docx", "pptx", "txt", "md", "xlsx"):
            written["skipped_type"] += 1
            logger.info("rag.sskb_skip", extra={"key": key})
            continue
        subdir = _subdir(key)
        if subdir not in _SSKB_DIR:
            written["skipped_type"] += 1
            logger.info("rag.sskb_skip_unknown_dir", extra={"key": key, "subdir": subdir})
            continue
        if etag and indexed.get(key) == etag:  # already indexed, unchanged
            written["skipped_existing"] += 1
            continue
        if key in indexed:  # changed doc → drop stale chunks before re-embed
            store.delete_by_doc_key("sskb", key)
        try:
            recs = _records_for_key(key)
            for r in recs:
                r["doc_key"] = key
                r["s3_etag"] = etag
            written[subdir] += embed_and_upsert("sskb", recs)
        except Exception:  # noqa: BLE001
            logger.exception("rag.sskb_ingest_failed", extra={"key": key})

    # Same safety-guarded prune as CSKB (see ingest/cskb.py:ingest_cskb) — only
    # when RAG_S3_BUCKET is actually configured, so an unconfigured bucket can
    # never be mistaken for "everything was deleted".
    if config.RAG_S3_BUCKET:
        stale = [k for k in indexed if k not in seen]
        for key in stale:
            store.delete_by_doc_key("sskb", key)
            logger.info("rag.sskb_pruned", extra={"key": key})
        written["pruned"] = len(stale)

    logger.info("rag.sskb_ingested", extra=written)
    return written
