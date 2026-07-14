"""Extraction registry — the business-editable definition of the 9 extractions
(Extract1-8 + the LearningAidSelect composite).

This is the "criteria, not hard-coded" layer. Each extraction declares, in one
place: which KB, the trigger condition, how to build the query from turn state,
which metadata filters to apply, top-k, whether it needs the cheap extraction LLM,
whether it carries a source link, its output fields, and the placeholder token it
is bound to in the prompts.

Source of truth, in priority order:
  1. An `extraction` sheet in the SAME workbook as the prompts (S3, versioned,
     reversible) — so non-technical users edit queries/criteria in Excel.
  2. The in-code SEED below — used when the sheet is absent/empty, and as the
     schema the sheet overrides field-by-field.

The framework (RAG_Extraction_Framework) is the spec these rows encode. Source
links are PER-EXTRACTION: present for CSKB (2/3/4/5) + SSKB curated (7), absent
for SSKB concept (1), micro-learning (6), and SSKB competencies (8).
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field, replace
from functools import lru_cache
from typing import Dict, List, Optional

from app import config

logger = logging.getLogger("cerebrozen.rag")

# Trigger conditions (evaluated against turn state by the extractor).
COND_ALWAYS = "always"            # Extract runs unconditionally.
COND_ORG_AVAILABLE = "org_available"  # Only if the user's org has CSKB content.


@dataclass(frozen=True)
class Extraction:
    extract_id: str
    kb: str                       # "sskb" | "cskb"
    placeholder: str              # prompt token bound to this extraction (no braces)
    condition: str = COND_ALWAYS
    # State fields concatenated (in order) to form the semantic query text.
    query_params: List[str] = field(default_factory=list)
    # column -> state-field: equality filters applied to the vector search.
    filters: Dict[str, str] = field(default_factory=dict)
    top_k: int = 5
    needs_llm: bool = False       # cheap extraction LLM in the pre-step?
    source_required: bool = False  # does this extraction carry a source link?
    output_fields: List[str] = field(default_factory=list)
    used_in: List[str] = field(default_factory=list)  # CIM/CH/CBT (informational)
    # Text substituted when the extraction returns NULL (framework's skip line).
    null_text: str = ""
    enabled: bool = True
    note: str = ""


# Canonical placeholder tokens (adjust freely in the workbook). Prompt authors put
# these in the coaching prompts; their presence is what triggers retrieval.
TOK_CONCEPT = "SSKB_Concept"
TOK_FRAMEWORK = "CSKB_Framework"
TOK_VALUES = "CSKB_Values"
TOK_COMPETENCIES = "CSKB_Competencies"
TOK_CLIENT_AID = "CSKB_LearningAid"
TOK_MICRO = "SSKB_MicroLearning"
TOK_CURATED = "SSKB_CuratedContent"
TOK_SSKB_COMPETENCIES = "SSKB_Competencies"
# Composite: pick ONE aid (micro OR curated OR client) for the learning_aid_agent.
TOK_LEARNING_AID = "LearningAid"

_SKIP_PHASE3 = "No relevant knowledge found. Skip Phase 3 – Client Specific Learning Framework."

SEED: List[Extraction] = [
    Extraction(
        extract_id="Extract1",
        kb="sskb",
        placeholder=TOK_CONCEPT,
        condition=COND_ALWAYS,
        query_params=["session_goal", "conversation_history"],
        filters={"source": "_const:concept"},
        top_k=8,
        needs_llm=True,  # source is now a chunked PDF (sskb_concept/), not atomic
        # concept rows, so the LLM must extract concept_name/description from the
        # chunk text itself (like Extract2) — no deterministic verbatim-from-metadata
        # shortcut anymore (see extractors._DETERMINISTIC_MAP).
        source_required=False,
        output_fields=["concept_name", "concept_description"],
        used_in=["CIM"],
        note="Evidence-Based Concept Library (sskb_concept/, PDF). Embeds the user's "
        "goal/challenge + conversation to find the best-fit concept.",
    ),
    Extraction(
        extract_id="Extract2",
        kb="cskb",
        placeholder=TOK_FRAMEWORK,
        condition=COND_ORG_AVAILABLE,
        # orgID scopes via `filters`, NOT the semantic query (embedding a hex id is
        # noise). Query = goal/challenge + conversation history (per framework).
        query_params=["session_goal", "conversation_history"],
        # Org-scoped AND doc_type-scoped: the real bucket has typed subfolders per
        # org (cskb_org_framework/, cskb_values/, cskb_competency/, cskb_learning_aid/
        # — classify_doc_type() tags each chunk from its subfolder), so pin this
        # extraction to its own subfolder rather than letting any org doc compete.
        filters={"org_id": "org_id", "doc_type": "_const:frameworks"},
        top_k=5,
        needs_llm=True,  # verbatim extract of one framework
        source_required=True,
        output_fields=["retrieved_knowledge", "framework_topic", "relevant_skills", "source_link"],
        used_in=["CIM"],
    ),
    Extraction(
        extract_id="Extract3",
        kb="cskb",
        placeholder=TOK_VALUES,
        condition=COND_ORG_AVAILABLE,
        # Deterministic fetch-all-for-org from Mongo `org` (not a vector query), so
        # no semantic query_params — org_id is read directly by the handler.
        query_params=[],
        # doc_type scopes the CSKB search to cskb_values/ specifically; the custom
        # handler (_extract_values) applies this itself rather than via _build_filters.
        filters={"org_id": "org_id", "doc_type": "_const:values"},
        top_k=50,
        needs_llm=False,  # deterministic fetch-all-for-org (from Mongo `org`, not S3)
        source_required=True,
        output_fields=["values", "source_link"],
        used_in=["CIM", "CH", "CBT"],
        note="Values lookup order: CSKB docs (framework: CSKB-RAG, cskb_values/ subfolder) "
        "→ Mongo `org` collection fallback → null. Custom handler (ignores needs_llm).",
    ),
    Extraction(
        extract_id="Extract4",
        kb="cskb",
        placeholder=TOK_COMPETENCIES,
        condition=COND_ORG_AVAILABLE,
        query_params=["user_level", "session_goal"],  # org scopes via filter
        # Org-scoped AND pinned to the org's cskb_competency/ subfolder. `level`
        # filter still dropped until competency docs are level-tagged.
        filters={"org_id": "org_id", "doc_type": "_const:competencies"},
        top_k=5,
        needs_llm=True,  # best-effort: LLM extracts cluster/competency verbatim from
        # the org's competency doc (like Extract2/5). No "why it matters" anymore.
        source_required=True,
        output_fields=["cluster_pillar", "competency", "source_link"],
        used_in=["CH"],
        note="Best-effort CSKB extraction (org- and doc_type-scoped to cskb_competency/). "
        "Sibling to Extract8 (SSKB master competencies) — these are independent "
        "placeholders, not an automatic fallback chain; a CH prompt referencing "
        "both {CSKB_Competencies} and {SSKB_Competencies} gets whichever resolves. "
        "Re-add the level filter once ingestion tags level.",
    ),
    Extraction(
        extract_id="Extract5",
        kb="cskb",
        placeholder=TOK_CLIENT_AID,
        condition=COND_ORG_AVAILABLE,
        # Framework input #6 is "goal/challenge/outcome": the raw goal/challenge
        # (CIM/CBT) AND the CH user-stated outcome, both carried by session_goal.
        # Plus role, level, and session history. orgID scopes via filter (not the
        # query text).
        query_params=["userRoleContext", "user_level", "session_goal", "conversation_history"],
        # Org- and doc_type-scoped to the org's cskb_learning_aid/ subfolder.
        filters={"org_id": "org_id", "doc_type": "_const:learning_aids"},
        top_k=5,
        needs_llm=True,  # select + verbatim
        source_required=True,
        output_fields=["tool_name", "retrieved_knowledge", "learning_aid_type", "source_link"],
        used_in=["CIM", "CH", "CBT"],
    ),
    Extraction(
        extract_id="Extract6",
        kb="sskb",
        placeholder=TOK_MICRO,
        condition=COND_ALWAYS,
        # confirmed_competency/coaching_shift_summary are CH-only (empty in CIM until
        # populated, so they drop from the query — see extractors._PARAM_ALIASES).
        query_params=["userRoleContext", "session_goal", "user_level", "confirmed_competency", "coaching_shift_summary"],
        filters={"source": "_const:micro_learning"},
        top_k=5,
        needs_llm=True,  # source is now a chunked PDF (sskb_microlearning/); LLM
        # extracts/selects the best-fit bite from chunk text (like Extract2).
        source_required=False,  # micro-learning has NO link — content only
        output_fields=["micro_learning_topic", "retrieved_content"],
        used_in=["CIM", "CH"],
        note="sskb_microlearning/ (PDF, chunked). CIM: query = goal+role+level. "
        "CH: adds confirmed competency + shift summary.",
    ),
    Extraction(
        extract_id="Extract7",
        kb="sskb",
        placeholder=TOK_CURATED,
        condition=COND_ALWAYS,
        query_params=["userRoleContext", "user_level", "coaching_shift_summary"],
        filters={"source": "_const:curated"},
        top_k=5,
        needs_llm=False,  # deterministic top-1 (works once sskb_curated/ has
        # structured xlsx content; degrades to sparse fields for any other file type)
        source_required=True,
        output_fields=["heading", "sub_heading", "author", "source_link", "content_format"],
        used_in=["CIM", "CH"],
        note="sskb_curated/ — no content uploaded yet (returns null until it is, same "
        "as Extract4 before cskb_competency/ was populated). CIM: query = role+level. "
        "CH: adds shift summary.",
    ),
    # Composite selection (NOT a framework ExtractN — it's the upstream "pick one"
    # the deck + orchestrator prompt require):
    #   - Architecture deck, Learning Aid Agent: "Retrieve Either 1 micro learning
    #     item or 1 curated content item from SSKB-RAG + CSKB RAG (if available)".
    #   - Learning_aid_agent_orchestrator_prompt.docx: "the upstream selection chose
    #     EITHER a client aid OR one SSKB item … Honor that choice. Show exactly ONE."
    # Retrieve from 5/6/7 and surface exactly ONE aid for the learning_aid_agent
    # (which honours this selection — it never re-queries).
    Extraction(
        extract_id="LearningAidSelect",
        kb="sskb",
        placeholder=TOK_LEARNING_AID,
        condition=COND_ALWAYS,
        query_params=["user_challenge", "skill_to_develop", "user_role", "user_level", "conversation"],
        filters={},
        top_k=5,
        needs_llm=True,  # pick ONE: micro vs curated vs client aid
        source_required=False,  # depends on which item is chosen
        output_fields=["item_type", "item_title", "retrieved_item", "source_link", "content_format"],
        used_in=["CIM", "CH"],
        note="Runs Extract6/7 (+Extract5 if org has CSKB) and selects one item.",
    ),
    Extraction(
        extract_id="Extract8",
        kb="sskb",
        placeholder=TOK_SSKB_COMPETENCIES,
        condition=COND_ALWAYS,
        query_params=["user_level", "session_goal"],
        # source filter added (empty in the raw workbook data) — without it this
        # would search all of SSKB (concepts/microlearning too), not just the
        # master competency framework in sskb_competency/.
        filters={"source": "_const:competency"},
        top_k=5,
        needs_llm=True,  # LLM extracts cluster/competency verbatim from the chunked
        # master-competency PDF (sskb_competency/), same pattern as Extract4.
        source_required=False,
        output_fields=["cluster_pillar", "competency"],
        used_in=["CH"],
        note="CereBroZen's global/master competency framework (sskb_competency/) — a "
        "sibling to Extract4 (CSKB, org-specific), not an automatic fallback.",
    ),
]


class ExtractionRegistry:
    """Loads the seed, then overlays the workbook `extraction` sheet when present."""

    def __init__(self) -> None:
        self._by_id: Dict[str, Extraction] = {}
        self._by_token: Dict[str, str] = {}
        self.reload()

    def reload(self) -> None:
        rows = {e.extract_id: e for e in SEED}
        try:
            overrides = _load_sheet()
            for ex_id, patch in overrides.items():
                base = rows.get(ex_id)
                rows[ex_id] = replace(base, **patch) if base else None  # type: ignore[arg-type]
                if rows[ex_id] is None:  # sheet defined a brand-new extraction
                    rows.pop(ex_id)
                    logger.warning("rag.registry_unknown_row", extra={"extract_id": ex_id})
        except Exception as exc:  # noqa: BLE001 — never let a bad sheet break retrieval
            logger.warning("rag.registry_sheet_failed", extra={"error": str(exc)})
        self._by_id = {k: v for k, v in rows.items() if v}
        self._by_token = {
            e.placeholder: e.extract_id for e in self._by_id.values() if e.enabled
        }
        logger.info(
            "rag.registry_loaded",
            extra={"count": len(self._by_id), "tokens": sorted(self._by_token)},
        )

    def by_id(self, extract_id: str) -> Optional[Extraction]:
        return self._by_id.get(extract_id)

    def by_token(self, token: str) -> Optional[Extraction]:
        ex_id = self._by_token.get(token)
        return self._by_id.get(ex_id) if ex_id else None

    def binding_tokens(self) -> List[str]:
        """All placeholder tokens bound to an enabled extraction."""
        return list(self._by_token.keys())

    def all(self) -> List[Extraction]:
        return list(self._by_id.values())


# The NULL skip text is exposed so prompt.py / extractors reuse one canonical line.
SKIP_PHASE3 = _SKIP_PHASE3


def _load_sheet() -> Dict[str, dict]:
    """Read the `extraction` sheet from the prompt workbook, if it exists.

    Tolerant: a missing sheet/columns yields {} (seed wins). Returns a per-id dict
    of field overrides. Cells are simple Excel-friendly types (TRUE/FALSE, ints,
    comma-separated lists, `col=field` filter pairs)."""
    from openpyxl import load_workbook

    from app.llm.prompt_store import WORKBOOK_CACHE_PATH, resolve_workbook_path

    # Prefer the already-downloaded cache (written during prompt registry startup)
    # to avoid a redundant S3 round-trip on every session start.
    path = str(WORKBOOK_CACHE_PATH) if WORKBOOK_CACHE_PATH.is_file() else resolve_workbook_path()
    wb = load_workbook(path, read_only=True, data_only=True)
    lookup = {s.strip(): s for s in wb.sheetnames}
    actual = lookup.get(config.RAG_REGISTRY_SHEET.strip())
    if actual is None:
        wb.close()
        return {}

    ws = wb[actual]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return {}
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    out: Dict[str, dict] = {}
    for raw in rows[1:]:
        record = {header[i]: raw[i] for i in range(min(len(header), len(raw)))}
        ex_id = str(record.get("extract_id", "") or "").strip()
        if not ex_id:
            continue
        out[ex_id] = _coerce_overrides(record)
    return out


def _coerce_overrides(record: dict) -> dict:
    """Turn a raw sheet row into typed Extraction field overrides (only set keys)."""
    patch: dict = {}

    def _str(v):
        return str(v).strip() if v is not None else ""

    def _list(v):
        return [p.strip() for p in str(v).split(",") if p.strip()] if v else []

    def _bool(v):
        return str(v).strip().lower() in ("true", "1", "yes", "y")

    if "placeholder" in record and record["placeholder"]:
        patch["placeholder"] = _str(record["placeholder"])
    if "kb" in record and record["kb"]:
        patch["kb"] = _str(record["kb"]).lower()
    if "condition" in record and record["condition"]:
        patch["condition"] = _str(record["condition"])
    if "top_k" in record and record["top_k"] not in (None, ""):
        try:
            patch["top_k"] = max(1, int(record["top_k"]))
        except (TypeError, ValueError):
            pass
    if "needs_llm" in record and record["needs_llm"] not in (None, ""):
        patch["needs_llm"] = _bool(record["needs_llm"])
    if "source_required" in record and record["source_required"] not in (None, ""):
        patch["source_required"] = _bool(record["source_required"])
    if "enabled" in record and record["enabled"] not in (None, ""):
        patch["enabled"] = _bool(record["enabled"])
    if "query_params" in record and record["query_params"]:
        patch["query_params"] = _list(record["query_params"])
    if "output_fields" in record and record["output_fields"]:
        patch["output_fields"] = _list(record["output_fields"])
    if "used_in" in record and record["used_in"]:
        patch["used_in"] = _list(record["used_in"])
    if "null_text" in record and record["null_text"]:
        patch["null_text"] = _str(record["null_text"])
    if "filters" in record and record["filters"]:
        # "col=field, col2=field2"
        pairs = {}
        for token in str(record["filters"]).split(","):
            if "=" in token:
                col, fld = token.split("=", 1)
                pairs[col.strip()] = fld.strip()
        if pairs:
            patch["filters"] = pairs
    return patch


@lru_cache(maxsize=1)
def get_registry() -> ExtractionRegistry:
    return ExtractionRegistry()


def reload_registry() -> ExtractionRegistry:
    """Hot-reload after a workbook edit (mirrors the prompt registry reload)."""
    reg = get_registry()
    reg.reload()
    return reg
