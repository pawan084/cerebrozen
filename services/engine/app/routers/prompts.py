"""Prompt-admin HTTP surface: view/edit/reload the registry + upload a workbook to S3.

Every endpoint requires the INTERNAL_ADMIN role (`require_internal_admin`; dev bypass
only when ENV=local /
AUTH_DEV_BYPASS). Mutations (PUT edit, POST upload) emit a `prompts.audit`
log with the caller's JWT subject and the registry version before/after, so
every prompt change is attributable. Upload validates the .xlsx, backs up the
current S3 object to a timestamped key, replaces it, then reloads the live
registry so edits take effect without a redeploy."""

from __future__ import annotations

import io
import logging
import threading

from pathlib import Path

from fastapi import APIRouter, Body, Depends, File, HTTPException, UploadFile
from fastapi.concurrency import run_in_threadpool
from fastapi.responses import FileResponse
from openpyxl import load_workbook

from app import config
from app.auth import require_internal_admin
from app.graph.runtime import get_registry, reload_prompts
from app.llm.prompt_store import (
    WORKBOOK_CACHE_PATH,
    resolve_workbook,
    resolve_workbook_path,
    upload_workbook_to_s3,
    workbook_checksum,
)
from app.stores import prompt_versions
from app.llm.prompts import (
    _ORPHAN_SCAN_ROWS,
    ALWAYS_ENABLED,
    CATALOG_SHEET,
    PROMPT_COL,
    PROMPT_START_ROW,
    STAGE_SHEET,
    validate_prompt_text,
)

# Excel hard cap per cell; a longer prompt continues down B8, B9, ... (loader concats).
_EXCEL_CELL_MAX = 32767

_XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

logger = logging.getLogger("cerebrozen.prompts_api")
router = APIRouter()

# Serializes workbook read-modify-write edits (PUT): two concurrent edits would
# otherwise load the same base workbook and the second save would silently drop
# the first one's change. Registry reload has its own internal lock.
_EDIT_LOCK = threading.Lock()


def _audit(action: str, claims: dict, **fields) -> None:
    """One structured log line per prompt mutation: who did what, and the registry
    version before/after — the minimal audit trail every change must leave."""
    logger.info(
        "prompts.audit",
        extra={"action": action, "actor": (claims or {}).get("sub", "unknown"), **fields},
    )


@router.post("/v1/prompts/reload")
async def reload_endpoint(claims: dict = Depends(require_internal_admin)) -> dict:
    """Re-read the workbook (re-downloads from S3 in s3 mode) into the registry."""
    reg = get_registry()
    version_before = reg.version
    await run_in_threadpool(reload_prompts)
    _audit("reload", claims, version_before=version_before, version_after=reg.version)
    return {
        "status": "reloaded",
        "source": config.PROMPT_SOURCE,
        "sizes": reg.sizes(),
        **reg.info(),
    }


# ─────────────────────────── registry interface (view + edit) ───────────────────────────
@router.get("/v1/prompts")
async def list_prompts(_claims: dict = Depends(require_internal_admin)) -> dict:
    """Full registry snapshot for the admin UI: every stage agent with its enabled
    flag, model, prompt size, and prompt text."""
    reg = get_registry()
    agents = []
    for stage, sheet in STAGE_SHEET.items():
        text = reg.get(stage)
        agents.append({
            "stage": stage,
            "sheet": sheet,
            "enabled": reg.is_enabled(stage),
            "always_on": stage in ALWAYS_ENABLED,
            "model": reg.model_for(stage) or "",
            "size": len(text),
            "prompt": text,
        })
    return {"source": config.PROMPT_SOURCE, "editable": config.PROMPT_SOURCE == "codebase",
            "count": len(agents), "agents": agents, **reg.info()}


@router.get("/v1/prompts/validate")
async def validate_endpoint(_claims: dict = Depends(require_internal_admin)) -> dict:
    """The registry's load-time validation report: missing sheets, enabled agents
    with no prompt/model, orphaned continuation rows (silent-truncation hazard),
    oversize prompts, and placeholders no data source can resolve. Advisory —
    the same report is computed on every reload and logged when non-empty."""
    reg = get_registry()
    return {"version": reg.version, "degraded": reg.degraded,
            "degraded_reason": reg.degraded_reason, "validation": reg.validation}


def _find_sheet(wb, target: str) -> str:
    """Resolve a sheet by stripped name (the workbook has trailing-space sheet names)."""
    t = target.strip()
    for name in wb.sheetnames:
        if name.strip() == t:
            return name
    raise HTTPException(404, f"sheet not found in workbook: {target}")


def _write_prompt_edit(stage: str, prompt, enabled, model) -> dict:
    """Write one agent's edit back to the workbook, then hot-reload.

    prompt → the stage sheet's B7 (spilling into B8.. when >32k, clearing leftovers);
    enabled/model → the matching Catalog row.

    In s3 mode the edit is applied to a FRESH download of the canonical object and
    uploaded back (which backs the prior object up to a timestamped key), so a
    single-prompt edit no longer requires round-tripping the whole workbook by hand.
    Serialized by _EDIT_LOCK: two concurrent edits would otherwise load the same
    base workbook and the later save would silently drop the earlier change.
    Runs in a threadpool (blocking IO)."""
    with _EDIT_LOCK:
        if config.PROMPT_SOURCE == "codebase":
            path = config.PROMPT_WORKBOOK
        else:
            src = resolve_workbook()
            if src["fallback"]:
                # S3 is unreachable: the file on disk is the BUNDLED workbook, not the
                # canonical object. Editing + uploading it would clobber S3 with stale
                # content. Refuse rather than corrupt the source of truth.
                raise HTTPException(
                    503,
                    "S3 is unreachable — the registry is serving the bundled fallback "
                    f"workbook. Refusing to edit ({src['error']}).",
                )
            path = src["path"]

        wb = load_workbook(path)  # full load (not read_only) so we can save edits
        sheet_name = STAGE_SHEET[stage]

        if prompt is not None:
            ws = wb[_find_sheet(wb, sheet_name)]
            # Clear the old B7.. continuation. Scan PAST the first blank cell so a
            # stray/orphaned fragment below a blank row can't survive the edit and
            # get silently concatenated into a later read.
            row = PROMPT_START_ROW
            blanks = 0
            while blanks < _ORPHAN_SCAN_ROWS:
                if ws.cell(row=row, column=PROMPT_COL).value in (None, ""):
                    blanks += 1
                else:
                    blanks = 0
                    ws.cell(row=row, column=PROMPT_COL).value = None
                row += 1
            chunks = [prompt[i:i + _EXCEL_CELL_MAX] for i in range(0, len(prompt), _EXCEL_CELL_MAX)] or [""]
            for i, chunk in enumerate(chunks):
                ws.cell(row=PROMPT_START_ROW + i, column=PROMPT_COL).value = chunk

        if enabled is not None or model is not None:
            cat = wb[_find_sheet(wb, CATALOG_SHEET)]
            header = [str(c.value).strip().lower() if c.value else "" for c in cat[1]]
            i_sheet = header.index("sheet_name") if "sheet_name" in header else 4
            i_en = header.index("enabled") if "enabled" in header else 2
            i_model = header.index("model") if "model" in header else 3
            target = sheet_name.strip()
            for r in range(2, cat.max_row + 1):
                cell = cat.cell(row=r, column=i_sheet + 1).value
                if cell is not None and str(cell).strip() == target:
                    if enabled is not None:
                        cat.cell(row=r, column=i_en + 1).value = "TRUE" if enabled else "FALSE"
                    if model is not None:
                        cat.cell(row=r, column=i_model + 1).value = model
                    break

        wb.save(path)
        published: dict = {}
        if config.PROMPT_SOURCE != "codebase":
            with open(path, "rb") as fh:
                published = upload_workbook_to_s3(fh.read())  # backs up the prior object
        reload_prompts()
        return published


@router.put("/v1/prompts/{stage}")
async def edit_prompt(
    stage: str,
    body: dict = Body(...),
    claims: dict = Depends(require_internal_admin),
) -> dict:
    """Edit one agent's prompt text / enabled / model in the workbook, then hot-reload.

    Body: {"prompt"?: str, "enabled"?: bool, "model"?: str}. The prompt body is
    validated BEFORE the write (validate-on-save): hard errors block the save,
    warnings are returned so the author sees them at edit time. In s3 mode the
    edit is published back to the canonical object with an automatic backup."""
    if stage not in STAGE_SHEET:
        raise HTTPException(404, f"unknown stage: {stage}")
    prompt = body.get("prompt")
    enabled = body.get("enabled")
    model = body.get("model")
    if prompt is None and enabled is None and model is None:
        raise HTTPException(400, "Provide at least one of: prompt, enabled, model.")

    reg = get_registry()
    version_before = reg.version
    # Validate-on-save. `enabled` for the check is the value being written, or the
    # current one when the edit doesn't touch it.
    will_be_enabled = reg.is_enabled(stage) if enabled is None else bool(enabled)
    report = validate_prompt_text(
        stage, reg.get(stage) if prompt is None else prompt, enabled=will_be_enabled,
    )
    if report["errors"]:
        raise HTTPException(422, {"message": "prompt failed validation", **report})

    # Snapshot the CURRENT text before overwriting it. The ordering is the whole design:
    # snapshot-then-write means the thing you want back is already saved. Snapshot after,
    # and the one edit you most need to undo — the one that just destroyed the text — is
    # the one that never got recorded. Best-effort: a store that is down must not block a
    # prompt fix, which would invert the point of an undo log.
    version_id = None
    if prompt is not None:
        version_id = await run_in_threadpool(
            prompt_versions.snapshot, stage, reg.get(stage),
            actor=(claims or {}).get("sub", ""), reason="edit",
        )

    try:
        published = await run_in_threadpool(_write_prompt_edit, stage, prompt, enabled, model)
    except HTTPException:
        raise
    except Exception as exc:  # noqa: BLE001
        logger.exception("prompts.edit_error")
        raise HTTPException(500, f"Failed to save edit: {exc}")

    _audit(
        "edit", claims, stage=stage, version_before=version_before,
        version_after=reg.version, fields=[
            f for f, v in (("prompt", prompt), ("enabled", enabled), ("model", model))
            if v is not None
        ],
        warnings=report["warnings"], published=published or None, snapshot=version_id,
    )
    return {"status": "saved", "stage": stage, "enabled": reg.is_enabled(stage),
            "model": reg.model_for(stage) or "", "size": len(reg.get(stage)),
            "version": reg.version, "warnings": report["warnings"],
            # The id of what this edit REPLACED — so a client can offer "undo" immediately
            # rather than making the author go and find it.
            **({"replaced_version": version_id} if version_id else {}),
            **({"published": published} if published else {})}


@router.get("/v1/prompts/checksum")
async def checksum_endpoint(_claims: dict = Depends(require_internal_admin)) -> dict:
    """Confirm the server cache matches the S3 object.

    Compares the MD5 of the file the registry loaded (WORKBOOK_CACHE_PATH) with
    the ETag of the canonical S3 object.  For workbooks uploaded via this API
    (single-part put_object) the ETag is always the plain MD5, so ``match: true``
    means the server is running exactly what is in S3.

    Typical use: call this immediately after POST /v1/prompts/upload to confirm
    the new sheet is live on the server.
    """
    result = await run_in_threadpool(workbook_checksum)
    return result


@router.get("/v1/prompts/download")
async def download_endpoint(_claims: dict = Depends(require_internal_admin)) -> FileResponse:
    """Return the workbook currently loaded in the registry.

    Serves from the server-side cache (the file the registry last loaded) so the
    download is always consistent with what the server is actually running.
    In codebase mode this is the bundled local file.  In s3 mode it is the temp
    cache written during the last reload — call POST /v1/prompts/reload first if
    you want to pull a fresher S3 object without uploading a new file.
    """
    if config.PROMPT_SOURCE == "codebase":
        path = Path(config.PROMPT_WORKBOOK)
    else:
        path = WORKBOOK_CACHE_PATH
        if not path.is_file():
            # Cache not yet populated (cold start before first reload). Pull from S3
            # once so the download is not a 404, then the registry will use the same
            # file on its next reload.
            try:
                await run_in_threadpool(resolve_workbook_path)
            except Exception as exc:
                logger.warning("prompts.download_cache_miss", extra={"error": str(exc)})
            if not path.is_file():
                raise HTTPException(
                    status_code=404,
                    detail="Workbook cache not populated. Call POST /v1/prompts/reload first.",
                )
    if not path.is_file():
        raise HTTPException(status_code=404, detail="Workbook not found.")
    return FileResponse(
        str(path), media_type=_XLSX_MEDIA_TYPE, filename="agent_prompts.xlsx",
    )


@router.post("/v1/prompts/upload")
async def upload_endpoint(
    file: UploadFile = File(...),
    claims: dict = Depends(require_internal_admin),
) -> dict:
    """Upload a new agent_prompts.xlsx: validate -> back up current S3 object to a
    timestamped key -> replace -> reload."""
    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="Empty file.")

    # Validate it's a real .xlsx with the expected agent sheets before S3 writes.
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=False)
        present = {s.strip() for s in wb.sheetnames}
        wb.close()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"Not a valid .xlsx workbook: {exc}")

    # Sanity-check it's the prompts workbook, but mirror the loader's tolerance:
    # some stage sheets (e.g. CH_coaching_agent) are Phase-3 paths that
    # are intentionally "not authored yet" — the loader treats a missing sheet as ""
    # and falls back to CIM, so the upload guard must NOT hard-require them. Reject
    # only a file that carries none of the known stage sheets (i.e. the wrong file);
    # report any absent sheets as informational.
    required = set(STAGE_SHEET.values())
    if not (required & present):
        raise HTTPException(
            status_code=400,
            detail=(
                "This doesn't look like the prompts workbook — it contains none of "
                f"the known stage sheets: {sorted(required)}."
            ),
        )
    missing = sorted(required - present)  # not-yet-authored stages; allowed

    reg = get_registry()
    version_before = reg.version
    try:
        result = await run_in_threadpool(upload_workbook_to_s3, data)
    except Exception as exc:  # noqa: BLE001
        logger.exception("prompts.upload_error")
        raise HTTPException(status_code=500, detail=f"S3 upload failed: {exc}")

    # Re-fetch so the new prompts go live (effective when PROMPT_SOURCE=s3).
    await run_in_threadpool(reload_prompts)
    _audit(
        "upload", claims, version_before=version_before, version_after=reg.version,
        backup_key=result.get("backup_key"), md5=result.get("md5"),
        validation_issues=reg.validation.get("issue_count", 0),
    )
    return {
        "status": "uploaded",
        "source": config.PROMPT_SOURCE,
        **result,
        "missing_sheets": missing,  # not-yet-authored stages, if any (informational)
        "sizes": reg.sizes(),
        "version": reg.version,
        # The freshly-loaded workbook's issue report — an upload that breaks an
        # agent's contract is visible in the upload response, not only in logs.
        "validation": reg.validation,
    }


# ── catch-all LAST ────────────────────────────────────────────────────────────
# Registered after every literal /v1/prompts/* path: FastAPI matches routes in
# registration order, so declaring this earlier would shadow /checksum,
# /download, /validate and /reload (a GET /v1/prompts/checksum would resolve
# here and 404 as "unknown stage: checksum").
@router.get("/v1/prompts/{stage}/versions")
async def list_versions(stage: str, _claims: dict = Depends(require_internal_admin)) -> dict:
    """What this prompt used to say, newest first — metadata only.

    No bodies: 60 versions of a 39,000-character prompt is 2.3MB to render a table of
    dates. Fetch one at a time.
    """
    if stage not in STAGE_SHEET:
        raise HTTPException(404, f"unknown stage: {stage}")
    rows = await run_in_threadpool(prompt_versions.history, stage)
    reg = get_registry()
    return {
        "stage": stage,
        # So the UI can show "current" against the live text without a second call.
        "current_hash": prompt_versions.content_hash(reg.get(stage)),
        "current_size": len(reg.get(stage)),
        "count": len(rows),
        "versions": rows,
    }


@router.get("/v1/prompts/{stage}/versions/{version_id}")
async def get_version(
    stage: str, version_id: str, _claims: dict = Depends(require_internal_admin)
) -> dict:
    """One prior version, body included — for reading before restoring."""
    if stage not in STAGE_SHEET:
        raise HTTPException(404, f"unknown stage: {stage}")
    row = await run_in_threadpool(prompt_versions.get, stage, version_id)
    if not row:
        raise HTTPException(404, "no such version for this stage")
    return row


@router.post("/v1/prompts/{stage}/revert/{version_id}")
async def revert_prompt(
    stage: str, version_id: str, claims: dict = Depends(require_internal_admin)
) -> dict:
    """Put an old version back.

    A restore is an ordinary edit, and is treated as one: it snapshots the text it is
    about to replace (so an accidental revert is itself undoable), it goes through
    validate-on-save (a version that was valid when written may not be now — the validator
    moves), and it hot-reloads. There is no privileged path that skips the checks just
    because the text used to live here.
    """
    if stage not in STAGE_SHEET:
        raise HTTPException(404, f"unknown stage: {stage}")
    row = await run_in_threadpool(prompt_versions.get, stage, version_id)
    if not row:
        raise HTTPException(404, "no such version for this stage")

    reg = get_registry()
    text = row.get("text") or ""
    report = validate_prompt_text(stage, text, enabled=reg.is_enabled(stage))
    if report["errors"]:
        # A version can rot: the validator gains rules, the workbook's contract moves.
        # Better to refuse and say why than to restore something the graph will choke on.
        raise HTTPException(422, {"message": "that version no longer passes validation", **report})

    replaced = await run_in_threadpool(
        prompt_versions.snapshot, stage, reg.get(stage),
        actor=(claims or {}).get("sub", ""), reason="revert",
    )
    try:
        published = await run_in_threadpool(_write_prompt_edit, stage, text, None, None)
    except HTTPException:
        raise
    except Exception as exc:  # noqa: BLE001
        logger.exception("prompts.revert_error")
        raise HTTPException(500, f"Failed to revert: {exc}")

    _audit("revert", claims, stage=stage, restored=version_id, replaced_version=replaced,
           version_after=reg.version)
    return {"status": "reverted", "stage": stage, "restored": version_id,
            "replaced_version": replaced, "size": len(reg.get(stage)),
            "version": reg.version, "warnings": report["warnings"],
            **({"published": published} if published else {})}


@router.get("/v1/prompts/{stage}")
async def get_prompt(stage: str, _claims: dict = Depends(require_internal_admin)) -> dict:
    reg = get_registry()
    if stage not in STAGE_SHEET:
        raise HTTPException(404, f"unknown stage: {stage}")
    text = reg.get(stage)
    return {"stage": stage, "sheet": STAGE_SHEET[stage], "enabled": reg.is_enabled(stage),
            "always_on": stage in ALWAYS_ENABLED, "model": reg.model_for(stage) or "",
            "size": len(text), "prompt": text, "version": reg.version}
