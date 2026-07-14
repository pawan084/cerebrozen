"""Runtime registry for which variables to capture and how to persist them.

The canonical source of truth is:
  business_requirements/variable_capture_registry.xlsx
  Sheet: "Variable Capture Registry"

Columns the loader reads:
  variable_name    - exact key used in variables_set JSON blocks and MongoDB
  update_frequency - once_in_lifetime | every_session | only_on_shift
  capture_enabled  - TRUE / FALSE (business users can toggle without code change)
  source_agent     - documentation only; not used as a filter

The registry is a lazy-loaded singleton. On import it is NOT loaded — it loads
on the first call to get() so unit tests can call reload() with a custom path.

Design decisions:
  - Variables NOT in the registry are still captured (every_session behaviour).
    This means a new variables_set key from any agent is stored immediately
    without requiring a sheet edit — the sheet is used to CONSTRAIN, not to
    gate discovery.
  - capture_enabled=FALSE suppresses a variable even if the agent emits it.
    This lets business users temporarily disable a variable without code deploy.
  - only_on_shift maps to every_session in code because the agent prompt is
    the correct place to control emission frequency for shift-based vars.
"""

from __future__ import annotations

import logging

from app import config
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, Optional

logger = logging.getLogger("cerebrozen.variable_registry")

# The variable capture registry lives as a tab in the same workbook as the
# agent prompts so product/coaching-design teams have one place to edit.
REGISTRY_PATH = (
    Path(__file__).parent.parent.parent / "agent_prompts.xlsx"
)
REGISTRY_SHEET = "dynamic_variables_persistent"

FREQ_ONCE    = "once_in_lifetime"


# The Coachable Index: eight dimensions and a weighted composite. This is the "score the
# worker" surface, enumerated in one place so the regulated-workplace switch has something
# precise to act on rather than a prefix match somebody will get wrong later.
_PERSON_SCORE_VARS = frozenset({
    "coachability_score",
    "ci_openness", "ci_accountability", "ci_growth_mindset", "ci_action_bias",
    "ci_honesty", "ci_consistency", "ci_specificity", "ci_reflectiveness",
})


def _is_person_score(variable_name: str) -> bool:
    return variable_name in _PERSON_SCORE_VARS
FREQ_SESSION = "every_session"
FREQ_SHIFT   = "only_on_shift"

# update_frequency values that behave like "overwrite every time" at the
# persistence layer (the agent controls when to emit for shift-based vars).
_OVERWRITE_FREQS = frozenset({FREQ_SESSION, FREQ_SHIFT})


@dataclass(frozen=True)
class VarConfig:
    variable_name: str
    update_frequency: str       # once_in_lifetime | every_session | only_on_shift
    capture_enabled: bool
    source_agent: str = ""
    prompt_placeholder: str = ""
    description: str = ""
    notes: str = ""

    @property
    def is_once_in_lifetime(self) -> bool:
        return self.update_frequency == FREQ_ONCE


class VariableCaptureRegistry:
    """Singleton that drives save/skip decisions for every variables_set key."""

    _instance: Optional["VariableCaptureRegistry"] = None

    def __init__(self, path: Path = REGISTRY_PATH, sheet: str = REGISTRY_SHEET) -> None:
        self._vars: Dict[str, VarConfig] = {}
        self._once_in_lifetime: frozenset[str] = frozenset()
        self._disabled: frozenset[str] = frozenset()
        self._load(path, sheet)

    # ── singleton access ──────────────────────────────────────────────────────

    @classmethod
    def get(cls) -> "VariableCaptureRegistry":
        if cls._instance is None:
            cls._instance = cls()
        return cls._instance

    @classmethod
    def reload(
        cls,
        path: Optional[Path] = None,
        sheet: str = REGISTRY_SHEET,
    ) -> "VariableCaptureRegistry":
        """Force a fresh load from disk. Useful after the sheet is edited or
        in tests that supply a fixture path / sheet name."""
        cls._instance = cls(path or REGISTRY_PATH, sheet)
        logger.info(
            "variable_registry.reloaded",
            extra={
                "path": str(path or REGISTRY_PATH),
                "sheet": sheet,
                "total": len(cls._instance._vars),
                "once_in_lifetime": len(cls._instance._once_in_lifetime),
                "disabled": len(cls._instance._disabled),
            },
        )
        return cls._instance

    # ── public API ─────────────────────────────────────────────────────────────

    def is_once_in_lifetime(self, variable_name: str) -> bool:
        """True when the variable must never be overwritten once it has a
        non-empty value in MongoDB. False for unknowns (default: overwrite)."""
        return variable_name in self._once_in_lifetime

    def is_capture_enabled(self, variable_name: str) -> bool:
        """False when the variable is explicitly disabled in the sheet.
        Unknown variables return True (captured by default)."""
        return variable_name not in self._disabled

    def config(self, variable_name: str) -> Optional[VarConfig]:
        return self._vars.get(variable_name)

    @property
    def once_in_lifetime_vars(self) -> frozenset[str]:
        return self._once_in_lifetime

    @property
    def all_vars(self) -> Dict[str, VarConfig]:
        return dict(self._vars)

    def top_level_vars_for_agent(self, source_agent: str) -> frozenset[str]:
        """Top-level context_update/variables_set keys registered for this
        agent — each variable_name's segment up to (not including) its first
        '.', deduped. E.g. "coaching_style_context.selected_style" and
        "user_strengths.self_reported" both contribute their parent key
        ("coaching_style_context", "user_strengths").

        Lets a node bridge an agent's captured fields into persistence purely
        from the sheet, without a hand-maintained per-stage whitelist in code —
        see app/graph/nodes.py `_bridge_registry_vars`."""
        return frozenset(
            v.variable_name.split(".", 1)[0]
            for v in self._vars.values()
            if v.source_agent == source_agent
        )

    # ── loader ────────────────────────────────────────────────────────────────

    def _load(self, path: Path, sheet: str = REGISTRY_SHEET) -> None:
        if not path.exists():
            logger.warning(
                "variable_registry.file_not_found",
                extra={"path": str(path)},
            )
            return

        try:
            import openpyxl  # optional dependency — only for loading the sheet

            wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)

            if sheet not in wb.sheetnames:
                logger.error(
                    "variable_registry.sheet_not_found",
                    extra={"sheet": sheet, "available": wb.sheetnames},
                )
                return

            ws = wb[sheet]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return

            # Resolve column positions from header row (defensive — tolerates
            # column reordering in the sheet).
            header = [str(h).strip().lower() if h else "" for h in rows[0]]
            col = {name: header.index(name) for name in header if name}

            required = {"variable_name", "update_frequency", "capture_enabled"}
            missing = required - col.keys()
            if missing:
                logger.error(
                    "variable_registry.missing_columns",
                    extra={"missing": sorted(missing)},
                )
                return

            loaded: Dict[str, VarConfig] = {}
            for raw in rows[1:]:
                var_name = raw[col["variable_name"]] if raw[col["variable_name"]] else None
                if not var_name or not str(var_name).strip():
                    continue
                var_name = str(var_name).strip()
                # Skip section-header rows (they have no update_frequency)
                freq_raw = raw[col["update_frequency"]] if raw[col["update_frequency"]] else ""
                freq = str(freq_raw).strip().lower()
                if freq not in (FREQ_ONCE, FREQ_SESSION, FREQ_SHIFT):
                    continue

                enabled_raw = raw[col["capture_enabled"]] if raw[col["capture_enabled"]] is not None else True
                if isinstance(enabled_raw, bool):
                    enabled = enabled_raw
                else:
                    enabled = str(enabled_raw).strip().upper() == "TRUE"

                # A durable RATING OF A PERSON, captured in an employment context, is what
                # the EU AI Act's high-risk rules for worker management are actually about
                # — and `coachability_score` plus its eight `ci_*` dimensions is precisely
                # that: a once-in-a-lifetime score about an employee, computed from a
                # coaching conversation they were told was confidential.
                #
                # A tenant that must not hold such a score gets the guarantee here, at LOAD:
                # the variable is never registered, so no agent can capture it, no store can
                # persist it, and no later prompt edit can quietly bring it back. Turning it
                # off in the workbook would be a setting; turning it off here is a property.
                if enabled and not config.PERSON_SCORING_ENABLED and _is_person_score(var_name):
                    logger.info(
                        "variable_registry.person_scoring_suppressed",
                        extra={"variable": var_name,
                               "reason": "durable person-scoring is off for this tenant"},
                    )
                    enabled = False

                def _get(c: str) -> str:
                    idx = col.get(c)
                    if idx is None:
                        return ""
                    v = raw[idx]
                    return str(v).strip() if v else ""

                loaded[var_name] = VarConfig(
                    variable_name=var_name,
                    update_frequency=freq,
                    capture_enabled=enabled,
                    source_agent=_get("source_agent"),
                    prompt_placeholder=_get("prompt_placeholder"),
                    description=_get("description"),
                    notes=_get("notes"),
                )

            self._vars = loaded
            self._once_in_lifetime = frozenset(
                k for k, v in loaded.items() if v.is_once_in_lifetime
            )
            self._disabled = frozenset(
                k for k, v in loaded.items() if not v.capture_enabled
            )

            logger.info(
                "variable_registry.loaded",
                extra={
                    "path": str(path),
                    "total": len(loaded),
                    "once_in_lifetime": sorted(self._once_in_lifetime),
                    "disabled": sorted(self._disabled),
                },
            )

        except Exception as exc:  # noqa: BLE001
            logger.error(
                "variable_registry.load_failed",
                extra={"path": str(path), "error": str(exc)},
            )
