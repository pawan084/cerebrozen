"""Seed the editable prompt workbook (``agent_prompts.xlsx``).

Generates one worksheet per agent (system prompt in cell B7), a ``Catalog`` tab
(agent_name | role | enabled | model | sheet_name | description) and an
``extraction`` tab. The layout matches the loader contract in ``app.llm.prompts``
and ``app.rag.registry`` exactly — those modules are imported here so the sheet
list / extraction defaults stay in sync with the code.

The Catalog is keyed by **sheet_name** and MUST carry the `sheet_name` and
`enabled` headers: the loader looks those columns up by name and, when they are
absent, disables every agent that isn't ALWAYS_ENABLED. Seeding a Catalog the
loader can't read is therefore a silent "nothing runs" failure, so the shape is
asserted at the end of main().

Run from the repo root::

    python -m scripts.seed_prompts
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

# Import the contract from the app so this stays in sync (no divergent hardcoding).
from app import config
from app.llm.prompts import (
    ALWAYS_ENABLED,
    CATALOG_SHEET,
    PROMPT_COL,
    PROMPT_START_ROW,
    STAGE_SHEET,
)
from app.rag.registry import SEED as EXTRACTION_SEED

# Cell where the prompt label / body live (A6 label, B7 body).
_COL_LETTER = chr(ord("A") + PROMPT_COL - 1)    # 2 -> "B"
LABEL_ROW = PROMPT_START_ROW - 1                # 6
LABEL_CELL = f"A{LABEL_ROW}"                    # A6
PROMPT_CELL = f"{_COL_LETTER}{PROMPT_START_ROW}"  # B7
LABEL_TEXT = f"System prompt (edit below, cell {PROMPT_CELL}):"

# Default model for any agent not overridden below. The Catalog is the single
# source of truth for model selection and a blank cell raises at turn time, so
# every seeded row carries a real model id.
DEFAULT_MODEL = "gpt-5-mini"

# Per-agent model overrides for the Catalog tab.
MODEL_OVERRIDES: dict[str, str] = {
    # Light / classification agents -> fast, cheap model.
    "repeat_user_checkin_agent": "gpt-5-nano",
    "simulation_decision_agent": "gpt-5-nano",
    # Structured generation agent -> mini RAG-grade model.
    "dynamic_actions_insights_agent": "gpt-4o-mini",
}

# Catalog `role` column — documentation for the prompt team, not read by the loader.
_ROLES: dict[str, str] = {
    "environment": "guardrail layer (always on)",
    "repeat_user_checkin_agent": "pre-session check-in",
    "coaching_intake_agent": "one-time intake",
    "challenge_context_agent": "path decision (CIM/CBT/CH)",
    "core_coaching_agent": "coaching (CIM + CBT)",
    "CH_coaching_agent": "coaching (capability)",
    "simulation_decision_agent": "simulation gate",
    "role_play_agent": "simulation",
    "SJT_simulation_agent": "simulation",
    "learning_aid_agent": "support",
    "feedback_mood_capture_agent": "closing layer (always on)",
    "dynamic_actions_insights_agent": "builder",
    "user_context_builder_agent": "builder",
    "pattern_agent": "builder",
    "action_checkin_agent": "standalone check-in",
}

# The default system prompt text for each agent. Every prompt is a real, method-
# appropriate coaching instruction (2-6 sentences), not a placeholder. Prompt
# placeholders like {currentChallenge} are resolved by the app at runtime.
PROMPTS: dict[str, str] = {
    "environment": (
        "You are CereBroZen, a warm, evidence-based coach. Always coach by asking, not "
        "telling: lead with open questions and reflective listening, and let the user "
        "reach their own insight. Never give medical, legal, or financial advice, and "
        "never diagnose. Protect the user's privacy — do not repeat sensitive details "
        "beyond what is needed to help. If the user expresses crisis, self-harm, or "
        "risk to others, stop coaching immediately and refer them to appropriate "
        "professional or emergency support. Stay strictly within the coaching method "
        "for the current stage and keep a calm, respectful, non-judgemental tone."
    ),
    "repeat_user_checkin_agent": (
        "The user, {name}, is returning to a session. Give a brief, warm welcome-back "
        "and ask one simple question to gauge whether they have the energy and focus to "
        "coach today. Keep it to one or two sentences — do not start coaching yet. Read "
        "their reply to decide if they are ready to proceed or would prefer to pause."
    ),
    "coaching_intake_agent": (
        "You are running a short intake to understand {name} as a coachee. Administer an "
        "8-question 'Coachable Index', asking exactly one question at a time and keeping "
        "each turn brief. Across the conversation, capture the user's role (userRole), "
        "their prior coaching history (coachingHistory), their primary motivation "
        "(primaryMotivation), and their preferred coaching style (stylePreference). "
        "Acknowledge each answer briefly before moving on, and do not offer advice — "
        "this stage is only about getting to know them."
    ),
    "challenge_context_agent": (
        "Over 2-3 turns, help the user articulate the challenge they want to work on "
        "({currentChallenge}). Ask focused, open questions to understand the situation, "
        "what matters to them about it, and whether a specific person is involved. From "
        "their answers, infer the most fitting coaching_path — CIM, CBT, or CH — and "
        "note whether an interpersonal counterpart is central. Stay brief and reflective; "
        "do not begin deep coaching yet."
    ),
    "core_coaching_agent": (
        "You are delivering unified reflective coaching on the {coaching_path} path "
        "(CIM or CBT) for the user's challenge: {currentChallenge}. Guide the user "
        "through 6-7 thoughtful questions, one at a time, that surface their own "
        "thinking, assumptions, and options. Coach by asking, not telling — mirror what "
        "you hear, then deepen it with the next question. Let the retrieved "
        "{SSKB_Concept} shape the ANGLE of your questions, but never teach it: the user "
        "must reach their own insight, and the concept stays invisible in the prose."
    ),
    "CH_coaching_agent": (
        "You are coaching on the Capability/CH path for {currentChallenge}. Move the "
        "user through three phases, asking guided questions within each. At the end of "
        "each phase, confirm completion with the user, help them name a concrete action, "
        "and mark the milestone before advancing. Keep an ask-not-tell stance and pace "
        "the phases to the user's readiness rather than rushing to the next one."
    ),
    "simulation_decision_agent": (
        "Coaching has reached a natural point to practise. Briefly offer the user a "
        "choice: a live role-play rehearsal, a situational judgment test (SJT), or "
        "skipping practice for now. Present the options plainly in one short message and "
        "wait for their pick — do not begin any simulation until they choose."
    ),
    "role_play_agent": (
        "Run a role-play rehearsal for {currentChallenge}. First set up the persona: "
        "confirm who you will play and the scene. Then run several rehearsal rounds in "
        "character, staying realistic and responsive to the user's approach. Finish with "
        "a short wrap-up that reflects back what the user tried and what seemed to land, "
        "without lecturing."
    ),
    "SJT_simulation_agent": (
        "Deliver a situational judgment test grounded in the user's challenge "
        "({currentChallenge}). Present one concise, realistic scenario followed by "
        "several plausible response options. Ask the user to rank or choose among them, "
        "then reflect on their reasoning. Keep options credible and avoid an obvious "
        "single 'right' answer."
    ),
    "pattern_agent": (
        "The user has just completed a simulation. Offer a single, well-observed pattern "
        "mirror: reflect back one recurring theme, tendency, or strength you noticed in "
        "how they engaged. Frame it as an observation to consider, not a verdict, and "
        "keep it to one short reflection."
    ),
    "learning_aid_agent": (
        "Deliver the one retrieved learning aid ({learning_aid}) using the "
        "Grasp -> Practise -> Apply -> Commit structure. First help the user grasp the "
        "core idea, then invite a small practice, then connect it to their real situation "
        "({currentChallenge}), and finally prompt a concrete commitment. Keep each step "
        "brief and interactive."
    ),
    "dynamic_actions_insights_agent": (
        "From everything in this session, generate concrete, personalised action cards "
        "and a few sharp insights for the user. Each action should be specific, doable, "
        "and tied to {currentChallenge}; each insight should name something the user "
        "discovered about themselves. Prefer a small number of high-quality items over a "
        "long list."
    ),
    "final_action_check": (
        "Before the session closes, make sure the user has saved at least one action. "
        "If they have, briefly affirm it. If not, gently nudge them to commit to one "
        "concrete next step now. Keep it short and encouraging — a single prompt, not a "
        "recap of the whole session."
    ),
    "feedback_mood_capture_agent": (
        "You are the closing layer and the sole path to ending the session. Guide the "
        "user through a brief close: capture their mood on a simple scale, help them name "
        "the feeling behind it, confirm their commitment, and verify at least one action "
        "is saved. Be warm and concise, then bring the session to a clean, supportive "
        "close."
    ),
    "action_checkin_agent": (
        "The user tapped a single action card to reflect on it. Run a focused 15-step "
        "reflection on that one action: how it went, what helped or blocked it, what they "
        "learned, and what they will adjust next. Ask one step at a time, stay curious "
        "and non-judgemental, and keep the whole check-in centred on this single action."
    ),
    "user_context_builder_agent": (
        "You are a background data processor, not a coach — ignore all conversational "
        "rules and return strict JSON only. From the session transcript "
        "({session_transcript}) and the prior context model ({previousUserContext}), "
        "produce the updated User Context Model. Do NOT invent facts: if a field is not "
        "explicitly stated or strongly anchored in the transcript, omit it rather than "
        "guessing. Never include sensitive personal details beyond what the coaching "
        "context requires."
    ),
    # NOTE: `greeting` and `final_action_check` are deliberately NOT in STAGE_SHEET —
    # the greeting is generated by app/llm/greeting_generator.py and final_action_check
    # is a code node with a fixed reply. They are seeded here only for reference and
    # are skipped when the workbook is written (the writer iterates STAGE_SHEET).
    "greeting": (
        "Generate a short, personalised home-screen greeting for {name}. Keep it to one "
        "or two warm sentences that feel human and invite them into today's session, "
        "without asking a coaching question. Vary the wording so it does not feel "
        "templated."
    ),
}


def _write_prompt_sheet(ws, agent: str) -> None:
    """Write the A6 label and B7 prompt body for one agent worksheet."""
    ws[LABEL_CELL] = LABEL_TEXT
    ws[LABEL_CELL].font = Font(bold=True)
    # A missing prompt would still produce a valid sheet, but we expect coverage.
    ws[PROMPT_CELL] = PROMPTS.get(agent, "")
    ws[PROMPT_CELL].alignment = Alignment(wrap_text=True, vertical="top")
    # A little breathing room so the prompt is readable when opened.
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 100


def _write_catalog(wb) -> None:
    """Write the Catalog tab in the LOADER's schema.

    Columns: agent_name | role | enabled | model | sheet_name | description.
    `sheet_name` is the key the loader joins on and `enabled`/`model` are read by
    name — an agent missing from this tab is DISABLED (except ALWAYS_ENABLED)."""
    ws = wb.create_sheet(CATALOG_SHEET)
    ws.append(["agent_name", "role", "enabled", "model", "sheet_name", "description"])
    for c in ws[1]:
        c.font = Font(bold=True)
    for agent, sheet in STAGE_SHEET.items():
        # All agents enabled by default. Every agent that makes its own LLM call
        # needs a non-blank model — a blank cell raises at turn time by design
        # (the Catalog is the single source of truth for model selection), so seed
        # the default model rather than leaving it empty.
        model = MODEL_OVERRIDES.get(agent, DEFAULT_MODEL)
        ws.append([agent, _ROLES.get(agent, ""), True, model, sheet, ""])
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 34


def _write_extractions(wb) -> None:
    """Write the `extraction` tab from the RAG registry SEED.

    Column names are read by name in registry._coerce_overrides and overlay the
    in-code seed field-by-field, so they must match those keys exactly."""
    ws = wb.create_sheet(config.RAG_REGISTRY_SHEET)
    headers = [
        "extract_id", "kb", "placeholder", "condition", "query_params",
        "filters", "top_k", "needs_llm", "source_required", "null_text", "enabled",
    ]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    for spec in EXTRACTION_SEED:
        ws.append([
            spec.extract_id,
            spec.kb,
            spec.placeholder,
            spec.condition,
            ", ".join(spec.query_params),
            ", ".join(f"{col}={fld}" for col, fld in spec.filters.items()),
            spec.top_k,
            spec.needs_llm,
            spec.source_required,
            spec.null_text,
            spec.enabled,
        ])
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["E"].width = 46


def build_workbook() -> Workbook:
    """Assemble the full workbook in the loader-expected layout."""
    wb = Workbook()
    # Remove the default sheet openpyxl creates; we add named sheets explicitly.
    wb.remove(wb.active)

    # One prompt worksheet per agent, in STAGE_SHEET order.
    for agent, sheet_name in STAGE_SHEET.items():
        ws = wb.create_sheet(sheet_name)
        _write_prompt_sheet(ws, agent)

    _write_catalog(wb)
    _write_extractions(wb)
    return wb


def main() -> None:
    out_path = Path(config.PROMPT_WORKBOOK)
    # Ensure the output directory (e.g. prompts/) exists.
    out_path.parent.mkdir(parents=True, exist_ok=True)

    assert ALWAYS_ENABLED <= set(STAGE_SHEET), "ALWAYS_ENABLED agents must exist in STAGE_SHEET"

    wb = build_workbook()
    wb.save(out_path)

    # Load the file we just wrote through the REAL registry. A seed script that
    # emits a workbook the loader can't read (e.g. a Catalog missing the
    # `sheet_name` header → every agent silently disabled) is worse than useless,
    # so prove the round-trip here instead of discovering it at boot.
    from app.llm.prompts import PromptRegistry

    reg = PromptRegistry(path=str(out_path))
    missing = [s for s in STAGE_SHEET if not reg.get(s).strip()]
    disabled = [s for s in STAGE_SHEET if not reg.is_enabled(s)]
    assert not missing, f"seeded workbook has empty prompts: {missing}"
    assert not disabled, f"seeded workbook loads these agents as DISABLED: {disabled}"

    print(f"Saved prompt workbook to {out_path.resolve()}")
    print(f"  loaded back: {len(STAGE_SHEET)} agents, all enabled, version {reg.version}")
    report = reg.validation
    if not report["ok"]:
        print(f"  validation issues ({report['issue_count']}):")
        for key in ("missing_sheets", "not_in_catalog", "enabled_no_prompt",
                    "enabled_no_model", "oversize", "unknown_placeholders"):
            if report[key]:
                print(f"    {key}: {report[key]}")


if __name__ == "__main__":
    main()
