"""The API layer: routers, the service that drives a turn, the read models, and the
7-day check-in rule that decides what a returning user gets asked about.

Everything here drives the REAL app through TestClient (or calls the real pure
function). The only things stubbed are the boundaries this service genuinely does
not own: the LLM provider (the offline mock provider the app already ships), S3,
SSM, ElevenLabs' HTTP API, and the vector store. Mongo is a real pymongo API
(mongomock) via the shared `mongo` / `agentic_coll` fixtures, so the store queries
these endpoints depend on are the real ones.

The rule of thumb applied throughout: a test that only proves "a mock was called"
proves nothing. Every assertion below is on a status code and a response body.
"""

from __future__ import annotations

import importlib.util
import io
import json
import sys
from datetime import date, timedelta
from pathlib import Path

import pytest
from fastapi import HTTPException

from app import config
from app.stores import agentic, conversation

# ─────────────────────────────────────────────────────────────────────────────
# Fixtures
# ─────────────────────────────────────────────────────────────────────────────


@pytest.fixture(autouse=True)
def offline_llm(monkeypatch):
    """Pin the LLM to the bundled offline mock provider for EVERY test in this file.

    conftest empties OPENAI_API_KEY so the provider factory falls back to the mock,
    but `app.main` calls `load_local_env()` at import, which re-populates an EMPTY
    OPENAI_API_KEY from a developer's `.env` (that is deliberate for `uvicorn
    app.main:app`). On a machine with a real key in `.env` the "offline" suite would
    quietly start billing OpenAI and become non-deterministic. Pinning the provider
    singleton here makes offline a property of the test, not of the developer's
    filesystem.
    """
    from app.llm import providers
    from app.llm.providers.mock import MockLLMClient

    monkeypatch.setenv("CEREBROZEN_LLM_PROVIDER", "mock")
    monkeypatch.setattr(providers, "_provider", MockLLMClient())


@pytest.fixture
def client():
    """The real app, dev-auth-bypassed (conftest sets AUTH_DEV_BYPASS)."""
    from fastapi.testclient import TestClient

    from app.main import create_app

    return TestClient(create_app(), raise_server_exceptions=False)


@pytest.fixture
def authed_client(monkeypatch):
    """The real app with auth ENFORCED, plus a helper that mints a valid HS512 JWT.

    Several endpoints (DELETE /v1/sessions, /v1/greeting) take the user id ONLY from
    the token — under the dev bypass the claims are empty, so those paths are simply
    not reachable without a real signed token.

    Yields ``(client, token_for)``.
    """
    import jwt
    from fastapi.testclient import TestClient

    from app.main import create_app

    monkeypatch.setenv("AUTH_DEV_BYPASS", "")
    monkeypatch.setattr(config, "ENV", "production")
    monkeypatch.setattr(config, "JWT_SECRET", "s3cret")

    def token_for(user_id: str) -> dict:
        raw = jwt.encode({"user": {"username": user_id}, "org_id": "default"}, "s3cret", algorithm=config.JWT_ALGORITHM)
        return {"Authorization": f"Bearer {raw}"}

    return TestClient(create_app(), raise_server_exceptions=False), token_for


@pytest.fixture
def workbook_sandbox(tmp_path, monkeypatch):
    """A throw-away copy of the prompt workbook, wired in as config.PROMPT_WORKBOOK.

    The PUT endpoint SAVES the workbook. Without this the edit tests would rewrite the
    repository's own agent_prompts.xlsx. Teardown restores the real path and reloads
    the (process-global) registry, so an edited prompt cannot leak into another test.
    """
    import shutil

    from app.graph.runtime import reload_prompts

    original = config.PROMPT_WORKBOOK
    copy = tmp_path / "agent_prompts.xlsx"
    shutil.copy(original, copy)
    monkeypatch.setattr(config, "PROMPT_WORKBOOK", str(copy))
    reload_prompts()
    try:
        yield copy
    finally:
        config.PROMPT_WORKBOOK = original
        reload_prompts()


# ── helpers ──────────────────────────────────────────────────────────────────


def _sse(client, url, **kw) -> list:
    """Collect the parsed SSE events from a streaming POST."""
    events = []
    with client.stream("POST", url, **kw) as response:
        assert response.status_code == 200, response.status_code
        for line in response.iter_lines():
            if line.startswith("data: "):
                events.append(json.loads(line[6:]))
    return events


def _seed_turn(session_id, user_id, *, user_message="I struggle with delegation",
               bot_text="Here is a thought.", request_id="", **kw) -> None:
    """Record one exchange through the REAL transcript writer (mongomock behind it)."""
    from app.request_context import request_id as rid_var

    token = rid_var.set(request_id)
    try:
        conversation.record_turn(
            session_id=session_id, user_id=user_id, user_message=user_message,
            bot_text=bot_text, **kw,
        )
    finally:
        rid_var.reset(token)


def _seed_actions(user_id, session_id, actions, insights=(), request_id="") -> list:
    """Write actions/insights through the REAL agentic writer; returns what was added."""
    from app.request_context import request_id as rid_var

    token = rid_var.set(request_id)
    try:
        added, _ = agentic.append_actions_insights(
            user_id, list(actions), list(insights),
            session_id=session_id, agent_name="dynamic_actions_insights_agent",
        )
        return added
    finally:
        rid_var.reset(token)


def _import_fresh(module_name: str, relpath: str, blocked=(), monkeypatch=None):
    """Execute a source file as a FRESH module object with some imports made to fail.

    Used to exercise the optional-dependency fallbacks (no livekit, no
    prometheus_client) that are decided at import time and can therefore never be
    reached by patching a live module. Coverage attributes the lines to the real file.
    """
    for name in blocked:
        monkeypatch.setitem(sys.modules, name, None)
    path = Path(__file__).resolve().parent.parent / relpath
    spec = importlib.util.spec_from_file_location(module_name, path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


# ═════════════════════════════════════════════════════════════════════════════
# app/checkin_scheduler.py — the 7-day rule (R1–R6)
#
# The most valuable file in this layer: it decides WHICH of a returning user's prior
# commitments get re-opened. Get it wrong in one direction and the coach nags about
# something the user was already asked about (R3); wrong in the other and a whole
# session's actions are silently never followed up on. It is pure, so every rule and
# every boundary is testable exactly — `today` is passed in, and the tests pass the
# frozen clock, never the wall clock (a test that reads the wall clock fails on a
# Tuesday).
# ═════════════════════════════════════════════════════════════════════════════

from app.checkin_scheduler import (  # noqa: E402
    _action_date,
    eligible_checkin_actions,
    eligible_session_ids,
)


def _act(day: date, session_id="s-old", **kw) -> dict:
    """One stored action, stamped with the calendar date its session was held."""
    return {"full_text": f"I will do the thing ({day})", "session_id": session_id,
            "session_date": day.isoformat(), "status": "active", **kw}


@pytest.mark.parametrize(
    "days_ago, due",
    [(0, False), (1, False), (5, False), (6, False), (7, True), (8, True), (30, True)],
)
def test_an_action_is_due_for_checkin_exactly_seven_days_after_its_session(
    frozen_now, days_ago, due
):
    """R1, at the boundary. Day 6 is early and day 7 is due — the day the whole rule
    turns on. An off-by-one here either nags the user a day early forever or delays
    every check-in by a day, and neither is visible in any log."""
    today = frozen_now[0].date()
    actions = [_act(today - timedelta(days=days_ago))]

    eligible = eligible_checkin_actions(actions, today, current_session_id="s-now")

    assert bool(eligible) is due, f"{days_ago} days after the session: expected due={due}"


def test_the_checkin_window_is_configurable(frozen_now):
    """`due_days` is a parameter, not a constant baked into the comparison — a client
    on a 14-day cadence must not silently get the 7-day one."""
    today = frozen_now[0].date()
    week_old = [_act(today - timedelta(days=7))]

    assert eligible_checkin_actions(week_old, today, due_days=14) == []
    assert eligible_checkin_actions(week_old, today, due_days=1) == week_old


def test_every_overdue_session_is_batched_into_one_checkin(frozen_now):
    """R2. Two sessions went un-checked-in; the user gets ONE consolidated check-in
    covering both, not just the oldest. Returning only one batch would strand the
    other session's commitments forever — nothing ever comes back for them."""
    today = frozen_now[0].date()
    old = _act(today - timedelta(days=21), session_id="s-1")
    older = _act(today - timedelta(days=9), session_id="s-2")
    fresh = _act(today - timedelta(days=2), session_id="s-3")

    eligible = eligible_checkin_actions([old, older, fresh], today, current_session_id="s-now")

    assert eligible == [old, older], "input order must be preserved for the prompt"
    assert eligible_session_ids(eligible) == ["s-1", "s-2"]


def test_a_session_that_was_already_checked_in_is_never_offered_again(frozen_now):
    """R3. The check-in closes the loop ONCE. If a completed batch could come back,
    every later session would re-ask about the same commitments — the single most
    obvious way for the coach to look broken to a returning user."""
    today = frozen_now[0].date()
    done = _act(today - timedelta(days=10), session_id="s-done")
    open_ = _act(today - timedelta(days=10), session_id="s-open")

    eligible = eligible_checkin_actions(
        [done, open_], today, checked_in_sessions=["s-done"]
    )

    assert eligible == [open_]


def test_the_current_sessions_own_actions_are_never_due(frozen_now):
    """R4. Actions committed to minutes ago must not be "checked in on" in the same
    breath. (Belt and braces with R1: an action can carry a stale session_date.)"""
    today = frozen_now[0].date()
    mine = _act(today - timedelta(days=30), session_id="s-now")
    theirs = _act(today - timedelta(days=30), session_id="s-prior")

    assert eligible_checkin_actions([mine, theirs], today, current_session_id="s-now") == [theirs]


def test_a_first_time_user_has_nothing_to_check_in_on(frozen_now):
    """R5/R6. No actions at all (and the `None` a store returns for an unknown user)
    yields an empty list — the caller reads that as "no check-in due" and routes
    straight to intake. Anything else would open a session by asking a brand-new user
    how last week's commitments went."""
    today = frozen_now[0].date()

    assert eligible_checkin_actions([], today) == []
    assert eligible_checkin_actions(None, today) == []


@pytest.mark.parametrize("status", ["skipped", "deleted"])
def test_an_action_the_user_dismissed_is_never_due(frozen_now, status):
    """The user said no. A check-in on a commitment they explicitly skipped or deleted
    re-opens something they closed on purpose."""
    today = frozen_now[0].date()
    dropped = _act(today - timedelta(days=10), status=status)

    assert eligible_checkin_actions([dropped], today) == []


@pytest.mark.parametrize("raw", [None, "", "   "])
def test_an_action_with_no_date_at_all_is_never_due(frozen_now, raw):
    """No date → no window → not eligible. The alternative (treating it as "very old")
    would make every undated legacy action due on the user's next session."""
    today = frozen_now[0].date()
    undated = {"full_text": "I will do the thing", "session_id": "s-1"}
    if raw is not None:
        undated["session_date"] = raw

    assert eligible_checkin_actions([undated], today) == []


@pytest.mark.parametrize("bad", ["not-a-date", "2026-13-45", "yesterday", "2026/07/07"])
def test_a_malformed_date_is_skipped_rather_than_raising(frozen_now, bad):
    """A single corrupt stamp in a user's document must not take the turn down. The
    scheduler runs on the request path — an unparsed date is dropped, quietly."""
    today = frozen_now[0].date()
    good = _act(today - timedelta(days=10), session_id="s-good")

    eligible = eligible_checkin_actions(
        [{"full_text": "x", "session_id": "s-bad", "session_date": bad}, good], today
    )

    assert eligible == [good]


@pytest.mark.parametrize(
    "ts", ["2026-07-07T10:00:00Z", "2026-07-07T10:00:00+00:00", "2026-07-07 10:00:00+00:00",
           "20260707T101500Z"]
)
def test_the_write_timestamp_is_used_when_the_session_date_is_missing(frozen_now, ts):
    """Actions stored before `session_date` existed only carry an ISO-8601 `ts`. They
    are still real commitments; falling back to `ts` is what keeps a long-standing
    user's older actions from being permanently invisible to the check-in."""
    today = frozen_now[0].date()  # 2026-07-14 — exactly 7 days after 2026-07-07
    legacy = {"full_text": "I will do the thing", "session_id": "s-legacy", "ts": ts}

    assert eligible_checkin_actions([legacy], today) == [legacy]
    assert _action_date(legacy) == date(2026, 7, 7)


def test_the_session_date_wins_over_the_write_timestamp(frozen_now):
    """`session_date` is the authoritative stamp: an action re-written today (fresh
    `ts`) still belongs to the session it was committed in, and stays due."""
    today = frozen_now[0].date()
    action = {"full_text": "I will do the thing", "session_id": "s-1",
              "session_date": (today - timedelta(days=8)).isoformat(),
              "ts": today.isoformat() + "T09:00:00Z"}

    assert eligible_checkin_actions([action], today) == [action]


def test_junk_entries_in_the_actions_array_are_skipped(frozen_now):
    """Mongo arrays are not schema-checked. A stray string/None among the actions must
    not crash the scheduler (and with it, the turn)."""
    today = frozen_now[0].date()
    good = _act(today - timedelta(days=10))

    assert eligible_checkin_actions([None, "oops", 42, good], today) == [good]


def test_an_action_with_no_session_id_still_gets_checked_in(frozen_now):
    """A blank session_id can't be matched against the current session or the
    already-checked-in list, so it must not be *excluded* by those rules either —
    otherwise the oldest (pre-session_id) actions would never come back."""
    today = frozen_now[0].date()
    orphan = _act(today - timedelta(days=10), session_id="")

    eligible = eligible_checkin_actions(
        [orphan], today, current_session_id="s-now", checked_in_sessions=["s-old"]
    )

    assert eligible == [orphan]
    assert eligible_session_ids(eligible) == [], "a blank session_id is not a batch"


def test_the_batches_to_close_are_distinct_and_stable(frozen_now):
    """`eligible_session_ids` names the batches the check-in will mark complete (R3).
    Duplicates or an unstable order would mark the wrong set / churn the prompt."""
    today = frozen_now[0].date()
    actions = [
        _act(today - timedelta(days=10), session_id="s-b"),
        _act(today - timedelta(days=11), session_id="s-a"),
        _act(today - timedelta(days=12), session_id="s-b"),
    ]

    assert eligible_session_ids(actions) == ["s-a", "s-b"]


def test_the_scheduler_reaches_the_graph_through_the_user_context(
    mongo, agentic_coll, frozen_now, monkeypatch
):
    """End of the wire: an overdue action in the user's document surfaces in the context
    the graph routes on (`checkinDue` / `checkinEligibleActions` / `checkinSessionIds`),
    and an already-checked-in one does not. The rule is worth nothing if it is not
    actually consulted — and `checkinDue` is what decides whether the check-in agent runs
    at all. The clock is frozen (the store reads `datetime.now`), so this is not a test
    that passes on a Monday."""
    from app.stores import mongo as mongo_store

    fixed, frozen_datetime = frozen_now
    monkeypatch.setattr(mongo_store, "datetime", frozen_datetime)
    today = fixed.date()
    agentic_coll.insert_one({
        "user_id": "u-repeat",
        "actions": [
            _act(today - timedelta(days=8), session_id="s-due"),
            _act(today - timedelta(days=8), session_id="s-closed"),
            _act(today - timedelta(days=1), session_id="s-recent"),
        ],
        "checkin_complete_sessions": ["s-closed"],
    })

    ctx = mongo_store.read_user_context("u-repeat", session_id="s-now")

    assert ctx["checkinDue"] is True
    assert ctx["checkinSessionIds"] == ["s-due"]
    assert ctx["checkinEligibleActions"] == [f"I will do the thing ({today - timedelta(days=8)})"]


# ═════════════════════════════════════════════════════════════════════════════
# app/roi_metrics.py — the Development-Area catalogue, parsed from the live prompt
# ═════════════════════════════════════════════════════════════════════════════

from app.roi_metrics import (  # noqa: E402
    _parse_from_prompt,
    canonical_roi_metric,
    canonical_roi_metrics,
    get_roi_metrics,
)


def test_the_catalogue_is_a_non_empty_list_of_names():
    """Whatever the source, the picker must never come back empty — an action card with
    no Development Areas to choose from cannot be saved with one."""
    metrics = get_roi_metrics()

    assert metrics and all(isinstance(m, str) and m.strip() for m in metrics)


def test_a_json_array_under_the_heading_becomes_the_catalogue(monkeypatch):
    """The parse the module is built around: the array under "ROI Metric Mapping" IS the
    catalogue, so a workbook edit reaches the UI picker without a deploy. (The live
    prompt doesn't currently write it in this shape — see the bug pinned below.)"""
    from app.graph import runtime

    prompt = 'ROI Metric Mapping\n["Delegation", "Clarity of purpose"]'
    monkeypatch.setattr(runtime, "get_registry", lambda: type("R", (), {"get": lambda *_: prompt})())

    assert _parse_from_prompt(prompt) == ["Delegation", "Clarity of purpose"]
    assert get_roi_metrics() == ["Delegation", "Clarity of purpose"]
    assert canonical_roi_metric("delegation") == "Delegation", "the picker drives canonicalisation"


def test_the_workbook_really_is_the_source_of_the_roi_catalogue(monkeypatch):
    """The whole point of this module: ONE source of truth for the Development-Area list —
    the `dynamic_actions_insights_agent` prompt, editable in the workbook by a
    non-technical author.

    It parsed only a JSON array. The live prompt writes a PIPE-SEPARATED run of quoted
    strings (`"Mental & emotional state" | "Inspiration" | …`), so the regex never matched
    and `get_roi_metrics()` had ALWAYS fallen through to the hardcoded
    `config.ROI_METRICS`. The feature had never once worked.

    Nothing looked broken, because the two lists happen to be identical today — which is
    precisely the danger. The day an author edited the list in the workbook (the entire
    reason the feature exists), the agent would have started assigning metrics from the new
    list while the UI picker kept offering the old one, with no error and no obvious cause.
    The parser now accepts both shapes.
    """
    from app.graph.runtime import get_registry
    from app.roi_metrics import ACTIONS_INSIGHTS_AGENT

    prompt = get_registry().get(ACTIONS_INSIGHTS_AGENT)

    assert "ROI Metric Mapping" in prompt, "the heading moved — re-check the parser"
    parsed = _parse_from_prompt(prompt)
    assert parsed, "the workbook's ROI list does not parse — the feature is dead again"
    assert "Mental & emotional state" in parsed
    assert get_roi_metrics() == parsed, "the catalogue must come from the WORKBOOK, not the default"


def test_the_roi_block_is_read_from_after_the_heading():
    """The prompt contains other JSON arrays; the catalogue is the one under the
    "ROI Metric Mapping" heading. Taking the first array in the file would ship
    whatever unrelated list happened to be written above it."""
    text = '["not", "the", "list"]\n ROI Metric Mapping\n ["Influence", "Clarity"]'

    assert _parse_from_prompt(text) == ["Influence", "Clarity"]


@pytest.mark.parametrize(
    "text",
    ["", "ROI Metric Mapping — see the deck", 'ROI Metric Mapping ["broken" "json"]'],
)
def test_an_unparseable_prompt_falls_back_to_the_bundled_catalogue(text, monkeypatch):
    """A prompt edit is made by non-technical authors in Excel. A typo in the ROI block
    must not take the Development-Area picker down — it falls back to the bundled list."""
    from app.graph import runtime

    assert _parse_from_prompt(text) == []
    monkeypatch.setattr(runtime, "get_registry", lambda: type("R", (), {"get": lambda *_: text})())
    assert get_roi_metrics() == list(config.ROI_METRICS)


def test_a_registry_outage_falls_back_to_the_bundled_catalogue(monkeypatch, caplog):
    """Same contract when the registry itself is unavailable (a failed S3 reload):
    degrade to the bundled list, don't raise into the endpoint."""
    from app.graph import runtime

    def _boom():
        raise RuntimeError("registry down")

    monkeypatch.setattr(runtime, "get_registry", _boom)

    assert get_roi_metrics() == list(config.ROI_METRICS)
    assert "roi.parse_failed" in caplog.text


def test_a_metric_is_canonicalised_to_the_catalogues_casing():
    """The UI posts back whatever it rendered; storage must hold one canonical spelling
    or the same Development Area shows up twice in reporting."""
    known = get_roi_metrics()[0]

    assert canonical_roi_metric(known.upper()) == known
    assert canonical_roi_metric(known.lower()) == known


@pytest.mark.parametrize("empty", [None, "", "   "])
def test_an_empty_metric_resolves_to_nothing(empty):
    assert canonical_roi_metric(empty) is None


def test_an_unknown_metric_passes_through_untouched():
    """Lenient by design: an unknown value is stored as-is rather than failing the
    user's save. A prompt edit that adds a metric must not 500 the save endpoint
    for anyone whose UI hasn't reloaded the catalogue yet."""
    assert canonical_roi_metric("Something Nobody Configured") == "Something Nobody Configured"


def test_the_list_form_accepts_a_bare_string_and_drops_the_empties():
    known = get_roi_metrics()[0]

    assert canonical_roi_metrics(known.lower()) == [known]
    assert canonical_roi_metrics([known, "", None]) == [known]


@pytest.mark.parametrize("junk", [None, ["", None], {"a": 1}, 7])
def test_the_list_form_returns_none_when_nothing_survives(junk):
    """None means "the caller sent no re-tag" — distinct from an empty list, which the
    store would happily write over the agent's own tagging."""
    assert canonical_roi_metrics(junk) is None


# ═════════════════════════════════════════════════════════════════════════════
# app/routers/prompts.py
# ═════════════════════════════════════════════════════════════════════════════


def test_the_literal_prompt_routes_are_not_shadowed_by_the_stage_catch_all(client):
    """THE regression pin.

    `/v1/prompts/{stage}` is a catch-all. It was once registered BEFORE the literal
    paths, and FastAPI matches in registration order — so GET /v1/prompts/checksum
    resolved to the catch-all and came back `404 unknown stage: checksum`. Every
    literal admin route was dead, and nothing in the app failed loudly: the endpoints
    existed, they just answered 404 forever.

    This asserts BOTH the behaviour (each literal path answers as itself) and the
    property that produces it (the catch-all is registered last), so re-ordering the
    decorators fails here instead of in someone's browser.
    """
    from app.main import create_app

    routes = [r for r in create_app().routes if str(getattr(r, "path", "")).startswith("/v1/prompts")]
    order = [(r.path, sorted(r.methods)) for r in routes]
    catch_all = next(i for i, (p, m) in enumerate(order) if p == "/v1/prompts/{stage}" and "GET" in m)
    for literal in ("/v1/prompts/validate", "/v1/prompts/checksum", "/v1/prompts/download"):
        idx = next(i for i, (p, _) in enumerate(order) if p == literal)
        assert idx < catch_all, f"{literal} is registered after the catch-all and will 404"

    for path in ("/v1/prompts/checksum", "/v1/prompts/download", "/v1/prompts/validate"):
        body = client.get(path)
        assert body.status_code != 404, f"{path} resolved to the catch-all stage route"
        assert "unknown stage" not in body.text


def test_openapi_schema_builds_and_docs_serve(client):
    """The auto-generated schema and docs must not 500. Regression: the browser-UI
    routes ("/", "/chat", "/prompts", "/flow") return HTMLResponse; when they were
    left in the schema, FastAPI tried to build a response model from that annotation
    and raised PydanticUserError, 500-ing /openapi.json and /docs. They are now
    include_in_schema=False — served, but absent from the API schema."""
    schema = client.get("/openapi.json")
    assert schema.status_code == 200, schema.text
    paths = schema.json()["paths"]
    for ui in ("/", "/chat", "/prompts", "/flow"):
        assert ui not in paths, f"{ui} is a UI page and must not appear in the API schema"
    assert client.get("/docs").status_code == 200
    # The pages themselves still serve HTML (they were excluded from the schema, not removed).
    for ui in ("/", "/chat", "/prompts", "/flow"):
        r = client.get(ui)
        assert r.status_code == 200, ui
        assert "text/html" in r.headers["content-type"]


def test_checksum_says_so_when_there_is_no_s3_to_compare_against(client):
    """In codebase mode there is no canonical S3 object — the endpoint says that
    rather than pretending the cache matches."""
    body = client.get("/v1/prompts/checksum").json()

    assert "codebase mode" in body["error"]
    assert "match" not in body


def test_download_serves_the_workbook_the_registry_is_running(client):
    """The admin downloads, edits and re-uploads this file. Serving anything other
    than what the server actually loaded means the next upload silently reverts
    whatever the last one changed."""
    response = client.get("/v1/prompts/download")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert response.content[:2] == b"PK" and len(response.content) > 10_000


def test_download_404s_with_a_recovery_hint_when_the_s3_cache_is_cold(client, tmp_path, monkeypatch):
    """s3 mode, first request after a boot: there is no cached workbook to serve. The
    caller gets a 404 that tells them how to fix it, not an empty file."""
    from app.routers import prompts as prompts_router

    monkeypatch.setattr(config, "PROMPT_SOURCE", "s3")
    monkeypatch.setattr(config, "PROMPT_S3_BUCKET", "")
    monkeypatch.setattr(prompts_router, "WORKBOOK_CACHE_PATH", tmp_path / "cold.xlsx")

    response = client.get("/v1/prompts/download")

    assert response.status_code == 404
    assert "reload" in response.json()["detail"]


def test_download_serves_the_s3_cache_once_it_is_warm(client, tmp_path, monkeypatch):
    """s3 mode with a populated cache: the download comes from the file the registry
    loaded, not a fresh S3 GET, so it is consistent with what the server is running."""
    from app.routers import prompts as prompts_router

    cache = tmp_path / "agent_prompts.xlsx"
    cache.write_bytes(Path(config.PROMPT_WORKBOOK).read_bytes())
    monkeypatch.setattr(config, "PROMPT_SOURCE", "s3")
    monkeypatch.setattr(prompts_router, "WORKBOOK_CACHE_PATH", cache)

    response = client.get("/v1/prompts/download")

    assert response.status_code == 200
    assert response.content == cache.read_bytes()


def test_validate_returns_the_registrys_issue_report(client):
    """Advisory, but it is the only place a silently-truncated prompt or an
    unresolvable placeholder is visible before it reaches a user."""
    body = client.get("/v1/prompts/validate").json()

    assert body["degraded"] is False
    assert isinstance(body["validation"]["issue_count"], int)
    assert body["version"]


def test_the_registry_snapshot_lists_every_stage_with_its_prompt(client):
    """The admin UI renders this. A stage missing from it can't be edited at all."""
    from app.llm.prompts import STAGE_SHEET

    body = client.get("/v1/prompts").json()

    assert body["count"] == len(STAGE_SHEET)
    assert {a["stage"] for a in body["agents"]} == set(STAGE_SHEET)
    environment = next(a for a in body["agents"] if a["stage"] == "environment")
    assert environment["always_on"] is True and environment["size"] == len(environment["prompt"])


def test_one_stage_can_be_fetched_by_name(client):
    body = client.get("/v1/prompts/core_coaching_agent").json()

    assert body["stage"] == "core_coaching_agent"
    assert body["sheet"] == "core_coaching_agent"
    assert body["size"] == len(body["prompt"]) > 0


def test_an_unknown_stage_is_a_404(client):
    assert client.get("/v1/prompts/no_such_agent").status_code == 404


def test_reload_re_reads_the_workbook(client):
    body = client.post("/v1/prompts/reload").json()

    assert body["status"] == "reloaded"
    assert body["source"] == "codebase"
    assert body["sizes"]["core_coaching_agent"] > 0


def test_editing_an_unknown_stage_is_a_404(client):
    assert client.put("/v1/prompts/nope", json={"prompt": "x"}).status_code == 404


def test_an_edit_must_actually_change_something(client):
    response = client.put("/v1/prompts/pattern_agent", json={})

    assert response.status_code == 400
    assert "at least one" in response.json()["detail"]


def test_an_empty_prompt_for_an_enabled_agent_is_refused(client, workbook_sandbox):
    """Validate-on-save. An enabled agent with an empty prompt cannot run a turn — the
    save is BLOCKED rather than accepted and discovered as a dead stage in production."""
    response = client.put("/v1/prompts/core_coaching_agent", json={"prompt": "   "})

    assert response.status_code == 422
    assert response.json()["detail"]["errors"]
    assert client.get("/v1/prompts/core_coaching_agent").json()["size"] > 0, "the old prompt survived"


def test_an_edit_is_written_to_the_workbook_and_goes_live_without_a_redeploy(
    client, workbook_sandbox
):
    """The whole point of the prompt admin surface: a prompt author edits a prompt and
    the NEXT turn uses it. If the save didn't hot-reload the registry, the change would
    only appear after a deploy — which is exactly the workflow this replaces."""
    before = client.get("/v1/prompts/pattern_agent").json()["version"]

    saved = client.put("/v1/prompts/pattern_agent", json={"prompt": "You are a pattern agent."})

    assert saved.status_code == 200
    assert saved.json()["status"] == "saved"
    assert saved.json()["size"] == len("You are a pattern agent.")
    assert saved.json()["version"] != before, "the registry did not reload the edit"

    live = client.get("/v1/prompts/pattern_agent").json()
    assert live["prompt"] == "You are a pattern agent."

    from openpyxl import load_workbook
    sheet = load_workbook(workbook_sandbox)["pattern_agent"]
    assert sheet.cell(row=7, column=2).value == "You are a pattern agent."


def test_an_edit_clears_the_stray_rows_below_the_prompt_cell(client, workbook_sandbox):
    """The loader concatenates B7, B8, … and STOPS at the first blank cell. A fragment
    left below a blank row is invisible on load — until a later edit spills far enough
    to reach it and silently concatenates it into the live prompt. The edit path scans
    PAST the first blank and clears the strays; this pins that."""
    from openpyxl import load_workbook

    workbook = load_workbook(workbook_sandbox)
    workbook["pattern_agent"].cell(row=9, column=2).value = "STRAY-FRAGMENT"
    workbook.save(workbook_sandbox)

    client.put("/v1/prompts/pattern_agent", json={"prompt": "Fresh prompt."})

    assert client.get("/v1/prompts/pattern_agent").json()["prompt"] == "Fresh prompt."
    sheet = load_workbook(workbook_sandbox)["pattern_agent"]
    assert sheet.cell(row=9, column=2).value is None, "the stray fragment survived the edit"


def test_an_edit_can_disable_an_agent_and_change_its_model(client, workbook_sandbox):
    """enabled/model live in the Catalog tab, not the stage sheet — a different write
    path from the prompt body, and the one an operator reaches for to turn an agent off."""
    response = client.put(
        "/v1/prompts/role_play_agent", json={"enabled": False, "model": "gpt-4.1-mini"}
    )

    assert response.status_code == 200
    live = client.get("/v1/prompts/role_play_agent").json()
    assert live["enabled"] is False
    assert live["model"] == "gpt-4.1-mini"


def test_an_edit_returns_the_warnings_the_author_needs_to_see(client, workbook_sandbox):
    """Warnings do not block the save, but they must come back in the response: a
    placeholder no data source can resolve is blanked at runtime, and the author
    would otherwise only find out from a user."""
    response = client.put(
        "/v1/prompts/pattern_agent", json={"prompt": "Use {no_such_placeholder} here."}
    )

    assert response.status_code == 200
    assert any("no_such_placeholder" in w for w in response.json()["warnings"])


def test_an_edit_is_refused_when_s3_is_unreachable(client, monkeypatch):
    """s3 mode + an S3 that can't be read: the file on disk is the BUNDLED fallback, not
    the canonical object. Editing it and uploading would overwrite everyone's prompts
    with stale content. Refuse (503) rather than corrupt the source of truth."""
    monkeypatch.setattr(config, "PROMPT_SOURCE", "s3")
    monkeypatch.setattr(config, "PROMPT_S3_BUCKET", "")  # → download raises → fallback

    response = client.put("/v1/prompts/pattern_agent", json={"prompt": "hello"})

    assert response.status_code == 503
    assert "Refusing to edit" in response.json()["detail"]


def test_an_edit_that_fails_to_save_is_a_500_not_a_silent_success(client, monkeypatch):
    """If the workbook write blows up, the caller must learn the prompt was NOT saved."""
    from app.routers import prompts as prompts_router

    def _boom(*_a, **_k):
        raise OSError("disk full")

    monkeypatch.setattr(prompts_router, "_write_prompt_edit", _boom)

    response = client.put("/v1/prompts/pattern_agent", json={"prompt": "hello"})

    assert response.status_code == 500
    assert "disk full" in response.json()["detail"]


def test_uploading_an_empty_file_is_refused(client):
    response = client.post("/v1/prompts/upload", files={"file": ("x.xlsx", b"", _XLSX)})

    assert response.status_code == 400
    assert response.json()["detail"] == "Empty file."


def test_uploading_something_that_is_not_a_workbook_is_refused(client):
    """Guard BEFORE any S3 write: a corrupt file must never become the canonical object."""
    response = client.post("/v1/prompts/upload", files={"file": ("x.xlsx", b"not-a-zip", _XLSX)})

    assert response.status_code == 400
    assert "Not a valid .xlsx" in response.json()["detail"]


def test_uploading_the_wrong_workbook_is_refused(client):
    """A real .xlsx with none of the stage sheets is somebody's spreadsheet, not the
    prompt registry. Replacing the canonical object with it would take every agent down."""
    from openpyxl import Workbook

    buffer = io.BytesIO()
    workbook = Workbook()
    workbook.active.title = "Budget"
    workbook.save(buffer)

    response = client.post(
        "/v1/prompts/upload", files={"file": ("budget.xlsx", buffer.getvalue(), _XLSX)}
    )

    assert response.status_code == 400
    assert "doesn't look like the prompts workbook" in response.json()["detail"]


def test_an_s3_failure_during_upload_surfaces_as_a_500(client):
    """No bucket configured → the upload genuinely cannot happen. The caller must not
    be told "uploaded"."""
    data = Path(config.PROMPT_WORKBOOK).read_bytes()

    response = client.post("/v1/prompts/upload", files={"file": ("p.xlsx", data, _XLSX)})

    assert response.status_code == 500
    assert "S3 upload failed" in response.json()["detail"]


def test_a_valid_upload_is_backed_up_replaced_and_reloaded(client, monkeypatch, fake_s3):
    """The happy path, end to end against a stubbed S3: the PRIOR object is copied to a
    timestamped backup key before it is replaced (a bad prompt upload must always be
    reversible), and the registry reloads so the new sheet is live immediately."""
    monkeypatch.setattr(config, "PROMPT_S3_BUCKET", "sys-config")
    data = Path(config.PROMPT_WORKBOOK).read_bytes()

    response = client.post("/v1/prompts/upload", files={"file": ("p.xlsx", data, _XLSX)})
    body = response.json()

    assert response.status_code == 200
    assert body["status"] == "uploaded"
    assert body["missing_sheets"] == []
    assert body["sizes"]["core_coaching_agent"] > 0
    assert body["backup_key"] and body["backup_key"] != config.PROMPT_S3_KEY
    assert fake_s3.copied == [(config.PROMPT_S3_KEY, body["backup_key"])], "no backup was taken"
    assert fake_s3.put[-1][0] == config.PROMPT_S3_KEY


def test_an_edit_in_s3_mode_is_applied_to_the_canonical_object_and_published_back(
    client, s3_workbook
):
    """s3 mode, end to end against a stubbed S3. The edit must be applied to a FRESH
    download of the canonical object (not to whatever stale copy is on this box's disk),
    published back WITH a backup, and reloaded — otherwise a single-prompt edit made on
    one instance silently reverts everything another instance uploaded."""
    response = client.put("/v1/prompts/pattern_agent", json={"prompt": "S3-authored prompt."})
    body = response.json()

    assert response.status_code == 200
    assert body["published"]["backup_key"] != config.PROMPT_S3_KEY, "no backup was taken"
    assert s3_workbook.copied == [(config.PROMPT_S3_KEY, body["published"]["backup_key"])]
    assert client.get("/v1/prompts/pattern_agent").json()["prompt"] == "S3-authored prompt."

    republished = s3_workbook.objects[config.PROMPT_S3_KEY]
    from openpyxl import load_workbook
    assert load_workbook(io.BytesIO(republished))["pattern_agent"].cell(row=7, column=2).value == (
        "S3-authored prompt."
    ), "S3 still holds the pre-edit workbook"


def test_an_agent_with_no_catalog_row_can_still_have_its_prompt_edited(client, workbook_sandbox):
    """`environment` is always-on and has no row in the Catalog tab. The catalog scan must
    simply find nothing to update rather than writing the flags onto a random row."""
    response = client.put(
        "/v1/prompts/environment", json={"prompt": "Guardrails.", "enabled": True, "model": "gpt-4.1"}
    )

    assert response.status_code == 200
    assert client.get("/v1/prompts/environment").json()["prompt"] == "Guardrails."


@pytest.mark.parametrize("field, value", [("enabled", False), ("model", "gpt-4.1-nano")])
def test_a_single_catalog_field_can_be_changed_on_its_own(client, workbook_sandbox, field, value):
    """Changing the model must not silently re-enable a disabled agent, and toggling
    enabled must not blank the model — each field is written independently."""
    before = client.get("/v1/prompts/learning_aid_agent").json()

    response = client.put("/v1/prompts/learning_aid_agent", json={field: value})

    assert response.status_code == 200
    after = client.get("/v1/prompts/learning_aid_agent").json()
    assert after[field] == value
    untouched = "model" if field == "enabled" else "enabled"
    assert after[untouched] == before[untouched], f"editing {field} also changed {untouched}"
    assert after["prompt"] == before["prompt"], "a catalog-only edit rewrote the prompt"


def test_download_404s_when_the_bundled_workbook_is_missing(client, monkeypatch, tmp_path):
    """A container built without the workbook: the endpoint must 404, not serve a
    zero-byte file the admin would happily re-upload over the real prompts."""
    monkeypatch.setattr(config, "PROMPT_WORKBOOK", str(tmp_path / "gone.xlsx"))

    response = client.get("/v1/prompts/download")

    assert response.status_code == 404
    assert response.json()["detail"] == "Workbook not found."


def test_download_pulls_the_workbook_once_when_the_cache_is_cold(
    client, s3_workbook, tmp_path, monkeypatch
):
    """Cold start in s3 mode: rather than 404-ing on the first download after a boot, the
    endpoint pulls the canonical object ONCE — and the registry then loads that same
    cached file, so the admin's download and the running prompts cannot diverge."""
    from app.llm import prompt_store
    from app.routers import prompts as prompts_router

    cache = prompt_store.WORKBOOK_CACHE_PATH
    monkeypatch.setattr(prompts_router, "WORKBOOK_CACHE_PATH", cache)
    assert not cache.is_file(), "the cache must start cold for this test to mean anything"

    response = client.get("/v1/prompts/download")

    assert response.status_code == 200
    assert response.content == s3_workbook.objects[config.PROMPT_S3_KEY]
    assert cache.is_file(), "the cold-start download did not populate the cache"


def test_download_reports_a_cold_cache_when_s3_cannot_be_reached(
    client, monkeypatch, tmp_path, caplog
):
    """Cold cache AND S3 down. The endpoint must still answer with the actionable 404
    rather than propagating the S3 exception as a 500."""
    from app.routers import prompts as prompts_router

    def _boom():
        raise RuntimeError("s3 unreachable")

    monkeypatch.setattr(config, "PROMPT_SOURCE", "s3")
    monkeypatch.setattr(prompts_router, "WORKBOOK_CACHE_PATH", tmp_path / "cold.xlsx")
    monkeypatch.setattr(prompts_router, "resolve_workbook_path", _boom)

    response = client.get("/v1/prompts/download")

    assert response.status_code == 404
    assert "reload" in response.json()["detail"]
    assert "prompts.download_cache_miss" in caplog.text


def test_a_sheet_missing_from_the_workbook_is_a_404_not_a_traceback():
    """`_find_sheet` resolves the workbook's trailing-space sheet names. When the sheet
    genuinely isn't there the edit must fail as a 404, not a KeyError 500."""
    from openpyxl import Workbook

    from app.routers.prompts import _find_sheet

    workbook = Workbook()
    workbook.active.title = "core_coaching_agent "  # trailing space, as in the real file

    assert _find_sheet(workbook, "core_coaching_agent") == "core_coaching_agent "
    with pytest.raises(HTTPException) as exc:
        _find_sheet(workbook, "pattern_agent")
    assert exc.value.status_code == 404


_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class _FakeS3:
    """An in-memory S3 — the one true external boundary of the upload/publish path.

    Faithful enough that a download-edit-upload-reload round trip really round-trips:
    put_object stores the bytes, download_file hands the SAME bytes back.
    """

    class exceptions:  # noqa: N801 — mirrors botocore's client.exceptions namespace
        class ClientError(Exception):
            def __init__(self, code="500"):
                self.response = {"Error": {"Code": code}}

    def __init__(self, objects=None):
        self.objects: dict = dict(objects or {})
        self.put: list = []
        self.copied: list = []

    def head_object(self, Bucket, Key, **kw):  # noqa: N803 — boto3's kwarg names
        if Key not in self.objects:
            raise self.exceptions.ClientError("404")
        return {"ETag": '"deadbeef"'}

    def copy_object(self, Bucket, Key, CopySource):  # noqa: N803
        self.objects[Key] = self.objects[CopySource["Key"]]
        self.copied.append((CopySource["Key"], Key))

    def put_object(self, Bucket, Key, Body, **kw):  # noqa: N803
        self.objects[Key] = Body
        self.put.append((Key, len(Body)))

    def download_file(self, Bucket, Key, dest, ExtraArgs=None):  # noqa: N803
        Path(dest).write_bytes(self.objects[Key])


@pytest.fixture
def fake_s3(monkeypatch):
    from app.llm import prompt_store

    s3 = _FakeS3({config.PROMPT_S3_KEY: Path(config.PROMPT_WORKBOOK).read_bytes()})
    monkeypatch.setattr(prompt_store, "_s3_client", lambda: s3)
    return s3


@pytest.fixture
def s3_workbook(fake_s3, tmp_path, monkeypatch):
    """PROMPT_SOURCE=s3 against the in-memory S3, with the workbook cache in tmp.

    Restores the source AND reloads the registry itself (rather than relying on
    monkeypatch's undo ordering) — a registry left pointing at a temp S3 cache would
    poison every test that runs after this one.
    """
    from app.graph.runtime import reload_prompts
    from app.llm import prompt_store

    original_source, original_bucket = config.PROMPT_SOURCE, config.PROMPT_S3_BUCKET
    monkeypatch.setattr(prompt_store, "WORKBOOK_CACHE_PATH", tmp_path / "cache.xlsx")
    config.PROMPT_SOURCE, config.PROMPT_S3_BUCKET = "s3", "sys-config"
    try:
        yield fake_s3
    finally:
        config.PROMPT_SOURCE, config.PROMPT_S3_BUCKET = original_source, original_bucket
        reload_prompts()


# ═════════════════════════════════════════════════════════════════════════════
# app/routers/sessions.py + app/service.py — the coaching turn
# ═════════════════════════════════════════════════════════════════════════════


def test_starting_a_session_mints_an_id_and_answers(client):
    """The first turn is the whole product. The response must carry the session_id the
    client will send on every subsequent turn — losing it strands the conversation."""
    response = client.post("/v1/sessions/start", json={"user_id": "u1", "text": "delegation is hard"})
    body = response.json()

    assert response.status_code == 200
    assert len(body["session_id"]) == 32
    assert body["response_to_user"].strip()
    assert body["served_by"] == "graph"
    assert body["is_first_turn"] is True


def test_a_caller_supplied_session_id_is_adopted(client):
    """Voice reconnects and resumes re-enter start_session with an EXISTING id. Minting
    a new one there would fork the conversation into a second checkpoint thread."""
    response = client.post(
        "/v1/sessions/start", json={"user_id": "u1", "session_id": "sess-adopted", "text": "hi"}
    )

    assert response.json()["session_id"] == "sess-adopted"


def test_the_reply_streams_token_by_token_and_the_done_event_is_authoritative(client):
    """The SSE contract the chat UI renders against: the concatenated `token` events ARE
    the final reply, and `done` carries the same payload the non-streaming call returns.
    If they diverge, the user watches one answer appear and a different one get saved."""
    events = _sse(client, "/v1/sessions/start?stream=true",
                  json={"user_id": "u1", "text": "delegation is hard"})

    kinds = [e["type"] for e in events]
    assert "token" in kinds and kinds[-1] == "done"
    assert "status" in kinds and "node" in kinds, "the flow view animates from these"
    streamed = "".join(e["text"] for e in events if e["type"] == "token")
    done = events[-1]
    assert streamed == done["response_to_user"]
    assert done["session_id"] and done["served_by"] == "graph"


def test_a_turn_continues_the_same_session(client):
    start = client.post("/v1/sessions/start", json={"user_id": "u1", "text": "hi"}).json()

    response = client.post(
        f"/v1/sessions/{start['session_id']}/turn", json={"user_id": "u1", "text": "tell me more"}
    )
    body = response.json()

    assert response.status_code == 200
    assert body["session_id"] == start["session_id"]
    assert body["is_first_turn"] is False, "the second turn re-ran as a first turn"


def test_a_turn_can_stream_too(client):
    start = client.post("/v1/sessions/start", json={"user_id": "u1", "text": "hi"}).json()

    events = _sse(client, f"/v1/sessions/{start['session_id']}/turn?stream=true",
                  json={"user_id": "u1", "text": "more"})

    assert events[-1]["type"] == "done"
    assert events[-1]["session_id"] == start["session_id"]


def test_ending_a_session_closes_it(client, mongo):
    """/endconversation is a terminal close: it is recorded, and the session is marked
    ended so the UI stops offering to resume it."""
    start = client.post("/v1/sessions/start", json={"user_id": "u1", "text": "hi"}).json()
    session_id = start["session_id"]

    response = client.post(
        f"/v1/sessions/{session_id}/turn", json={"user_id": "u1", "message": "/endconversation"}
    )
    body = response.json()

    assert body["stage"] == "close"
    assert body["handoff_ready"] is True
    assert conversation.get_session(session_id)["ended"] is True


def test_a_message_to_a_closed_session_is_refused_without_an_llm_call(client, mongo):
    """A finished session stays finished. Re-opening it would resume a checkpoint that
    already reached its terminal close — and would bill an LLM call to do it. The user
    is told to start a new session; served_by proves no graph run happened."""
    session_id = "sess-closed"
    _seed_turn(session_id, "u1", ended=True)

    response = client.post(f"/v1/sessions/{session_id}/turn", json={"user_id": "u1", "text": "hello?"})
    body = response.json()

    assert response.status_code == 200
    assert body["served_by"] == "closed"
    assert body["route_reason"] == "session_already_closed"
    assert body["prompt_tokens"] == 0 and body["completion_tokens"] == 0
    assert "session has ended" in body["response_to_user"].lower()


def test_restart_is_allowed_through_on_a_closed_session(client, mongo):
    """/restart is not a coaching turn — it must not be caught by the closed-session
    gate, or a user could never restart a finished chat."""
    session_id = "sess-restart"
    _seed_turn(session_id, "u1", ended=True)

    body = client.post(
        f"/v1/sessions/{session_id}/turn", json={"user_id": "u1", "message": "/restart"}
    ).json()

    assert body["served_by"] == "graph"


def test_the_rollout_gate_can_turn_the_graph_off_for_a_request(client):
    """The strangler-fig valve. `use_graph: false` returns the configured disabled
    message and spends nothing — this is the switch that gets flipped in an incident."""
    body = client.post(
        "/v1/sessions/start",
        json={"user_id": "u1", "text": "hi", "metadata": {"use_graph": False}},
    ).json()

    assert body["served_by"] == "disabled"
    assert body["route_reason"] == "request_override"
    assert body["response_to_user"] == config.GRAPH_DISABLED_MESSAGE
    assert body["prompt_tokens"] == 0


def test_a_second_turn_on_a_busy_session_is_rejected_not_interleaved(client, monkeypatch):
    """Two turns racing one session_id would interleave writes to the SAME checkpoint
    thread and corrupt the conversation state. The second is rejected with a "give me a
    moment" reply. Driven by actually holding the per-session lock (fakeredis), with the
    wait window set to 0 so the test doesn't sleep."""
    import app.stores.redis_state as redis_state

    monkeypatch.setattr(config, "REDIS_LOCK_WAIT_MS", 0)
    redis_state.get_redis().set("cerebrozen:lock:default:sess-busy", "held-by-another-turn")
    try:
        body = client.post(
            "/v1/sessions/sess-busy/turn", json={"user_id": "u1", "text": "hello"}
        ).json()
    finally:
        redis_state.get_redis().delete("cerebrozen:lock:default:sess-busy")

    assert body["served_by"] == "busy"
    assert body["stage"] == "busy"
    assert body["route_reason"] == "session_in_flight"
    assert body["prompt_tokens"] == 0


def test_an_action_checkin_tap_with_an_unresolvable_id_fails_loudly(client, mongo):
    """The UI passes the card's action_checkin_id. When it doesn't resolve (the known
    id-format mismatch: a 24-char ObjectId sent where the 12-char stable id belongs),
    running the check-in anyway would produce a hollow reflection on a blank action.
    Refuse, and say so."""
    body = client.post(
        "/v1/sessions/start",
        json={"user_id": "u1", "text": "hi", "metadata": {"action_checkin_id": "0" * 24}},
    ).json()

    assert body["served_by"] == "action_checkin_not_found"
    assert body["route_reason"] == "action_checkin_action_not_found"
    assert body["stage"] == "close"
    assert "couldn't find that action" in body["response_to_user"]


def test_an_action_checkin_tap_runs_against_the_tapped_action(client, mongo):
    """The resolved action's text is what the check-in is about — it must reach the
    graph as its own mini-session (it runs regardless of the rollout gate)."""
    added = _seed_actions("u1", "s-prior", [
        {"full_text": "I will run a 1:1 with Priya", "expected_outcome": "She feels heard"},
    ])
    action_id = added[0]["action_id"]

    body = client.post(
        "/v1/sessions/start",
        json={"user_id": "u1", "text": "hi", "metadata": {"action_checkin_id": action_id}},
    ).json()

    assert body["served_by"] == "graph"
    assert body["active_node"] == "action_checkin_agent"


def test_an_action_checkin_on_an_in_use_session_gets_a_fresh_one(client, mongo):
    """Live incident: the UI re-sent an OPEN chat's session_id with the check-in tap.
    The engine only seeds the stage on a first turn, so the graph resumed that chat's
    coaching stage instead — the check-in never ran, and its turns were recorded into
    someone else's transcript. The service re-mints a session_id when the given one
    already has turns; the response carries the new id so the client continues on it."""
    added = _seed_actions("u1", "s-prior", [{"full_text": "I will delegate the report"}])
    started = client.post("/v1/sessions/start", json={"user_id": "u1", "text": "hi"}).json()
    used_session = started["session_id"]

    body = client.post(
        "/v1/sessions/start",
        json={"user_id": "u1", "session_id": used_session, "text": "hi",
              "metadata": {"action_checkin_id": added[0]["action_id"]}},
    ).json()

    assert body["session_id"] != used_session, "the check-in ran inside the open chat"
    assert body["active_node"] == "action_checkin_agent"


def test_a_mid_arc_action_checkin_turn_continues_the_same_check_in(client, mongo):
    """The other half of the re-mint rule: when the session the tap carries IS the
    in-flight check-in (the UI re-sent the id mid-arc), it must be CONTINUED — re-minting
    there would restart the check-in from scratch and lose the turns already taken."""
    added = _seed_actions("u1", "s-prior", [{"full_text": "I will delegate the report"}])
    payload = {"user_id": "u1", "text": "hi",
               "metadata": {"action_checkin_id": added[0]["action_id"]}}
    first = client.post("/v1/sessions/start", json=payload).json()

    second = client.post(
        "/v1/sessions/start", json={**payload, "session_id": first["session_id"]}
    ).json()

    assert second["session_id"] == first["session_id"], "the in-flight check-in was restarted"
    assert second["active_node"] == "action_checkin_agent"


def test_a_checkpoint_probe_failure_does_not_kill_the_action_checkin(client, mongo, monkeypatch):
    """The stage probe that decides re-mint-vs-continue is a best-effort read. If the
    checkpointer is unreachable the turn must still run (as a fresh check-in), not 500."""
    from app.graph.engine import CereBroZenEngine

    added = _seed_actions("u1", "s-prior", [{"full_text": "I will delegate the report"}])
    started = client.post("/v1/sessions/start", json={"user_id": "u1", "text": "hi"}).json()
    monkeypatch.setattr(
        CereBroZenEngine, "session_state",
        lambda *_a, **_k: (_ for _ in ()).throw(RuntimeError("checkpointer down")),
    )

    body = client.post(
        "/v1/sessions/start",
        json={"user_id": "u1", "session_id": started["session_id"], "text": "hi",
              "metadata": {"action_checkin_id": added[0]["action_id"]}},
    ).json()

    assert body["session_id"] != started["session_id"], "an unreadable stage must re-mint"
    assert body["active_node"] == "action_checkin_agent"


def test_a_continue_button_is_recorded_by_the_turn_that_carries_it(client, mongo):
    """Pressing "Continue" used to cost the client TWO calls (the turn + a separate
    /phase-selection). The turn now stamps the selection on the phase-completion message
    itself — if it didn't, the conversation document would have no record of which
    button the user pressed."""
    _seed_turn("s-ch", "u1", bot_text="Phase 1 complete", active_phase="phase_1",
               phase_buttons=[{"label": "Continue", "user_selection": "continue_to_phase_2"}])

    client.post(
        "/v1/sessions/s-ch/turn",
        json={"user_id": "u1", "text": "let's go on",
              "metadata": {"session_continued": "continue_to_phase_2"}},
    )

    messages = conversation.get_session("s-ch")["messages"]
    assert messages[1]["phase_user_selection"] == "continue_to_phase_2"


def test_a_close_still_answers_when_the_background_builders_cannot_be_dispatched(
    client, mongo, monkeypatch, caplog
):
    """/endconversation fires the session-close context + pattern builders off-path. They
    are best-effort: if the checkpointer can't hand over the transcript, the user must
    still get their goodbye — the close is not allowed to fail on a background write."""
    from app.graph.engine import CereBroZenEngine

    monkeypatch.setattr(
        CereBroZenEngine, "session_state",
        lambda *_a, **_k: (_ for _ in ()).throw(RuntimeError("checkpointer down")),
    )

    body = client.post(
        "/v1/sessions/s-close/turn", json={"user_id": "u1", "message": "/endconversation"}
    ).json()

    assert body["stage"] == "close"
    assert body["response_to_user"].startswith("Thanks for the session")
    assert "service.session_close_builders_dispatch_failed" in caplog.text


def test_editing_the_last_message_regenerates_the_reply_and_rewrites_history(client, mongo):
    """Time-travel edit: the last user message is REPLACED and the reply regenerated
    from the forked checkpoint. The transcript must be rewritten to match — a stale
    user bubble in history against a regenerated reply is a conversation that reads as
    a non-sequitur on resume."""
    start = client.post("/v1/sessions/start", json={"user_id": "u1", "text": "delegation"}).json()
    session_id = start["session_id"]

    response = client.post(
        f"/v1/sessions/{session_id}/turn?edit=true",
        json={"user_id": "u1", "text": "actually, it is about feedback"},
    )
    body = response.json()

    assert response.status_code == 200
    assert body["edited"] is True
    assert body["served_by"] == "graph"

    messages = conversation.get_session(session_id)["messages"]
    user_messages = [m["text"] for m in messages if m["role"] == "user"]
    assert user_messages == ["actually, it is about feedback"], "the old message survived the edit"


def test_an_edit_can_stream(client, mongo):
    start = client.post("/v1/sessions/start", json={"user_id": "u1", "text": "hi"}).json()

    events = _sse(client, f"/v1/sessions/{start['session_id']}/turn?stream=true&edit=true",
                  json={"user_id": "u1", "text": "edited"})

    assert events[-1]["type"] == "done"
    assert events[-1]["edited"] is True


def test_editing_an_unknown_session_is_a_404(client, mongo):
    """Validated synchronously, BEFORE any streaming starts: a 4xx must be a real HTTP
    status the client can branch on, not an `error` event inside a 200 SSE stream."""
    response = client.post("/v1/sessions/nope/turn?edit=true", json={"user_id": "u1", "text": "x"})

    assert response.status_code == 404
    assert response.json()["detail"] == "Session not found."


def test_editing_another_users_session_is_a_404(client, mongo):
    """Not a 403 — a caller must not even learn the session exists."""
    _seed_turn("s-theirs", "someone-else")

    response = client.post("/v1/sessions/s-theirs/turn?edit=true", json={"user_id": "u1", "text": "x"})

    assert response.status_code == 404


def test_editing_an_ended_session_is_a_409(client, mongo):
    """An ended session is read-only. Editing it would regenerate from a checkpoint past
    its terminal close."""
    _seed_turn("s-ended", "u1", ended=True)

    response = client.post("/v1/sessions/s-ended/turn?edit=true", json={"user_id": "u1", "text": "x"})

    assert response.status_code == 409
    assert "ended" in response.json()["detail"]


def test_editing_a_session_with_no_user_message_is_a_400(client, mongo, agentic_coll):
    """Nothing to edit. (A transcript that only holds a /restart system marker.)"""
    mongo[config.MONGO_RASA_DB][config.MONGO_USER_CONVERSATIONS_COLLECTION].insert_one({
        "session_id": "s-system-only", "user_id": "u1", "ended": False,
        "messages": [{"role": "system", "text": "User restarted the chat.", "message_num": 1}],
    })

    response = client.post(
        "/v1/sessions/s-system-only/turn?edit=true", json={"user_id": "u1", "text": "x"}
    )

    assert response.status_code == 400
    assert response.json()["detail"] == "No user message to edit."


def test_editing_a_transcript_with_no_checkpoint_behind_it_is_a_400_not_a_500(client, mongo):
    """A deliberate 4xx must survive the worker thread.

    The transcript says there is a turn, but the graph has no checkpoint to fork from (a
    session written before the checkpointer, or a wiped thread). `service.edit_last_message`
    raises HTTPException(400, "No turn to edit.") for exactly this — and it runs INSIDE
    `_run_or_stream`'s `run_in_threadpool`, whose broad `except Exception` used to catch the
    HTTPException and re-emit it as a 500 with the status stringified into the detail.

    The status code is the API's contract. A client can fix its request and retry a 400; all
    it can do with a 500 is page somebody. Reporting one as the other is not cosmetic — it
    misattributes the fault, and it would have sent an on-call engineer looking for a server
    outage that was never there. `_run_or_stream` now re-raises HTTPException ahead of the
    broad catch (as routers/prompts.py already did).
    """
    _seed_turn("s-orphan", "u1")

    response = client.post("/v1/sessions/s-orphan/turn?edit=true", json={"user_id": "u1", "text": "x"})

    assert response.status_code == 400, "a documented 400 was re-emitted as a server error"
    assert "No turn to edit" in response.json()["detail"]


def test_a_turn_that_blows_up_returns_a_500_with_the_reason(client, monkeypatch):
    """When the graph raises, the caller gets a 500 JSON error — not a hung request and
    not a 200 with an empty reply."""
    from app.graph.engine import CereBroZenEngine

    def _boom(*_a, **_k):
        raise RuntimeError("graph exploded")

    monkeypatch.setattr(CereBroZenEngine, "run_turn_stream", _boom)

    response = client.post("/v1/sessions/start", json={"user_id": "u1", "text": "hi"})

    assert response.status_code == 500
    assert response.json() == {"type": "error", "detail": "graph exploded"}


def test_a_stream_that_blows_up_emits_an_error_event_and_terminates(client, monkeypatch):
    """The SSE stream has already returned 200 by the time the worker thread fails, so
    the failure has to arrive as an `error` event — and the stream must END. A stream
    that just stops without a terminal event leaves the UI spinning forever."""
    from app.graph.engine import CereBroZenEngine

    def _boom(*_a, **_k):
        raise RuntimeError("graph exploded")

    monkeypatch.setattr(CereBroZenEngine, "run_turn_stream", _boom)

    events = _sse(client, "/v1/sessions/start?stream=true", json={"user_id": "u1", "text": "hi"})

    assert events[-1] == {"type": "error", "detail": "graph exploded"}


async def test_the_sse_plumbing_does_not_depend_on_a_request_context():
    """The SSE worker re-stamps the request's correlation IDs into its own thread (the
    middleware's ContextVars don't cross the thread boundary). When there is nothing to
    re-stamp — a caller outside the HTTP middleware — it must still stream and still
    terminate, not KeyError on a missing id."""
    from app.routers.sessions import _sse_response

    response = _sse_response(
        lambda on_status=None, on_token=None, on_node=None: (
            on_status("thinking"), on_token("hi"), {"session_id": "s-1"})[-1]
    )
    events = [json.loads(chunk[6:]) async for chunk in response.body_iterator]

    assert events == [
        {"type": "status", "msg": "thinking"},
        {"type": "token", "text": "hi"},
        {"type": "done", "session_id": "s-1"},
    ]


def test_a_message_over_the_length_cap_is_rejected_before_the_llm(client):
    """The first user message is also the fallback chat title. One cap bounds both, and
    it must be enforced as a clean 422 rather than a 21K-token prompt plus a Mongo doc
    the size of a novel."""
    response = client.post(
        "/v1/sessions/start",
        json={"user_id": "u1", "text": "x" * (config.MAX_USER_MESSAGE_CHARS + 1)},
    )

    assert response.status_code == 422
    assert "maximum allowed length" in response.text


# ── title ───────────────────────────────────────────────────────────────────


def test_the_title_endpoint_needs_something_to_title(client):
    response = client.post("/v1/sessions/s1/title", json={"user_id": "u1"})

    assert response.status_code == 400
    assert "required" in response.json()["detail"]


def test_the_generated_title_is_persisted_where_every_other_endpoint_reads_it(client, mongo):
    """The Recents sidebar reads `title` off the session document. If the title endpoint
    generated a title but didn't persist it there, the sidebar would keep showing the
    raw first user message — which is precisely the behaviour this endpoint replaced."""
    _seed_turn("s-title", "u1")

    response = client.post("/v1/sessions/s-title/title", json={"user_id": "u1", "text": "delegation"})
    body = response.json()

    assert response.status_code == 200
    assert body["session_id"] == "s-title"
    assert body["title"]
    assert conversation.get_session_title("s-title") == body["title"]

    listed = client.get("/v1/sessions?user_id=u1").json()["sessions"]
    assert [s["title"] for s in listed] == [body["title"]]


# ── delete ──────────────────────────────────────────────────────────────────


def test_deleting_a_session_needs_a_user_in_the_token(client):
    """The user id for a delete comes ONLY from the JWT — never the payload — so a
    caller can't delete somebody else's session by naming it. With no token there is
    no user, and nothing can be deleted."""
    response = client.request("DELETE", "/v1/sessions", json={"session_id": "s1"})

    assert response.status_code == 400
    assert "JWT" in response.json()["detail"]


def test_a_session_can_only_be_deleted_by_its_owner(authed_client, mongo):
    """The delete is scoped by session_id AND the token's user_id. Another user's
    session is a 404 — and, crucially, is still there afterwards."""
    client, token_for = authed_client
    _seed_turn("s-owned", "owner")

    response = client.request(
        "DELETE", "/v1/sessions", json={"session_id": "s-owned"}, headers=token_for("attacker")
    )

    assert response.status_code == 404
    assert conversation.get_session("s-owned") is not None, "another user's session was deleted"


def test_the_owner_can_delete_their_session(authed_client, mongo):
    client, token_for = authed_client
    _seed_turn("s-owned", "owner")

    response = client.request(
        "DELETE", "/v1/sessions", json={"session_id": "s-owned"}, headers=token_for("owner")
    )

    assert response.status_code == 200
    assert response.json() == {"session_id": "s-owned", "user_id": "owner", "ok": True}
    assert conversation.get_session("s-owned") is None


# ── action status ───────────────────────────────────────────────────────────


def test_saving_an_action_needs_a_user(client):
    response = client.post(
        "/v1/sessions/s1/actions/status", json={"action_id": "a1", "action": "save"}
    )

    assert response.status_code == 400
    assert "user_id required" in response.json()["detail"]


@pytest.mark.parametrize("action", ["", "SAVE-IT", "remove", "yes"])
def test_an_unknown_action_verb_is_refused(client, mongo, action):
    response = client.post(
        "/v1/sessions/s1/actions/status",
        json={"user_id": "u1", "action_id": "a1", "action": action},
    )

    assert response.status_code == 400
    assert 'must be "save", "skip", or "delete"' in response.json()["detail"]


def test_an_empty_batch_is_refused(client, mongo):
    response = client.post(
        "/v1/sessions/s1/actions/status", json={"user_id": "u1", "actions": []}
    )

    assert response.status_code == 400
    assert response.json()["detail"] == "actions must not be empty."


def test_saving_an_action_that_does_not_exist_is_a_404(client, mongo):
    response = client.post(
        "/v1/sessions/s1/actions/status",
        json={"user_id": "u1", "action_id": "nope", "action": "save"},
    )

    assert response.status_code == 404
    assert "not found" in response.json()["detail"]


def test_saving_an_action_applies_the_cards_inline_edits_and_canonicalises_its_metrics(
    client, mongo
):
    """The card's ✏️ edit + Development-Area picker post back through this one call. The
    action must KEEP its action_id even though its text changed, or the UI's card key
    (and every later save/delete) stops resolving it."""
    added = _seed_actions("u1", "s-1", [{"full_text": "I will delegate the report",
                                         "expected_outcome": "less overload"}])
    action_id = added[0]["action_id"]
    known_metric = get_roi_metrics()[0]

    response = client.post(
        "/v1/sessions/s-1/actions/status",
        json={"user_id": "u1", "action_id": action_id, "action": "save",
              "roi_metrics": [known_metric.upper()],
              "full_text": "I will delegate the Q3 report to Sam",
              "expected_outcome": "I get 4 hours back"},
    )
    body = response.json()

    assert response.status_code == 200
    assert body["status"] == "saved"
    assert body["roi_metrics"] == [known_metric], "the metric was stored in the UI's casing"

    stored = agentic.get_action("u1", action_id)
    assert stored["action_id"] == action_id, "the inline edit moved the action's id"
    assert stored["full_text"] == "I will delegate the Q3 report to Sam"
    assert stored["expected_outcome"] == "I get 4 hours back"
    assert stored["status"] == "saved"


def test_a_deleted_action_disappears_from_the_panel(client, mongo):
    """"Deleted" is a flag, not a purge (the row stays for audit) — so the read path is
    the only thing standing between a dismissed card and the user seeing it again."""
    added = _seed_actions("u1", "s-1", [{"full_text": "I will do the thing"}])

    deleted = client.post(
        "/v1/sessions/s-1/actions/status",
        json={"user_id": "u1", "action_id": added[0]["action_id"], "action": "delete"},
    )

    assert deleted.json()["status"] == "deleted"
    assert "roi_metrics" not in deleted.json()
    panel = client.get("/v1/sessions/s-1/actions-insights?user_id=u1").json()
    assert panel["actions"] == []


@pytest.mark.parametrize("verb", ["skip", "skipped"])
def test_skipping_an_action_keeps_it_visible(client, mongo, verb):
    """A skipped card is still offered in the final-action carousel — the user can still
    save it before closing. Only "delete" hides it."""
    added = _seed_actions("u1", "s-1", [{"full_text": f"I will {verb} this"}])

    response = client.post(
        "/v1/sessions/s-1/actions/status",
        json={"user_id": "u1", "action_id": added[0]["action_id"], "action": verb},
    )

    assert response.json()["status"] == "skipped"
    panel = client.get("/v1/sessions/s-1/actions-insights?user_id=u1").json()
    assert [a["status"] for a in panel["actions"]] == ["skipped"]


def test_a_batch_applies_every_item_and_reports_the_misses_per_item(client, mongo):
    """One Mongo round trip for the whole card panel. A single unknown id must not fail
    the other saves — the user pressed those buttons, and re-running the batch would
    have to be idempotent for them."""
    added = _seed_actions("u1", "s-1", [
        {"full_text": "I will do A"}, {"full_text": "I will do B"},
    ])

    response = client.post(
        "/v1/sessions/s-1/actions/status",
        json={"user_id": "u1", "actions": [
            {"action_id": added[0]["action_id"], "action": "save"},
            {"action_id": "ghost", "action": "save"},
            {"action_id": added[1]["action_id"], "action": "delete"},
        ]},
    )
    body = response.json()

    assert response.status_code == 200
    assert body["ok"] is False, "one item missed — the batch is not fully ok"
    assert [r["ok"] for r in body["results"]] == [True, False, True]
    assert [r["status"] for r in body["results"]] == ["saved", "saved", "deleted"]
    assert agentic.get_action("u1", added[0]["action_id"])["status"] == "saved"
    assert agentic.get_action("u1", added[1]["action_id"])["status"] == "deleted"


def test_a_batch_of_saves_reports_the_metrics_it_stored(client, mongo):
    added = _seed_actions("u1", "s-1", [{"full_text": "I will do A"}])
    known_metric = get_roi_metrics()[0]

    body = client.post(
        "/v1/sessions/s-1/actions/status",
        json={"user_id": "u1", "actions": [
            {"action_id": added[0]["action_id"], "action": "save",
             "roi_metrics": [known_metric.lower()]},
        ]},
    ).json()

    assert body["ok"] is True
    assert body["results"][0]["roi_metrics"] == [known_metric]


# ── phase selection ─────────────────────────────────────────────────────────


def test_a_phase_selection_must_name_a_button(client, mongo):
    response = client.post("/v1/sessions/s1/phase-selection", json={"user_selection": "  "})

    assert response.status_code == 400
    assert response.json()["detail"] == "user_selection is required."


def test_a_phase_selection_with_no_message_to_stamp_is_a_404(client, mongo):
    response = client.post("/v1/sessions/ghost/phase-selection", json={"user_selection": "save_and_exit"})

    assert response.status_code == 404
    assert "No bot message" in response.json()["detail"]


def test_save_and_exit_is_recorded_even_though_no_turn_follows_it(client, mongo):
    """This endpoint exists ONLY because "Save & Exit" has no next turn to carry the
    selection — without it, the single most important phase button the user can press
    would never be persisted anywhere."""
    _seed_turn("s-ch", "u1", active_phase="phase_1",
               phase_buttons=[{"label": "Save & Exit", "user_selection": "save_and_exit"}])

    response = client.post(
        "/v1/sessions/s-ch/phase-selection", json={"user_selection": " save_and_exit "}
    )

    assert response.status_code == 200
    assert response.json() == {"ok": True, "session_id": "s-ch", "user_selection": "save_and_exit"}

    history = client.post("/v1/sessions/history", json={"session_id": "s-ch", "user_id": "u1"}).json()
    bot = history["chat_history"][1]["bot"]
    assert bot["phase_user_selection"] == "save_and_exit"


# ═════════════════════════════════════════════════════════════════════════════
# app/history.py — the transcript read model
# ═════════════════════════════════════════════════════════════════════════════


def test_the_recents_list_is_newest_first_and_says_what_can_be_resumed(client, mongo):
    """The sidebar disables the input box on an ended session using `resumable`. Getting
    it wrong lets a user type into a conversation that will refuse every message."""
    _seed_turn("s-old", "u1", user_message="older")
    _seed_turn("s-new", "u1", user_message="newer", ended=True)

    body = client.get("/v1/sessions?user_id=u1").json()

    assert body["count"] == 2
    assert [s["session_id"] for s in body["sessions"]] == ["s-new", "s-old"]
    assert [s["resumable"] for s in body["sessions"]] == [False, True]
    assert body["sessions"][0]["ended"] is True
    assert body["sessions"][0]["created_at"] and body["sessions"][0]["updated_at"]


def test_a_session_that_was_never_titled_falls_back_to_the_first_user_message(client, mongo):
    """Legacy sessions predate the title endpoint. An empty title in the sidebar is a
    row the user cannot identify."""
    _seed_turn("s-untitled", "u1", user_message="How do I give hard feedback?")

    body = client.get("/v1/sessions?user_id=u1").json()

    assert body["sessions"][0]["title"] == "How do I give hard feedback?"


def test_the_title_fallback_skips_past_the_system_marker(client, mongo, agentic_coll):
    """The fallback title is the first USER message — not the first message. A session
    that opens with a `/restart` system marker (or has no user message at all) must not
    show "User restarted the chat." as its name in the sidebar."""
    sessions = mongo[config.MONGO_RASA_DB][config.MONGO_USER_CONVERSATIONS_COLLECTION]
    sessions.insert_many([
        {"session_id": "s-sys", "user_id": "u1", "ended": False, "updated_at": "2026-07-02",
         "messages": [{"role": "system", "text": "User restarted the chat."},
                      {"role": "user", "text": "How do I delegate?"}]},
        {"session_id": "s-botonly", "user_id": "u1", "ended": False, "updated_at": "2026-07-01",
         "messages": [{"role": "bot", "text": "Hello?"}]},
    ])

    listed = client.get("/v1/sessions?user_id=u1").json()["sessions"]

    assert [s["title"] for s in listed] == ["How do I delegate?", ""]


@pytest.mark.parametrize("path", ["/v1/sessions", "/v1/sessions/resumable"])
def test_the_read_endpoints_need_a_user(client, path):
    response = client.get(path)

    assert response.status_code == 400
    assert "user_id required" in response.json()["detail"]


def test_the_resume_pill_is_hidden_when_nothing_is_resumable(client, mongo):
    _seed_turn("s-done", "u1", ended=True)

    body = client.get("/v1/sessions/resumable?user_id=u1").json()

    assert body == {"user_id": "u1", "resumable": False, "session_id": None,
                    "title": None, "updated_at": None}


def test_the_resume_pill_points_at_the_most_recent_open_session(client, mongo):
    _seed_turn("s-closed", "u1", ended=True)
    _seed_turn("s-open", "u1", user_message="still going")

    body = client.get("/v1/sessions/resumable?user_id=u1").json()

    assert body["resumable"] is True
    assert body["session_id"] == "s-open"
    assert body["title"] == "still going"


def test_an_unknown_session_reads_as_a_new_conversation(client, mongo):
    """Not a 404 — the resume UI asks for a transcript before it knows whether one
    exists, and "new" is the answer that renders an empty chat."""
    body = client.post("/v1/sessions/history", json={"session_id": "ghost", "user_id": "u1"}).json()

    assert body == {"converstation_status": "new", "chat_history": []}


def test_another_users_transcript_is_not_leaked(client, mongo):
    """Ownership is enforced on the read. A wrong session_id (guessed or stale) must
    read as "new", never as somebody else's conversation."""
    _seed_turn("s-private", "owner", user_message="my private problem")

    body = client.post("/v1/sessions/history", json={"session_id": "s-private", "user_id": "u2"}).json()

    assert body == {"converstation_status": "new", "chat_history": []}
    assert "private problem" not in json.dumps(body)


def test_a_transcript_is_returned_in_full_and_role_keyed(client, mongo):
    """The wire contract the existing frontend renders against — including the legacy
    `converstation_status` spelling, which is part of the contract and not a typo to fix."""
    _seed_turn("s-1", "u1", user_message="hello", bot_text="hi there")

    body = client.post(
        "/v1/sessions/history", json={"session_id": "s-1", "user_id": "u1", "page": 1, "take": 2}
    ).json()

    assert body["converstation_status"] == "mid"
    assert body["total_messages"] == 2
    assert body["has_more"] is False, "page/take are accepted but must not slice the transcript"
    assert body["chat_history"][0]["user"] == {"text": "hello", "message_num": 1, "hidden": False}
    bot = body["chat_history"][1]["bot"]
    assert bot["text"] == "hi there" and bot["tot_messages"] == 2 and bot["message_num"] == 2


def test_the_same_transcript_is_available_by_path(client, mongo):
    """GET /{session_id}/history is the curl-friendly twin of the POST — same payload."""
    _seed_turn("s-1", "u1")

    by_path = client.get("/v1/sessions/s-1/history?user_id=u1").json()
    by_body = client.post("/v1/sessions/history", json={"session_id": "s-1", "user_id": "u1"}).json()

    assert by_path == by_body


def test_an_ended_transcript_is_tagged_ended(client, mongo):
    _seed_turn("s-1", "u1", ended=True)

    body = client.get("/v1/sessions/s-1/history?user_id=u1").json()

    assert body["converstation_status"] == "ended"


def test_history_replays_the_system_marker_and_the_hidden_user_bubble(client, mongo):
    """`/restart` writes a system message; the action-ack turn writes a HIDDEN user
    message (the bot replies to it, but it must not render as a chat bubble). Both
    shapes have to survive the read or the replayed chat doesn't match what happened."""
    _seed_turn("s-1", "u1", user_message="/restart")
    _seed_turn("s-1", "u1", user_message="saved|skipped", hidden=True)

    chat = client.get("/v1/sessions/s-1/history?user_id=u1").json()["chat_history"]

    assert chat[0]["system"]["text"] == "User restarted the chat."
    assert chat[1]["user"] == {"text": "saved|skipped", "message_num": 2, "hidden": True}


def test_replayed_phase_buttons_are_disabled_and_show_what_was_pressed(client, mongo):
    """A past turn's phase buttons are NOT actionable — re-pressing one would drive a
    phase transition from the middle of a replayed transcript."""
    _seed_turn("s-1", "u1", active_phase="phase_1",
               phase_buttons=[{"label": "Continue", "user_selection": "continue_to_phase_2"}])
    conversation.record_phase_selection("s-1", "continue_to_phase_2")

    bot = client.get("/v1/sessions/s-1/history?user_id=u1").json()["chat_history"][1]["bot"]

    assert bot["active_phase"] == "phase_1"
    assert bot["phase_buttons"] == [
        {"label": "Continue", "user_selection": "continue_to_phase_2", "disabled": True}
    ]
    assert bot["phase_user_selection"] == "continue_to_phase_2"


def test_history_carries_the_actions_each_turn_generated(client, mongo):
    """Actions are stored per-user and stamped with the request_id of the turn that made
    them. That stamp is the ONLY thing tying a card back to the message it came from —
    without it, a resumed chat renders every action under the last message."""
    _seed_turn("s-1", "u1", request_id="req-1")
    _seed_turn("s-1", "u1", user_message="and then?", request_id="req-2")
    _seed_actions("u1", "s-1", [{"full_text": "I will delegate the report"}], request_id="req-2")

    chat = client.get("/v1/sessions/s-1/history?user_id=u1").json()["chat_history"]

    assert "actions" not in chat[1]["bot"], "the action was attached to the wrong turn"
    assert [a["full_text"] for a in chat[3]["bot"]["actions"]] == ["I will delegate the report"]
    assert chat[3]["bot"]["actions"][0]["id"] == chat[3]["bot"]["actions"][0]["action_id"]


def test_history_shows_deleted_actions_too(client, mongo):
    """Unlike the live panel, the transcript shows every card that was generated —
    including ones the user dismissed. It is a record of what happened."""
    added = _seed_actions("u1", "s-1", [{"full_text": "I will do the thing"}], request_id="req-1")
    _seed_turn("s-1", "u1", request_id="req-1")
    agentic.set_action_status("u1", added[0]["action_id"], "deleted")

    chat = client.get("/v1/sessions/s-1/history?user_id=u1").json()["chat_history"]

    assert [a["status"] for a in chat[1]["bot"]["actions"]] == ["deleted"]


def test_an_action_that_cannot_be_tied_to_a_turn_is_left_out(client, mongo, agentic_coll):
    """An action with no request_id (or one belonging to another session) can't be
    placed against a message — it must not be smeared across every bot bubble."""
    agentic_coll.insert_one({"user_id": "u1", "actions": [
        {"full_text": "no request id", "session_id": "s-1", "action_id": "a1"},
        {"full_text": "other session", "session_id": "s-2", "request_id": "req-1", "action_id": "a2"},
        {"session_id": "s-1", "request_id": "req-1", "action_id": "a3"},  # no full_text
    ]})
    _seed_turn("s-1", "u1", request_id="req-1")

    chat = client.get("/v1/sessions/s-1/history?user_id=u1").json()["chat_history"]

    assert "actions" not in chat[1]["bot"]


def test_an_anonymous_transcript_is_readable_without_an_owner(client, mongo, agentic_coll):
    """A transcript written with no user_id (the legacy webhook accepted a blank sender)
    still has to render — and it has no actions to attach, because actions are keyed by
    user. The read must not blow up trying to load them for an empty user."""
    mongo[config.MONGO_RASA_DB][config.MONGO_USER_CONVERSATIONS_COLLECTION].insert_one({
        "session_id": "s-anon", "user_id": "", "ended": False,
        "messages": [
            {"role": "user", "text": "hello", "message_num": 1},
            {"role": "bot", "text": "hi", "message_num": 2, "request_id": "req-1",
             "buttons": [{"title": "Yes"}, {"title": "No"}]},
        ],
    })

    body = client.post("/v1/sessions/history", json={"session_id": "s-anon"}).json()

    assert body["converstation_status"] == "mid"
    assert body["chat_history"][1]["bot"]["buttons"] == [{"title": "Yes"}, {"title": "No"}]
    assert "actions" not in body["chat_history"][1]["bot"]


def test_the_latest_reply_can_be_fetched_without_a_page_number(client, mongo):
    """Voice barge-in: the client cut TTS off mid-sentence and has no reliable page
    number to re-fetch from. This returns whichever bot message is last, with its
    buttons still LIVE (pressable) — the disabled replay shape would strand the user
    with an un-pressable phase button."""
    _seed_turn("s-1", "u1", bot_text="first")
    _seed_turn("s-1", "u1", bot_text="second", active_phase="phase_1",
               phase_buttons=[{"label": "Continue", "user_selection": "continue_to_phase_2"}])

    body = client.get("/v1/sessions/s-1/latest-response?user_id=u1").json()

    assert body["session_id"] == "s-1"
    assert body["message"]["text"] == "second"
    assert body["message"]["phase_buttons"] == [
        {"label": "Continue", "user_selection": "continue_to_phase_2"}
    ], "the live turn's buttons came back disabled"


def test_the_latest_reply_is_disabled_once_its_button_was_pressed(client, mongo):
    """Live or not, a choice already made cannot be made again."""
    _seed_turn("s-1", "u1", phase_buttons=[{"label": "Continue", "user_selection": "continue_to_phase_2"}])
    conversation.record_phase_selection("s-1", "continue_to_phase_2")

    message = client.get("/v1/sessions/s-1/latest-response?user_id=u1").json()["message"]

    assert message["phase_buttons"][0]["disabled"] is True
    assert message["phase_user_selection"] == "continue_to_phase_2"


def test_the_latest_reply_carries_the_turns_actions(client, mongo):
    _seed_turn("s-1", "u1", request_id="req-1")
    _seed_actions("u1", "s-1", [{"full_text": "I will do the thing"}], request_id="req-1")

    message = client.get("/v1/sessions/s-1/latest-response?user_id=u1").json()["message"]

    assert [a["full_text"] for a in message["actions"]] == ["I will do the thing"]


@pytest.mark.parametrize("session_id, user_id", [("ghost", "u1"), ("s-private", "u2")])
def test_the_latest_reply_is_null_when_there_is_nothing_to_return(
    client, mongo, session_id, user_id
):
    """Unknown session, or one belonging to somebody else: `message` is null, never
    another user's reply."""
    _seed_turn("s-private", "owner", bot_text="their private reply")

    body = client.get(f"/v1/sessions/{session_id}/latest-response?user_id={user_id}").json()

    assert body == {"session_id": session_id, "message": None}


def test_the_latest_reply_is_null_when_the_bot_has_not_spoken_yet(client, mongo):
    """A session whose only message is the user's (the bot reply was never recorded)."""
    _seed_turn("s-1", "u1", bot_text=None)

    body = client.get("/v1/sessions/s-1/latest-response?user_id=u1").json()

    assert body["message"] is None


# ═════════════════════════════════════════════════════════════════════════════
# app/actions_insights.py — the right-panel read model
# ═════════════════════════════════════════════════════════════════════════════


def test_the_panel_needs_a_user(client):
    response = client.get("/v1/sessions/s-1/actions-insights")

    assert response.status_code == 400
    assert "user_id required" in response.json()["detail"]


def test_an_empty_panel_says_so_rather_than_failing(client, mongo):
    """An unknown user/session is not an error — it is the empty state the panel renders
    on every brand-new session."""
    body = client.get("/v1/sessions/s-new/actions-insights?user_id=nobody").json()

    assert body == {"session_id": "s-new", "data_present": False, "version": 0,
                    "actions": [], "insights": [], "message": "no data present"}


def test_the_panel_shows_only_this_sessions_items(client, mongo):
    """Actions accumulate per USER (that is what gives cross-session continuity), but the
    panel is per SESSION. Without the filter, a new session opens showing last week's
    cards."""
    _seed_actions("u1", "s-1", [{"full_text": "I will do this session's thing"}],
                  insights=[{"insight_title": "You avoid conflict"}])
    _seed_actions("u1", "s-2", [{"full_text": "I will do another session's thing"}])

    body = client.get("/v1/sessions/s-1/actions-insights?user_id=u1").json()

    assert body["data_present"] is True
    assert body["version"] == 2
    assert [a["full_text"] for a in body["actions"]] == ["I will do this session's thing"]
    assert [i["insight_title"] for i in body["insights"]] == ["You avoid conflict"]
    assert body["available_roi_metrics"] == get_roi_metrics()
    assert body["actions"][0]["id"] == body["actions"][0]["action_id"]


def test_the_catalogue_is_only_shipped_when_there_are_cards_to_tag(client, mongo):
    """The Development-Area picker only exists on an action card. Shipping the catalogue
    with an insights-only payload is dead weight on every poll."""
    _seed_actions("u1", "s-1", [], insights=[{"insight_title": "You avoid conflict"}])

    body = client.get("/v1/sessions/s-1/actions-insights?user_id=u1").json()

    assert body["insights"] and "available_roi_metrics" not in body


def test_the_panel_dedupes_and_drops_the_unrenderable(client, mongo, agentic_coll):
    """The store is append-only and not schema-checked. Two rows with the same id (a
    later beat re-extracted the same action) must render as ONE card."""
    agentic_coll.insert_one({"user_id": "u1", "actions": [
        {"full_text": "I will do the thing", "session_id": "s-1", "action_id": "dup"},
        {"full_text": "I will do the thing", "session_id": "s-1", "action_id": "dup"},
        {"full_text": "", "session_id": "s-1", "action_id": "blank"},
        "not-a-dict",
    ]})

    body = client.get("/v1/sessions/s-1/actions-insights?user_id=u1").json()

    assert [a["action_id"] for a in body["actions"]] == ["dup"]


def test_a_legacy_single_roi_metric_is_returned_as_a_list(client, mongo, agentic_coll):
    """Older documents stored `roi_metric` (one string). The UI now renders a list. The
    read path coerces, so an old action doesn't crash the panel that renders it."""
    agentic_coll.insert_one({"user_id": "u1", "actions": [
        {"full_text": "legacy", "session_id": "s-1", "action_id": "a1", "roi_metric": "Influence"},
        {"full_text": "empty legacy", "session_id": "s-1", "action_id": "a2", "roi_metric": ""},
        {"full_text": "both", "session_id": "s-1", "action_id": "a3",
         "roi_metric": "Influence", "roi_metrics": ["Clarity"]},
    ]})

    actions = client.get("/v1/sessions/s-1/actions-insights?user_id=u1").json()["actions"]

    by_id = {a["action_id"]: a for a in actions}
    assert by_id["a1"]["roi_metrics"] == ["Influence"]
    assert by_id["a2"]["roi_metrics"] == []
    assert by_id["a3"]["roi_metrics"] == ["Clarity"]
    assert all("roi_metric" not in a for a in actions), "the legacy field leaked to the UI"


def test_an_item_with_no_stored_id_gets_a_stable_one_derived_from_its_text(
    client, mongo, agentic_coll
):
    """Actions written before ids were stamped still have to render. The id is derived
    from the normalised text, so it is STABLE across polls — a card whose id changed
    between two polls would re-render (and lose the user's in-progress edit)."""
    agentic_coll.insert_one({"user_id": "u1", "actions": [
        {"full_text": "  I  will   Delegate the report ", "session_id": "s-1"},
    ], "insights": [
        {"insight_title": "You avoid conflict", "session_id": "s-1"},
    ]})

    first = client.get("/v1/sessions/s-1/actions-insights?user_id=u1").json()
    second = client.get("/v1/sessions/s-1/actions-insights?user_id=u1").json()

    action_id = first["actions"][0]["id"]
    assert len(action_id) == 12
    assert action_id == second["actions"][0]["id"], "the id is not stable across polls"
    assert len(first["insights"][0]["id"]) == 12


def test_an_unchanged_panel_answers_304_so_the_poll_is_free(client, mongo):
    """The panel polls. Re-serialising every action on every poll for a session that has
    not changed is pure waste — the ETag turns it into a 304."""
    _seed_actions("u1", "s-1", [{"full_text": "I will do the thing"}])

    first = client.get("/v1/sessions/s-1/actions-insights?user_id=u1")
    etag = first.headers["ETag"]
    again = client.get(
        "/v1/sessions/s-1/actions-insights?user_id=u1", headers={"If-None-Match": etag}
    )

    assert first.status_code == 200 and etag
    assert again.status_code == 304
    assert again.content == b""


def test_a_new_action_invalidates_the_etag(client, mongo):
    """A 304 on a panel that HAS changed is a card the user never sees."""
    _seed_actions("u1", "s-1", [{"full_text": "I will do the thing"}])
    etag = client.get("/v1/sessions/s-1/actions-insights?user_id=u1").headers["ETag"]

    _seed_actions("u1", "s-1", [{"full_text": "I will do a second thing"}])
    again = client.get(
        "/v1/sessions/s-1/actions-insights?user_id=u1", headers={"If-None-Match": etag}
    )

    assert again.status_code == 200
    assert len(again.json()["actions"]) == 2


# ═════════════════════════════════════════════════════════════════════════════
# app/routers/flow.py — graph introspection + the agent console
# ═════════════════════════════════════════════════════════════════════════════


def test_the_graph_renders_as_a_diagram_with_a_stage_to_node_map(client):
    """The flow screen animates the live path by matching a turn's `stage` to a node in
    this diagram. A stage missing from the map lights up nothing."""
    body = client.get("/v1/graph/mermaid").json()

    assert body["mermaid"].strip().startswith("---") or "graph" in body["mermaid"]
    assert body["stage_to_node"]["coaching_intake_agent"] == "intake"


def test_the_diagram_is_rendered_once_and_cached(client):
    """The flow screen re-requests this on every navigation. Re-walking and re-drawing
    the compiled graph each time is pure waste — it cannot change without a redeploy."""
    first = client.get("/v1/graph/mermaid").json()["mermaid"]
    second = client.get("/v1/graph/mermaid").json()["mermaid"]

    assert first == second and first


def test_a_graph_that_cannot_be_rendered_is_a_500(client, monkeypatch):
    from app.routers import flow
    from app.service import get_service

    class _Broken:
        def get_graph(self):
            raise RuntimeError("cannot draw")

    monkeypatch.setattr(flow, "_mermaid_cache", None)
    monkeypatch.setattr(get_service().engine, "graph", _Broken())

    response = client.get("/v1/graph/mermaid")

    assert response.status_code == 500
    assert "could not render graph" in response.json()["detail"]


def test_a_sessions_live_stage_is_readable_from_the_checkpoint(client):
    """The flow view polls this to light up the node the session is on."""
    start = client.post("/v1/sessions/start", json={"user_id": "u1", "text": "hi"}).json()

    body = client.get(f"/v1/sessions/{start['session_id']}/stage").json()

    assert body["stage"] == start["stage"]
    assert body["node_id"] == "intake"
    assert body["safety_flag"] == "ok"
    assert body["handoff_ready"] is False


def test_an_unknown_sessions_stage_is_empty_not_an_error(client):
    body = client.get("/v1/sessions/ghost/stage").json()

    assert body == {"session_id": "ghost", "stage": "", "active_node": "", "coaching_path": "",
                    "node_id": "", "safety_flag": "", "handoff_ready": False}


def test_a_sessions_turns_can_be_replayed_from_the_checkpoint(client):
    start = client.post("/v1/sessions/start", json={"user_id": "u1", "text": "delegation"}).json()

    body = client.get(f"/v1/sessions/{start['session_id']}/transcript").json()

    assert body["count"] >= 2
    assert body["turns"][0] == {"role": "user", "content": "delegation"}
    assert body["node_id"] == "intake"


@pytest.mark.parametrize("path", ["stage", "transcript"])
def test_another_users_session_is_invisible_to_the_flow_view(authed_client, mongo, path):
    """Transcripts (and stage/safety flags) are content. Org tenancy alone must not
    be enough: a colleague in the same org who guesses a session id gets a 404,
    exactly like the delete endpoint — while the owner still reads it fine."""
    client, token_for = authed_client
    _seed_turn("s-flow-owned", "owner")

    denied = client.get(f"/v1/sessions/s-flow-owned/{path}", headers=token_for("colleague"))
    allowed = client.get(f"/v1/sessions/s-flow-owned/{path}", headers=token_for("owner"))

    assert denied.status_code == 404
    assert allowed.status_code == 200
    assert allowed.json()["session_id"] == "s-flow-owned"


@pytest.mark.parametrize("path", ["stage", "transcript"])
def test_an_unknown_session_is_a_404_when_auth_is_enforced(authed_client, mongo, path):
    """Under enforced auth the flow view must not distinguish "no such session" from
    "not your session" — both are the same 404, so ids cannot be probed."""
    client, token_for = authed_client

    response = client.get(f"/v1/sessions/ghost/{path}", headers=token_for("owner"))

    assert response.status_code == 404


@pytest.mark.parametrize("path", ["stage", "transcript"])
def test_reading_the_flow_view_needs_a_user_in_the_token(authed_client, mongo, path):
    """Same rule as delete: the user id comes ONLY from the JWT. A token that names
    no user reads nothing."""
    import jwt as pyjwt

    client, _ = authed_client
    _seed_turn("s-flow-owned", "owner")
    userless = pyjwt.encode({"org_id": "default"}, "s3cret", algorithm=config.JWT_ALGORITHM)

    response = client.get(
        f"/v1/sessions/s-flow-owned/{path}", headers={"Authorization": f"Bearer {userless}"}
    )

    assert response.status_code == 400
    assert "JWT" in response.json()["detail"]


@pytest.mark.parametrize("path", ["stage", "transcript"])
def test_a_checkpointer_outage_degrades_the_flow_view_rather_than_erroring(
    client, monkeypatch, path
):
    """The flow view is diagnostics. If the checkpointer is unreachable it shows nothing
    — it must not 500 the page an operator opened *because* something is wrong."""
    from app.graph.engine import CereBroZenEngine

    def _boom(*_a, **_k):
        raise RuntimeError("checkpointer down")

    monkeypatch.setattr(CereBroZenEngine, "session_state", _boom)

    body = client.get(f"/v1/sessions/s1/{path}").json()

    assert body["stage"] == ""


def test_the_console_lists_the_agents_that_can_actually_run(client):
    """An agent with no authored prompt cannot be run — it must not appear in the picker
    (final_action_check is a code-only stage with no sheet)."""
    stages = {a["stage"] for a in client.get("/v1/agents").json()["agents"]}

    assert "core_coaching_agent" in stages
    assert "final_action_check" not in stages
    assert "session_complete" not in stages


def test_the_console_still_lists_an_agent_whose_model_cannot_be_resolved(client, monkeypatch):
    """A bad model id in the Catalog tab must not empty the whole console picker — the
    agent is listed with whatever the catalog says, and fails loudly only when run."""
    from app.llm import responses_client

    monkeypatch.setattr(responses_client, "model_for", lambda *_a: (_ for _ in ()).throw(ValueError("bad model")))

    agents = client.get("/v1/agents").json()["agents"]

    assert agents, "a bad model id emptied the console"


def test_the_console_needs_a_prompt_to_run(client):
    response = client.post("/v1/console/run", json={"system": "You are helpful."})

    assert response.status_code == 400
    assert "user prompt is required" in response.json()["detail"]


def test_the_console_runs_a_free_form_prompt_against_the_live_model(client):
    body = client.post("/v1/console/run", json={"user": "hello"}).json()

    assert body["model"] == "mock"
    assert body["reply"]
    assert body["prompt_tokens"] > 0
    assert body["control"]["coaching_path"] == "CIM", "the JSON envelope was not parsed"


def test_a_console_run_that_fails_is_a_500(client, monkeypatch):
    from app.graph import runtime

    class _Broken:
        def generate(self, *_a, **_k):
            raise RuntimeError("model refused")

    monkeypatch.setattr(runtime, "get_client", _Broken)

    response = client.post("/v1/console/run", json={"user": "hello"})

    assert response.status_code == 500
    assert "run failed" in response.json()["detail"]


def test_running_an_unknown_agent_is_a_404(client):
    assert client.post("/v1/agents/nope/run", json={"text": "hi"}).status_code == 404


def test_running_an_agent_with_no_authored_prompt_is_a_400(client):
    """`final_action_check` is a graph stage with no workbook sheet. Running it would
    send the guardrail layer alone to the model and bill for a meaningless answer."""
    response = client.post("/v1/agents/final_action_check/run", json={"text": "hi"})

    assert response.status_code == 400
    assert "no prompt authored" in response.json()["detail"]


def test_running_an_agent_with_an_unresolvable_model_is_a_400(client, monkeypatch):
    from app.llm import responses_client

    def _boom(*_a):
        raise ValueError("unknown model")

    monkeypatch.setattr(responses_client, "model_for", _boom)

    response = client.post("/v1/agents/pattern_agent/run", json={"text": "hi"})

    assert response.status_code == 400
    assert "no model for pattern_agent" in response.json()["detail"]


def test_one_agent_can_be_run_in_isolation_with_the_same_prompt_a_turn_would_build(client):
    """The console builds the SAME system prompt the graph builds (guardrail + agent
    prompt + placeholders). If it didn't, tuning a prompt here would prove nothing about
    how it behaves in a session."""
    response = client.post(
        "/v1/agents/core_coaching_agent/run",
        json={"text": "I avoid conflict", "coaching_path": "CIM",
              "user_context": {"user_name": "Sam"}, "history": [{"role": "user", "content": "hi"}]},
    )
    body = response.json()

    assert response.status_code == 200
    assert body["stage"] == "core_coaching_agent"
    assert body["reply"] and body["raw"]
    assert body["control"]["handoff_ready"] is False
    assert body["cost_usd"] == 0.0


def test_an_agent_run_that_fails_is_a_500(client, monkeypatch):
    from app.graph import runtime

    class _Broken:
        def generate(self, *_a, **_k):
            raise RuntimeError("model refused")

    monkeypatch.setattr(runtime, "get_client", _Broken)

    response = client.post("/v1/agents/pattern_agent/run", json={"text": "hi"})

    assert response.status_code == 500
    assert "agent run failed" in response.json()["detail"]


@pytest.mark.parametrize(
    "raw, expected",
    [
        ('{"next_question": "  What now?  "}', ("What now?", {"next_question": "  What now?  "})),
        ("just text", ("just text", {})),
        ("", ("", {})),
        ('{"handoff_ready": true}', ("", {"handoff_ready": True})),
    ],
)
def test_the_console_reads_the_reply_out_of_whatever_envelope_the_agent_returned(raw, expected):
    """Agents answer with a JSON control envelope whose user-facing field differs per
    stage (and is sometimes malformed). The console shows the human-readable reply AND
    the control fields — a raw JSON blob in the reply box is useless for prompt tuning."""
    from app.routers.flow import _parse_envelope

    assert _parse_envelope(raw) == expected


def test_an_envelope_the_repair_parser_chokes_on_is_shown_as_raw_text(monkeypatch):
    """The console is where a prompt author looks WHEN the output is malformed. If the
    repair parser itself throws, the reply box must fall back to the raw text — the one
    thing the author needs to see."""
    import json_repair

    from app.routers.flow import _parse_envelope

    monkeypatch.setattr(json_repair, "repair_json",
                        lambda *_a, **_k: (_ for _ in ()).throw(RecursionError("too deep")))

    assert _parse_envelope("  {broken  ") == ("{broken", {})


# ═════════════════════════════════════════════════════════════════════════════
# app/routers/api.py — /health, the greeting, the deprecated webhook
# ═════════════════════════════════════════════════════════════════════════════


def test_health_reports_the_prompt_registry_it_is_serving(client):
    """A silent S3 fallback (serving the stale bundled workbook in a configured env) is a
    config-drift incident. It is invisible in every other response — /health is where a
    load balancer or dashboard can see it."""
    body = client.get("/health").json()

    assert body["status"] == "ok"
    assert body["brand"] == config.BRAND_NAME
    assert body["prompts"]["degraded"] is False
    assert body["prompts"]["source"] == "codebase"
    assert body["prompts"]["version"]
    assert body["force_handoff"] == {"enabled": False, "all": False, "stages": []}


def test_health_goes_degraded_when_the_prompt_registry_is_unavailable(client, monkeypatch):
    """/health must still ANSWER when prompts can't load — a health check that throws
    tells the load balancer nothing except that the process is up."""
    from app.graph import runtime

    def _boom():
        raise RuntimeError("registry down")

    monkeypatch.setattr(runtime, "get_registry", _boom)

    body = client.get("/health").json()

    assert body["status"] == "degraded"
    assert body["prompts"] == {"degraded": True, "error": "registry unavailable"}


def test_health_surfaces_the_force_handoff_test_flag(client, monkeypatch):
    """The tester UI shows an indicator for it. Leaving it on in a real environment makes
    every stage auto-advance after one turn — it must be visible, not just in an env var
    the UI process cannot read."""
    monkeypatch.setattr(config, "FORCE_HANDOFF_STAGES", {"__all__", "intake"})

    body = client.get("/health").json()

    assert body["force_handoff"] == {"enabled": True, "all": True, "stages": ["intake"]}


def test_health_status_reports_a_keyless_sovereign_posture(client):
    """The engine's sovereignty self-check: in the offline test env (mock LLM, no
    Redis/Mongo/LiveKit) it reports a fully self-hostable posture. Posture, never data."""
    body = client.get("/health/status").json()
    assert body["service"] == "engine"
    assert body["llm_provider"] == "mock" and body["llm_local"] is True
    assert body["redis_external"] is False and body["voice_cloud"] is False
    assert body["sovereign_ready"] is True


def test_health_status_flags_a_non_sovereign_deployment(client, monkeypatch):
    """A cloud LLM + a cloud-voice provider flips sovereign_ready off — honestly."""
    monkeypatch.setenv("CEREBROZEN_LLM_PROVIDER", "openai")
    monkeypatch.setattr(config, "LIVEKIT_URL", "wss://x.livekit.cloud")
    body = client.get("/health/status").json()
    assert body["llm_local"] is False and body["voice_cloud"] is True
    assert body["sovereign_ready"] is False


def test_the_deprecated_webhook_still_runs_a_turn(client):
    """Legacy callers are still on this shim. It must keep answering with the same
    payload the session endpoints return, or migrating clients break in production."""
    response = client.post("/v1/webhook", json={"sender": "u1", "text": "delegation is hard"})
    body = response.json()

    assert response.status_code == 200
    assert body["response_to_user"]
    assert len(body["session_id"]) == 32, "the webhook must mint a session id, not reuse user.bot"


def test_the_webhook_adopts_a_session_id_and_streams(client):
    events = _sse(client, "/v1/webhook?stream=true",
                  json={"sender": "u1", "session_id": "s-legacy", "message": "hi"})

    done = events[-1]
    assert done["type"] == "done"
    assert done["session_id"] == "s-legacy"
    assert "".join(e["text"] for e in events if e["type"] == "token") == done["response_to_user"]


def test_a_webhook_turn_that_blows_up_is_a_500_and_a_stream_error(client, monkeypatch):
    from app.graph.engine import CereBroZenEngine

    def _boom(*_a, **_k):
        raise RuntimeError("graph exploded")

    monkeypatch.setattr(CereBroZenEngine, "run_turn_stream", _boom)

    response = client.post("/v1/webhook", json={"sender": "u1", "text": "hi"})
    events = _sse(client, "/v1/webhook?stream=true", json={"sender": "u1", "text": "hi"})

    assert response.status_code == 500
    assert response.json() == {"type": "error", "detail": "graph exploded"}
    assert events[-1] == {"type": "error", "detail": "graph exploded"}


def test_the_greeting_is_refused_without_a_user_in_the_token(client):
    """The greeting takes NOTHING from the caller but the token — there is no user_id
    parameter to spoof. With the dev bypass there are no claims, so there is no user."""
    response = client.get("/v1/greeting?stream=false")

    assert response.status_code == 400
    assert "user_id not found in JWT" in response.json()["detail"]


def test_the_greeting_comes_back_as_one_body_when_streaming_is_off(authed_client):
    client, token_for = authed_client

    response = client.get("/v1/greeting?stream=false", headers=token_for("u1"))

    assert response.status_code == 200
    assert response.json()["greeting"].strip()


def test_the_greeting_streams_by_default(authed_client):
    client, token_for = authed_client

    events = []
    with client.stream("GET", "/v1/greeting", headers=token_for("u1")) as response:
        assert response.headers["content-type"].startswith("text/event-stream")
        for line in response.iter_lines():
            if line.startswith("data: "):
                events.append(json.loads(line[6:]))

    assert events[-1]["type"] == "done"
    assert events[-1]["greeting"].strip()
    assert "".join(e["text"] for e in events if e["type"] == "token") == events[-1]["greeting"]


def test_a_greeting_that_blows_up_ends_the_stream_with_an_error(authed_client, monkeypatch):
    """Same contract as the coaching stream: a failed worker thread must emit a terminal
    `error` event, never leave the home screen waiting on a stream that stopped."""
    from app.routers import api

    def _boom(*_a, **_k):
        raise RuntimeError("greeting exploded")

    monkeypatch.setattr(api, "generate_greeting_stream", _boom)
    client, token_for = authed_client

    events = []
    with client.stream("GET", "/v1/greeting", headers=token_for("u1")) as response:
        for line in response.iter_lines():
            if line.startswith("data: "):
                events.append(json.loads(line[6:]))

    assert events == [{"type": "error", "detail": "greeting exploded"}]


# ═════════════════════════════════════════════════════════════════════════════
# app/routers/rag.py
# ═════════════════════════════════════════════════════════════════════════════


def test_rag_health_reports_the_row_counts_and_the_bound_tokens(client, monkeypatch):
    """The placeholder tokens are what a prompt author can write in a sheet. If a token
    is missing here it silently resolves to "" in a live prompt."""
    from app.rag import store

    monkeypatch.setattr(store, "count", lambda kb: {"sskb": 12, "cskb": 7}[kb])

    body = client.get("/rag/health").json()

    assert body["sskb_rows"] == 12 and body["cskb_rows"] == 7
    assert "SSKB_Concept" in body["tokens"]
    assert body["extractions"]


def test_an_unknown_extraction_returns_a_null_result_rather_than_raising(client):
    """The extract endpoint is the only way to test an extraction against real params. An
    id typo must come back as a result with status "null", not a 500."""
    body = client.post("/rag/extract", json={"extract_id": "NoSuchExtract", "params": {}}).json()

    assert body["extract_id"] == "NoSuchExtract"
    assert body["status"] == "null"
    assert body["formatted"] == ""


def test_the_extraction_registry_can_be_reloaded_after_a_workbook_edit(client):
    body = client.post("/rag/reload").json()

    assert body["status"] == "reloaded"
    assert body["count"] > 0
    assert body["tokens"]


# ═════════════════════════════════════════════════════════════════════════════
# app/routers/voice.py
#
# NOTE: the LiveKit SDK is NOT installed in this service (voice is a separate
# deployable). Every guard on the token endpoint is exercised below; the token MINT
# itself (`from livekit import api` onwards) cannot execute here and is not faked.
# ═════════════════════════════════════════════════════════════════════════════


def test_voice_is_refused_when_livekit_is_not_configured(client):
    """503 with the exact env vars to set — not a 500 out of a None credential."""
    response = client.post("/v1/sessions/s-1/voice/token", json={"user_id": "u1"})

    assert response.status_code == 503
    assert "LIVEKIT_URL" in response.json()["detail"]


@pytest.fixture
def livekit_configured(monkeypatch):
    monkeypatch.setattr(config, "LIVEKIT_URL", "wss://livekit.example")
    monkeypatch.setattr(config, "LIVEKIT_API_KEY", "key")
    monkeypatch.setattr(config, "LIVEKIT_API_SECRET", "secret")


def test_a_voice_token_needs_a_user(client, livekit_configured):
    response = client.post("/v1/sessions/s-1/voice/token", json={})

    assert response.status_code == 400
    assert "user_id required" in response.json()["detail"]


def test_a_voice_token_needs_a_session_in_the_path(client, livekit_configured):
    """The room is named after the session_id — a blank one would put every caller in
    the same room."""
    response = client.post("/v1/sessions/%20/voice/token", json={"user_id": "u1"})

    assert response.status_code == 400
    assert "session_id required" in response.json()["detail"]


def test_the_voice_lab_falls_back_to_the_shipped_defaults_when_ssm_is_empty(client, monkeypatch):
    """No SSM (local dev, or no AWS credentials): the lab pre-fills from config, so the
    sliders are never blank and an operator can't accidentally save an empty voice id."""
    from app.routers import voice

    monkeypatch.setattr(voice, "read_voice_params", lambda env: {})

    body = client.get("/v1/voice/config").json()

    assert body["source"] == "defaults"
    assert body["ssm_configured"] is False
    assert body["params"]["voice_id"] == config.VOICE_TTS_VOICE_ID
    assert body["params"]["model"] == config.VOICE_TTS_MODEL


def test_ssm_overrides_the_shipped_defaults(client, monkeypatch):
    """SSM is the live source of truth for the voice params — a value stored there must
    win over the code default, or the lab shows the operator something that isn't running."""
    from app.routers import voice

    monkeypatch.setattr(
        voice, "read_voice_params",
        lambda env: {"CEREBROZEN_VOICE_TTS_VOICE_ID": "from-ssm", "CEREBROZEN_VOICE_SPEED": "1.2"},
    )

    body = client.get("/v1/voice/config").json()

    assert body["source"] == "ssm" and body["ssm_configured"] is True
    assert body["params"]["voice_id"] == "from-ssm"
    assert body["params"]["speed"] == "1.2"
    assert body["params"]["model"] == config.VOICE_TTS_MODEL, "unset params keep the default"


def test_saving_no_voice_params_is_refused(client):
    response = client.post("/v1/voice/config", json={})

    assert response.status_code == 400
    assert "No params" in response.json()["detail"]


def test_saving_voice_params_writes_only_what_was_supplied(client, monkeypatch):
    from app.routers import voice

    written = {}

    def _write(env, payload):
        written.update(payload)
        return sorted(payload)

    monkeypatch.setattr(voice, "write_voice_params", _write)

    response = client.post("/v1/voice/config", json={"voice_id": "v1", "speed": "1.1"})

    assert response.status_code == 200
    assert written == {"CEREBROZEN_VOICE_TTS_VOICE_ID": "v1", "CEREBROZEN_VOICE_SPEED": "1.1"}
    assert response.json()["saved"] == ["CEREBROZEN_VOICE_SPEED", "CEREBROZEN_VOICE_TTS_VOICE_ID"]


def test_an_ssm_write_failure_is_a_503_not_a_silent_no_op(client, monkeypatch):
    """The operator must know the slider they moved did NOT take effect."""
    from app.routers import voice

    def _boom(env, payload):
        raise RuntimeError("no AWS credentials")

    monkeypatch.setattr(voice, "write_voice_params", _boom)

    response = client.post("/v1/voice/config", json={"voice_id": "v1"})

    assert response.status_code == 503
    assert "no AWS credentials" in response.json()["detail"]


def test_the_voice_lab_lists_the_configured_voices(client):
    body = client.get("/v1/voice/voices").json()

    assert body["voices"] == config.VOICE_AVAILABLE_IDS
    assert body["current"]["voice_id"] == config.VOICE_TTS_VOICE_ID


def test_a_preview_without_an_elevenlabs_key_is_a_503(client, monkeypatch):
    monkeypatch.setattr(config, "ELEVEN_API_KEY", "")

    response = client.post("/v1/voice/preview", json={"text": "hello", "voice_id": "v1"})

    assert response.status_code == 503
    assert "ELEVENLABS_API_KEY" in response.json()["detail"]


def test_a_preview_proxies_the_audio_back_with_the_key_kept_server_side(client, monkeypatch):
    """The browser must never see the ElevenLabs key, so the audio is proxied. The
    request also has to carry the tuning params — a preview that ignores the sliders is
    a preview of the wrong voice."""
    import requests

    sent = {}

    class _Response:
        ok = True
        status_code = 200

        def iter_content(self, chunk_size):
            # The empty chunk is what a real streamed response emits on a keep-alive;
            # forwarding it would end the audio stream early in some clients.
            return iter([b"ID3", b"", b"audio"])

    def _post(url, headers=None, json=None, **kw):
        sent.update(url=url, headers=headers, body=json)
        return _Response()

    monkeypatch.setattr(config, "ELEVEN_API_KEY", "sk-eleven")
    monkeypatch.setattr(requests, "post", _post)

    response = client.post(
        "/v1/voice/preview",
        json={"text": "hello", "voice_id": "v1", "model": "eleven_multilingual_v2",
              "stability": 0.4, "speed": 1.1},
    )

    assert response.status_code == 200
    assert response.headers["content-type"] == "audio/mpeg"
    assert response.content == b"ID3audio"
    assert sent["url"].endswith("/text-to-speech/v1")
    assert sent["headers"]["xi-api-key"] == "sk-eleven"
    assert sent["body"]["voice_settings"] == {
        "stability": 0.4, "similarity_boost": 0.75, "style": 0.0, "speed": 1.1,
        "use_speaker_boost": True,
    }


def test_a_preview_on_eleven_v3_sends_only_the_params_that_model_accepts(client, monkeypatch):
    """eleven_v3 400s if the other VoiceSettings are sent. Silently forwarding them would
    make every v3 preview fail with an upstream error the operator can't act on."""
    import requests

    sent = {}

    class _Response:
        ok = True

        def iter_content(self, chunk_size):
            return iter([b"mp3"])

    monkeypatch.setattr(config, "ELEVEN_API_KEY", "sk-eleven")
    monkeypatch.setattr(requests, "post", lambda url, headers=None, json=None, **kw: (
        sent.update(json), _Response())[1])

    client.post(
        "/v1/voice/preview",
        json={"text": "hi", "voice_id": "v1", "model": "eleven_v3", "stability": 0.3},
    )

    assert sent["voice_settings"] == {"stability": 0.3}


def test_an_elevenlabs_error_is_forwarded_with_its_status(client, monkeypatch):
    import requests

    class _Response:
        ok = False
        status_code = 401
        text = "invalid api key"

    monkeypatch.setattr(config, "ELEVEN_API_KEY", "sk-bad")
    monkeypatch.setattr(requests, "post", lambda *a, **kw: _Response())

    response = client.post("/v1/voice/preview", json={"text": "hi", "voice_id": "v1"})

    assert response.status_code == 401
    assert "invalid api key" in response.json()["detail"]


# ═════════════════════════════════════════════════════════════════════════════
# app/main.py — the app factory
# ═════════════════════════════════════════════════════════════════════════════


def test_every_surface_is_mounted(client):
    """One assembly test: if a router stops being included, its whole test section above
    would still pass against a client that never had it."""
    from app.main import create_app

    paths = {getattr(r, "path", "") for r in create_app().routes}

    for path in ("/health", "/v1/webhook", "/v1/prompts", "/rag/health", "/v1/sessions/start",
                 "/v1/graph/mermaid", "/v1/sessions/{session_id}/voice/token", "/v1/voice/config"):
        assert path in paths, f"{path} is not mounted"


def test_every_request_is_access_logged_under_its_route_template(client, caplog):
    """The API-hit dashboard groups by route. Logging the raw path would make every
    session_id its own line in the panel — the parameterised TEMPLATE is what makes the
    metric readable."""
    with caplog.at_level("INFO", logger="cerebrozen.main"):
        client.post("/v1/sessions/sess-abc/turn", json={"user_id": "u1", "text": "hi"})

    logged = [r for r in caplog.records if r.getMessage() == "http.request"]
    assert logged, "no access log was emitted"
    assert logged[-1].route == "/v1/sessions/{session_id}/turn"
    assert logged[-1].service_name == "POST/v1/sessions/{session_id}/turn"
    assert logged[-1].status_code == 200
    assert logged[-1].response_time_ms >= 0


def test_the_system_endpoints_are_not_access_logged(client, caplog):
    """/health and /metrics are scraped constantly. Logging them would bury every real
    request, and they intentionally carry no request context."""
    with caplog.at_level("INFO", logger="cerebrozen.main"):
        client.get("/health")

    assert not [r for r in caplog.records if r.getMessage() == "http.request"]


@pytest.mark.parametrize(
    "path, marker",
    [("/", "CereBroZen"), ("/chat", "<"), ("/prompts", "<"), ("/flow", "<")],
)
def test_the_browser_uis_are_served(client, path, marker):
    response = client.get(path)

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("text/html")
    assert marker in response.text


@pytest.mark.parametrize("path", ["/", "/chat", "/prompts", "/flow"])
def test_a_missing_static_bundle_degrades_to_a_placeholder(client, monkeypatch, path):
    """A container built without app/static must still boot and serve the API. The UI
    routes fall back to a placeholder rather than 500-ing on every page load."""
    import builtins

    real_open = builtins.open

    def _no_html(file, *a, **kw):
        if str(file).endswith(".html"):
            raise FileNotFoundError(file)
        return real_open(file, *a, **kw)

    monkeypatch.setattr(builtins, "open", _no_html)

    response = client.get(path)

    assert response.status_code == 200
    assert "UI not found" in response.text


def test_the_static_mount_is_skipped_when_the_bundle_is_absent(monkeypatch):
    import os

    from app.main import create_app

    real_isdir = os.path.isdir
    monkeypatch.setattr(
        os.path, "isdir",
        lambda p: False if str(p).endswith(("static", "testui")) else real_isdir(p),
    )

    paths = {getattr(r, "path", "") for r in create_app().routes}

    assert "/static" not in paths
    assert "/v1/sessions/start" in paths, "the API must still be served without the UI"


def test_prometheus_is_scrapeable(client):
    response = client.get("/metrics")

    assert response.status_code == 200
    assert "python_info" in response.text or "# HELP" in response.text


def test_metrics_are_simply_not_mounted_on_a_lean_install(monkeypatch):
    """prometheus_client is optional. Without it /metrics must be absent, not broken."""
    from fastapi.testclient import TestClient

    import app.metrics as metrics
    from app.main import create_app

    monkeypatch.setattr(metrics, "metrics_asgi_app", lambda: None)
    client = TestClient(create_app(), raise_server_exceptions=False)

    assert client.get("/metrics").status_code == 404
    assert client.get("/health").status_code == 200


def test_an_auth_failure_is_a_401_with_a_message(authed_client):
    """The other services in the estate return {"message": ...} on 401. A FastAPI
    default {"detail": ...} would break every client's error handling."""
    client, _ = authed_client

    response = client.get("/v1/sessions?user_id=u1")

    assert response.status_code == 401
    assert response.json() == {"message": "Missing Authentication Token"}
    assert client.get("/v1/sessions?user_id=u1",
                      headers={"Authorization": "Bearer nonsense"}).status_code == 401


def test_a_preflight_is_answered_for_a_browser_frontend(client):
    response = client.options(
        "/v1/sessions/start",
        headers={"Origin": "https://app.example.com",
                 "Access-Control-Request-Method": "POST"},
    )

    assert response.status_code == 200
    assert response.headers["access-control-allow-origin"] == "*"


def test_startup_warms_the_registry_and_the_client(monkeypatch, caplog):
    """The first turn must not pay the prompt-download + client-construction cost, and
    the boot log is where the prompt SOURCE is confirmed. Startup only runs under the
    TestClient context manager, so this is the one place it is exercised."""
    from fastapi.testclient import TestClient

    from app.main import create_app
    from app.rag import startup as rag_startup

    warmed = []
    monkeypatch.setattr(rag_startup, "run_startup", lambda reindex: warmed.append(reindex))

    with caplog.at_level("INFO", logger="cerebrozen.main"):
        with TestClient(create_app()) as client:
            assert client.get("/health").json()["status"] == "ok"

    assert "app.ready" in caplog.text
    assert warmed == [config.RAG_REINDEX], "the RAG index was not warmed at boot"


@pytest.mark.parametrize("victim", ["prompts", "rag"])
def test_a_failed_warm_up_does_not_stop_the_service_from_booting(monkeypatch, caplog, victim):
    """Warming is an optimisation. An S3 hiccup at boot must degrade to a cold first
    turn — never to a service that refuses to start."""
    from fastapi.testclient import TestClient

    from app.graph import runtime
    from app.main import create_app
    from app.rag import startup as rag_startup

    def _boom(*_a, **_k):
        raise RuntimeError(f"{victim} is down")

    monkeypatch.setattr(rag_startup, "run_startup", _boom if victim == "rag" else lambda _r: None)
    if victim == "prompts":
        monkeypatch.setattr(runtime, "get_registry", _boom)

    with caplog.at_level("INFO", logger="cerebrozen.main"):
        with TestClient(create_app()) as client:
            assert client.get("/health").status_code == 200

    assert "app.ready" in caplog.text
    assert f"{victim} is down" in caplog.text


def test_the_app_boots_with_the_voice_stack_absent(monkeypatch):
    """Voice depends on the (heavy, optional) livekit stack and is a separate deployable.
    The API must boot without it — the voice routes are simply not mounted, and every
    other surface still serves."""
    main = _import_fresh("app.main_no_voice", "app/main.py",
                         blocked=["app.routers.voice"], monkeypatch=monkeypatch)

    assert main.voice_router is None and main.voice_lab_router is None
    paths = {getattr(r, "path", "") for r in main.create_app().routes}
    assert not any("voice" in p for p in paths)
    assert "/v1/sessions/start" in paths and "/health" in paths


# ═════════════════════════════════════════════════════════════════════════════
# app/metrics.py — the Prometheus/OTEL counters
# ═════════════════════════════════════════════════════════════════════════════


def _counter(name: str, **labels) -> float:
    from prometheus_client import REGISTRY

    return REGISTRY.get_sample_value(name, labels) or 0.0


def test_an_llm_call_is_counted_by_stage_and_model():
    """The cost/latency dashboards and the spend alert read these. `session_id` must
    never become a label — unbounded cardinality would take Prometheus down."""
    from app.metrics import record_llm

    before = _counter("cerebrozen_llm_cost_usd_total", stage="core", model="gpt-4.1")

    record_llm(stage="core", model="gpt-4.1", latency_ms=1500, prompt_tokens=100,
               cached_tokens=40, completion_tokens=20, cost_usd=0.25)

    assert _counter("cerebrozen_llm_cost_usd_total", stage="core", model="gpt-4.1") == before + 0.25
    assert _counter("cerebrozen_llm_calls_total", stage="core", model="gpt-4.1") >= 1
    assert _counter("cerebrozen_llm_tokens_total", stage="core", model="gpt-4.1", kind="cached") >= 40
    assert _counter("cerebrozen_llm_latency_seconds_sum", stage="core", model="gpt-4.1") >= 1.5


def test_an_unlabelled_llm_call_is_still_counted():
    """A missing stage/model must not silently drop the sample (or blow up on a None
    label) — it lands under "unknown", where it is visible."""
    from app.metrics import record_llm

    before = _counter("cerebrozen_llm_calls_total", stage="unknown", model="unknown")

    record_llm(stage="", model="", latency_ms=0, prompt_tokens=0, cached_tokens=0,
               completion_tokens=0, cost_usd=0)

    assert _counter("cerebrozen_llm_calls_total", stage="unknown", model="unknown") == before + 1


def test_the_quiet_failures_are_counted():
    """Contract violations and watchdog force-advances are invisible by design — the
    graph's fallbacks keep the session running. These counters are the ONLY signal that
    a prompt has drifted from its contract."""
    from app.metrics import record_contract_violation, record_rate_limited, record_stage_watchdog

    record_contract_violation(stage="core", contract="coaching_path")
    record_stage_watchdog(stage="intake")
    record_rate_limited(bucket="turn")
    record_contract_violation(stage="", contract="")

    assert _counter("cerebrozen_agent_contract_violations_total", stage="core",
                    contract="coaching_path") >= 1
    assert _counter("cerebrozen_stage_watchdog_total", stage="intake") >= 1
    assert _counter("cerebrozen_rate_limited_total", bucket="turn") >= 1
    assert _counter("cerebrozen_agent_contract_violations_total", stage="unknown",
                    contract="unknown") >= 1


def test_the_otel_meter_is_skipped_when_otel_is_off(monkeypatch):
    from app import metrics

    monkeypatch.setattr(config, "OTEL_METRICS_ENABLED", False)
    monkeypatch.setattr(metrics, "_otel", {"init": False, "calls": None, "cost": None,
                                           "latency": None, "tokens": None})

    assert metrics._otel_instruments() is None
    assert metrics._otel_instruments() is None, "the negative result must be cached, not retried"


def test_the_otel_meter_records_when_otel_is_on(monkeypatch):
    """The OTLP path feeds the ADOT collector. It is lazy (the meter provider is set at
    startup), so the instruments must be built on first use, not at import."""
    from app import metrics

    monkeypatch.setattr(config, "OTEL_METRICS_ENABLED", True)
    monkeypatch.setattr(metrics, "_otel", {"init": False, "calls": None, "cost": None,
                                           "latency": None, "tokens": None})

    metrics.record_llm(stage="core", model="m", latency_ms=10, prompt_tokens=1,
                       cached_tokens=0, completion_tokens=1, cost_usd=0.01)

    assert metrics._otel["calls"] is not None
    assert metrics._otel_instruments() is metrics._otel


def test_a_missing_otel_package_leaves_the_meter_off(monkeypatch):
    """opentelemetry is optional. Its absence must make the OTEL meter a no-op, not an
    ImportError on every LLM call."""
    from app import metrics

    monkeypatch.setattr(config, "OTEL_METRICS_ENABLED", True)
    monkeypatch.setattr(metrics, "_otel", {"init": False, "calls": None, "cost": None,
                                           "latency": None, "tokens": None})
    monkeypatch.setitem(sys.modules, "opentelemetry", None)

    assert metrics._otel_instruments() is None
    metrics.record_llm(stage="core", model="m", latency_ms=1, prompt_tokens=1, cached_tokens=0,
                       completion_tokens=1, cost_usd=0.1)


def test_a_broken_meter_never_raises_into_the_turn(monkeypatch):
    """Metrics are on the LLM call path. An instrument that throws must never become a
    failed coaching turn."""
    from app import metrics

    class _Broken:
        def add(self, *_a, **_k):
            raise RuntimeError("meter exploded")

        record = add

    monkeypatch.setattr(config, "OTEL_METRICS_ENABLED", True)
    monkeypatch.setattr(metrics, "_otel", {"init": True, "calls": _Broken(), "cost": _Broken(),
                                           "latency": _Broken(), "tokens": _Broken()})

    metrics.record_llm(stage="core", model="m", latency_ms=10, prompt_tokens=1,
                       cached_tokens=0, completion_tokens=1, cost_usd=0.01)


def test_metrics_degrade_to_no_ops_without_prometheus(monkeypatch, caplog):
    """The lean install: prometheus_client absent. Every helper must become a no-op and
    /metrics must simply not exist — importing the app has to keep working."""
    lean = _import_fresh("app.metrics_lean", "app/metrics.py",
                         blocked=["prometheus_client"], monkeypatch=monkeypatch)

    assert lean._ENABLED is False
    assert lean.metrics_asgi_app() is None
    lean.record_llm(stage="core", model="m", latency_ms=1, prompt_tokens=1, cached_tokens=0,
                    completion_tokens=1, cost_usd=0.1)
    lean.record_rate_limited(bucket="turn")
    lean.record_contract_violation(stage="core", contract="c")
    lean.record_stage_watchdog(stage="core")


# ═════════════════════════════════════════════════════════════════════════════
# app/observability.py + app/request_context.py — the JSON log record
# ═════════════════════════════════════════════════════════════════════════════


def _format(record) -> dict:
    from app.observability import JsonFormatter

    return json.loads(JsonFormatter().format(record))


def _record(msg="event", **extra):
    import logging

    record = logging.LogRecord("cerebrozen.test", logging.INFO, __file__, 1, msg, (), None)
    for key, value in extra.items():
        setattr(record, key, value)
    return record


def test_a_log_record_carries_the_correlation_ids_from_the_request(client, caplog):
    """Every log line in one turn shares a request_id, and the turn's user/session ids
    are stamped on ALL of them — including the ones deep in the graph that never pass
    them explicitly. Without this, debugging a live session means grepping by timestamp."""
    from app.observability import RequestContextFilter
    from app.request_context import ctx_session_id, ctx_user_id, request_id

    tokens = [request_id.set("rid-1"), ctx_user_id.set("u-1"), ctx_session_id.set("s-1")]
    try:
        record = _record()
        RequestContextFilter().filter(record)
        payload = _format(record)

        explicit = _record(user_id="explicit-user")
        RequestContextFilter().filter(explicit)
    finally:
        request_id.reset(tokens[0])
        ctx_user_id.reset(tokens[1])
        ctx_session_id.reset(tokens[2])

    assert payload["request_id"] == "rid-1"
    assert payload["user_id"] == "u-1"
    assert payload["session_id"] == "s-1"
    assert list(payload)[:4] == ["ts", "level", "logger", "event"]
    assert explicit.user_id == "explicit-user", "an explicit extra must not be overwritten"


def test_a_log_record_outside_a_request_carries_no_correlation_fields():
    """Startup and health-check logs have no request context — they must stay clean, not
    carry empty strings."""
    from app.observability import RequestContextFilter

    record = _record()
    RequestContextFilter().filter(record)
    payload = _format(record)

    assert "request_id" not in payload and "user_id" not in payload


def test_an_unserialisable_extra_is_stringified_rather_than_losing_the_record():
    """A log line is not worth crashing a turn for. An object that json can't encode gets
    str()'d — the record still ships."""
    payload = _format(_record(obj=object(), stage="core"))

    assert payload["stage"] == "core"
    assert payload["obj"].startswith("<object object at")


def test_an_exception_is_carried_in_the_record():
    import logging

    try:
        raise ValueError("boom")
    except ValueError:
        record = logging.LogRecord("t", logging.ERROR, __file__, 1, "failed", (),
                                   sys.exc_info())

    payload = _format(record)

    assert "ValueError: boom" in payload["exc"]


def test_logging_is_configured_once(caplog):
    """configure_logging runs on every create_app(). Re-adding the handlers would
    duplicate every log line once per app instance."""
    import logging

    from app.observability import configure_logging

    configure_logging()
    before = len([h for h in logging.getLogger().handlers if getattr(h, "_cerebrozen", False)])
    configure_logging()

    assert before == len([h for h in logging.getLogger().handlers if getattr(h, "_cerebrozen", False)])
    assert logging.getLogger("pymongo").level == logging.WARNING, "noisy loggers stay quiet"


# ═════════════════════════════════════════════════════════════════════════════
# app/env_loader.py
# ═════════════════════════════════════════════════════════════════════════════


def test_a_powershell_env_file_populates_the_environment(tmp_path, monkeypatch):
    """Windows env vars are per-terminal; forgetting to source env-dev.ps1 left the
    server with no OPENAI_API_KEY. This parses it — but must NEVER override a var the
    operator set explicitly."""
    from app.env_loader import load_env_file

    import os

    ps1 = tmp_path / "env-dev.ps1"
    ps1.write_text(
        '$env:PS_QUOTED = "double"\n'
        "$env:PS_SINGLE = 'single'\n"
        "$env:PS_BARE = bare\n"
        '$env:PS_ALREADY_SET = "from-file"\n'
    )
    monkeypatch.setenv("PS_ALREADY_SET", "from-shell")
    for name in ("PS_QUOTED", "PS_SINGLE", "PS_BARE"):
        monkeypatch.delenv(name, raising=False)

    loaded = load_env_file(ps1)

    assert loaded == 3
    assert os.environ["PS_QUOTED"] == "double"
    assert os.environ["PS_SINGLE"] == "single"
    assert os.environ["PS_BARE"] == "bare"
    assert os.environ["PS_ALREADY_SET"] == "from-shell", "an explicitly-set var was clobbered"


def test_a_dotenv_file_fills_in_the_blanks(tmp_path, monkeypatch):
    """The empty-value rule matters: uvicorn inherits an EMPTY OPENAI_API_KEY, and the
    .env has to be able to fill it. Only a non-empty explicit value wins."""
    from app.env_loader import load_dotenv_file

    import os

    dotenv = tmp_path / ".env"
    dotenv.write_text(
        "# a comment\n"
        "\n"
        "DOT_PLAIN=value\n"
        'DOT_QUOTED="spaced value"\n'
        "export DOT_EXPORTED=exported\n"
        "DOT_WAS_EMPTY=filled\n"
        "DOT_WAS_SET=from-file\n"
        "not a valid line\n"
    )
    for name in ("DOT_PLAIN", "DOT_QUOTED", "DOT_EXPORTED"):
        monkeypatch.delenv(name, raising=False)
    monkeypatch.setenv("DOT_WAS_EMPTY", "")
    monkeypatch.setenv("DOT_WAS_SET", "from-shell")

    loaded = load_dotenv_file(dotenv)

    assert loaded == 4
    assert os.environ["DOT_PLAIN"] == "value"
    assert os.environ["DOT_QUOTED"] == "spaced value"
    assert os.environ["DOT_EXPORTED"] == "exported"
    assert os.environ["DOT_WAS_EMPTY"] == "filled", "an EMPTY var must be fillable from .env"
    assert os.environ["DOT_WAS_SET"] == "from-shell"


def test_a_missing_env_file_is_not_an_error(tmp_path):
    from app.env_loader import load_dotenv_file, load_env_file

    assert load_env_file(tmp_path / "nope.ps1") == 0
    assert load_dotenv_file(tmp_path / "nope.env") == 0


def test_loading_the_local_env_never_overrides_the_running_environment(monkeypatch, caplog):
    """It runs at import of `app.main`, INSIDE a process whose env is already set (ECS
    task definition, sourced shell, CI secrets). A local file that could overwrite those
    would let a stray `.env` on a build box change what a deployed service connects to.

    With every name already present and non-empty, nothing is loaded — and the module
    says so rather than logging a misleading "env.loaded"."""
    import os

    from app.env_loader import _DOTENV_LINE, load_local_env

    repo_root = Path(__file__).resolve().parent.parent
    dotenv = repo_root / ".env"
    if dotenv.exists():
        for line in dotenv.read_text(encoding="utf-8-sig", errors="replace").splitlines():
            match = _DOTENV_LINE.match(line)
            if match and not line.strip().startswith("#"):
                monkeypatch.setenv(match.group(1), "already-set-in-the-environment")

    before = dict(os.environ)
    with caplog.at_level("INFO", logger="cerebrozen.env"):
        load_local_env()

    assert dict(os.environ) == before, "a local file overrode the running environment"
    assert "env.skipped" in caplog.text


# ═════════════════════════════════════════════════════════════════════════════
# app/tracing_otel.py
#
# NOTE: `opentelemetry-sdk-extension-aws` (AwsXRayIdGenerator) and the gRPC OTLP
# exporter are NOT installed and are NOT in requirements.txt — only the HTTP
# exporter is. So `configure_tracing()` can never get past its import block in this
# deployment, and the provider setup that follows it is unreachable from any test
# here. That is a real gap in the app, not a testing one; it is reported, not faked.
# ═════════════════════════════════════════════════════════════════════════════


def test_tracing_is_off_and_free_when_no_collector_is_configured(monkeypatch, caplog):
    """OFF by default: no endpoint → no providers, no background exporter threads, and a
    no-op tracer, so a turn pays nothing for observability that isn't wired up."""
    from app import tracing_otel

    monkeypatch.setattr(tracing_otel, "_CONFIGURED", False)
    monkeypatch.setattr(tracing_otel, "_TRACER", None)
    monkeypatch.setattr(config, "OTEL_ENABLED", False)

    with caplog.at_level("INFO", logger="cerebrozen.otel"):
        assert tracing_otel.configure_tracing() is False

    assert "otel.disabled" in caplog.text
    assert tracing_otel.configure_tracing() is False, "configure_tracing must be idempotent"


def test_enabling_otel_without_its_packages_degrades_instead_of_failing_boot(
    monkeypatch, caplog
):
    """BUG SURFACE (pinned): turning OTEL on cannot work in this build. The X-Ray id
    generator (`opentelemetry-sdk-extension-aws`) and the default gRPC exporter are not
    installed and are not declared in requirements.txt — so the import block always
    fails and tracing silently stays OFF, however the collector is configured.

    What this test guarantees is the only thing that still holds: a misconfigured OTEL
    must never take the service down. It degrades, loudly, and the app still boots."""
    from app import tracing_otel

    monkeypatch.setattr(tracing_otel, "_CONFIGURED", False)
    monkeypatch.setattr(tracing_otel, "_TRACER", None)
    monkeypatch.setattr(config, "OTEL_ENABLED", True)

    with caplog.at_level("WARNING", logger="cerebrozen.otel"):
        enabled = tracing_otel.configure_tracing()

    assert enabled is False
    degraded = next(r for r in caplog.records if r.getMessage() == "otel.deps_missing")
    assert "pip install" in degraded.hint, "the hint must name what to install"
    assert "opentelemetry-sdk-extension-aws" in degraded.hint


@pytest.mark.parametrize(
    "protocol, is_http",
    [("grpc", False), ("http", True), ("http/protobuf", True), ("httpprotobuf", True)],
)
def test_the_exporter_follows_the_configured_protocol(monkeypatch, protocol, is_http):
    """Endpoint and protocol are standard OTEL env vars (set via SSM) — picking the
    exporter is a config choice, not a code change. Only the HTTP exporter is installed
    here, so only that one can be resolved."""
    from app import tracing_otel

    monkeypatch.setattr(config, "OTEL_EXPORTER_OTLP_PROTOCOL", protocol)

    assert tracing_otel._is_http_protocol() is is_http
    if is_http:
        assert tracing_otel._span_exporter_cls().__name__ == "OTLPSpanExporter"
        assert tracing_otel._metric_exporter_cls().__name__ == "OTLPMetricExporter"


def test_the_meter_provider_is_configured_without_touching_the_network(monkeypatch, caplog):
    """Metrics export must be ASYNC (a periodic reader on a background thread) — a
    synchronous export on the LLM call path would put the collector's latency into every
    coaching turn. The OTLP exporter (the network boundary) is stubbed out."""
    from opentelemetry.exporter.otlp.proto.http import metric_exporter
    from opentelemetry.sdk.resources import Resource

    from app import tracing_otel

    exported = []

    class _StubExporter:
        def __init__(self, *a, **kw):
            exported.append(True)

        def export(self, *_a, **_k):
            return None

        def shutdown(self, *_a, **_k):
            return None

        def force_flush(self, *_a, **_k):
            return True

        _preferred_temporality = {}
        _preferred_aggregation = {}

    monkeypatch.setattr(config, "OTEL_EXPORTER_OTLP_PROTOCOL", "http")
    monkeypatch.setattr(metric_exporter, "OTLPMetricExporter", _StubExporter)

    with caplog.at_level("INFO", logger="cerebrozen.otel"):
        tracing_otel._configure_metrics(Resource.create({"service.name": "test"}))

    assert exported == [True]
    assert "otel.metrics_enabled" in caplog.text


def test_a_broken_metrics_exporter_does_not_break_tracing(monkeypatch, caplog):
    """Metrics are the optional half. If the meter provider can't be built, traces must
    still export — one broken exporter must not cost you both signals."""
    from app import tracing_otel

    monkeypatch.setattr(
        tracing_otel, "_metric_exporter_cls",
        lambda: (_ for _ in ()).throw(RuntimeError("no exporter")),
    )

    with caplog.at_level("WARNING", logger="cerebrozen.otel"):
        tracing_otel._configure_metrics(object())

    assert "otel.metrics_skipped" in caplog.text


def test_the_tracer_is_a_no_op_when_otel_is_off(monkeypatch):
    """Every node opens a span. With OTEL off that has to be free and, above all, safe —
    a span context manager that raises would take down every turn."""
    from app import tracing_otel

    monkeypatch.setattr(tracing_otel, "_TRACER", None)
    tracer = tracing_otel.get_tracer()

    with tracer.start_as_current_span("turn") as span:
        span.set_attribute("stage", "core")
        span.record_exception(RuntimeError("x"))

    monkeypatch.setitem(sys.modules, "opentelemetry", None)
    stub = tracing_otel.get_tracer()
    with stub.start_as_current_span("turn") as span:
        span.set_attribute("stage", "core")
        span.record_exception(RuntimeError("x"))
    assert isinstance(stub, tracing_otel._NoopTracer)


def test_the_configured_tracer_is_reused(monkeypatch):
    from app import tracing_otel

    sentinel = object()
    monkeypatch.setattr(tracing_otel, "_TRACER", sentinel)

    assert tracing_otel.get_tracer() is sentinel


# ═════════════════════════════════════════════════════════════════════════════
# app/schemas.py — the request contract
# ═════════════════════════════════════════════════════════════════════════════


def test_a_slash_command_is_an_intent_not_a_user_message():
    """`/endconversation` must reach the service as a COMMAND (raw_message) and as an
    EMPTY user message — feeding it to the coaching agent as text would have the coach
    reply to the literal string "/endconversation"."""
    from app.schemas import SessionTurnRequest, WebhookRequest

    command = SessionTurnRequest(message="/endconversation")
    assert command.raw_message() == "/endconversation"
    assert command.user_text() == ""

    typed = SessionTurnRequest(message="/endconversation", text="  wait, one more thing  ")
    assert typed.user_text() == "wait, one more thing", "the typed text wins over the command"

    legacy = WebhookRequest(sender="u1", message="  hello  ")
    assert legacy.raw_message() == "hello" and legacy.user_text() == "hello"
    assert WebhookRequest(sender="u1", message="/restart").user_text() == ""
