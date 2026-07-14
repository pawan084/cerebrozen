"""The LLM client, the Redis hot-state tier, and the guards around them.

These are the layers where a defect is *invisible*: nothing throws, the coaching reply
still arrives, and the damage shows up on the OpenAI invoice, in a wedged session, or in
a stream that never ends. So the tests here are deliberately about the things nobody
watches —

  * the exact JSON body we put on the wire (a dropped `prompt_cache_key` costs ~48% more
    per turn and nothing fails);
  * the per-stage reasoning effort (a stage missing from the map silently reverts to the
    model's own default reasoning — that is the measured 31.9s→4.8s action_checkin defect);
  * the wall-clock deadline on a stream (live incident 2026-07-06: a generation ran
    forever, never errored, emitted no telemetry);
  * and the rule that a Redis hiccup must NEVER raise into a turn — the lock falls open,
    the cache misses, coaching continues.

The OpenAI SDK is exercised for real: only the network is replaced (an httpx
MockTransport), so every assertion below is against the bytes the real SDK actually
serialized, not against a mock we taught to agree with us.
"""

from __future__ import annotations

import json
import sys
import uuid
from pathlib import Path

import fakeredis
import httpx
import openai
import pytest
from openpyxl import Workbook

from app import config
from app.llm import prompts as prompts_mod
from app.llm import resilience
from app.llm import responses_client as rc
from app.stores import redis_state as rs

# ═════════════════════════════════════════════════════════════════════════════
# The OpenAI boundary: the real SDK, a fake network.
# ═════════════════════════════════════════════════════════════════════════════


def _response_body(text: str, usage: dict) -> dict:
    """A minimal but schema-valid Responses API payload the real SDK will parse."""
    return {
        "id": "resp_test",
        "object": "response",
        "created_at": 1,
        "model": "gpt-5-mini",
        "status": "completed",
        "output": [{
            "id": "msg_1", "type": "message", "role": "assistant", "status": "completed",
            "content": [{"type": "output_text", "text": text, "annotations": []}],
        }],
        "parallel_tool_calls": False,
        "tool_choice": "auto",
        "tools": [],
        "usage": {
            "input_tokens": usage["input_tokens"],
            "output_tokens": usage["output_tokens"],
            "total_tokens": usage["total_tokens"],
            "input_tokens_details": {"cached_tokens": usage["cached_tokens"]},
            "output_tokens_details": {"reasoning_tokens": 0},
        },
    }


def _sse(events: list[tuple[str, dict]]) -> bytes:
    out = []
    for i, (kind, payload) in enumerate(events, start=1):
        data = json.dumps({"type": kind, "sequence_number": i, **payload})
        out.append(f"event: {kind}\ndata: {data}\n\n")
    return "".join(out).encode()


class _Wire:
    """The network under the OpenAI SDK. Records the request bodies actually sent and
    replays scripted outcomes (success, or an HTTP status the SDK maps to a real
    openai exception — so our retry classification is tested against the SDK's own
    exception tree, not against a hand-rolled stand-in)."""

    def __init__(self) -> None:
        self.requests: list[dict] = []
        self.client_kwargs: dict = {}
        self.text = "hello there"
        self.deltas = ["hel", "lo"]
        self.usage = {"input_tokens": 1000, "output_tokens": 50,
                      "total_tokens": 1050, "cached_tokens": 0}
        self.script: list = []   # each item: an int HTTP status, or a callable(body)

    @property
    def models_tried(self) -> list[str]:
        return [r["model"] for r in self.requests]

    def handle(self, request: httpx.Request) -> httpx.Response:
        body = json.loads(request.content)
        self.requests.append(body)
        if self.script:
            outcome = self.script.pop(0)
            if isinstance(outcome, int):
                return httpx.Response(
                    outcome, json={"error": {"message": f"upstream {outcome}",
                                             "type": "server_error", "code": None}},
                )
            if callable(outcome):
                return outcome(body)
        if body.get("stream"):
            return self._stream_response()
        return httpx.Response(200, json=_response_body(self.text, self.usage))

    def _stream_response(self) -> httpx.Response:
        final = _response_body("".join(self.deltas), self.usage)
        events: list[tuple[str, dict]] = [
            ("response.created", {"response": {**final, "status": "in_progress",
                                               "output": [], "usage": None}}),
            ("response.output_item.added", {"output_index": 0, "item": {
                "id": "msg_1", "type": "message", "role": "assistant",
                "status": "in_progress", "content": []}}),
            ("response.content_part.added", {
                "item_id": "msg_1", "output_index": 0, "content_index": 0,
                "part": {"type": "output_text", "text": "", "annotations": []}}),
        ]
        for d in self.deltas:
            events.append(("response.output_text.delta", {
                "item_id": "msg_1", "output_index": 0, "content_index": 0,
                "delta": d, "logprobs": []}))
        events.append(("response.completed", {"response": final}))
        return httpx.Response(200, headers={"content-type": "text/event-stream"},
                              content=_sse(events))


class _Clock:
    """A deterministic stand-in for the `time` module *inside responses_client only*.

    Substituted for the module's `time` binding rather than patching stdlib globally, so
    the SDK and httpx keep their real clocks. `step` is how far perf_counter() jumps per
    call — that is what lets the stream deadline be tripped without any real waiting.
    """

    def __init__(self, step: float = 0.0) -> None:
        self.now = 0.0
        self.step = step
        self.sleeps: list[float] = []

    def perf_counter(self) -> float:
        self.now += self.step
        return self.now

    def sleep(self, seconds: float) -> None:
        self.sleeps.append(seconds)
        self.now += seconds


@pytest.fixture
def clock(monkeypatch):
    c = _Clock()
    monkeypatch.setattr(rc, "time", c)
    return c


@pytest.fixture
def wire(monkeypatch):
    """Replace only the socket: OpenAIResponsesClient builds a REAL openai.OpenAI."""
    w = _Wire()
    real_openai_cls = openai.OpenAI

    def _factory(**kwargs):
        w.client_kwargs = dict(kwargs)
        return real_openai_cls(
            api_key="test-key",
            http_client=httpx.Client(transport=httpx.MockTransport(w.handle)),
            **kwargs,
        )

    monkeypatch.setattr(openai, "OpenAI", _factory)
    return w


@pytest.fixture
def client(wire, clock):
    return rc.OpenAIResponsesClient()


@pytest.fixture(autouse=True)
def _isolated_breaker(monkeypatch):
    """A private circuit breaker per test. The breaker is a process-wide singleton, so
    without this a failure injected here would trip it for every other suite in the run."""
    monkeypatch.setattr(resilience, "_breaker", resilience.CircuitBreaker())


@pytest.fixture(autouse=True)
def _isolated_redis(monkeypatch):
    """A private Redis per test, and the module's connection singleton restored after —
    otherwise one test's exploding client becomes every later test's Redis."""
    monkeypatch.setattr(rs, "_client", fakeredis.FakeRedis(decode_responses=True))
    monkeypatch.setattr(rs, "_backend", "fakeredis")


@pytest.fixture(autouse=True)
def _restore_global_registries(monkeypatch):
    """The workbook version and the variable-capture singleton are module globals that
    other suites read. Snapshot them so a registry built here cannot leak."""
    from app.stores.variable_capture_registry import VariableCaptureRegistry

    monkeypatch.setattr(prompts_mod, "_CURRENT_VERSION", prompts_mod._CURRENT_VERSION)
    monkeypatch.setattr(VariableCaptureRegistry, "_instance",
                        VariableCaptureRegistry._instance, raising=False)


def _sent(wire: _Wire) -> dict:
    """The last request body the SDK actually put on the wire."""
    return wire.requests[-1]


# ═════════════════════════════════════════════════════════════════════════════
# Prompt-cache affinity: the single biggest cost lever on a turn.
# ═════════════════════════════════════════════════════════════════════════════


def test_the_cache_key_pins_a_stage_to_a_cache_slot_and_the_workbook_busts_it():
    """OpenAI only serves a cached prefix if the request lands on a machine that holds it.
    Without a key, routing is free and hits are sporadic — measured at 1 turn in 6 on a
    byte-identical 16.5K-token prompt, i.e. five turns re-paying for the whole prompt.

    The key must be stage-scoped (so every user of an agent SHARES one slot — a per-user
    key would fragment the cache and throw the cross-user reuse away) and must change when
    the workbook changes (so an edited prompt can never be served from the stale slot).
    """
    monkeypatch_version = "v-aaa111"
    prompts_mod._CURRENT_VERSION = monkeypatch_version

    key = rc._cache_key("core_coaching_agent")
    assert key == "core_coaching_agent:v-aaa111"

    # Same stage, different user/session -> same key. That sharing IS the saving.
    assert rc._cache_key("core_coaching_agent") == key
    # A different agent must not share a slot with a different system prompt.
    assert rc._cache_key("CH_coaching_agent") != key

    # A workbook reload changes the content hash -> a new slot, automatically.
    prompts_mod._CURRENT_VERSION = "v-bbb222"
    assert rc._cache_key("core_coaching_agent") == "core_coaching_agent:v-bbb222"


def test_the_cache_key_never_collapses_to_an_empty_string():
    """Before the registry has loaded (boot, or a degraded reload) the version is "".
    An empty key is not "no key" to OpenAI — every stage would then share one slot and
    serve each other's prompts. Both halves fall back to a literal instead."""
    prompts_mod._CURRENT_VERSION = ""
    assert rc._cache_key("") == "agent:dev"
    assert rc._cache_key("greeting_agent") == "greeting_agent:dev"


def test_with_caching_ON_the_prefix_is_byte_identical_across_turns(client, wire, monkeypatch):
    """The whole point of caching: the system prefix must not vary between requests, or the
    cache can never hit. It is stamped with the workbook revision so a prompt edit — and
    only a prompt edit — invalidates it."""
    monkeypatch.setattr(config, "LLM_PROMPT_CACHE_ENABLED", True)
    monkeypatch.setattr(prompts_mod, "_CURRENT_VERSION", "rev-9f9f")

    client.generate("SYSTEM", "turn one", model="gpt-5-mini", stage="core_coaching_agent")
    client.generate("SYSTEM", "turn two", model="gpt-5-mini", stage="core_coaching_agent")

    first, second = wire.requests[0], wire.requests[1]
    assert first["input"][0]["content"] == "<!-- prompt-rev:rev-9f9f -->\nSYSTEM"
    assert second["input"][0]["content"] == first["input"][0]["content"], (
        "the cached prefix changed between turns — the cache can never hit"
    )
    assert first["prompt_cache_key"] == "core_coaching_agent:rev-9f9f"


def test_with_caching_ON_but_no_workbook_version_the_prompt_is_left_unstamped(
    client, wire, monkeypatch
):
    """Before the registry loads there is nothing to stamp. Stamping an empty revision
    would still be a stable prefix, but it would pin the cache to a version that means
    nothing — so the prompt is passed through untouched instead."""
    monkeypatch.setattr(config, "LLM_PROMPT_CACHE_ENABLED", True)
    monkeypatch.setattr(prompts_mod, "_CURRENT_VERSION", "")

    client.generate("SYSTEM", "hi", model="gpt-5-mini", stage="core_coaching_agent")

    assert _sent(wire)["input"][0]["content"] == "SYSTEM"


def test_with_caching_OFF_every_request_gets_a_fresh_nonce_and_no_cache_key(client, wire):
    """The default. Caching off must mean a 0% hit rate, not "mostly off" — so each request
    carries a unique nonce, and the affinity key is omitted entirely. If the nonce ever
    stopped varying, stale replies would survive a prompt edit for 5-10 minutes and nobody
    would see an error."""
    assert config.LLM_PROMPT_CACHE_ENABLED is False, "the default must stay off"

    client.generate("SYSTEM", "one", model="gpt-5-mini", stage="core_coaching_agent")
    client.generate("SYSTEM", "two", model="gpt-5-mini", stage="core_coaching_agent")

    first = wire.requests[0]["input"][0]["content"]
    second = wire.requests[1]["input"][0]["content"]
    assert first != second, "caching is off but the prefix is stable — it will be cached"
    assert first.endswith("\nSYSTEM") and second.endswith("\nSYSTEM")
    assert "prompt_cache_key" not in wire.requests[0], (
        "a cache-affinity key was sent while caching is disabled"
    )
    # The nonce must be a real nonce, not a constant.
    assert uuid.UUID(first.split("<!-- ")[1].split(" -->")[0], version=4)


# ═════════════════════════════════════════════════════════════════════════════
# Per-stage reasoning effort — the latency lever.
# ═════════════════════════════════════════════════════════════════════════════


@pytest.mark.parametrize("stage", [
    "coaching_intake_agent", "challenge_context_agent", "repeat_user_checkin_agent",
    "action_checkin_agent", "feedback_mood_capture_agent", "simulation_decision_agent",
    "learning_aid_agent", "chat_title_agent", "greeting_agent",
])
def test_the_conversational_stages_are_all_on_the_fast_path(stage):
    """action_checkin and feedback_mood_capture were both ABSENT from this map and so fell
    to "medium" — which is omitted from the request, which means the model's own default
    reasoning. Measured cost: 31.9s -> 4.8s per turn once mapped. Every conversational
    stage must stay pinned to `minimal`, and the regression is silent: nothing errors, the
    turn is just slow."""
    assert rc.STAGE_REASONING_EFFORT[stage] == "minimal"


@pytest.mark.parametrize("stage", [
    "core_coaching_agent", "CH_coaching_agent", "role_play_agent",
    "SJT_simulation_agent", "builder",
])
def test_the_coaching_grade_stages_get_real_reasoning(stage):
    """The flip side: the stages that actually do the coaching must NOT be starved down to
    `minimal` in a latency push — 'low' is the floor for reasoning quality here."""
    assert rc.STAGE_REASONING_EFFORT[stage] == "low"


def test_every_llm_stage_is_listed_so_none_can_silently_fall_to_the_default():
    """The defect class itself: a stage absent from the table gets DEFAULT_REASONING_EFFORT
    ("medium"), which the resolver then OMITS — handing the model its own default reasoning
    and a ~24-32s turn. A new agent added to the workbook without a row here reproduces the
    incident exactly, so pin the map against the graph's stage list."""
    missing = {
        stage for stage in prompts_mod.STAGE_SHEET
        # These never make their own LLM call from a mapped stage name.
        if stage not in {"environment", "dynamic_actions_insights_agent",
                         "user_context_builder_agent", "pattern_agent"}
    } - set(rc.STAGE_REASONING_EFFORT)
    assert not missing, f"stages with no reasoning effort — they will run at the model default: {missing}"


def test_a_stage_effort_is_overridable_per_environment(monkeypatch):
    """The tuning knob promised in the module docstring (CEREBROZEN_REASONING_<STAGE>).
    If it stopped being read, an operator raising effort at 3am would change nothing."""
    monkeypatch.setenv("CEREBROZEN_REASONING_ACTION_CHECKIN", "  HIGH  ")
    assert rc._effort("ACTION_CHECKIN", "minimal") == "high"
    assert rc._effort("NOT_SET_ANYWHERE", "minimal") == "minimal"


@pytest.mark.parametrize("model,effort,expected", [
    # gpt-5-mini / o-series accept `minimal`; the gpt-5.x point releases do not.
    ("gpt-5-mini", "minimal", "minimal"),
    ("gpt-5.4",    "minimal", "low"),      # translated, not dropped
    ("gpt-5-mini", "xhigh",   "high"),     # translated the other way
    ("gpt-5.4",    "xhigh",   "xhigh"),
    ("gpt-5-mini", "none",    None),       # no equivalent -> omit rather than 400
    ("gpt-5.4",    "none",    "none"),
    ("gpt-5-mini", "medium",  None),       # medium == the model's own default -> omit
    ("gpt-5-mini", "nonsense", None),      # a typo in the env var must not 400 the turn
    ("o3-mini",    "low",     "low"),
])
def test_an_effort_valid_for_one_model_family_is_translated_for_another(
    monkeypatch, model, effort, expected
):
    """The workbook's Catalog can change a stage's model at any time, and the effort
    vocabularies differ between families. Sending `minimal` to gpt-5.4 is a 400 — the whole
    turn fails — so the resolver translates to the nearest supported tier instead of making
    the model column and the CEREBROZEN_REASONING_* vars change together."""
    monkeypatch.setitem(rc.STAGE_REASONING_EFFORT, "test_stage", effort)
    assert rc.reasoning_effort_for("test_stage", model) == expected


def test_a_non_reasoning_model_is_never_sent_a_reasoning_block(monkeypatch):
    """gpt-4o and friends reject `reasoning` outright. The model is workbook-controlled, so
    this is one Catalog edit away from being live."""
    monkeypatch.setitem(rc.STAGE_REASONING_EFFORT, "test_stage", "low")
    assert rc.reasoning_effort_for("test_stage", "gpt-4o-mini") is None
    assert rc.reasoning_effort_for("test_stage", "") is None


def test_an_unmapped_stage_omits_the_effort_rather_than_guessing():
    """The silent-slowness path, asserted directly."""
    assert rc.reasoning_effort_for("a_stage_nobody_mapped", "gpt-5-mini") is None


def test_the_revert_flag_collapses_every_stage_back_to_the_pre_optimisation_baseline(
    monkeypatch,
):
    """CEREBROZEN_STAGE_OPT=false is the rollback if the fast path ever degrades coaching
    quality. A rollback switch that doesn't switch anything is worse than none."""
    monkeypatch.setattr(rc, "STAGE_OPT", False)
    assert rc.reasoning_effort_for("coaching_intake_agent", "gpt-5-mini") == "low"
    assert rc.reasoning_effort_for("action_checkin_agent", "gpt-5-mini") == "low"


def test_the_resolved_effort_is_what_actually_reaches_the_api(client, wire):
    """Resolution is only half of it — the value has to survive into the request body."""
    effort = rc.reasoning_effort_for("coaching_intake_agent", "gpt-5-mini")
    client.generate("sys", "hi", model="gpt-5-mini", reasoning_effort=effort,
                    stage="coaching_intake_agent")
    assert _sent(wire)["reasoning"] == {"effort": "minimal"}

    client.generate("sys", "hi", model="gpt-5-mini", reasoning_effort=None, stage="x")
    assert "reasoning" not in _sent(wire), "an omitted effort must not be sent as null"


# ═════════════════════════════════════════════════════════════════════════════
# Temperature: only where the API accepts it.
# ═════════════════════════════════════════════════════════════════════════════


def test_temperature_is_never_sent_to_a_reasoning_model(monkeypatch, client, wire):
    """The Responses API 400s on `temperature` for gpt-5/o-series rather than ignoring it —
    so an operator setting CEREBROZEN_TEMP_CHALLENGE on the default model would take the
    routing stage down completely."""
    monkeypatch.setitem(rc.STAGE_TEMPERATURE, "challenge_context_agent", 0.0)

    assert rc._temperature_for("challenge_context_agent", "gpt-5-mini") is None
    client.generate("sys", "hi", model="gpt-5-mini", stage="challenge_context_agent")
    assert "temperature" not in _sent(wire)


def test_temperature_reaches_a_model_that_accepts_it(monkeypatch, client, wire):
    """The opt-in must actually work on a non-reasoning model, or the escape hatch for
    flaky routing is decorative. 0.0 must survive as 0.0 (a falsy value is easy to drop)."""
    monkeypatch.setitem(rc.STAGE_TEMPERATURE, "challenge_context_agent", 0.0)

    client.generate("sys", "hi", model="gpt-4o-mini", stage="challenge_context_agent")
    assert _sent(wire)["temperature"] == 0.0

    client.generate("sys", "hi", model="gpt-4o-mini", stage="core_coaching_agent")
    assert "temperature" not in _sent(wire), "an unconfigured stage must use the provider default"


def test_the_default_ships_with_temperature_pinning_OFF():
    """Measured and reverted: pinning the routing stage to 0 made CH routing WORSE (1/6,
    down from 3/6) because the prompt's true answer is wrong and randomness was masking it.
    Re-enabling it by default would make the wrong answer consistent."""
    assert rc.STAGE_TEMPERATURE == {}, "temperature pinning was re-enabled by default"


# ═════════════════════════════════════════════════════════════════════════════
# The rest of the request: history, output cap, structured output.
# ═════════════════════════════════════════════════════════════════════════════


def test_the_system_prompt_leads_and_history_is_replayed_in_order(client, wire):
    """A misordered or dropped history turn makes the agent lose the conversation without
    any error — it just starts repeating itself. Junk roles must be dropped, not forwarded
    (the API rejects them and the turn dies)."""
    history = [
        {"role": "user", "content": "first"},
        {"role": "assistant", "content": "second"},
        {"role": "tool", "content": "not a chat role"},   # must be dropped
        {"role": "user", "content": ""},                  # empty -> dropped
        {"role": "assistant", "content": "third"},
    ]
    client.generate("SYSTEM", "now", model="gpt-5-mini", history=history)

    roles_and_text = [(m["role"], m["content"]) for m in _sent(wire)["input"]]
    assert roles_and_text[0][0] == "system"
    assert roles_and_text[0][1].endswith("SYSTEM")
    assert roles_and_text[1:] == [
        ("user", "first"), ("assistant", "second"), ("assistant", "third"), ("user", "now"),
    ]


def test_the_output_token_cap_is_actually_sent(client, wire, monkeypatch):
    """An uncapped generation is an unbounded bill and an unbounded latency — this is the
    only thing bounding how much a single runaway turn can produce."""
    monkeypatch.setattr(config, "OPENAI_MAX_OUTPUT_TOKENS", 4096)
    client.generate("sys", "hi", model="gpt-5-mini")
    assert _sent(wire)["max_output_tokens"] == 4096

    monkeypatch.setattr(config, "OPENAI_MAX_OUTPUT_TOKENS", 0)
    client.generate("sys", "hi", model="gpt-5-mini")
    assert "max_output_tokens" not in _sent(wire), "0 means uncapped, not max_output_tokens=0"


def test_the_sdk_is_given_a_timeout_and_is_forbidden_from_retrying_behind_our_back(client, wire):
    """A 341s CIM call once broke the UI — hence the per-request timeout. And max_retries=0
    is load-bearing: if the SDK retried too, our retry loop would multiply with its own and
    a single failing turn could fire 9 paid requests."""
    assert wire.client_kwargs["max_retries"] == 0, (
        "the SDK will silently retry on top of our resilience layer"
    )
    assert wire.client_kwargs["timeout"] == config.OPENAI_TIMEOUT


def test_json_object_mode_makes_a_prose_reply_impossible(client, wire):
    """CH and friends emit a JSON envelope but DROP it for plain prose on long narrative
    turns — which loses phase/current_step/milestone and silently breaks phase routing.
    json_object mode removes the model's ability to do that. It is deliberately NOT strict
    json_schema: `context_update` is open-ended, and grammar-constrained decoding would make
    any key we forgot to enumerate literally un-emittable — coaching context vanishing on
    every turn, with no error anywhere."""
    client.generate_stream("sys", "hi", model="gpt-5-mini", json_output=True)
    assert _sent(wire)["text"] == {"format": {"type": "json_object"}}

    client.generate_stream("sys", "hi", model="gpt-5-mini", json_output=False)
    assert "text" not in _sent(wire), "structured output must stay opt-in per stage"


def test_the_streamed_request_carries_the_same_cache_key_and_caps_as_the_blocking_one(
    client, wire, monkeypatch
):
    """Streaming is the hot path — every coaching turn. If the cache key or the token cap
    were only wired into the non-streaming call, the savings and the cap would apply to
    almost nothing."""
    monkeypatch.setattr(config, "LLM_PROMPT_CACHE_ENABLED", True)
    monkeypatch.setattr(config, "OPENAI_MAX_OUTPUT_TOKENS", 2048)
    monkeypatch.setattr(prompts_mod, "_CURRENT_VERSION", "rev-1")
    monkeypatch.setitem(rc.STAGE_TEMPERATURE, "challenge_context_agent", 0.3)

    client.generate_stream("sys", "hi", model="gpt-4o-mini", reasoning_effort=None,
                           stage="challenge_context_agent")

    body = _sent(wire)
    assert body["prompt_cache_key"] == "challenge_context_agent:rev-1"
    assert body["max_output_tokens"] == 2048
    assert body["temperature"] == 0.3
    assert body["stream"] is True


# ═════════════════════════════════════════════════════════════════════════════
# Telemetry off the response: tokens, cache hits, cost, latency.
# ═════════════════════════════════════════════════════════════════════════════


def test_a_cache_hit_is_visible_in_the_returned_telemetry_and_costs_less(client, wire):
    """cached_tokens is the ONLY evidence the prompt cache is working. If it stopped being
    read off the usage block, a cache that silently stopped hitting would look identical to
    one that never stopped — and the ~48% saving would evaporate unnoticed."""
    wire.usage = {"input_tokens": 16500, "output_tokens": 200,
                  "total_tokens": 16700, "cached_tokens": 16000}
    hit = client.generate("sys", "hi", model="gpt-5-mini", stage="core_coaching_agent")

    assert hit.cached_tokens == 16000
    assert hit.prompt_tokens == 16500
    assert hit.completion_tokens == 200
    assert hit.total_tokens == 16700
    assert hit.text == wire.text
    assert hit.model == "gpt-5-mini"

    wire.usage = {"input_tokens": 16500, "output_tokens": 200,
                  "total_tokens": 16700, "cached_tokens": 0}
    miss = client.generate("sys", "hi", model="gpt-5-mini", stage="core_coaching_agent")

    assert miss.cached_tokens == 0
    assert hit.cost_usd < miss.cost_usd, (
        "cached tokens were billed at the full rate — the cache saves nothing"
    )


def test_a_response_with_no_usage_block_still_returns_the_reply(client, wire):
    """Some errors and some models return no usage. Reading it defensively is the
    difference between a coaching reply and an AttributeError in the middle of a turn."""
    wire.script = [lambda body: httpx.Response(200, json={
        "id": "r", "object": "response", "created_at": 1, "model": "gpt-5-mini",
        "status": "completed", "parallel_tool_calls": False, "tool_choice": "auto",
        "tools": [],
        "output": [{"id": "m", "type": "message", "role": "assistant", "status": "completed",
                    "content": [{"type": "output_text", "text": "still here",
                                 "annotations": []}]}],
    })]
    resp = client.generate("sys", "hi", model="gpt-5-mini")

    assert resp.text == "still here"
    assert (resp.prompt_tokens, resp.completion_tokens, resp.cached_tokens) == (0, 0, 0)
    assert resp.cost_usd == 0.0


def test_latency_is_measured_over_the_whole_call(wire, monkeypatch):
    """A turn with no latency number is a turn nobody can debug. The clock is stepped 0.5s
    per read, so the measured span is exactly one step."""
    c = _Clock(step=0.5)
    monkeypatch.setattr(rc, "time", c)
    resp = rc.OpenAIResponsesClient().generate("sys", "hi", model="gpt-5-mini")
    assert resp.model_latency_ms == 500.0


def test_the_request_and_the_reply_are_logged_with_the_session_that_produced_them(
    client, wire, caplog, monkeypatch
):
    """Content logging is how a bad coaching turn is reconstructed after the fact. The
    session_id has to be on the line or the log is a pile of anonymous prompts."""
    monkeypatch.setattr(config, "CEREBROZEN_LLM_LOG_CONTENT", True)
    monkeypatch.setattr(config, "CEREBROZEN_LLM_LOG_CONTENT_CHARS", 0)
    caplog.set_level("INFO", logger="cerebrozen.llm")

    client.generate("SYSTEM-PROMPT", "the user said this", model="gpt-5-mini",
                    stage="core_coaching_agent", session_id="sess-77", user_id="u-9")

    req = next(r for r in caplog.records if r.message == "llm.request")
    assert req.session_id == "sess-77"
    assert req.user_message == "the user said this"
    assert req.system_prompt.endswith("SYSTEM-PROMPT")
    assert req.stream is False

    resp = next(r for r in caplog.records if r.message == "openai.response")
    assert (resp.session_id, resp.user_id) == ("sess-77", "u-9")
    assert resp.response_text == wire.text
    assert resp.response_chars == len(wire.text)
    assert resp.cached_tokens == 0


def test_logged_content_is_truncated_when_a_limit_is_configured(client, wire, caplog, monkeypatch):
    """PII in CloudWatch is a liability and a 16K-token prompt per turn is a bill. The
    truncation limit is the control for both, and it is silently useless if it is not
    applied to BOTH the request and the response."""
    monkeypatch.setattr(config, "CEREBROZEN_LLM_LOG_CONTENT", True)
    monkeypatch.setattr(config, "CEREBROZEN_LLM_LOG_CONTENT_CHARS", 8)
    caplog.set_level("INFO", logger="cerebrozen.llm")
    wire.text = "a very long assistant reply indeed"

    client.generate("sys", "the user said something long", model="gpt-5-mini")

    req = next(r for r in caplog.records if r.message == "llm.request")
    assert req.user_message == "the user…"
    resp = next(r for r in caplog.records if r.message == "openai.response")
    assert resp.response_text == "a very l…"
    assert resp.response_chars == len(wire.text), "the true length must survive truncation"


def test_content_logging_can_be_switched_off_entirely(client, wire, caplog, monkeypatch):
    """The privacy switch. If it leaked content anyway, a customer with a no-PII-in-logs
    clause would be in breach and nothing would look broken."""
    monkeypatch.setattr(config, "CEREBROZEN_LLM_LOG_CONTENT", False)
    caplog.set_level("INFO", logger="cerebrozen.llm")

    client.generate("sys", "confidential", model="gpt-5-mini")
    client.generate_stream("sys", "confidential", model="gpt-5-mini")

    assert not [r for r in caplog.records if r.message == "llm.request"]
    for rec in (r for r in caplog.records if r.message == "openai.response"):
        assert not hasattr(rec, "response_text"), "content was logged with logging disabled"
        assert rec.total_tokens is not None, "telemetry must survive the privacy switch"


# ═════════════════════════════════════════════════════════════════════════════
# Streaming.
# ═════════════════════════════════════════════════════════════════════════════


def test_streamed_deltas_reach_the_caller_as_they_arrive(client, wire, caplog, monkeypatch):
    """The user watches the reply appear. Buffering it (or dropping a delta) is the
    difference between a coach and a 20-second blank screen."""
    monkeypatch.setattr(config, "CEREBROZEN_LLM_LOG_CONTENT", True)
    monkeypatch.setattr(config, "CEREBROZEN_LLM_LOG_CONTENT_CHARS", 0)
    caplog.set_level("INFO", logger="cerebrozen.llm")
    wire.deltas = ["Let ", "me ", "ask ", "you"]
    wire.usage = {"input_tokens": 900, "output_tokens": 12,
                  "total_tokens": 912, "cached_tokens": 800}

    got: list[str] = []
    resp = client.generate_stream("sys", "hi", model="gpt-5-mini", on_token=got.append,
                                  stage="core_coaching_agent", session_id="s-1")

    assert got == ["Let ", "me ", "ask ", "you"], "tokens were coalesced or lost"
    assert resp.text == "Let me ask you"
    assert resp.cached_tokens == 800, "the stream's cache telemetry was dropped"
    assert resp.total_tokens == 912
    assert resp.cost_usd > 0

    logged = next(r for r in caplog.records
                  if r.message == "openai.response" and r.stream is True)
    assert logged.response_text == "Let me ask you"
    assert logged.session_id == "s-1"


def test_a_stream_with_no_on_token_still_returns_the_full_text(client, wire):
    """Background stages (title, insights) stream with no sink. Dropping the accumulated
    text there would silently produce empty titles."""
    wire.deltas = ["Career ", "goals"]
    resp = client.generate_stream("sys", "hi", model="gpt-5-mini", on_token=None)
    assert resp.text == "Career goals"


def test_the_streaming_path_sends_the_reasoning_effort_too(client, wire):
    """Streaming IS the coaching path. An effort that only reached the blocking call would
    leave every real turn running at the model's default reasoning — the 24-32s defect,
    reintroduced everywhere it actually matters."""
    client.generate_stream("sys", "hi", model="gpt-5-mini", reasoning_effort="minimal",
                           stage="action_checkin_agent")
    assert _sent(wire)["reasoning"] == {"effort": "minimal"}


def test_an_uncapped_stream_omits_the_token_cap(client, wire, monkeypatch):
    """0 means uncapped on the streaming path as well; sending max_output_tokens=0 would be
    a 400 and every coaching turn would fail."""
    monkeypatch.setattr(config, "OPENAI_MAX_OUTPUT_TOKENS", 0)
    client.generate_stream("sys", "hi", model="gpt-5-mini")
    assert "max_output_tokens" not in _sent(wire)


def test_empty_deltas_and_a_missing_usage_block_do_not_break_a_stream(client, wire):
    """Keep-alive frames arrive as empty deltas, and a stream can complete with no usage
    block at all. Either one raising would kill a turn whose reply the user has already
    half-read."""
    wire.deltas = ["Say ", "", "more"]
    wire.script = [lambda body: httpx.Response(
        200, headers={"content-type": "text/event-stream"}, content=_sse([
            ("response.created", {"response": {
                "id": "r", "object": "response", "created_at": 1, "model": "gpt-5-mini",
                "status": "in_progress", "output": [], "parallel_tool_calls": False,
                "tool_choice": "auto", "tools": [], "usage": None}}),
            ("response.output_item.added", {"output_index": 0, "item": {
                "id": "msg_1", "type": "message", "role": "assistant",
                "status": "in_progress", "content": []}}),
            ("response.content_part.added", {
                "item_id": "msg_1", "output_index": 0, "content_index": 0,
                "part": {"type": "output_text", "text": "", "annotations": []}}),
            ("response.output_text.delta", {
                "item_id": "msg_1", "output_index": 0, "content_index": 0,
                "delta": "Say ", "logprobs": []}),
            ("response.output_text.delta", {
                "item_id": "msg_1", "output_index": 0, "content_index": 0,
                "delta": "", "logprobs": []}),          # keep-alive
            ("response.output_text.delta", {
                "item_id": "msg_1", "output_index": 0, "content_index": 0,
                "delta": "more", "logprobs": []}),
            ("response.completed", {"response": {
                "id": "r", "object": "response", "created_at": 1, "model": "gpt-5-mini",
                "status": "completed", "parallel_tool_calls": False, "tool_choice": "auto",
                "tools": [],
                "output": [{"id": "msg_1", "type": "message", "role": "assistant",
                            "status": "completed",
                            "content": [{"type": "output_text", "text": "Say more",
                                         "annotations": []}]}],
                "usage": None}}),
        ]))]

    got: list[str] = []
    resp = client.generate_stream("sys", "hi", model="gpt-5-mini", on_token=got.append)

    assert resp.text == "Say more"
    assert got == ["Say ", "more"], "an empty keep-alive delta was forwarded as a token"
    assert (resp.prompt_tokens, resp.cached_tokens, resp.cost_usd) == (0, 0, 0.0)


def test_a_stream_that_runs_past_the_deadline_is_killed_loudly(wire, monkeypatch):
    """LIVE INCIDENT 2026-07-06: an action_checkin stream never completed and never errored.
    OPENAI_TIMEOUT is httpx's gap-BETWEEN-chunks timeout, so a generation that keeps
    drip-feeding tokens can run forever inside it — no error, no telemetry, a wedged turn.

    The wall-clock deadline is the only thing that ends it, and it must raise (so the
    resilience/worker error path sees it) rather than return a truncated reply as if
    nothing happened. The failure has to name the stage and how far it got, or the next
    incident is just as blind as the last one.
    """
    # 40s of wall clock per stream event: the deadline lands between the 1st and 2nd
    # text delta, so the raise happens with tokens already streamed.
    c = _Clock(step=40.0)
    monkeypatch.setattr(rc, "time", c)
    monkeypatch.setattr(config, "OPENAI_STREAM_DEADLINE_S", 180.0)
    wire.deltas = ["tick ", "tock ", "forever"]

    got: list[str] = []
    client = rc.OpenAIResponsesClient()
    with pytest.raises(TimeoutError) as exc:
        client._generate_stream_once(
            "sys", "hi", model="gpt-5-mini", on_token=got.append,
            reasoning_effort=None, history=None, stage="action_checkin_agent",
        )

    msg = str(exc.value)
    assert "OPENAI_STREAM_DEADLINE_S=180s" in msg
    assert "action_checkin_agent" in msg, "the incident log must name the stage that hung"
    assert "chars_so_far=5" in msg, "the deadline must report how far the stream got"
    assert got == ["tick "], "the deadline fired late — tokens kept flowing past it"


def test_the_deadline_can_be_disabled_without_disabling_streaming(wire, monkeypatch):
    """0 means off. An operator who needs a genuinely long generation must be able to turn
    the wall clock off without the stream itself breaking."""
    c = _Clock(step=100.0)
    monkeypatch.setattr(rc, "time", c)
    monkeypatch.setattr(config, "OPENAI_STREAM_DEADLINE_S", 0)
    wire.deltas = ["a", "b", "c"]

    resp = rc.OpenAIResponsesClient()._generate_stream_once(
        "sys", "hi", model="gpt-5-mini", on_token=None, reasoning_effort=None, history=None,
    )
    assert resp.text == "abc"


# ═════════════════════════════════════════════════════════════════════════════
# Retry, model cascade, circuit breaker — composed.
# ═════════════════════════════════════════════════════════════════════════════


def test_a_transient_upstream_error_is_retried_on_the_same_model_with_backoff(
    client, wire, clock
):
    """A 500 from OpenAI is a blip, not a coaching failure. Retrying without backoff is a
    stampede; not retrying at all turns every blip into a failed turn for a real user."""
    wire.script = [500, 500]        # then success
    resp = client.generate("sys", "hi", model="gpt-5-mini")

    assert resp.text == wire.text, "a retryable blip surfaced to the user"
    assert wire.models_tried == ["gpt-5-mini"] * 3, "it cascaded instead of retrying"
    assert len(clock.sleeps) == 2, "retries fired with no backoff between them"
    assert clock.sleeps[0] <= config.LLM_BACKOFF_BASE_S
    assert all(s >= 0 for s in clock.sleeps)
    assert resilience.get_breaker().state == "closed"


def test_retries_are_exhausted_before_the_cascade_falls_to_the_next_model(
    client, wire, clock
):
    """The cascade is the second line of defence, not the first: falling to a weaker model
    on the first blip would quietly degrade coaching quality for a hiccup that a retry
    would have fixed."""
    wire.script = [500, 500, 500]   # gpt-5-mini exhausts its retries, then fall over
    resp = client.generate("sys", "hi", model="gpt-5-mini")

    assert resp.text == wire.text
    assert wire.models_tried == ["gpt-5-mini", "gpt-5-mini", "gpt-5-mini", "gpt-5-nano"], (
        "the cascade did not preserve retry-then-fall-over ordering"
    )
    assert resp.model == "gpt-5-nano", "the response must report the model that ANSWERED"


def test_a_non_retryable_error_fails_immediately_instead_of_burning_quota(
    client, wire, clock
):
    """A 401 or a 400 will fail identically every time. Retrying it three times across two
    models turns one broken config into six paid, doomed requests — and delays the loud
    failure an operator needs to see."""
    wire.script = [401, 401, 401, 401, 401, 401]

    with pytest.raises(openai.AuthenticationError):
        client.generate("sys", "hi", model="gpt-5-mini")

    assert wire.models_tried == ["gpt-5-mini"], "a permanent failure was retried"
    assert clock.sleeps == []


def test_the_breaker_short_circuits_when_openai_is_persistently_down(client, wire, clock):
    """When OpenAI is down, hammering it is not resilience — it is a slow outage for every
    user at once. The breaker fails the turn FAST so the caller degrades to a safe reply,
    and it must do so WITHOUT making a request."""
    breaker = resilience.get_breaker()
    for _ in range(breaker.fail_threshold):
        breaker.record_failure()
    assert breaker.state == "open"

    before = len(wire.requests)
    with pytest.raises(resilience.BreakerOpen):
        client.generate("sys", "hi", model="gpt-5-mini")

    assert len(wire.requests) == before, "the breaker was open but a request was still sent"


def test_a_call_that_exhausts_every_model_records_a_failure_and_re_raises_the_last_error(
    client, wire, clock
):
    """The whole cascade failing is the signal the breaker is counting. If the failure were
    swallowed here the breaker would never open and a real outage would just look like slow
    coaching, forever."""
    wire.script = [503] * 12
    breaker = resilience.get_breaker()

    with pytest.raises(openai.APIStatusError):
        client.generate("sys", "hi", model="gpt-5-mini")

    assert wire.models_tried == ["gpt-5-mini"] * 3 + ["gpt-5-nano"] * 3, (
        "every model must be given its full retry budget before the call is abandoned"
    )
    assert breaker._consecutive_failures == 1, (
        "one failed turn must count as ONE breaker failure, not one per HTTP attempt"
    )


def test_a_call_with_no_model_to_try_fails_loudly_instead_of_returning_nothing(
    client, wire, clock, monkeypatch
):
    """A blank model and an empty cascade means the loop body never runs — there is no
    exception to re-raise, and an early version of this would have fallen off the end and
    returned None into the graph. A node receiving None as an LLMResponse fails somewhere
    far away from the cause."""
    monkeypatch.setattr(config, "MODEL_CASCADE", [])

    with pytest.raises(RuntimeError, match="no exception captured"):
        client.generate("sys", "hi", model="")

    assert wire.requests == [], "a request was sent with no model"


def test_a_successful_call_clears_the_breakers_failure_count(client, wire, clock):
    """Consecutive failures only. A single success means OpenAI is back, and a breaker that
    kept counting would trip on a handful of unrelated blips spread over an hour."""
    breaker = resilience.get_breaker()
    breaker.record_failure()
    breaker.record_failure()

    client.generate("sys", "hi", model="gpt-5-mini")

    assert breaker._consecutive_failures == 0


def test_a_failure_BEFORE_the_first_token_is_retried_transparently(client, wire, clock):
    """Nothing has reached the user yet, so a retry is invisible and free. Refusing to retry
    here would surface a blip that the user never needed to see."""
    wire.script = [500]
    got: list[str] = []
    resp = client.generate_stream("sys", "hi", model="gpt-5-mini", on_token=got.append)

    assert resp.text == "hello"
    assert got == ["hel", "lo"], "the user saw duplicated or missing text"
    assert len(wire.requests) == 2


def test_a_failure_AFTER_the_first_token_is_never_retried(client, wire, clock, caplog):
    """THE streaming safety rule. Once text is on the user's screen, a retry re-streams the
    reply from the start — the user watches the coach say the first half of a sentence
    twice, then something different. Garbled output is worse than a clean failure, so a
    mid-stream failure must raise and stop.
    """
    def _die_midstream(body):
        # A well-formed stream that ends after one delta, with no `response.completed` —
        # exactly what a dropped connection looks like to the SDK.
        return httpx.Response(200, headers={"content-type": "text/event-stream"},
                              content=_sse([
                                  ("response.created", {"response": {
                                      "id": "r", "object": "response", "created_at": 1,
                                      "model": "gpt-5-mini", "status": "in_progress",
                                      "output": [], "parallel_tool_calls": False,
                                      "tool_choice": "auto", "tools": [], "usage": None}}),
                                  ("response.output_item.added", {"output_index": 0, "item": {
                                      "id": "msg_1", "type": "message", "role": "assistant",
                                      "status": "in_progress", "content": []}}),
                                  ("response.content_part.added", {
                                      "item_id": "msg_1", "output_index": 0, "content_index": 0,
                                      "part": {"type": "output_text", "text": "",
                                               "annotations": []}}),
                                  ("response.output_text.delta", {
                                      "item_id": "msg_1", "output_index": 0,
                                      "content_index": 0, "delta": "I hear ", "logprobs": []}),
                              ]))

    wire.script = [_die_midstream]
    caplog.set_level("ERROR", logger="cerebrozen.llm")
    got: list[str] = []

    with pytest.raises(Exception):
        client.generate_stream("sys", "hi", model="gpt-5-mini", on_token=got.append)

    assert got == ["I hear "]
    assert len(wire.requests) == 1, (
        "a stream was retried after tokens reached the user — the reply will be garbled"
    )
    assert "openai.fail_midstream" in caplog.text
    assert resilience.get_breaker()._consecutive_failures == 1


# ═════════════════════════════════════════════════════════════════════════════
# model_for + tracing wiring.
# ═════════════════════════════════════════════════════════════════════════════


def test_the_workbook_catalog_is_the_only_source_of_the_model(monkeypatch):
    """A default model hidden in code is how an agent ends up silently running on the wrong
    model after a Catalog typo — its JSON contract then breaks in production, not here."""
    monkeypatch.delenv("CEREBROZEN_MODEL_OVERRIDE", raising=False)
    assert rc.model_for("core_coaching_agent", "gpt-5-mini") == "gpt-5-mini"

    with pytest.raises(RuntimeError, match="core_coaching_agent"):
        rc.model_for("core_coaching_agent", None)


def test_the_env_override_pins_one_model_for_every_stage(monkeypatch):
    """The escape hatch for a workbook full of placeholder ids (gpt-5.4). It has to beat the
    Catalog, or it cannot rescue a broken workbook."""
    monkeypatch.setenv("CEREBROZEN_MODEL_OVERRIDE", "gpt-4o-mini")
    assert rc.model_for("core_coaching_agent", "gpt-5.4") == "gpt-4o-mini"
    assert rc.model_for("any_stage", None) == "gpt-4o-mini"


@pytest.mark.parametrize("var", ["LANGSMITH_TRACING", "LANGCHAIN_TRACING_V2"])
@pytest.mark.parametrize("value,expected", [
    ("true", True), ("1", True), ("YES", True), ("false", False), ("", False),
])
def test_tracing_is_enabled_by_either_env_flag(monkeypatch, var, value, expected):
    """Two flag names because LangChain renamed it. Honouring only one means tracing
    silently stays off for half the deployments that asked for it."""
    for name in ("LANGSMITH_TRACING", "LANGCHAIN_TRACING_V2"):
        monkeypatch.delenv(name, raising=False)
    monkeypatch.setenv(var, value)
    assert rc._tracing_enabled() is expected


def test_with_tracing_on_the_openai_client_is_wrapped_for_langsmith(wire, monkeypatch, caplog):
    """Tracing wraps the SDK so each call is an LLM span under the graph-node span. The
    wrapper must not change the client we then use — the constitution keeps the raw
    Responses API — so prove the wrapped client still makes a real, correct call."""
    monkeypatch.setenv("LANGSMITH_TRACING", "true")
    monkeypatch.setenv("LANGCHAIN_TRACING_V2", "")
    monkeypatch.setenv("LANGCHAIN_API_KEY", "ls-test")
    caplog.set_level("INFO", logger="cerebrozen.llm")

    client = rc.OpenAIResponsesClient()
    resp = client.generate("sys", "hi", model="gpt-5-mini")

    assert "langsmith.wrap_openai" in caplog.text
    assert resp.text == wire.text, "the langsmith wrapper broke the Responses call"


def test_a_broken_langsmith_install_degrades_to_an_unwrapped_client(wire, monkeypatch, caplog):
    """Observability is not worth an outage. If the wrapper cannot be imported, the turn
    must still run — untraced."""
    monkeypatch.setenv("LANGSMITH_TRACING", "true")
    monkeypatch.setitem(sys.modules, "langsmith.wrappers", None)
    caplog.set_level("WARNING", logger="cerebrozen.llm")

    resp = rc.OpenAIResponsesClient().generate("sys", "hi", model="gpt-5-mini")

    assert "langsmith.wrap_failed" in caplog.text
    assert resp.text == wire.text, "a failed tracing import took the LLM call down with it"


# ═════════════════════════════════════════════════════════════════════════════
# UserTextStreamer — what the user is allowed to see.
# ═════════════════════════════════════════════════════════════════════════════


def _stream_into(chunks: list[str]) -> str:
    out: list[str] = []
    s = rc.UserTextStreamer(out.append)
    for c in chunks:
        s.feed(c)
    return "".join(out)


@pytest.mark.parametrize("key", [
    "response_to_user", "next_question", "clarifying_question", "message", "response",
])
def test_the_user_facing_field_is_extracted_from_the_json_envelope(key):
    """Agents reply with a JSON envelope; only one field of it is the coaching message. Emit
    the wrong one (or the whole blob) and the user reads backend control fields."""
    payload = f'{{"{key}": "What is on your mind?", "handoff_ready": false}}'
    assert _stream_into([payload]) == "What is on your mind?"


def test_the_reply_streams_character_by_character_across_arbitrary_chunk_boundaries():
    """The deltas arrive wherever the network cuts them — including mid-key and mid-escape.
    A streamer that only works on whole tokens produces garbled text under load."""
    payload = '{"next_question": "Tell me more.", "handoff_ready": false}'
    for size in (1, 2, 3, 7, 13):
        chunks = [payload[i:i + size] for i in range(0, len(payload), size)]
        assert _stream_into(chunks) == "Tell me more.", f"broke at chunk size {size}"


def test_json_escapes_are_unescaped_before_the_user_sees_them():
    """Raw \\n and \\u2019 in a chat bubble look like the coach is leaking code."""
    payload = (
        r'{"message": "Line one\nLine two\ttabbed’s \"quoted\" \\ back\/slash", "x": 1}'
    )
    assert _stream_into([payload]) == 'Line one\nLine two\ttabbed’s "quoted" \\ back/slash'


def test_a_unicode_escape_split_across_two_deltas_is_not_mangled():
    """\\u2019 is six characters and the network will happily cut it in half. Emitting the
    fragment produces a broken glyph mid-word."""
    payload = '{"message": "it\\u2019s fine"}'
    assert _stream_into([payload[:16], payload[16:]]) == "it’s fine"
    assert _stream_into(list(payload)) == "it’s fine"


def test_a_malformed_unicode_escape_degrades_instead_of_crashing():
    """A model emitting \\uZZZZ must not take the turn down — the reply is already streaming
    to the user, so a ValueError here is a half-sent message and a 500.

    It degrades lossily: the bad escape and its four hex characters are dropped and the
    stream carries on. That is the deliberate trade (a mangled word beats a dead turn), and
    it is pinned here because the alternative — raising — is the thing that must never
    come back."""
    assert _stream_into([r'{"message": "bad \uZZZZ escape"}']) == "bad u escape"


def test_leading_whitespace_alone_never_decides_the_mode():
    """A model that opens with a newline before its JSON would be misread as a plain-text
    reply, and the whole envelope would be streamed to the user as prose."""
    assert _stream_into(["  \n ", '{"message": "hi"}']) == "hi"


def test_a_plain_prose_reply_is_forwarded_as_is(caplog):
    """Not every agent honours the envelope. Emitting nothing when the JSON key is absent
    would leave the user staring at an empty bubble."""
    caplog.set_level("WARNING", logger="cerebrozen.llm")
    assert _stream_into(["Hello, ", "how are you?"]) == "Hello, how are you?"
    assert "streamer.raw_mode" in caplog.text


def test_a_fenced_json_envelope_is_unwrapped_instead_of_leaking_the_fence():
    """Without this the first backtick flips the streamer into raw mode and ```json is
    printed into the chat bubble ahead of the reply."""
    payload = '```json\n{"next_question": "Where shall we start?"}\n```'
    assert _stream_into([payload]) == "Where shall we start?"
    # The fence header can arrive on its own delta, before any JSON exists to inspect.
    assert _stream_into(["```json", "\n", '{"message": "ok"}']) == "ok"
    assert _stream_into(["```", '{"message": "ok"}']) == "ok"


def test_a_fence_around_NON_json_is_treated_as_prose(caplog):
    """A model that fences a plain answer (```\\nSome advice```) is still trying to talk to
    the user. Suppressing it would blank the bubble."""
    caplog.set_level("WARNING", logger="cerebrozen.llm")
    out = _stream_into(["```\nSome advice for you"])
    assert "Some advice for you" in out
    reasons = [getattr(r, "reason", None) for r in caplog.records
               if r.message == "streamer.raw_mode"]
    assert "code_fence_non_json" in reasons


def test_a_trailing_FENCED_control_block_is_suppressed_mid_stream(caplog):
    """Round-2 bug: a model answers in prose, then appends a ```json control envelope. Raw
    mode forwarded every later delta verbatim, so the fence and the backend JSON appeared
    in the user's chat bubble."""
    caplog.set_level("WARNING", logger="cerebrozen.llm")
    out = _stream_into(['That sounds hard.\n```json\n{"handoff_ready": true}\n```'])
    assert out == "That sounds hard."
    assert "streamer.raw_control_suppressed" in caplog.text


def test_a_trailing_BARE_control_envelope_is_suppressed(caplog):
    """The action_checkin final-turn incident: prose, then a bare {"...": ...} envelope with
    no fence at all. The user watched backend JSON flash on screen."""
    caplog.set_level("WARNING", logger="cerebrozen.llm")
    out = _stream_into(['Nice work today.\n{"handoff_ready": true, "agent": "x"}'])
    assert out == "Nice work today."


def test_a_bare_envelope_that_leads_with_a_DATA_field_is_also_suppressed():
    """The regression that reopened the bug: the suppressor originally matched only KNOWN
    control keys, so an envelope starting with a data field ({"action_item": ...}) sailed
    through and leaked. Any `{"key":` is a JSON object start, not prose."""
    out = _stream_into(['Well done.\n{"action_item": "call Sam", "handoff_ready": true}'])
    assert out == "Well done."


def test_an_envelope_split_across_a_chunk_boundary_never_leaks_its_first_characters():
    """The half-arrived envelope: "...today.\\n{" lands in one delta and '"action_item":' in
    the next. If the trailing `{` were flushed as prose the moment it arrived, the user would
    see a stray brace — and the suppressor would have already lost its chance to hide what
    follows it. The tail is held back until the next delta confirms or denies it.

    (The blank line separating prose from the envelope does survive when the split lands
    exactly here — the rstrip that removes it only sees a single buffer. Cosmetic; the
    envelope itself, which is the incident, cannot get out.)"""
    out = _stream_into(['All set for today.\n{', '"action_item": "call Sam"}'])
    assert out.strip() == "All set for today."
    assert "{" not in out and "action_item" not in out, "a JSON envelope leaked to the user"

    # And the mirror case: a lone brace that turns out NOT to be an envelope must be
    # released as prose, not swallowed.
    out = _stream_into(["Try the {plan} ", "we discussed."])
    assert out == "Try the {plan} we discussed."


def test_a_newline_inside_a_prose_reply_is_held_back_and_then_released():
    """A "\\n" is the first character of the "\\n```" fence trigger, so raw mode has to hold
    it back until the next delta proves it is not one. It must then be RELEASED — a streamer
    that swallowed newlines would run every paragraph of the coach's reply together."""
    assert _stream_into(["First point.", "\n", "Second point."]) == (
        "First point.\nSecond point.")

    # The same held-back tail at the very end of a stream is simply never released — it
    # could still have been a fence, and a trailing newline is not worth the risk.
    assert _stream_into(["Done.", "\n"]) == "Done."


def test_prose_that_merely_MENTIONS_a_brace_is_preserved():
    """Over-suppression is its own bug: silently truncating a legitimate reply at the first
    `{` would drop half the coach's message with no error."""
    assert _stream_into(["Use the {goal} template."]) == "Use the {goal} template."


def test_nothing_is_emitted_when_no_user_facing_key_ever_appears():
    """A control-only envelope has nothing to say to the user. Guessing a field would put
    raw state into the chat; the authoritative reply is the parsed payload anyway."""
    assert _stream_into(['{"handoff_ready": true, "phase": "closing"}']) == ""


def test_the_streamer_stops_at_the_closing_quote_and_ignores_the_rest():
    """Everything after the user-facing value is control state. Streaming it would print
    the envelope's internals into the bubble."""
    payload = '{"next_question": "Ready?", "handoff_ready": true, "phase": "wrap"}'
    assert _stream_into([payload]) == "Ready?"
    # ...and a delta arriving after the value has closed changes nothing.
    s_out: list[str] = []
    s = rc.UserTextStreamer(s_out.append)
    s.feed(payload)
    s.feed('{"more": "junk"}')
    assert "".join(s_out) == "Ready?"


def test_an_empty_delta_is_a_no_op():
    """Keep-alive frames arrive as empty deltas; treating one as end-of-stream would
    truncate the reply."""
    s_out: list[str] = []
    s = rc.UserTextStreamer(s_out.append)
    s.feed("")
    s.feed('{"message": "still here"}')
    assert "".join(s_out) == "still here"


def test_a_streamer_with_no_sink_never_raises():
    """The nodes construct one before they know whether anyone is listening."""
    s = rc.UserTextStreamer(None)
    s.feed('{"message": "hi"}')
    s.feed("plain prose")


# ═════════════════════════════════════════════════════════════════════════════
# Redis hot-state: the per-session turn lock.
#
# The contract in one line: a Redis hiccup must NEVER raise into a turn.
# ═════════════════════════════════════════════════════════════════════════════


class _ExplodingRedis:
    """A Redis that is up enough to be called and broken enough to fail every call — the
    realistic outage (a failover, a maxmemory-OOM, a severed link), not a clean absence."""

    def __init__(self, exc: Exception | None = None) -> None:
        self.exc = exc or ConnectionError("redis is on fire")

    def _boom(self, *a, **kw):
        raise self.exc

    set = get = delete = exists = incr = expire = eval = ping = _boom


def test_the_lock_serialises_turns_on_one_session(monkeypatch):
    """Art. 8.4. Two turns on one session race the checkpoint: the second write clobbers the
    first, and the user loses a turn of coaching context with nothing in any log. Only one
    turn may hold a session at a time."""
    monkeypatch.setattr(config, "REDIS_LOCK_WAIT_MS", 0)

    with rs.session_lock("sess-1"):
        with pytest.raises(rs.SessionBusyError, match="sess-1"):
            with rs.session_lock("sess-1"):
                pytest.fail("two turns held the same session at once")

        # A DIFFERENT session is unaffected — the lock must not become a global mutex.
        with rs.session_lock("sess-2"):
            pass

    # Released on exit: the next turn gets straight in.
    with rs.session_lock("sess-1"):
        pass


def test_the_lock_is_released_even_when_the_turn_blows_up(monkeypatch):
    """A crashed turn that keeps its lock wedges the session until the TTL expires — the
    user retries, gets "another turn is in progress", and is stuck for two minutes."""
    monkeypatch.setattr(config, "REDIS_LOCK_WAIT_MS", 0)

    with pytest.raises(ValueError):
        with rs.session_lock("sess-crash"):
            raise ValueError("the graph exploded")

    with rs.session_lock("sess-crash"):
        pass  # not wedged


def test_a_contended_lock_waits_and_then_proceeds(monkeypatch):
    """The second turn is meant to BLOCK briefly, not fail instantly — a user double-tapping
    Send should wait, not get an error."""
    class _Clock2:
        def __init__(self):
            self.t = 0.0
            self.sleeps: list[float] = []

        def monotonic(self):
            return self.t

        def sleep(self, s):
            self.sleeps.append(s)
            self.t += s

    clock2 = _Clock2()
    monkeypatch.setattr(rs, "time", clock2)
    monkeypatch.setattr(config, "REDIS_LOCK_WAIT_MS", 30000)

    real = rs.get_redis()
    busy = {"n": 2}

    class _Contended:
        def set(self, *a, **kw):
            if busy["n"] > 0:          # held by another turn for two polls
                busy["n"] -= 1
                return None
            return real.set(*a, **kw)

        def __getattr__(self, name):
            return getattr(real, name)

    monkeypatch.setattr(rs, "_client", _Contended())

    with rs.session_lock("sess-wait"):
        pass

    assert clock2.sleeps == [0.05, 0.05], "the lock spun hot instead of backing off"


def test_a_redis_that_throws_lets_the_turn_through(monkeypatch, caplog):
    """THE degrade-safe property. Redis holds hot, regenerable state; Mongo is the durable
    truth. A Redis outage must therefore cost us the LOCK, not the session — the turn
    proceeds unlocked (worst case: the pre-Redis behaviour) rather than raising into the
    user's face."""
    monkeypatch.setattr(rs, "_client", _ExplodingRedis())
    caplog.set_level("WARNING", logger="cerebrozen.redis")

    ran = False
    with rs.session_lock("sess-x"):
        ran = True

    assert ran, "a Redis failure raised into the turn instead of degrading open"
    assert "redis.lock_error_degrade_open" in caplog.text


def test_no_redis_at_all_also_lets_the_turn_through(monkeypatch):
    """Local dev with no Redis, and the "not even fakeredis" case. Neither may gate coaching."""
    monkeypatch.setattr(rs, "_client", None)
    monkeypatch.setattr(rs, "get_redis", lambda: None)

    ran = []
    with rs.session_lock("sess-y"):
        ran.append(True)
    with rs.session_lock(""):          # no session id -> nothing to lock
        ran.append(True)
    assert ran == [True, True]


def test_a_turn_only_ever_releases_its_OWN_lock(monkeypatch):
    """Compare-and-delete. If our lock TTL expires mid-turn and a SECOND turn acquires the
    session, our `finally` must not delete the second turn's lock — that would put two turns
    into the session simultaneously, which is the exact race the lock exists to prevent."""
    monkeypatch.setattr(config, "REDIS_LOCK_WAIT_MS", 0)
    client = rs.get_redis()
    key = "cerebrozen:lock:default:sess-steal"

    with rs.session_lock("sess-steal"):
        # Our lock expired and another turn took the session while we were still running.
        client.set(key, "another-turns-token")

    assert client.get(key) == "another-turns-token", (
        "a finishing turn deleted a lock it did not own — two turns can now run at once"
    )


def test_a_failure_while_RELEASING_the_lock_is_swallowed(monkeypatch):
    """The TTL will clean it up. Raising here would fail a turn that has already succeeded
    and whose reply the user has already read."""
    monkeypatch.setattr(config, "REDIS_LOCK_WAIT_MS", 0)
    real = rs.get_redis()

    class _FailsOnRelease:
        def set(self, *a, **kw):
            return real.set(*a, **kw)

        def eval(self, *a, **kw):
            raise ConnectionError("gone")

        def get(self, *a, **kw):
            raise ConnectionError("gone")

    monkeypatch.setattr(rs, "_client", _FailsOnRelease())

    with rs.session_lock("sess-release-fail"):
        pass  # must not raise


# ═════════════════════════════════════════════════════════════════════════════
# Redis hot-state: caches. Every one of them must MISS, never raise.
# ═════════════════════════════════════════════════════════════════════════════


def test_the_profile_cache_round_trips_and_expires(monkeypatch):
    """A user opening several sessions in a minute re-reads the same profile from Mongo
    every time without this. The TTL is what stops a stale profile outliving an edit."""
    monkeypatch.setattr(config, "REDIS_PROFILE_TTL_S", 60)
    profile = {"name": "Ada", "level": "senior", "sessions": 3}

    assert rs.get_cached_profile("u-1") is None, "a cold cache must miss, not invent"

    rs.set_cached_profile("u-1", profile)
    assert rs.get_cached_profile("u-1") == profile

    ttl = rs.get_redis().ttl("cerebrozen:profile:default:u-1")
    assert 0 < ttl <= 60, "the profile was cached with no expiry — a stale profile forever"


def test_a_profile_with_unserialisable_values_is_still_cached(monkeypatch):
    """Profiles carry datetimes and ObjectIds straight out of Mongo. A TypeError from
    json.dumps here would raise into a turn — from the CACHE."""
    from datetime import datetime, timezone

    monkeypatch.setattr(config, "REDIS_PROFILE_TTL_S", 60)
    rs.set_cached_profile("u-2", {"last_seen": datetime(2026, 7, 14, tzinfo=timezone.utc)})

    cached = rs.get_cached_profile("u-2")
    assert cached is not None and "2026-07-14" in cached["last_seen"]


def test_a_zero_ttl_switches_the_profile_cache_off(monkeypatch):
    """The kill switch for a cache suspected of serving stale profiles."""
    monkeypatch.setattr(config, "REDIS_PROFILE_TTL_S", 60)
    rs.set_cached_profile("u-3", {"name": "Ada"})

    monkeypatch.setattr(config, "REDIS_PROFILE_TTL_S", 0)
    assert rs.get_cached_profile("u-3") is None, "the cache kept serving after being disabled"
    rs.set_cached_profile("u-3", {"name": "Grace"})   # writes must stop too

    monkeypatch.setattr(config, "REDIS_PROFILE_TTL_S", 60)
    assert rs.get_cached_profile("u-3") == {"name": "Ada"}, "a disabled cache still wrote"


def test_a_blank_user_id_never_touches_the_cache(monkeypatch):
    """An anonymous request must not read or poison a shared cache entry keyed on ""."""
    monkeypatch.setattr(config, "REDIS_PROFILE_TTL_S", 60)
    rs.set_cached_profile("", {"name": "nobody"})
    assert rs.get_cached_profile("") is None
    assert rs.get_redis().get("cerebrozen:profile:default:") is None


def test_a_broken_redis_makes_the_profile_cache_MISS_not_raise(monkeypatch, caplog):
    """Degrade safe: a cache miss costs one Mongo read. A raise costs the turn."""
    monkeypatch.setattr(config, "REDIS_PROFILE_TTL_S", 60)
    monkeypatch.setattr(rs, "_client", _ExplodingRedis())
    caplog.set_level("WARNING", logger="cerebrozen.redis")

    assert rs.get_cached_profile("u-1") is None
    rs.set_cached_profile("u-1", {"name": "Ada"})     # must not raise

    assert "redis.profile_get_failed" in caplog.text
    assert "redis.profile_set_failed" in caplog.text


def test_a_missing_redis_makes_the_profile_cache_MISS_not_raise(monkeypatch):
    monkeypatch.setattr(config, "REDIS_PROFILE_TTL_S", 60)
    monkeypatch.setattr(rs, "get_redis", lambda: None)

    assert rs.get_cached_profile("u-1") is None
    rs.set_cached_profile("u-1", {"name": "Ada"})     # must not raise


def test_the_generic_json_cache_round_trips_with_a_ttl():
    """Backs the RAG result cache. A cached retrieval that never expires serves deleted
    documents forever."""
    rs.cache_set_json("rag:q1", {"hits": [{"id": 1}]}, ttl_s=30)

    assert rs.cache_get_json("rag:q1") == {"hits": [{"id": 1}]}
    assert 0 < rs.get_redis().ttl("rag:q1") <= 30
    assert rs.cache_get_json("rag:never-written") is None


def test_the_generic_cache_refuses_a_blank_key_or_a_non_positive_ttl():
    """ttl_s<=0 into redis SET EX is a redis error, not a no-op — it would raise into
    whatever asked to cache something."""
    rs.cache_set_json("", {"a": 1}, ttl_s=30)
    rs.cache_set_json("k", {"a": 1}, ttl_s=0)
    rs.cache_set_json("k", {"a": 1}, ttl_s=-5)

    assert rs.cache_get_json("") is None
    assert rs.cache_get_json("k") is None


def test_a_broken_redis_makes_the_generic_cache_MISS_not_raise(monkeypatch, caplog):
    monkeypatch.setattr(rs, "_client", _ExplodingRedis())
    caplog.set_level("WARNING", logger="cerebrozen.redis")

    assert rs.cache_get_json("rag:q1") is None
    rs.cache_set_json("rag:q1", {"hits": []}, ttl_s=30)   # must not raise

    assert "redis.cache_get_failed" in caplog.text
    assert "redis.cache_set_failed" in caplog.text


def test_a_missing_redis_makes_the_generic_cache_MISS_not_raise(monkeypatch):
    monkeypatch.setattr(rs, "get_redis", lambda: None)
    assert rs.cache_get_json("k") is None
    rs.cache_set_json("k", {"a": 1}, ttl_s=30)


def test_the_seen_marker_is_a_fast_path_that_can_only_ever_say_no(monkeypatch, caplog):
    """It exists to SKIP a Mongo get_state on continuing turns. Every failure mode must
    answer "not seen", because that answer falls back to the authoritative get_state — a
    false "seen" would skip the state load and run a continuing turn as if it were fresh."""
    assert rs.is_session_seen("s-1") is False

    rs.mark_session_seen("s-1")
    assert rs.is_session_seen("s-1") is True
    assert 0 < rs.get_redis().ttl("cerebrozen:seen:default:s-1") <= rs._SEEN_TTL_S

    assert rs.is_session_seen("") is False
    rs.mark_session_seen("")
    assert rs.get_redis().get("cerebrozen:seen:default:") is None

    monkeypatch.setattr(rs, "_client", _ExplodingRedis())
    caplog.set_level("WARNING", logger="cerebrozen.redis")
    assert rs.is_session_seen("s-1") is False, "a broken Redis claimed to know the session"
    rs.mark_session_seen("s-1")           # must not raise
    assert "redis.seen_get_failed" in caplog.text
    assert "redis.seen_set_failed" in caplog.text

    monkeypatch.setattr(rs, "get_redis", lambda: None)
    assert rs.is_session_seen("s-1") is False
    rs.mark_session_seen("s-1")


# ═════════════════════════════════════════════════════════════════════════════
# Redis connection: which backend, and what happens when there isn't one.
# ═════════════════════════════════════════════════════════════════════════════


class _FakeRedisPy:
    """Stands in for a real redis-py connection (the external boundary)."""

    def __init__(self, url, **kwargs):
        self.url = url
        self.kwargs = kwargs
        self.pinged = False

    def ping(self):
        self.pinged = True
        return True


def test_a_configured_redis_url_is_connected_with_timeouts(monkeypatch, caplog):
    """A Redis client with no socket timeout turns a network partition into a hung turn —
    the exact failure the whole degrade-safe design is built to avoid. And decode_responses
    is load-bearing: without it every cached value comes back as bytes."""
    import redis as redis_py

    created: list[_FakeRedisPy] = []

    def _from_url(url, **kwargs):
        c = _FakeRedisPy(url, **kwargs)
        created.append(c)
        return c

    monkeypatch.setattr(redis_py.Redis, "from_url", staticmethod(_from_url))
    monkeypatch.setattr(config, "REDIS_URL", "redis://cache:6379/0")
    monkeypatch.setattr(rs, "_client", None)
    caplog.set_level("INFO", logger="cerebrozen.redis")

    client = rs.get_redis()

    assert client is created[0]
    assert created[0].url == "redis://cache:6379/0"
    assert created[0].kwargs["socket_timeout"] == 2
    assert created[0].kwargs["socket_connect_timeout"] == 2
    assert created[0].kwargs["decode_responses"] is True
    assert created[0].pinged, "the connection was never verified — a dead Redis looks alive"
    assert rs.backend() == "redis"

    assert rs.get_redis() is client, "a new connection per call would exhaust the pool"


def test_an_unreachable_redis_falls_back_to_an_in_process_one(monkeypatch, caplog):
    """A misconfigured REDIS_URL must not stop the service booting. It degrades to a
    process-local cache — which loses cross-instance locking, and that is exactly why the
    fallback is logged loudly rather than silently."""
    import redis as redis_py

    class _Dead(_FakeRedisPy):
        def ping(self):
            raise ConnectionError("connection refused")

    monkeypatch.setattr(redis_py.Redis, "from_url", staticmethod(
        lambda url, **kw: _Dead(url, **kw)))
    monkeypatch.setattr(config, "REDIS_URL", "redis://nope:6379")
    monkeypatch.setattr(rs, "_client", None)
    caplog.set_level("WARNING", logger="cerebrozen.redis")

    client = rs.get_redis()

    assert client is not None
    assert rs.backend() == "fakeredis"
    assert "redis.connect_failed_fallback_fakeredis" in caplog.text
    client.set("k", "v")                       # and it genuinely works
    assert client.get("k") == "v"


def test_with_no_redis_library_at_all_everything_becomes_a_no_op(monkeypatch, caplog):
    """The floor of the degradation ladder: no Redis, no fakeredis. The lock falls open, the
    caches miss, and coaching still runs."""
    monkeypatch.setattr(config, "REDIS_URL", "")
    monkeypatch.setattr(rs, "_client", None)
    monkeypatch.setitem(sys.modules, "fakeredis", None)
    caplog.set_level("WARNING", logger="cerebrozen.redis")

    assert rs.get_redis() is None
    assert rs.backend() == "none"
    assert "redis.unavailable" in caplog.text

    with rs.session_lock("s-1"):
        pass
    assert rs.get_cached_profile("u-1") is None
    assert rs.is_session_seen("s-1") is False


# ═════════════════════════════════════════════════════════════════════════════
# Rate limiting — the gaps the security suite leaves.
# ═════════════════════════════════════════════════════════════════════════════


class _Req:
    def __init__(self, ip="1.2.3.4", headers=None):
        self.headers = headers or {}
        self.client = type("C", (), {"host": ip})() if ip else None


def test_the_limiter_keys_on_the_signed_token_subject(monkeypatch):
    """The turn body carries a client-supplied user_id; keying on it means an attacker
    rotates it per request and the limit evaporates. The token subject is signed, so it
    cannot be. This asserts the KEY, not just that the code mentions decode_token."""
    import jwt

    import app.ratelimit as rl

    monkeypatch.setenv("AUTH_DEV_BYPASS", "")
    monkeypatch.setattr(config, "ENV", "production")
    monkeypatch.setattr(config, "JWT_SECRET", b"top-secret")
    monkeypatch.setattr(config, "JWT_ALGORITHM", "HS512")

    token = jwt.encode({"sub": "user-42"}, b"top-secret", algorithm="HS512")
    assert rl._caller_id(_Req(headers={"Authorization": f"Bearer {token}"})) == "sub:user-42"

    # Some issuers put the identity in `user_id` instead of `sub`.
    alt = jwt.encode({"user_id": "user-99"}, b"top-secret", algorithm="HS512")
    assert rl._caller_id(_Req(headers={"Authorization": f"Bearer {alt}"})) == "sub:user-99"


def test_an_unusable_token_falls_back_to_the_peer_rather_than_crashing(monkeypatch):
    """A bad token is 401's problem, not the limiter's. If the limiter raised on it, a
    malformed Authorization header would 500 the endpoint instead of 401-ing it."""
    import jwt

    import app.ratelimit as rl

    monkeypatch.setenv("AUTH_DEV_BYPASS", "")
    monkeypatch.setattr(config, "ENV", "production")
    monkeypatch.setattr(config, "JWT_SECRET", b"top-secret")
    monkeypatch.setattr(config, "JWT_ALGORITHM", "HS512")

    assert rl._caller_id(_Req(headers={"Authorization": "Bearer garbage"})) == "ip:1.2.3.4"
    assert rl._caller_id(_Req()) == "ip:1.2.3.4"          # no header at all

    # A validly-signed token carrying no identity is not an identity.
    anon = jwt.encode({"scope": "read"}, b"top-secret", algorithm="HS512")
    assert rl._caller_id(_Req(headers={"Authorization": f"Bearer {anon}"})) == "ip:1.2.3.4"


def test_behind_a_load_balancer_the_client_ip_comes_from_the_forwarded_chain():
    """Every request through an LB has the LB as its peer. Without this, all callers share
    one bucket and the first busy user rate-limits everybody else."""
    import app.ratelimit as rl

    caller = rl._caller_id(_Req(ip="10.0.0.1", headers={
        "X-Forwarded-For": " 203.0.113.7 , 10.0.0.1 ",
    }))
    assert caller == "ip:203.0.113.7"
    assert rl._caller_id(_Req(ip=None)) == "ip:unknown"


def test_a_limit_misconfigured_as_a_word_falls_back_to_the_default():
    """CEREBROZEN_RATE_LIMIT_TURNS_PER_MIN=twenty must not crash the process at import — and
    must not silently become 0 either, which would switch the limit OFF."""
    import app.ratelimit as rl

    assert rl._int_env("A_VAR_THAT_IS_NOT_SET", 20) == 20
    import os
    os.environ["_CEREBROZEN_TEST_BAD_INT"] = "twenty"
    try:
        assert rl._int_env("_CEREBROZEN_TEST_BAD_INT", 20) == 20
    finally:
        del os.environ["_CEREBROZEN_TEST_BAD_INT"]


def test_with_no_counter_backend_at_all_requests_are_allowed(monkeypatch):
    """Fail open. Failing closed converts a cache outage into a total coaching outage — a
    strictly worse day than an unbounded bill for the minutes it takes to notice."""
    import app.ratelimit as rl

    monkeypatch.setattr(rs, "_client", None)
    monkeypatch.setattr(rs, "get_redis", lambda: None)

    assert rl._hit("turn", "sub:x", limit=1, window_s=60) is None
    assert rl._hit("turn", "sub:x", limit=1, window_s=60) is None


def test_the_counter_window_expires_so_it_cannot_be_held_open(monkeypatch):
    """Only the FIRST hit sets the TTL. If every hit reset it, a caller who keeps hitting a
    key would keep extending its window and the count would never reset."""
    import time as _time

    import app.ratelimit as rl

    monkeypatch.setenv("CEREBROZEN_RATE_LIMIT", "true")
    # Freeze the clock mid-window: the key embeds int(time)//window_s, so a
    # real-clock hour boundary between hits would split them across two
    # windows and the third hit would not be over-limit (a 1-in-3600 flake).
    monkeypatch.setattr(_time, "time", lambda: 1_000_000_000.5)
    client = rs.get_redis()

    assert rl._hit("turn", "sub:a", limit=2, window_s=3600) is None
    key = next(k for k in client.keys("rl:turn:sub:a:*"))
    first_ttl = client.ttl(key)
    assert 0 < first_ttl <= 3600

    assert rl._hit("turn", "sub:a", limit=2, window_s=3600) is None
    retry_after = rl._hit("turn", "sub:a", limit=2, window_s=3600)

    assert retry_after is not None and 1 <= retry_after <= 3600
    assert client.ttl(key) <= first_ttl, "the window was extended by a later hit"


@pytest.mark.asyncio
async def test_the_limiter_can_be_switched_off_without_a_deploy(monkeypatch):
    """The 3am escape hatch. If CEREBROZEN_RATE_LIMIT stopped being read at call time, an
    operator could not turn the limit off in an incident."""
    import app.ratelimit as rl

    monkeypatch.setitem(rl.LIMITS, "turn", (1, 3600))
    monkeypatch.setattr(rl, "_caller_id", lambda r: "ip:off-switch")
    dep = rl._limiter("turn")

    monkeypatch.setenv("CEREBROZEN_RATE_LIMIT", "false")
    for _ in range(10):
        await dep(_Req())            # no raise, no counting

    monkeypatch.setenv("CEREBROZEN_RATE_LIMIT", "true")
    await dep(_Req())
    with pytest.raises(Exception):   # HTTPException 429
        await dep(_Req())


# ═════════════════════════════════════════════════════════════════════════════
# Prompt registry — the workbook defects that become live incidents.
# ═════════════════════════════════════════════════════════════════════════════

_CAT_HEADER = ["agent_name", "role", "enabled", "model", "sheet_name", "description"]


def _workbook(tmp_path: Path, sheets: dict, catalog=None, catalog_header=None,
              name="wb.xlsx") -> str:
    """Build a prompt workbook. `sheets` maps sheet name -> prompt text, or a dict of
    {row_offset: text} for continuation/orphan layouts."""
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, body in sheets.items():
        ws = wb.create_sheet(sheet_name)
        if isinstance(body, dict):
            for offset, text in body.items():
                ws.cell(row=prompts_mod.PROMPT_START_ROW + offset,
                        column=prompts_mod.PROMPT_COL, value=text)
        elif body:
            ws.cell(row=prompts_mod.PROMPT_START_ROW,
                    column=prompts_mod.PROMPT_COL, value=body)
    if catalog is not None:
        ws = wb.create_sheet(prompts_mod.CATALOG_SHEET)
        ws.append(catalog_header if catalog_header is not None else _CAT_HEADER)
        for row in catalog:
            ws.append(row)
    path = tmp_path / name
    wb.save(path)
    return str(path)


def _clean_catalog() -> list[list]:
    """A Catalog row for every stage that needs one; only the always-on closer is armed."""
    rows = []
    for stage, sheet in prompts_mod.STAGE_SHEET.items():
        if stage == "environment":
            continue
        model = "gpt-5-mini" if stage == "feedback_mood_capture_agent" else ""
        rows.append([stage, "coach", "FALSE", model, sheet, ""])
    return rows


def test_a_workbook_with_no_defects_reports_ok(tmp_path):
    """The clean baseline. If the validator could never return ok=True it would be a
    permanent red light that everyone learns to ignore — and then the real defects below
    would be ignored with it."""
    sheets = {sheet: "" for sheet in prompts_mod.STAGE_SHEET.values()}
    sheets["environment_system_agent"] = "Stay safe and kind."
    sheets["feedback_mood_capture_agent"] = "How did that land?"

    reg = prompts_mod.PromptRegistry(_workbook(tmp_path, sheets, _clean_catalog()))

    assert reg.validation["issue_count"] == 0
    assert reg.validation["ok"] is True
    assert reg.degraded is False
    assert reg.info()["version"] == reg.version


def test_an_absent_stage_sheet_is_reported_not_guessed(tmp_path, caplog):
    """A stage whose sheet was renamed or deleted loads as "" and the node quietly falls
    back to CIM. Nobody notices the agent stopped existing unless the load says so."""
    caplog.set_level("WARNING", logger="cerebrozen.prompts")
    sheets = {"environment_system_agent": "Be safe.",
              "feedback_mood_capture_agent": "Wrap up.",
              "core_coaching_agent": "Coach them."}

    reg = prompts_mod.PromptRegistry(_workbook(tmp_path, sheets, _clean_catalog()))

    assert reg.get("core_coaching_agent") == "Coach them."
    assert reg.get("CH_coaching_agent") == "", "a missing sheet must not raise or invent text"
    assert "CH_coaching_agent" in reg.validation["missing_sheets"]
    assert "prompt.missing_sheet" in caplog.text


def test_an_enabled_agent_with_no_prompt_or_no_model_is_flagged_at_load(tmp_path):
    """Both are a turn-time RuntimeError in production — the user gets an error mid-session.
    Catching them at load turns a live incident into a line in the validation report."""
    sheets = {sheet: "" for sheet in prompts_mod.STAGE_SHEET.values()}
    sheets["environment_system_agent"] = "Be safe."
    sheets["feedback_mood_capture_agent"] = "Wrap up."
    sheets["role_play_agent"] = "Play the role."      # has a prompt, but no model
    catalog = [r for r in _clean_catalog()
               if r[4] not in ("core_coaching_agent", "role_play_agent")]
    catalog.append(["core_coaching_agent", "coach", "TRUE", "gpt-5-mini",
                    "core_coaching_agent", ""])        # enabled, but its sheet is empty
    catalog.append(["role_play_agent", "coach", "TRUE", "", "role_play_agent", ""])

    reg = prompts_mod.PromptRegistry(_workbook(tmp_path, sheets, catalog))

    assert "core_coaching_agent" in reg.validation["enabled_no_prompt"]
    assert "role_play_agent" in reg.validation["enabled_no_model"]
    assert reg.validation["ok"] is False


def test_a_stage_missing_from_the_catalog_entirely_is_flagged_and_disabled(tmp_path):
    """An agent with no Catalog row is OFF — silently. The report is the only place that
    difference between "deliberately disabled" and "forgotten" is visible."""
    sheets = {sheet: "" for sheet in prompts_mod.STAGE_SHEET.values()}
    sheets["environment_system_agent"] = "Be safe."
    sheets["feedback_mood_capture_agent"] = "Wrap up."
    catalog = [r for r in _clean_catalog() if r[4] != "learning_aid_agent"]

    reg = prompts_mod.PromptRegistry(_workbook(tmp_path, sheets, catalog))

    assert "learning_aid_agent" in reg.validation["not_in_catalog"]
    assert reg.is_enabled("learning_aid_agent") is False
    assert reg.model_for("learning_aid_agent") is None


def test_an_oversize_prompt_is_flagged_before_it_becomes_a_latency_bill(tmp_path):
    """~6K tokens of system prompt on every turn of every session. The Excel cell cap forces
    a spill long after the cost has become the problem."""
    sheets = {sheet: "" for sheet in prompts_mod.STAGE_SHEET.values()}
    sheets["environment_system_agent"] = "x" * (prompts_mod.PROMPT_SIZE_WARN_CHARS + 1)
    sheets["feedback_mood_capture_agent"] = "Wrap up."

    reg = prompts_mod.PromptRegistry(_workbook(tmp_path, sheets, _clean_catalog()))

    assert reg.validation["oversize"]["environment"] == prompts_mod.PROMPT_SIZE_WARN_CHARS + 1


def test_a_placeholder_nothing_can_resolve_is_flagged_rather_than_blanked(tmp_path):
    """The resolver blanks an unknown {token} at runtime — the agent then reads an
    instruction with a hole in it and nobody sees an error. Known tokens (context, RAG,
    capture-registry, state fields) must NOT be flagged, or the report becomes noise."""
    sheets = {sheet: "" for sheet in prompts_mod.STAGE_SHEET.values()}
    sheets["environment_system_agent"] = (
        "Hello {userName}, your goal is {session_goal}. Now {this_resolves_to_nothing}."
    )
    sheets["feedback_mood_capture_agent"] = "Wrap up."

    reg = prompts_mod.PromptRegistry(_workbook(tmp_path, sheets, _clean_catalog()))

    assert reg.validation["unknown_placeholders"]["environment"] == ["this_resolves_to_nothing"]


def test_a_dotted_placeholder_is_audited_on_its_root(tmp_path):
    """{ic_profile.style} resolves against a nested context dict. Auditing the full path
    would flag every legitimate nested token as unknown."""
    sheets = {sheet: "" for sheet in prompts_mod.STAGE_SHEET.values()}
    sheets["environment_system_agent"] = "Profile: {ic_profile.thinking} {nope.deep}"
    sheets["feedback_mood_capture_agent"] = "Wrap up."

    reg = prompts_mod.PromptRegistry(_workbook(tmp_path, sheets, _clean_catalog()))

    assert reg.validation["unknown_placeholders"]["environment"] == ["nope.deep"]


@pytest.mark.parametrize("cell,expected", [
    (True, True), (False, False), ("TRUE", True), (" yes ", True), ("1", True),
    ("enabled", True), ("on", True), ("FALSE", False), ("no", False), ("", False),
    (None, False), ("maybe", False),
])
def test_the_enabled_cell_is_read_the_same_whether_excel_typed_it_as_text_or_bool(
    cell, expected
):
    """openpyxl returns a real bool for a TRUE cell and a string for a text-typed one. A
    business user retyping "TRUE" as text must not silently disable an agent — and a blank
    cell must never mean enabled."""
    assert prompts_mod._parse_enabled(cell) is expected


def test_a_workbook_with_no_catalog_tab_disables_every_agent(tmp_path, caplog):
    """Fail safe: an unreadable catalog must not enable everything. The always-on closer
    still runs, so a session can still end."""
    caplog.set_level("WARNING", logger="cerebrozen.prompts")
    reg = prompts_mod.PromptRegistry(_workbook(
        tmp_path, {"environment_system_agent": "Be safe.",
                   "feedback_mood_capture_agent": "Wrap up."}, catalog=None))

    assert reg.is_enabled("core_coaching_agent") is False
    assert reg.is_enabled("feedback_mood_capture_agent") is True
    assert "prompt.missing_catalog" in caplog.text


def test_a_catalog_with_no_rows_at_all_is_survivable(tmp_path):
    """An empty tab (someone cleared it) must not raise during boot."""
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("environment_system_agent")
    ws.cell(row=7, column=2, value="Be safe.")
    wb.create_sheet("feedback_mood_capture_agent").cell(row=7, column=2, value="Wrap up.")
    wb.create_sheet(prompts_mod.CATALOG_SHEET)
    path = tmp_path / "empty_catalog.xlsx"
    wb.save(path)

    reg = prompts_mod.PromptRegistry(str(path))
    assert reg.is_enabled("core_coaching_agent") is False


def test_a_catalog_missing_its_key_columns_disables_everything_loudly(tmp_path, caplog):
    """A renamed `sheet_name` column would otherwise map nothing — every agent off, with no
    explanation anywhere."""
    caplog.set_level("WARNING", logger="cerebrozen.prompts")
    reg = prompts_mod.PromptRegistry(_workbook(
        tmp_path,
        {"environment_system_agent": "Be safe.", "feedback_mood_capture_agent": "Wrap up."},
        catalog=[["core_coaching_agent", "TRUE"]],
        catalog_header=["agent_name", "is_on"],       # no sheet_name / enabled
    ))

    assert reg.is_enabled("core_coaching_agent") is False
    assert "prompt.catalog_columns_missing" in caplog.text


def test_ragged_and_blank_catalog_rows_are_skipped_not_crashed_on(tmp_path):
    """Excel hands back short/None rows for anything a human has touched. One stray row must
    not take the whole catalog — and every agent — down."""
    catalog = [
        ["core_coaching_agent"],                       # ragged: no enabled/sheet columns
        [None, None, "TRUE", "gpt-5-mini", None, ""],  # blank sheet name
        ["", "", "", "", "   ", ""],                   # whitespace sheet name
        ["ch", "coach", "TRUE", None, "CH_coaching_agent", ""],        # model cell empty
        ["core", "coach", "TRUE", "gpt-5-mini", "core_coaching_agent", ""],
    ]
    reg = prompts_mod.PromptRegistry(_workbook(
        tmp_path,
        {"environment_system_agent": "Be safe.", "feedback_mood_capture_agent": "Wrap up.",
         "core_coaching_agent": "Coach.", "CH_coaching_agent": "CH."},
        catalog=catalog,
    ))

    assert reg.is_enabled("core_coaching_agent") is True
    assert reg.model_for("core_coaching_agent") == "gpt-5-mini"
    assert reg.is_enabled("CH_coaching_agent") is True
    assert reg.model_for("CH_coaching_agent") is None, "a blank model cell must not be invented"


def test_a_catalog_with_no_model_column_still_switches_agents_on(tmp_path):
    """Older workbooks predate the model column. They must still load — with no models."""
    reg = prompts_mod.PromptRegistry(_workbook(
        tmp_path,
        {"environment_system_agent": "Be safe.", "feedback_mood_capture_agent": "Wrap up.",
         "core_coaching_agent": "Coach."},
        catalog=[["core", "coach", "TRUE", "core_coaching_agent"]],
        catalog_header=["agent_name", "role", "enabled", "sheet_name"],
    ))

    assert reg.is_enabled("core_coaching_agent") is True
    assert reg.model_for("core_coaching_agent") is None


def test_a_dev_override_can_arm_a_stage_without_touching_the_workbook(tmp_path):
    """The test/dev switch. It must NOT be able to disarm the always-on closer, or a session
    could never reach its terminal close and every conversation would strand."""
    reg = prompts_mod.PromptRegistry(_workbook(
        tmp_path,
        {"environment_system_agent": "Be safe.", "feedback_mood_capture_agent": "Wrap up."},
        catalog=_clean_catalog(),
    ))

    reg.set_enabled("role_play_agent", True)
    assert reg.is_enabled("role_play_agent") is True

    reg.set_enabled("feedback_mood_capture_agent", False)
    assert reg.is_enabled("feedback_mood_capture_agent") is True, (
        "the closing layer was switched off — sessions can no longer end"
    )


def test_validate_on_save_warns_about_an_oversize_body_without_blocking_it():
    """Size is a smell, not an error. Blocking the save would stop an author fixing a
    genuinely long prompt; saying nothing lets it grow."""
    result = prompts_mod.validate_prompt_text(
        "core_coaching_agent", "x" * (prompts_mod.PROMPT_SIZE_WARN_CHARS + 10))

    assert result["errors"] == [], "an oversize prompt must still be savable"
    assert any("budget" in w for w in result["warnings"])
    assert result["size"] == prompts_mod.PROMPT_SIZE_WARN_CHARS + 10


def test_a_disabled_agent_may_be_saved_empty():
    """Emptying a prompt is how an author parks an agent they have switched off."""
    assert prompts_mod.validate_prompt_text("role_play_agent", "  ", enabled=False)["errors"] == []
    assert prompts_mod.validate_prompt_text("role_play_agent", "  ", enabled=True)["errors"]


def test_the_placeholder_audit_survives_every_registry_being_broken(monkeypatch):
    """Validation is advisory. If a broken RAG/state/registry import could raise here, a
    prompt LOAD would fail — the workbook would not serve at all — because of an
    *advisory report*. Every source is best-effort."""
    monkeypatch.setitem(sys.modules, "app.rag.registry", None)
    monkeypatch.setitem(sys.modules, "app.stores.variable_capture_registry", None)
    monkeypatch.setitem(sys.modules, "app.graph.state", None)

    tokens = prompts_mod.PromptRegistry._known_tokens_from_registries()

    assert tokens == frozenset()
    # And a prompt still validates against the built-in context tokens alone.
    assert prompts_mod.validate_prompt_text("x", "Hi {userName}")["warnings"] == []


# ═════════════════════════════════════════════════════════════════════════════
# Variable capture registry — what gets written to the user's permanent record.
# ═════════════════════════════════════════════════════════════════════════════

_VAR_HEADER = ["variable_name", "update_frequency", "capture_enabled", "source_agent",
               "prompt_placeholder", "description", "notes"]


def _var_workbook(tmp_path: Path, rows: list, header=None,
                  sheet=None, name="vars.xlsx") -> Path:
    from app.stores.variable_capture_registry import REGISTRY_SHEET

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(sheet or REGISTRY_SHEET)
    if header is not None or rows:
        ws.append(header if header is not None else _VAR_HEADER)
    for row in rows:
        ws.append(row)
    path = tmp_path / name
    wb.save(path)
    return path


def test_a_once_in_lifetime_variable_is_never_overwritten_by_a_later_session(tmp_path):
    """Some facts about a person are established once — their thinking preference, their
    NBI profile. If a later session could overwrite them, one throwaway remark reshapes the
    user's permanent record and every future session is coached off it."""
    from app.stores.variable_capture_registry import VariableCaptureRegistry as VCR

    path = _var_workbook(tmp_path, [
        ["user_thinking_preference", "once_in_lifetime", True, "coaching_intake_agent",
         "{userThinkingPreference}", "how they think", ""],
        ["session_goal", "every_session", True, "challenge_context_agent",
         "{session_goal}", "this session's goal", ""],
        ["coaching_shift_summary", "only_on_shift", True, "core_coaching_agent", "", "", ""],
    ])
    reg = VCR.reload(path)

    assert reg.is_once_in_lifetime("user_thinking_preference") is True
    assert reg.is_once_in_lifetime("session_goal") is False
    assert reg.is_once_in_lifetime("coaching_shift_summary") is False, (
        "only_on_shift must behave like every_session at the persistence layer — the "
        "agent prompt decides when to emit it"
    )
    assert reg.once_in_lifetime_vars == frozenset({"user_thinking_preference"})

    # An unregistered variable is still captured: the sheet CONSTRAINS, it does not gate
    # discovery — a new key from any agent is stored without a sheet edit.
    assert reg.is_once_in_lifetime("brand_new_key") is False
    assert reg.is_capture_enabled("brand_new_key") is True


def test_a_disabled_variable_is_suppressed_even_when_the_agent_emits_it(tmp_path):
    """The no-deploy kill switch for a variable that is capturing garbage. If capture_enabled
    stopped being honoured, a business user's only recourse would be an engineering release."""
    from app.stores.variable_capture_registry import VariableCaptureRegistry as VCR

    path = _var_workbook(tmp_path, [
        ["good_var", "every_session", True, "core_coaching_agent", "{good_var}", "", ""],
        ["bad_var", "every_session", False, "core_coaching_agent", "{bad_var}", "", ""],
        ["also_bad", "every_session", "FALSE", "core_coaching_agent", "", "", ""],
        ["still_on", "every_session", "TRUE", "core_coaching_agent", "", "", ""],
    ])
    reg = VCR.reload(path)

    assert reg.is_capture_enabled("good_var") is True
    assert reg.is_capture_enabled("bad_var") is False
    assert reg.is_capture_enabled("also_bad") is False, (
        "a text-typed FALSE was read as enabled — the kill switch does not work"
    )
    assert reg.is_capture_enabled("still_on") is True
    assert reg.config("bad_var").capture_enabled is False
    assert reg.config("not_in_the_sheet") is None


def test_the_sheet_alone_decides_which_keys_a_node_bridges_into_persistence(tmp_path):
    """`_bridge_registry_vars` reads this instead of a hand-maintained per-stage whitelist in
    code. Nested variables must contribute their PARENT key, or the bridge writes
    "coaching_style_context.selected_style" as a literal top-level field."""
    from app.stores.variable_capture_registry import VariableCaptureRegistry as VCR

    path = _var_workbook(tmp_path, [
        ["coaching_style_context.selected_style", "every_session", True,
         "coaching_intake_agent", "", "", ""],
        ["user_strengths.self_reported", "every_session", True, "coaching_intake_agent",
         "", "", ""],
        ["session_goal", "every_session", True, "challenge_context_agent", "", "", ""],
    ])
    reg = VCR.reload(path)

    assert reg.top_level_vars_for_agent("coaching_intake_agent") == frozenset(
        {"coaching_style_context", "user_strengths"})
    assert reg.top_level_vars_for_agent("challenge_context_agent") == frozenset({"session_goal"})
    assert reg.top_level_vars_for_agent("nobody") == frozenset()
    assert set(reg.all_vars) == {
        "coaching_style_context.selected_style", "user_strengths.self_reported", "session_goal"}


def test_section_headers_and_junk_rows_in_the_sheet_are_ignored(tmp_path):
    """The sheet is edited by humans: it has section-header rows, blank spacers and typos in
    the frequency column. Any of them becoming a "variable" would write junk keys into every
    user's permanent record."""
    from app.stores.variable_capture_registry import VariableCaptureRegistry as VCR

    path = _var_workbook(tmp_path, [
        ["=== INTAKE VARIABLES ===", None, None, None, None, None, None],   # section header
        [None, "every_session", True, "a", "", "", ""],                     # no name
        ["   ", "every_session", True, "a", "", "", ""],                    # blank name
        ["typo_freq", "evry_session", True, "a", "", "", ""],               # unknown frequency
        ["real_var", "every_session", True, "a", "{real_var}", "d", "n"],
    ])
    reg = VCR.reload(path)

    assert set(reg.all_vars) == {"real_var"}
    cfg = reg.config("real_var")
    assert (cfg.source_agent, cfg.prompt_placeholder, cfg.description, cfg.notes) == (
        "a", "{real_var}", "d", "n")


def test_the_registry_placeholders_are_accepted_by_the_prompt_validator(tmp_path):
    """The two registries have to agree, or every capture-registry placeholder in a prompt is
    reported as unresolvable and the validation report becomes noise nobody reads."""
    from app.stores.variable_capture_registry import VariableCaptureRegistry as VCR

    path = _var_workbook(tmp_path, [
        ["with_placeholder", "every_session", True, "a", "{PrettyName}", "", ""],
        ["no_placeholder", "every_session", True, "a", "", "", ""],
    ])
    VCR.reload(path)

    known = prompts_mod.PromptRegistry._known_tokens_from_registries()

    assert "with_placeholder" in known
    assert "PrettyName" in known, "a prompt using the sheet's own placeholder would be flagged"
    assert "no_placeholder" in known


@pytest.mark.parametrize("broken", ["missing_file", "missing_sheet", "empty_sheet",
                                    "missing_columns", "not_a_workbook"])
def test_an_unreadable_registry_degrades_to_capture_everything(tmp_path, caplog, broken):
    """A broken sheet must not stop the app, and must not stop capture either — with an empty
    registry every variable is captured with every_session semantics, which is the documented
    default. Failing shut would silently stop persisting the user's profile."""
    from app.stores.variable_capture_registry import VariableCaptureRegistry as VCR

    caplog.set_level("WARNING", logger="cerebrozen.variable_registry")
    good_row = ["v", "every_session", True, "a", "", "", ""]

    if broken == "missing_file":
        path = tmp_path / "nope.xlsx"
    elif broken == "missing_sheet":
        path = _var_workbook(tmp_path, [good_row], sheet="some_other_tab")
    elif broken == "empty_sheet":
        path = _var_workbook(tmp_path, [], header=None)
    elif broken == "missing_columns":
        path = _var_workbook(tmp_path, [["v", "every_session"]],
                             header=["variable_name", "update_frequency"])
    else:
        path = tmp_path / "corrupt.xlsx"
        path.write_bytes(b"this is not a workbook")

    reg = VCR.reload(path)

    assert reg.all_vars == {}
    assert reg.once_in_lifetime_vars == frozenset()
    assert reg.is_capture_enabled("anything") is True, (
        "a broken sheet stopped capture — the user's profile silently stops updating"
    )
    assert reg.is_once_in_lifetime("user_thinking_preference") is False
    if broken != "empty_sheet":
        # A cleared sheet is indistinguishable from a sheet with no rows, so it loads
        # quietly; every other way of being broken must say so.
        assert [r for r in caplog.records if r.levelname in ("WARNING", "ERROR")], (
            f"a broken registry ({broken}) failed silently"
        )


def test_the_registry_is_a_lazily_built_singleton(tmp_path):
    """It reads a 180KB workbook off disk. Rebuilding it per turn would be a file read on
    every coaching message."""
    from app.stores.variable_capture_registry import VariableCaptureRegistry as VCR

    VCR._instance = None
    first = VCR.get()
    assert VCR.get() is first

    reloaded = VCR.reload(_var_workbook(tmp_path, [["v", "every_session", True, "a", "", "", ""]]))
    assert VCR.get() is reloaded, "reload() did not replace the live singleton"


def test_the_bundled_workbook_actually_declares_the_lifetime_variables():
    """The shipped sheet is the one production runs on. An empty or renamed tab there means
    every 'once in a lifetime' fact silently becomes overwritable."""
    from app.stores.variable_capture_registry import VariableCaptureRegistry as VCR

    reg = VCR.reload()   # the real, bundled agent_prompts.xlsx

    assert reg.all_vars, "the bundled variable-capture sheet loaded EMPTY"
    assert reg.once_in_lifetime_vars, "no once-in-lifetime variables are protected"


# ═════════════════════════════════════════════════════════════════════════════
# History shaping + the prompt-workbook download.
# ═════════════════════════════════════════════════════════════════════════════


def test_a_bot_message_shaped_with_no_action_index_carries_no_actions():
    """The /latest-response and /history readers both pass an index; a caller that has no
    actions to attach (a transcript with no action store) must still get a renderable
    message rather than a KeyError mid-render."""
    from app.history import _shape_message

    shaped = _shape_message({
        "role": "assistant", "text": "Here is your plan.", "message_num": 4,
        "request_id": "req-1", "bot_name": "CereBroZen", "agent_name": "core_coaching_agent",
    }, total=9)

    assert shaped["bot"]["text"] == "Here is your plan."
    assert shaped["bot"]["tot_messages"] == 9
    assert "actions" not in shaped["bot"]


def test_the_workbook_download_populates_a_cold_cache_from_s3(tmp_path, monkeypatch):
    """Cold start in s3 mode: the server cache is empty, so the endpoint used to 404 on the
    first download after a deploy. It must pull the object once and serve it — and serve the
    SAME bytes the registry will load, or an admin downloads one workbook and edits another.
    """
    from fastapi.testclient import TestClient

    from app.llm import prompt_store
    from app.main import create_app
    from app.routers import prompts as prompts_router

    cache = tmp_path / "agent_prompts.xlsx"
    workbook_bytes = Path(config.PROMPT_WORKBOOK).read_bytes()

    class _S3:
        def download_file(self, bucket, key, dest, ExtraArgs=None):
            Path(dest).write_bytes(workbook_bytes)

    monkeypatch.setattr(config, "PROMPT_SOURCE", "s3")
    monkeypatch.setattr(config, "PROMPT_S3_BUCKET", "sys-config")
    monkeypatch.setattr(prompt_store, "WORKBOOK_CACHE_PATH", cache)
    monkeypatch.setattr(prompts_router, "WORKBOOK_CACHE_PATH", cache)
    monkeypatch.setattr(prompt_store, "_s3_client", lambda: _S3())

    assert not cache.exists()
    resp = TestClient(create_app()).get("/v1/prompts/download")

    assert resp.status_code == 200
    assert resp.content == workbook_bytes
    assert cache.is_file(), "the cache was not populated — the next request re-downloads"
