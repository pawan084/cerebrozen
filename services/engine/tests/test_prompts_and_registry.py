"""Prompt registry: loading, the Catalog gate, content versioning, validate-on-save,
and loud degradation. The registry is the seam non-technical users edit through, so
every failure mode here is a live-incident class — it gets a test, not a log line.
"""
import openpyxl
import pytest

from app.graph.runtime import get_registry
from app.llm.prompts import (
    ALWAYS_ENABLED,
    CATALOG_SHEET,
    PROMPT_COL,
    PROMPT_START_ROW,
    STAGE_SHEET,
    PromptRegistry,
    validate_prompt_text,
)
from app.rag.registry import get_registry as get_extraction_registry


# ── loading ──────────────────────────────────────────────────────────────────

def test_registry_loads_every_stage():
    reg = get_registry()
    for stage in STAGE_SHEET:
        assert stage in reg.sizes()
    # The guardrail layer must always carry text — everything composes on top of it.
    assert reg.environment.strip()
    assert reg.get("core_coaching_agent").strip()


def test_always_enabled_agents_cannot_be_gated_off():
    """The closing layer is the sole path to a terminal close: if it could be
    disabled, sessions could never end."""
    reg = get_registry()
    for stage in ALWAYS_ENABLED:
        assert reg.is_enabled(stage)
        reg.set_enabled(stage, False)  # explicit attempt to switch it off
        assert reg.is_enabled(stage), f"{stage} must be always-on"


def test_unknown_stage_is_disabled_not_crashing():
    reg = get_registry()
    assert reg.get("no_such_agent") == ""
    assert reg.is_enabled("no_such_agent") is False


def test_model_comes_from_the_catalog():
    reg = get_registry()
    assert reg.model_for("core_coaching_agent")  # a real model id from the Catalog
    assert reg.model_for("no_such_agent") is None  # absent → None (caller fails loudly)


def test_extraction_registry_binds_rag_tokens():
    tokens = set(get_extraction_registry().binding_tokens())
    assert tokens, "extractions must bind placeholder tokens"
    assert any("SSKB" in t or "CSKB" in t for t in tokens)


# ── content versioning (the prompt-cache buster) ─────────────────────────────

def test_version_is_a_stable_content_hash():
    reg = get_registry()
    assert reg.version
    before = reg.version
    reg.reload()
    assert reg.version == before, "identical content must produce an identical version"


def test_version_changes_when_content_changes(tmp_path):
    """The version is what busts the LLM prompt cache — an edit MUST change it, or
    a reloaded prompt keeps serving from the stale cache slot."""
    path = _workbook_copy(tmp_path)
    reg = PromptRegistry(path=str(path))
    before = reg.version

    wb = openpyxl.load_workbook(path)
    ws = wb[_sheet(wb, STAGE_SHEET["core_coaching_agent"])]
    ws.cell(row=PROMPT_START_ROW, column=PROMPT_COL).value = "a materially different prompt"
    wb.save(path)

    reg.reload()
    assert reg.version != before


# ── the Catalog gate ─────────────────────────────────────────────────────────

def test_catalog_disables_an_agent(tmp_path):
    path = _workbook_copy(tmp_path)
    _set_catalog_enabled(path, STAGE_SHEET["learning_aid_agent"], False)

    reg = PromptRegistry(path=str(path))
    assert reg.is_enabled("learning_aid_agent") is False
    assert reg.is_enabled("core_coaching_agent") is True  # others unaffected


# ── validation ───────────────────────────────────────────────────────────────

def test_validation_report_shape():
    report = get_registry().validation
    for key in ("missing_sheets", "not_in_catalog", "enabled_no_prompt",
                "enabled_no_model", "orphaned_continuation", "oversize",
                "unknown_placeholders", "issue_count", "ok"):
        assert key in report


def test_bundled_workbook_has_no_blocking_defects():
    """An enabled agent with no prompt or no model cannot run a turn — these are
    hard defects, not style warnings, so the shipped workbook must have none."""
    report = get_registry().validation
    assert report["enabled_no_prompt"] == []
    assert report["enabled_no_model"] == []
    assert report["missing_sheets"] == []


def test_validation_flags_an_orphaned_continuation(tmp_path):
    """A blank row inside a spilled prompt silently truncates it (the reader stops
    at the first blank) — the classic 'my edit vanished' corruption."""
    path = _workbook_copy(tmp_path)
    wb = openpyxl.load_workbook(path)
    ws = wb[_sheet(wb, STAGE_SHEET["role_play_agent"])]
    ws.cell(row=PROMPT_START_ROW, column=PROMPT_COL).value = "first half"
    ws.cell(row=PROMPT_START_ROW + 1, column=PROMPT_COL).value = None      # the blank
    ws.cell(row=PROMPT_START_ROW + 2, column=PROMPT_COL).value = "orphaned tail"
    wb.save(path)

    reg = PromptRegistry(path=str(path))
    assert reg.get("role_play_agent") == "first half"      # the tail is NOT loaded
    assert "role_play_agent" in reg.validation["orphaned_continuation"]


def test_validate_on_save_blocks_an_empty_prompt_for_an_enabled_agent():
    report = validate_prompt_text("core_coaching_agent", "   ", enabled=True)
    assert report["errors"]
    # A disabled agent may legitimately have no prompt yet.
    assert not validate_prompt_text("core_coaching_agent", "", enabled=False)["errors"]


def test_validate_on_save_warns_on_unresolvable_placeholders():
    report = validate_prompt_text("core_coaching_agent", "Coach {userName} on {totally_made_up}.")
    assert not report["errors"]           # advisory, never blocks
    assert any("totally_made_up" in w for w in report["warnings"])
    # A real context token and a real RAG token both resolve → no warning.
    clean = validate_prompt_text("core_coaching_agent", "Coach {userName} using {SSKB_Concept}.")
    assert clean["warnings"] == []


# ── degradation ──────────────────────────────────────────────────────────────

def test_a_failed_reload_keeps_serving_the_previous_prompts(tmp_path):
    """A bad upload must never take live prompts down: keep the last good ones and
    mark the registry degraded."""
    path = _workbook_copy(tmp_path)
    reg = PromptRegistry(path=str(path))
    good_version, good_prompt = reg.version, reg.get("core_coaching_agent")
    assert not reg.degraded

    path.write_bytes(b"this is not a workbook")  # corrupt it
    reg.reload()

    assert reg.degraded and reg.degraded_reason
    assert reg.get("core_coaching_agent") == good_prompt  # still serving the good text
    assert reg.version == good_version


def test_first_load_of_a_broken_workbook_fails_fast(tmp_path):
    """There is nothing to serve on a cold start — fail loudly rather than boot an
    app with no prompts."""
    bad = tmp_path / "broken.xlsx"
    bad.write_bytes(b"not a workbook")
    with pytest.raises(Exception):
        PromptRegistry(path=str(bad))


# ── helpers ──────────────────────────────────────────────────────────────────

def _workbook_copy(tmp_path):
    """A private copy of the real workbook — tests must never mutate the repo file."""
    import shutil

    from app import config

    dst = tmp_path / "agent_prompts.xlsx"
    shutil.copy(config.PROMPT_WORKBOOK, dst)
    return dst


def _sheet(wb, target: str) -> str:
    """Resolve a sheet by stripped name (the workbook has trailing-space sheet names)."""
    return next(n for n in wb.sheetnames if n.strip() == target.strip())


def _set_catalog_enabled(path, sheet_name: str, value: bool) -> None:
    wb = openpyxl.load_workbook(path)
    cat = wb[_sheet(wb, CATALOG_SHEET)]
    header = [str(c.value).strip().lower() if c.value else "" for c in cat[1]]
    i_sheet, i_enabled = header.index("sheet_name"), header.index("enabled")
    for row in range(2, cat.max_row + 1):
        cell = cat.cell(row=row, column=i_sheet + 1).value
        if cell and str(cell).strip() == sheet_name.strip():
            cat.cell(row=row, column=i_enabled + 1).value = "TRUE" if value else "FALSE"
            break
    wb.save(path)


def test_no_shipping_prompt_has_a_placeholder_nothing_can_resolve():
    """An unknown `{token}` is BLANKED at runtime, silently.

    That is not theoretical. Until 2026-07-17 `feedback_mood_capture_agent` told the model
    to build its reply "in this shape:" and the shape was a template — but two of its three
    slots were single-word `{Question}` / `{Examples}`, which the resolver ate. What the
    model actually received was:

        \\n\\n{One-line context/framing sentence}\\n\\n

    ...an instruction pointing at a template with its parts erased. The author's own
    MULTI-WORD slots ({low anchor}, {high anchor}) survived, because the resolver's regex
    has no space in its character class — so the bug was invisible unless you diffed the
    composed prompt against the workbook.

    It hid for another reason too: the size warning fired on five prompts permanently, so
    the validation report was never clean and nobody read it. This asserts it stays clean.
    """
    from app.graph.runtime import get_registry

    unknown = get_registry().validation["unknown_placeholders"]
    # No known gaps left. `{time}` was the last one — coaching_intake says "greet the user
    # based on {time}" five times and nothing resolved it, so the model was told to vary by
    # time of day and handed a blank. The client now sends `local_hour` and
    # guardrails.time_of_day turns it into a phrase (2026-07-17).
    assert unknown == {}, (
        f"a prompt gained a placeholder nothing resolves — it will be blanked at runtime "
        f"and no error will be raised: {unknown}"
    )


def test_the_reply_templates_survive_placeholder_resolution():
    """The two prompts that show the model a template must still show it one.

    Checks the composed prompt, not the workbook — the workbook was always fine; the damage
    happened in resolution, which is the only place it was visible.
    """
    from app.graph.guardrails import build_system_prompt
    from app.graph.runtime import get_registry

    reg = get_registry()
    for stage, slots in (
        ("feedback_mood_capture_agent", ("{the question}", "{the examples}")),
        ("coaching_intake_agent", ("{the question}", "{low anchor}", "{high anchor}")),
    ):
        composed = build_system_prompt(
            reg.environment, reg.get(stage), None, {},
            {"user_message": "x", "conversation_history": "x"}, invoking_agent=stage,
        )
        for slot in slots:
            assert slot in composed, (
                f"{stage}: the template slot {slot} was blanked before the model saw it — "
                f"the reply-shape instruction now points at a hole"
            )
