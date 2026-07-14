"""The RAG layer and the Postgres store shim — the two places this app fails SILENTLY.

Both layers are built to never raise into the caller, which is right for uptime and
terrible for observability: when RAG breaks, every extraction returns null, the coach
loses its evidence base, and the only symptom is blander advice. When the Postgres shim
mis-emulates a Mongo operator, the app keeps serving — with wrong data.

So these tests run against the REAL dependencies (a live Postgres+pgvector and a real,
local LanceDB in tmp_path) and assert on the DATA that lands in them. Nothing under test
is mocked. Exactly two boundaries are faked, both of them networks we refuse to touch:

  * S3          → an in-process object store (`_FakeS3`), so ingestion is exercised for
                  real: chunking, metadata, ETag idempotency, pruning.
  * embeddings  → `_fake_vector`, a deterministic bag-of-words hash embedding. Shared
                  words ⇒ closer vectors, so relevance assertions are exact and repeatable
                  and no API key is ever needed.

Postgres tests skip cleanly when no server is reachable; everything else always runs.
"""
from __future__ import annotations

import hashlib
import importlib.machinery
import io
import json
import math
import os
import re
import sys
import threading
import uuid
import zipfile
from functools import lru_cache
from pathlib import Path

import pytest

PG_URL = os.environ.get("POSTGRES_URL") or "postgresql://postgres:pg@localhost:55432/cerebrozen"


# ── the deterministic fake embedder ──────────────────────────────────────────
# Hash-based, so it is stable across processes (Python's built-in hash() is NOT —
# PYTHONHASHSEED would make every relevance assertion flaky).

_DIM = 16


def _fake_vector(text: str) -> list[float]:
    """A bag-of-words hash embedding: same text ⇒ same vector, shared words ⇒ closer."""
    vec = [0.0] * _DIM
    for word in re.findall(r"[a-z0-9]+", (text or "").lower()):
        vec[int(hashlib.sha1(word.encode()).hexdigest(), 16) % _DIM] += 1.0
    if not any(vec):
        vec[0] = 1.0
    norm = math.sqrt(sum(v * v for v in vec))
    return [v / norm for v in vec]


@pytest.fixture
def fake_embedder(monkeypatch):
    """Swap the embedding API for a deterministic local function.

    Retrieval quality is not what these tests measure — the plumbing is. A real embedding
    call would make them slow, non-deterministic, and dependent on a key we do not have.
    """
    from app.rag import embedder

    monkeypatch.setattr(embedder, "embed", lambda texts: [_fake_vector(t) for t in texts])
    monkeypatch.setattr(embedder, "embed_one", lambda text: _fake_vector(text))
    monkeypatch.setattr(embedder, "embedding_dim", lambda: _DIM)
    return _fake_vector


# ── a real, local LanceDB (no S3) ────────────────────────────────────────────


@pytest.fixture
def lance(tmp_path, monkeypatch):
    """The REAL LanceDB, in tmp_path. `_connect` is lru_cached, so the cache is cleared
    on both sides — a leaked connection would silently point the next test at this dir."""
    from app import config
    from app.rag import store

    monkeypatch.delenv("CEREBROZEN_RAG_BACKEND", raising=False)
    monkeypatch.setattr(config, "RAG_LANCEDB_URI", str(tmp_path / "lancedb"))
    store._connect.cache_clear()
    yield store
    store._connect.cache_clear()


def _rec(rec_id: str, text: str, **extra) -> dict:
    row = {"id": rec_id, "text": text, "vector": _fake_vector(text)}
    row.update(extra)
    return row


# ── a real Postgres ──────────────────────────────────────────────────────────


@lru_cache(maxsize=1)
def _pg_ready() -> bool:
    """Is a Postgres with pgvector reachable? The suite must still pass without one."""
    try:
        import psycopg

        with psycopg.connect(PG_URL, connect_timeout=3) as conn:
            conn.execute("CREATE EXTENSION IF NOT EXISTS vector")
        return True
    except Exception:  # noqa: BLE001
        return False


requires_pg = pytest.mark.skipif(
    not _pg_ready(), reason=f"no Postgres/pgvector reachable at {PG_URL}"
)


class _PgHarness:
    """Hands out uniquely-named collections and drops their tables afterwards, so tests
    can never see each other's rows. It holds its OWN pool reference, so a test that
    re-points `pg._pool` (to simulate an unconfigured Postgres) can still be cleaned up."""

    def __init__(self, pg, pool):
        self.pg, self.pool = pg, pool
        self.tables: list[str] = []

    def collection(self, prefix: str = "t"):
        name = f"{prefix}_{uuid.uuid4().hex[:10]}"
        self.tables.append(name)
        return self.pg.collection(name)

    def rows(self, table: str) -> list[dict]:
        """Read a table straight from Postgres — never through the shim under test."""
        with self.pool.connection() as conn:
            return [r[0] for r in conn.execute(f'SELECT doc FROM "{table}"').fetchall()]

    def sql(self, statement: str, args=()):
        with self.pool.connection() as conn:
            return conn.execute(statement, args).fetchall()


@pytest.fixture
def pgdb(monkeypatch):
    """A live Postgres, wired into app.stores.pg through its own env/globals."""
    from app.stores import pg

    monkeypatch.setenv("POSTGRES_URL", PG_URL)
    monkeypatch.setattr(pg, "_pool", None)
    monkeypatch.setattr(pg, "_ensured", set())
    monkeypatch.setattr(pg, "_collections", {})

    pool = pg.get_pool()
    assert pool is not None, "the fixture is guarded by requires_pg — this cannot be None"
    harness = _PgHarness(pg, pool)
    yield harness

    with pool.connection() as conn:
        for table in harness.tables:
            conn.execute(f'DROP TABLE IF EXISTS "{table}"')
    pool.close()


@pytest.fixture
def pgvec(pgdb, monkeypatch):
    """The pgvector RAG backend, switched on and pointed at the live Postgres."""
    from app.rag import pgvector_store as pgv

    monkeypatch.setenv("CEREBROZEN_RAG_BACKEND", "pgvector")
    monkeypatch.setattr(pgv, "_ready", set())
    pgdb.tables.extend(["rag_sskb", "rag_cskb"])
    with pgdb.pool.connection() as conn:
        conn.execute("DROP TABLE IF EXISTS rag_sskb")
        conn.execute("DROP TABLE IF EXISTS rag_cskb")
    return pgv


# ── the faked network boundaries ─────────────────────────────────────────────


class _FakeS3:
    """An in-process S3. The ONLY thing faked in the ingest tests: everything downstream
    (chunking, metadata, ETags, embedding, LanceDB) is the real code path."""

    def __init__(self, objects: dict[str, bytes]):
        self.objects = dict(objects)

    def etag(self, key: str) -> str:
        return hashlib.md5(self.objects[key]).hexdigest()

    def get_paginator(self, name: str):
        assert name == "list_objects_v2"
        outer = self

        class _Paginator:
            def paginate(self, Bucket, Prefix):  # noqa: N803 — boto3's kwarg names
                yield {
                    "Contents": [
                        {"Key": k, "ETag": f'"{outer.etag(k)}"'}
                        for k in sorted(outer.objects)
                        if k.startswith(Prefix)
                    ]
                }

        return _Paginator()

    def download_file(self, bucket, key, dest):
        Path(dest).write_bytes(self.objects[key])


class _FakeBoto3:
    """Just enough boto3 for the credential/region chain, with no network at all."""

    def __init__(self, location="us-west-2", creds=True, location_raises=False):
        self._location, self._creds, self._raises = location, creds, location_raises
        # startup._missing_ingest_deps() calls importlib.util.find_spec("boto3"), which
        # rejects a module object with no __spec__.
        self.__spec__ = importlib.machinery.ModuleSpec("boto3", loader=None)
        self.__name__ = "boto3"

    def client(self, name, **kwargs):
        outer = self

        class _S3:
            def get_bucket_location(self, Bucket):  # noqa: N803
                if outer._raises:
                    raise RuntimeError("no such bucket")
                return {"LocationConstraint": outer._location}

        return _S3()

    def Session(self):  # noqa: N802 — boto3's own capitalisation
        outer = self

        class _Frozen:
            access_key, secret_key, token = "AK", "SK", "TOK"

        class _Creds:
            def get_frozen_credentials(self):
                return _Frozen()

        class _Session:
            def get_credentials(self):
                return _Creds() if outer._creds else None

        return _Session()


# --- real document bytes (built here, so ingestion parses genuine files) ------


def _pdf_bytes(text: str) -> bytes:
    content = f"BT /F1 12 Tf 20 100 Td ({text}) Tj ET".encode()
    objects = [
        b"<</Type/Catalog/Pages 2 0 R>>",
        b"<</Type/Pages/Kids[3 0 R]/Count 1>>",
        b"<</Type/Page/Parent 2 0 R/MediaBox[0 0 300 300]/Contents 4 0 R"
        b"/Resources<</Font<</F1 5 0 R>>>>>>",
        b"<</Length " + str(len(content)).encode() + b">>stream\n" + content + b"\nendstream",
        b"<</Type/Font/Subtype/Type1/BaseFont/Helvetica>>",
    ]
    out, offsets = bytearray(b"%PDF-1.4\n"), []
    for i, obj in enumerate(objects, 1):
        offsets.append(len(out))
        out += f"{i} 0 obj".encode() + obj + b"endobj\n"
    xref = len(out)
    out += f"xref\n0 {len(objects) + 1}\n0000000000 65535 f \n".encode()
    for off in offsets:
        out += f"{off:010d} 00000 n \n".encode()
    out += f"trailer<</Size {len(objects) + 1}/Root 1 0 R>>\nstartxref\n{xref}\n%%EOF".encode()
    return bytes(out)


def _docx_bytes(paragraphs: list[str]) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as z:
        body = "".join(f"<w:p><w:r><w:t>{p}</w:t></w:r></w:p>" for p in paragraphs)
        z.writestr("word/document.xml", f"<w:document><w:body>{body}</w:body></w:document>")
    return buf.getvalue()


def _pptx_bytes(slides: dict[int, str]) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as z:
        for num, text in slides.items():
            z.writestr(f"ppt/slides/slide{num}.xml", f"<p:sld><a:p><a:t>{text}</a:t></a:p></p:sld>")
    return buf.getvalue()


def _curated_xlsx_bytes() -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Content Library"
    ws.append(["Name", "Author Name", "Synopsis", "Link", "Main Skill",
               "Level 1", "Level 2 - Final"])
    ws.append(["Radical Candor", "Kim Scott", "Care personally, challenge directly",
               "click here", "feedback", "manager", "senior"])
    ws["D2"].hyperlink = "https://www.youtube.com/watch?v=candor"
    ws.append(["", "Nobody", "a row with no name is not a record", "x", "", "", ""])
    ws.append(["Deep Work", "Cal Newport", "Focus without distraction",
               "https://example.com/deep-work.pdf", "focus", "ic", "mid"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def fake_s3(monkeypatch):
    """Patch the S3 CLIENT FACTORY only — `iter_s3_objects`, `download_to_temp`,
    `public_url` and every ingest code path above them stay real."""
    from app.rag import ingest

    def _install(objects: dict[str, bytes]) -> _FakeS3:
        s3 = _FakeS3(objects)
        monkeypatch.setattr(ingest, "s3_client", lambda: s3)
        return s3

    return _install


@pytest.fixture
def registry_isolation():
    """The registry is a process-wide lru_cached singleton. Any test that overlays a
    workbook must put it back, or it poisons every test that runs after it."""
    from app.rag.registry import get_registry

    yield
    get_registry().reload()


# ════════════════════════════════════════════════════════════════════════════
#  registry — the 9 extractions
# ════════════════════════════════════════════════════════════════════════════


def test_the_nine_extraction_tokens_are_all_bound():
    """The placeholder tokens ARE the trigger for retrieval: a prompt author writes
    {SSKB_Concept} and that presence is what fires Extract1. A token that loses its
    binding does not error — it is simply blanked out of the prompt, and the coach
    quietly stops receiving that evidence. Pin all nine bindings and their KBs."""
    from app.rag.registry import get_registry

    reg = get_registry()
    expected = {
        "SSKB_Concept": ("Extract1", "sskb"),
        "CSKB_Framework": ("Extract2", "cskb"),
        "CSKB_Values": ("Extract3", "cskb"),
        "CSKB_Competencies": ("Extract4", "cskb"),
        "CSKB_LearningAid": ("Extract5", "cskb"),
        "SSKB_MicroLearning": ("Extract6", "sskb"),
        "SSKB_CuratedContent": ("Extract7", "sskb"),
        "LearningAid": ("LearningAidSelect", "sskb"),
        "SSKB_Competencies": ("Extract8", "sskb"),
    }
    assert set(reg.binding_tokens()) == set(expected), "an extraction token lost its binding"
    for token, (extract_id, kb) in expected.items():
        ex = reg.by_token(token)
        assert ex is not None and ex.extract_id == extract_id and ex.kb == kb
        assert reg.by_id(extract_id) is ex
    assert len(reg.all()) == 9
    assert reg.by_token("NotAToken") is None and reg.by_id("Extract99") is None


def test_source_links_are_carried_only_where_the_framework_says():
    """A source link is shown to the user as "here's where this came from". The CSKB
    extractions + curated SSKB content have one; concepts, micro-learning and the SSKB
    competency framework do NOT — surfacing a link for those means citing a document the
    user has no right (or no way) to open."""
    from app.rag.registry import get_registry

    reg = get_registry()
    with_link = {e.extract_id for e in reg.all() if e.source_required}
    assert with_link == {"Extract2", "Extract3", "Extract4", "Extract5", "Extract7"}
    # CSKB is org-gated; SSKB always runs. A CSKB extraction that lost its gate would
    # search another client's knowledge base.
    for ex in reg.all():
        assert (ex.condition == "org_available") == (ex.kb == "cskb")
        if ex.kb == "cskb":
            assert ex.filters.get("org_id") == "org_id", "CSKB must be org-scoped"


def test_workbook_sheet_overrides_the_seed(tmp_path, monkeypatch, registry_isolation):
    """The registry is business-editable: an `extraction` sheet in the prompt workbook
    overrides the in-code seed field-by-field. If the overlay silently no-ops, every
    change the business team makes in Excel is ignored while the UI says it saved."""
    from openpyxl import Workbook

    from app.llm import prompt_store
    from app.rag import registry

    wb = Workbook()
    ws = wb.active
    ws.title = "extraction"
    ws.append(["extract_id", "top_k", "needs_llm", "filters", "query_params", "enabled",
               "null_text", "placeholder", "kb", "condition", "source_required",
               "output_fields", "used_in"])
    ws.append(["Extract1", 3, "FALSE", "source=_const:concept, level=user_level",
               "session_goal", "TRUE", "no concept found", "SSKB_Concept", "SSKB",
               "always", "no", "concept_name, concept_description", "CIM, CH"])
    ws.append(["Extract7", "not-a-number", "", "", "", "FALSE", "", "", "", "", "", "", ""])
    ws.append(["Extract99", 1, "TRUE", "", "", "TRUE", "", "", "", "", "", "", ""])
    ws.append([None, 5, "TRUE", "", "", "", "", "", "", "", "", "", ""])  # no id → ignored
    path = tmp_path / "workbook.xlsx"
    wb.save(path)

    monkeypatch.setattr(prompt_store, "WORKBOOK_CACHE_PATH", tmp_path / "absent.xlsx")
    monkeypatch.setattr(prompt_store, "resolve_workbook_path", lambda: str(path))
    monkeypatch.setattr(registry.config, "RAG_REGISTRY_SHEET", "extraction")

    reg = registry.reload_registry()

    ex1 = reg.by_id("Extract1")
    assert ex1.top_k == 3 and ex1.needs_llm is False
    assert ex1.filters == {"source": "_const:concept", "level": "user_level"}
    assert ex1.query_params == ["session_goal"], "a single value is still a one-item list"
    assert ex1.null_text == "no concept found"
    assert ex1.kb == "sskb", "kb is lower-cased so `_table_for` matches"
    assert ex1.source_required is False and ex1.used_in == ["CIM", "CH"]

    # An UNPARSEABLE top_k must be ignored, not crash the sheet and not become 0 — a
    # top_k of 0 would silently retrieve nothing at all.
    assert reg.by_id("Extract7").top_k == 5

    # enabled=FALSE unbinds the token: the extraction still exists but stops firing.
    assert reg.by_id("Extract7").enabled is False
    assert reg.by_token("SSKB_CuratedContent") is None
    assert "SSKB_CuratedContent" not in reg.binding_tokens()

    # A row for an extraction the code does not implement is DROPPED, not half-created.
    assert reg.by_id("Extract99") is None
    assert len(reg.all()) == 9, "the 9 seeded extractions, patched — never a 10th from a sheet"
    assert len(reg.binding_tokens()) == 8, "only the DISABLED one lost its token"


def test_a_broken_workbook_never_takes_retrieval_down(tmp_path, monkeypatch, registry_isolation):
    """A corrupt/renamed workbook is an editing accident, not an outage: the seed must
    keep serving all 9 extractions. The alternative is a bad Excel upload silently
    unbinding every RAG token in production."""
    from app.llm import prompt_store
    from app.rag import registry

    broken = tmp_path / "broken.xlsx"
    broken.write_bytes(b"this is not a zip, let alone a workbook")
    monkeypatch.setattr(prompt_store, "WORKBOOK_CACHE_PATH", tmp_path / "absent.xlsx")
    monkeypatch.setattr(prompt_store, "resolve_workbook_path", lambda: str(broken))

    reg = registry.reload_registry()
    assert len(reg.all()) == 9 and len(reg.binding_tokens()) == 9


def test_a_workbook_without_an_extraction_sheet_leaves_the_seed_alone(
    tmp_path, monkeypatch, registry_isolation
):
    """The sheet name is matched EXACTLY (config.RAG_REGISTRY_SHEET). A workbook that
    doesn't carry it must fall back to the seed silently — this is the normal case for a
    client who never edits extractions."""
    from openpyxl import Workbook

    from app.llm import prompt_store
    from app.rag import registry

    wb = Workbook()
    wb.active.title = "some_other_sheet"
    path = tmp_path / "no_extraction_tab.xlsx"
    wb.save(path)
    monkeypatch.setattr(prompt_store, "WORKBOOK_CACHE_PATH", tmp_path / "absent.xlsx")
    monkeypatch.setattr(prompt_store, "resolve_workbook_path", lambda: str(path))

    assert registry._load_sheet() == {}
    assert len(registry.reload_registry().all()) == 9

    # An `extraction` sheet with only a header (the state right after someone creates the
    # tab) must also be a no-op rather than wiping every extraction.
    wb2 = Workbook()
    wb2.active.title = "extraction"
    empty = tmp_path / "header_only.xlsx"
    wb2.save(empty)
    monkeypatch.setattr(prompt_store, "resolve_workbook_path", lambda: str(empty))
    assert registry._load_sheet() == {}


def test_sheet_cells_are_coerced_from_what_excel_actually_produces():
    """Business users type `TRUE`, `yes`, `1`, and comma-separated lists into Excel. The
    coercion layer is where a typo becomes a silently-disabled extraction."""
    from app.rag.registry import _coerce_overrides

    patch = _coerce_overrides({
        "extract_id": "Extract2", "needs_llm": "yes", "enabled": "1",
        "source_required": "nope", "top_k": "0", "kb": " CSKB ",
        "query_params": "a, b ,, c", "output_fields": "x", "used_in": "CIM",
        "filters": "org_id=org_id, doc_type=_const:frameworks, junk-no-equals",
        "placeholder": " CSKB_Framework ", "condition": "always", "null_text": " none ",
    })
    assert patch["needs_llm"] is True and patch["enabled"] is True
    assert patch["source_required"] is False, "anything but true/1/yes/y is False"
    assert patch["top_k"] == 1, "top_k is floored at 1 — 0 would retrieve nothing"
    assert patch["kb"] == "cskb" and patch["placeholder"] == "CSKB_Framework"
    assert patch["query_params"] == ["a", "b", "c"], "empty list entries are dropped"
    assert patch["filters"] == {"org_id": "org_id", "doc_type": "_const:frameworks"}
    assert patch["null_text"] == "none"

    # Blank/None cells must NOT be written as overrides — an empty Excel cell means
    # "leave the seed alone", never "set this field to empty".
    assert _coerce_overrides({"extract_id": "Extract1", "top_k": None, "kb": "",
                              "needs_llm": "", "filters": "", "query_params": None}) == {}


# ════════════════════════════════════════════════════════════════════════════
#  placeholders — the prompt/data contract
# ════════════════════════════════════════════════════════════════════════════


def test_time_is_resolved_at_HOUR_granularity():
    """{Time} sits ~130 tokens into a ~27,000-token system prompt. With microsecond
    precision the prefix changed on EVERY turn, so the LLM prompt cache could never hit:
    ~0% cached, ~21,000 tokens re-encoded per turn (measured). Rounding to the hour makes
    the instruction body byte-identical across a session.

    This is a COST test disguised as a formatting test: any drift back to sub-hour
    precision (seconds, microseconds, an ISO timestamp) re-breaks the cache silently.
    """
    from datetime import datetime, timezone

    from app.rag.placeholders import PlaceholderResolver

    resolver = PlaceholderResolver({}, rag_enabled=False)
    first = resolver.resolve_text("now: {Time}")
    second = resolver.resolve_text("now: {Time}")

    assert first == second, "two resolutions in the same hour must be byte-identical"
    stamp = first.removeprefix("now: ")
    assert re.fullmatch(r"\d{4}-\d{2}-\d{2}T\d{2}:00Z", stamp), (
        f"{stamp!r} is not hour-granular — sub-hour precision destroys the prompt cache"
    )
    assert stamp == datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:00Z")
    assert ":00Z" in stamp and stamp.count(":") == 1


def test_an_unresolvable_placeholder_is_blanked_never_leaked():
    """A prompt token with no value source must be BLANKED. Leaving `{user_name}` in the
    text ships prompt scaffolding into the coach's own reply — the user sees the template.
    Retrieval is disabled here, so even a RAG-shaped token is just an unknown token."""
    from app.rag.placeholders import PlaceholderResolver

    resolver = PlaceholderResolver({"userPosition": "CTO"}, rag_enabled=False)
    out = resolver.resolve_text("Hi {user_name}, you are {userPosition}. {SSKB_Concept}{Nope}")

    assert out == "Hi , you are CTO. "
    assert "{" not in out and "}" not in out, "a raw token must never survive resolution"


def test_context_tokens_resolve_from_the_turn_context():
    """Every non-RAG token comes from the turn context: exact key, case-insensitive key,
    Mongo dot-path (dynamic_vars writes nested dicts), the {userName} alias, and the
    non-string types the stores actually hold (lists, dicts, ints)."""
    from app.rag.placeholders import PlaceholderResolver

    ctx = {
        "name": "Ada",
        "userPosition": "Engineering Manager",
        "level": 3,
        "values": ["trust", "candor"],
        "coaching_style_context": {"selected_style": "socratic", "depth": {"n": 2}},
        "profile": {"tags": {"a": 1}},
    }
    resolver = PlaceholderResolver(ctx, rag_enabled=False)

    assert resolver.resolve_text("{userName}") == "Ada", "the {userName} alias maps to `name`"
    assert resolver.resolve_text("{userposition}") == "Engineering Manager", "case-insensitive"
    assert resolver.resolve_text("{level}") == "3", "non-strings are stringified"
    assert resolver.resolve_text("{values}") == "trust, candor", "lists join with commas"
    assert resolver.resolve_text("{coaching_style_context.selected_style}") == "socratic"
    assert resolver.resolve_text("{coaching_style_context.DEPTH.n}") == "2", "dot-path is case-insensitive"
    assert resolver.resolve_text("{profile.tags}") == "a: 1", "dicts render as k: v"
    # A dot-path that bottoms out on a scalar, or misses a segment, is blanked — not left.
    assert resolver.resolve_text("{name.missing}") == ""
    assert resolver.resolve_text("{coaching_style_context.absent}") == ""
    # Empty text short-circuits (no tokens, no work).
    assert resolver.resolve_text("") == "" and resolver.resolve_text("no tokens here") == "no tokens here"


def test_user_context_alias_and_precedence():
    """`user_context` is the back-compat alias for the old call sites; `context` wins when
    both carry the same key, or a stale profile would override the live turn."""
    from app.rag.placeholders import PlaceholderResolver

    resolver = PlaceholderResolver(
        {"userName": "live"}, user_context={"userName": "stale", "org": "acme"},
        rag_enabled=False,
    )
    assert resolver.resolve_text("{userName} at {org}") == "live at acme"


@pytest.fixture
def rag_e2e(lance, fake_embedder, monkeypatch):
    """A real curated-content corpus in a real LanceDB, ready for the resolver.

    The extraction result cache is switched OFF: it is backed by a process-wide fakeredis,
    so a cached hit from an earlier test would answer for a store that is now empty and
    the degradation assertions below would pass for the wrong reason.
    """
    from app import config
    from app.rag import store

    monkeypatch.setattr(config, "RAG_CACHE_TTL_S", 0)
    store.upsert("sskb", [
        _rec("c1", "delegation trust engineering manager",
             kb="sskb", source="curated", item_type="curated_content",
             title="Letting Go", author="Kim Scott",
             source_link="https://example.com/letting-go",
             content_format="article", synopsis="How managers delegate"),
        _rec("c2", "spreadsheets accounting quarterly close",
             kb="sskb", source="curated", item_type="curated_content",
             title="Closing The Books", author="Someone Else",
             source_link="https://example.com/books", content_format="pdf"),
    ])
    return store


def test_a_rag_token_is_replaced_by_retrieved_evidence(rag_e2e):
    """The whole point of the layer: {SSKB_CuratedContent} in a prompt must come back as
    the retrieved item's fields. Extract7 is the deterministic path (needs_llm=False), so
    this exercises registry → resolver → embedder → LanceDB → formatter with no LLM."""
    from app.rag.placeholders import PlaceholderResolver

    resolver = PlaceholderResolver({
        "userRoleContext": "engineering manager",
        "user_level": "delegation trust",
        "user_id": "u1",
        "invoking_agent": "core_coaching_agent",
    })
    out = resolver.resolve_text("Suggest this:\n{SSKB_CuratedContent}")

    assert "{SSKB_CuratedContent}" not in out
    assert "Heading: Letting Go" in out, "the nearest curated item must win, not the first row"
    assert "Author: Kim Scott" in out
    assert "Source Link: https://example.com/letting-go" in out
    assert "Content Format: article" in out
    assert "Sub Heading: How managers delegate" in out, (
        "sub_heading is mapped from the `meta` blob (meta:synopsis) — if the store stops "
        "round-tripping meta, this field silently empties"
    )
    assert "Closing The Books" not in out


def test_rag_degrades_SILENTLY_when_the_knowledge_base_is_unreachable(
    lance, fake_embedder, monkeypatch
):
    """THE production failure mode. When the vector store is empty/unreachable, every
    extraction returns null, the token is blanked, and NOTHING raises: the prompt is still
    well-formed, the coach still answers — with no evidence base at all.

    This test exists to make that behaviour explicit and intentional. The prompt must stay
    valid (no leaked token, no exception), and the evidence must be provably absent."""
    from app import config
    from app.rag.placeholders import PlaceholderResolver

    monkeypatch.setattr(config, "RAG_CACHE_TTL_S", 0)
    assert lance.count("sskb") == 0, "no table at all — the store is 'down'"

    resolver = PlaceholderResolver({"userRoleContext": "engineering manager",
                                    "user_level": "senior"})
    out = resolver.resolve_text("Suggest this:\n{SSKB_CuratedContent}\n-- end")

    assert out == "Suggest this:\n\n-- end", "a null extraction blanks its token"
    assert "{" not in out, "the token must never leak to the model/user"
    assert lance.search("sskb", _fake_vector("anything")) == [], "search degrades to []"


def test_several_rag_tokens_in_one_prompt_resolve_concurrently(rag_e2e, monkeypatch):
    """Two+ RAG tokens fan out to a thread pool with a FRESH copied context per task.
    Sharing one contextvars.Context across threads raises RuntimeError (it is not
    reentrant) and silently drops a token's retrieval — which shows up as a blank section
    in the prompt, never as an error."""
    from app.rag.placeholders import PlaceholderResolver

    resolver = PlaceholderResolver({
        "userRoleContext": "engineering manager",
        "user_level": "delegation trust",
        "session_goal": "delegation trust",
        "org_id": "",  # no org → the CSKB tokens are condition-gated to null
    })
    out = resolver.resolve_text(
        "A:{SSKB_CuratedContent}\nB:{SSKB_Concept}\nC:{CSKB_Framework}\nD:{userName}"
    )

    assert "Letting Go" in out, "the curated token resolved"
    assert "{SSKB_Concept}" not in out and "{CSKB_Framework}" not in out
    assert "{" not in out


def test_rag_disabled_turns_every_token_into_a_context_lookup(rag_e2e):
    """`rag_enabled=False` is the kill switch (voice/latency paths use it). It must not
    retrieve at all — the RAG token becomes an ordinary unresolved token and is blanked."""
    from app.rag.placeholders import PlaceholderResolver

    resolver = PlaceholderResolver({"userRoleContext": "engineering manager"}, rag_enabled=False)
    assert resolver.resolve_text("x{SSKB_CuratedContent}y") == "xy"


def test_a_repeated_token_is_retrieved_once_and_substituted_everywhere(rag_e2e):
    """The same token can appear several times in one prompt. Each occurrence must be
    filled, but the (embedding + vector search) work must happen ONCE — a prompt that
    mentions a token five times would otherwise cost five retrievals per turn."""
    from app.rag.placeholders import PlaceholderResolver

    calls: list[str] = []
    resolver = PlaceholderResolver({"userRoleContext": "engineering manager",
                                    "user_level": "delegation trust"})
    original = resolver._resolve_rag
    resolver._resolve_rag = lambda token: (calls.append(token), original(token))[1]

    out = resolver.resolve_text("{SSKB_CuratedContent} … again: {SSKB_CuratedContent}")

    assert out.count("Letting Go") == 2, "every occurrence is substituted"
    assert calls == ["{SSKB_CuratedContent}"], "but the retrieval ran only once"


@requires_pg
def test_a_failed_extraction_leaves_its_token_for_the_NEXT_turn(pgvec, fake_embedder, monkeypatch):
    """Null and ERROR are different outcomes and must behave differently:

      * null  → the knowledge base genuinely has nothing → blank the token.
      * error → the retrieval itself broke → LEAVE the token, so the next turn retries it.

    Blanking on error would make a transient outage look like "no evidence exists" and the
    token would never be retried. This drives a REAL error: an index built with one
    embedder, queried after the embedder changed dimensions."""
    from app import config
    from app.rag import embedder
    from app.rag.placeholders import PlaceholderResolver

    monkeypatch.setattr(config, "RAG_CACHE_TTL_S", 0)
    pgvec.upsert("sskb", [{"id": "a", "doc_key": "k", "text": "delegation",
                           "embedding": _fake_vector("delegation"), "source": "curated"}])

    pgvec._ready.clear()  # a restart, with a DIFFERENT embedder configured
    monkeypatch.setattr(embedder, "embed_one", lambda text: [0.1] * 8)

    resolver = PlaceholderResolver({"userRoleContext": "manager", "user_level": "senior"})
    out = resolver.resolve_text("Aid: {SSKB_CuratedContent}")

    assert out == "Aid: {SSKB_CuratedContent}", (
        "a BROKEN retrieval must leave its token untouched to be retried next turn"
    )


def test_resolution_never_raises_into_the_coaching_turn(rag_e2e, monkeypatch):
    """The resolver runs inside the turn. Whatever the extraction layer does — including
    blowing up in a worker thread — the turn must still produce a prompt. (Simulated by
    breaking `extract`, the resolver's one dependency; the resolver itself is real.)"""
    from app.rag import placeholders

    def _boom(extract_id, params):
        raise RuntimeError("the extraction layer is on fire")

    monkeypatch.setattr(placeholders, "extract", _boom)
    resolver = placeholders.PlaceholderResolver({"userRoleContext": "manager"})

    out = resolver.resolve_text("Aid: {SSKB_CuratedContent} for {userName}")
    assert out == "Aid: {SSKB_CuratedContent} for ", "the token is kept, nothing propagates"

    # A token that is not in the registry can never be resolved as RAG.
    assert resolver._resolve_rag("{NotRegistered}") is None


# ════════════════════════════════════════════════════════════════════════════
#  store.py — the LanceDB path
# ════════════════════════════════════════════════════════════════════════════


def test_upsert_and_search_round_trip_through_a_real_lancedb(lance):
    """Nearest-neighbour ordering, the fixed schema, and the meta blob. `_score` is a
    LanceDB DISTANCE (lower = closer) — callers rank on it."""
    written = lance.upsert("sskb", [
        _rec("a", "psychological safety in teams", kb="sskb", source="concept",
             title="Safety", extra_field="rides in meta"),
        _rec("b", "quarterly revenue forecasting", kb="sskb", source="concept", title="Revenue"),
    ])
    assert written == 2 and lance.count("sskb") == 2

    hits = lance.search("sskb", _fake_vector("psychological safety in teams"), top_k=2)
    assert [h["id"] for h in hits] == ["a", "b"], "the nearest row must come first"
    assert hits[0]["_score"] < hits[1]["_score"], "LanceDB _score is a distance: lower is closer"
    assert hits[0]["title"] == "Safety" and hits[0]["kb"] == "sskb"
    assert hits[0]["meta"] == {"extra_field": "rides in meta"}, "unknown keys ride in `meta`"
    assert hits[1]["meta"] == {}, "no extras → an empty dict, never a crash"
    assert hits[0]["org_id"] == "", "an unset scalar column is '' — never None"

    # top_k is a real limit, not a suggestion (it bounds the LLM's candidate pool).
    assert len(lance.search("sskb", _fake_vector("anything"), top_k=1)) == 1
    # An empty query vector (embedding failed upstream) must degrade to [], not explode.
    assert lance.search("sskb", []) == []


def test_reingesting_the_same_doc_replaces_rather_than_duplicates(lance):
    """Ingestion is re-run on every boot. If upsert appended instead of replacing, the
    corpus would double on each restart and the same passage would win top-k twice."""
    lance.upsert("sskb", [_rec("a", "first version", kb="sskb", source="concept")])
    lance.upsert("sskb", [_rec("a", "second version", kb="sskb", source="concept"),
                          _rec("b", "another row", kb="sskb", source="concept")])

    assert lance.count("sskb") == 2
    hits = lance.search("sskb", _fake_vector("second version"), top_k=1)
    assert hits[0]["id"] == "a" and hits[0]["text"] == "second version"

    # A record with no id can't be de-duplicated — it must still be written, not dropped.
    assert lance.upsert("sskb", [{"text": "no id", "vector": _fake_vector("no id")}]) == 1
    assert lance.count("sskb") == 3


def test_the_metadata_prefilter_is_what_keeps_the_two_KBs_and_orgs_apart(lance):
    """Filters run BEFORE ranking. Extract1 filters source=concept; a CSKB extraction
    filters org_id. A filter that silently no-ops means another client's document can
    surface inside this client's coaching prompt."""
    lance.upsert("sskb", [
        _rec("concept", "leadership presence", kb="sskb", source="concept"),
        _rec("micro", "leadership presence", kb="sskb", source="micro_learning"),
    ])
    lance.upsert("cskb", [
        _rec("acme", "leadership presence", kb="cskb", org_id="acme", doc_type="frameworks"),
        _rec("rival", "leadership presence", kb="cskb", org_id="rival", doc_type="frameworks"),
    ])
    query = _fake_vector("leadership presence")

    only_concept = lance.search("sskb", query, filters={"source": "concept"}, top_k=10)
    assert [h["id"] for h in only_concept] == ["concept"]

    acme = lance.search("cskb", query, filters={"org_id": "acme", "doc_type": "frameworks"}, top_k=10)
    assert [h["id"] for h in acme] == ["acme"], "an org filter must never leak another org"

    # The two tables are structurally isolated: an SSKB query cannot reach CSKB rows.
    assert {h["id"] for h in lance.search("sskb", query, top_k=10)} == {"concept", "micro"}

    # A list filter becomes IN(...); an unknown column is ignored rather than injected.
    both = lance.search("sskb", query, filters={"source": ["concept", "micro_learning"],
                                                "not_a_column": "x"}, top_k=10)
    assert {h["id"] for h in both} == {"concept", "micro"}


def test_a_filter_value_with_a_quote_cannot_break_the_where_clause(lance):
    """Filter values come from org names and doc types — i.e. from data. An unescaped
    apostrophe would either error out (retrieval dies) or, worse, alter the predicate."""
    from app.rag.store import _where_clause

    assert _where_clause({"org_id": "o'brien"}) == "org_id = 'o''brien'"
    assert _where_clause({"source": ["a'b", "c"]}) == "source IN ('a''b','c')"
    assert _where_clause({}) is None and _where_clause(None) is None
    assert _where_clause({"org_id": "", "source": None, "level": []}) is None, "empty → no filter"
    assert _where_clause({"source": ["", None]}) is None, "a list of empty values is no filter"

    lance.upsert("cskb", [_rec("x", "team norms", kb="cskb", org_id="o'brien")])
    hits = lance.search("cskb", _fake_vector("team norms"), filters={"org_id": "o'brien"})
    assert [h["id"] for h in hits] == ["x"]


def test_records_are_coerced_into_the_fixed_schema(lance):
    """LanceDB's schema is fixed on first write. A record whose columns are None/ints/dicts
    would either be rejected or poison the column types for the whole table."""
    from app.rag.store import _norm_record

    row = _norm_record({"id": 7, "text": "t", "vector": [0.1], "level": None,
                        "source": "concept", "nested": {"a": 1}})
    assert row["id"] == "7" and row["level"] == "" and row["source"] == "concept"
    assert json.loads(row["meta"]) == {"nested": {"a": 1}}
    assert set(row) == set(__import__("app.rag.store", fromlist=["x"]).SCALAR_COLUMNS) | {"vector", "meta"}

    assert lance.upsert("sskb", [{"id": 7, "text": "t", "vector": _fake_vector("t"),
                                  "level": None, "count": 3}]) == 1
    hit = lance.search("sskb", _fake_vector("t"))[0]
    assert hit["id"] == "7" and hit["meta"] == {"count": 3}


def test_changing_the_embedder_makes_lancedb_return_NOTHING_silently(lance):
    """A 16-dim index cannot be searched with an 8-dim vector. LanceDB raises — and
    store.search SWALLOWS it and returns []. So swapping the embedding model without
    re-ingesting doesn't error: the whole knowledge base just disappears, one null
    extraction at a time. (pgvector's guard, below, is the loud version of this.)"""
    lance.upsert("sskb", [_rec("a", "psychological safety", kb="sskb", source="concept")])
    assert lance.search("sskb", _fake_vector("psychological safety")) != []

    wrong_dim = [0.1] * 8
    assert lance.search("sskb", wrong_dim, top_k=3) == [], "the failure is INVISIBLE to callers"
    assert lance.upsert("sskb", [{"id": "b", "text": "x", "vector": wrong_dim}]) == 0, (
        "a failed upsert reports 0 rows written rather than raising"
    )
    assert lance.count("sskb") == 1, "and the existing index is left intact"


def test_a_missing_table_and_a_dead_store_both_degrade_to_empty(tmp_path, monkeypatch):
    """Two ways the store can be absent: the table was never created (nothing ingested
    yet), or the URI itself is unusable. Both must return neutral values so the graph
    keeps running — this is the contract the extractors rely on."""
    from app import config
    from app.rag import store

    monkeypatch.delenv("CEREBROZEN_RAG_BACKEND", raising=False)
    monkeypatch.setattr(config, "RAG_LANCEDB_URI", str(tmp_path / "empty"))
    store._connect.cache_clear()
    assert store.search("sskb", _fake_vector("q")) == []
    assert store.count("cskb") == 0 and store.indexed_docs("cskb") == {}
    store.delete_by_doc_key("sskb", "nothing")  # must not raise
    store.drop_tables()  # must not raise

    # A URI that cannot be opened (a FILE where a directory must be) → _connect() is None.
    blocker = tmp_path / "not-a-dir"
    blocker.write_text("x")
    monkeypatch.setattr(config, "RAG_LANCEDB_URI", str(blocker / "lancedb"))
    store._connect.cache_clear()
    assert store._connect() is None
    assert store.search("sskb", _fake_vector("q")) == [] and store.count("sskb") == 0
    assert store.upsert("sskb", [_rec("a", "x")]) == 0
    assert store.indexed_docs("sskb") == {}
    store.delete_by_doc_key("sskb", "k")
    store.drop_tables()
    store._connect.cache_clear()


def test_indexed_docs_and_delete_by_doc_key_drive_incremental_ingestion(lance):
    """Ingestion re-embeds a doc only when its S3 ETag changed. That decision is made from
    indexed_docs() — if it returns {} the whole corpus is re-embedded on every boot; if it
    returns stale ETags, a changed doc is never picked up."""
    lance.upsert("sskb", [
        _rec("a1", "chunk one", kb="sskb", doc_key="sskb/x.pdf", s3_etag="E1"),
        _rec("a2", "chunk two", kb="sskb", doc_key="sskb/x.pdf", s3_etag="E1"),
        _rec("b1", "other doc", kb="sskb", doc_key="sskb/y.pdf", s3_etag="E2"),
    ])
    assert lance.indexed_docs("sskb") == {"sskb/x.pdf": "E1", "sskb/y.pdf": "E2"}

    lance.delete_by_doc_key("sskb", "sskb/x.pdf")
    assert lance.count("sskb") == 1, "EVERY chunk of the doc must go, not just the first"
    assert lance.indexed_docs("sskb") == {"sskb/y.pdf": "E2"}

    lance.delete_by_doc_key("sskb", "")  # no key → no-op, never a full-table delete
    assert lance.count("sskb") == 1


def test_drop_tables_clears_both_knowledge_bases(lance):
    """RAG_REINDEX=true drops before rebuilding. A drop that misses one table leaves half
    the corpus at the old embedding model — i.e. unsearchable but present."""
    lance.upsert("sskb", [_rec("a", "x", kb="sskb")])
    lance.upsert("cskb", [_rec("b", "y", kb="cskb", org_id="acme")])
    assert lance.count("sskb") == 1 and lance.count("cskb") == 1

    lance.drop_tables()
    assert lance.count("sskb") == 0 and lance.count("cskb") == 0
    lance.drop_tables()  # idempotent: dropping an already-dropped table must not raise


def test_the_s3_index_is_opened_with_the_BUCKETs_region_not_the_apps(tmp_path, monkeypatch):
    """LanceDB's object store takes its SigV4 signing region from AWS_REGION and ignores
    storage_options when that env var is set. With the RAG bucket in us-east-1 and the app
    in ap-south-1, every S3 request was rejected (AuthorizationHeaderMalformed) and RAG
    went dark. The fix pins the env to the BUCKET's detected region — assert it holds.

    boto3 and lancedb.connect are stubbed here (they are the network); the region logic
    under test is real.
    """
    import lancedb

    from app import config
    from app.rag import store

    monkeypatch.setitem(sys.modules, "boto3", _FakeBoto3(location="us-east-2"))
    monkeypatch.setenv("AWS_REGION", "ap-south-1")
    monkeypatch.setattr(config, "RAG_LANCEDB_URI", "s3://rag-bucket/lancedb")
    captured = {}

    def _fake_connect(uri, **kwargs):
        captured["uri"], captured["kwargs"] = uri, kwargs
        return "connected"

    monkeypatch.setattr(lancedb, "connect", _fake_connect)
    store._connect.cache_clear()

    assert store._connect() == "connected"
    opts = captured["kwargs"]["storage_options"]
    assert opts["region"] == "us-east-2", "the BUCKET's region, not the app's"
    assert os.environ["AWS_REGION"] == "us-east-2" == os.environ["AWS_DEFAULT_REGION"]
    assert opts["endpoint"] == "https://s3.us-east-2.amazonaws.com"
    assert opts["aws_access_key_id"] == "AK" and opts["aws_session_token"] == "TOK", (
        "credentials must be resolved through boto3 so LanceDB authenticates the same way "
        "as the rest of the service (instance role / ~/.aws / env)"
    )
    store._connect.cache_clear()


def test_bucket_region_detection_covers_s3s_historical_quirks(monkeypatch):
    """us-east-1 answers with a NULL LocationConstraint and eu-west-1 with the legacy "EU".
    Mapping either one wrong points LanceDB at the wrong endpoint → every request fails."""
    from app.rag import store

    monkeypatch.setitem(sys.modules, "boto3", _FakeBoto3(location=None))
    assert store._resolve_bucket_region("b") == "us-east-1"
    monkeypatch.setitem(sys.modules, "boto3", _FakeBoto3(location="EU"))
    assert store._resolve_bucket_region("b") == "eu-west-1"
    monkeypatch.setitem(sys.modules, "boto3", _FakeBoto3(location="eu-central-1"))
    assert store._resolve_bucket_region("b") == "eu-central-1"
    # Detection failure → None, so the caller falls back to the configured region.
    monkeypatch.setitem(sys.modules, "boto3", _FakeBoto3(location_raises=True))
    assert store._resolve_bucket_region("b") is None


def test_a_corrupted_index_degrades_to_empty_instead_of_taking_the_turn_down(lance, tmp_path):
    """The index is a set of files in S3/on disk. They can be half-written, truncated, or
    deleted underneath a running instance. Every entry point must survive that: retrieval
    returns nothing, health reports 0 rows, and the coaching turn still completes."""
    lance.upsert("sskb", [_rec("a", "text", kb="sskb", doc_key="sskb/a.pdf", s3_etag="E1")])
    assert lance.count("sskb") == 1

    for manifest in (tmp_path / "lancedb" / "sskb.lance" / "_versions").glob("*"):
        manifest.write_bytes(b"a truncated / corrupt manifest")

    assert lance.count("sskb") == 0, "count degrades to 0, it does not raise"
    assert lance.search("sskb", _fake_vector("text")) == []
    assert lance.indexed_docs("sskb") == {}, "→ ingestion sees an empty index and rebuilds"
    lance.delete_by_doc_key("sskb", "sskb/a.pdf")  # must not raise


def test_an_index_written_by_an_older_schema_does_not_break_ingestion(lance, tmp_path):
    """indexed_docs() projects two columns. A table written before those columns existed
    makes BOTH the fast path and the full-scan fallback fail — which must mean "nothing is
    indexed" (re-ingest everything), never a crash at boot."""
    db = lance._connect()
    db.create_table("sskb", data=[{"id": "old", "text": "t", "vector": _fake_vector("t")}])

    assert lance.indexed_docs("sskb") == {}


def test_a_row_with_an_unparseable_meta_blob_still_returns_its_columns(lance):
    """`meta` is free-form JSON written by whatever ingested the row. A bad blob must not
    take out the whole search result — the scalar columns are what most extractions map."""
    lance.upsert("sskb", [_rec("a", "psychological safety", kb="sskb", title="Safety")])
    table = lance._connect().open_table("sskb")
    row = {col: "" for col in lance.SCALAR_COLUMNS}
    row.update({"id": "b", "text": "psychological safety", "title": "Broken",
                "meta": "{not json", "vector": _fake_vector("psychological safety")})
    table.add([row])

    hits = {h["id"]: h for h in lance.search("sskb", _fake_vector("psychological safety"), top_k=2)}
    assert hits["b"]["meta"] == {}, "an unparseable meta degrades to {}"
    assert hits["b"]["title"] == "Broken", "and the row's real columns still come through"


def test_lancedb_is_opened_without_credentials_when_boto3_has_none(tmp_path, monkeypatch):
    """Credential resolution is best-effort: on a box with no AWS creds (local dev) the
    store must still open, letting LanceDB's own object-store chain try. Raising here would
    make an offline install fail at import time."""
    import lancedb

    from app import config
    from app.rag import store

    class _NoCreds(_FakeBoto3):
        def Session(self):  # noqa: N802
            raise RuntimeError("no credential provider")

    monkeypatch.setattr(config, "RAG_LANCEDB_URI", "s3://rag-bucket/lancedb")
    captured = {}
    monkeypatch.setattr(lancedb, "connect",
                        lambda uri, **kw: captured.update(kw) or "connected")

    for boto3_stub in (_NoCreds(location="us-east-1"),          # the chain raises
                       _FakeBoto3(location="us-east-1", creds=False)):  # …or resolves nothing
        monkeypatch.setitem(sys.modules, "boto3", boto3_stub)
        store._connect.cache_clear()
        assert store._connect() == "connected"
        assert "aws_access_key_id" not in captured["storage_options"]
        assert captured["storage_options"]["region"] == "us-east-1"
    store._connect.cache_clear()


def test_the_store_disables_itself_when_lancedb_is_not_installed(monkeypatch):
    """`lancedb` is an optional dep. Its absence must disable RAG, not break the import of
    a coaching graph that would otherwise run fine without retrieval."""
    from app.rag import store

    monkeypatch.setitem(sys.modules, "lancedb", None)  # `import lancedb` → ImportError
    store._connect.cache_clear()
    assert store._connect() is None
    store._connect.cache_clear()


# ════════════════════════════════════════════════════════════════════════════
#  ingest — S3 → chunks → vectors
# ════════════════════════════════════════════════════════════════════════════


def test_stable_ids_and_content_format_detection():
    """`make_id` is what makes re-ingestion an UPSERT: an id that shifts between runs
    duplicates the corpus. `detect_content_format` drives what the UI renders (video vs
    article), and the content library has no explicit type column."""
    from app.rag.ingest import detect_content_format, make_id

    assert make_id("sskb", "k", "0") == make_id("sskb", "k", "0")
    assert make_id("sskb", "k", "0") != make_id("sskb", "k", "1")
    assert make_id("a", "", "b") == make_id("a", "b"), "empty parts are dropped, not encoded"

    assert detect_content_format("https://youtu.be/x") == "video"
    assert detect_content_format("https://www.youtube.com/watch?v=x") == "video"
    assert detect_content_format("https://vimeo.com/x") == "video"
    assert detect_content_format("https://open.spotify.com/e/x") == "audio"
    assert detect_content_format("https://anchor.fm/podcast/x") == "audio"
    assert detect_content_format("https://soundcloud.com/x") == "audio"
    assert detect_content_format("https://x.com/paper.PDF") == "pdf"
    assert detect_content_format("https://hbr.org/2020/01/x") == "article"
    assert detect_content_format("") == "" and detect_content_format(None) == ""


def test_chunking_keeps_an_overlap_so_a_sentence_is_never_cut_in_half():
    """Chunks are what get embedded. Without the overlap, a passage that straddles a
    boundary is un-retrievable from either half."""
    from app.rag.ingest import chunk_text

    assert chunk_text("") == [] and chunk_text("   ") == []
    assert chunk_text("short") == ["short"], "text under the window is one chunk"

    text = "".join(chr(ord("a") + i % 26) for i in range(3000))
    chunks = chunk_text(text, size=1200, overlap=150)
    assert [len(c) for c in chunks] == [1200, 1200, 900], "the window advances by size-overlap"
    assert chunks[0][-150:] == chunks[1][:150], "consecutive chunks overlap by `overlap`"
    assert chunks[1][-150:] == chunks[2][:150]
    assert chunks[0][:1050] + chunks[1][:1050] + chunks[2] == text, "no character is lost"


def test_every_document_format_the_kbs_actually_hold_is_parseable(tmp_path):
    """The KBs are PDFs, DOCX and PPTX. A format that extracts to "" is silently skipped by
    embed_and_upsert — the doc looks ingested (no error) and is simply never retrievable."""
    from app.rag.ingest import extract_text

    pdf = tmp_path / "a.pdf"
    pdf.write_bytes(_pdf_bytes("Psychological safety at work"))
    assert "Psychological safety" in extract_text(str(pdf), "a.pdf")

    docx = tmp_path / "b.docx"
    docx.write_bytes(_docx_bytes(["Growth mindset", "Feedback &amp; trust"]))
    text = extract_text(str(docx), "b.docx")
    assert "Growth mindset" in text and "Feedback & trust" in text, "XML entities are unescaped"

    pptx = tmp_path / "c.pptx"
    pptx.write_bytes(_pptx_bytes({10: "TENTH", 2: "SECOND", 1: "FIRST"}))
    assert extract_text(str(pptx), "c.pptx").splitlines() == ["FIRST", "SECOND", "TENTH"], (
        "slides must be ordered NUMERICALLY — a lexical sort puts slide10 before slide2 "
        "and scrambles the document"
    )

    txt = tmp_path / "d.txt"
    txt.write_text("plain text", encoding="utf-8")
    assert extract_text(str(txt), "d.txt") == "plain text"
    assert extract_text(str(txt), "d.md") == "plain text"

    # A corrupt file and an unsupported extension both yield "" instead of raising: one
    # bad document must never abort the ingestion of the whole knowledge base.
    corrupt = tmp_path / "e.pdf"
    corrupt.write_bytes(b"definitely not a pdf")
    assert extract_text(str(corrupt), "e.pdf") == ""
    assert extract_text(str(txt), "e.zip") == ""


def test_the_curated_content_library_is_parsed_from_the_cells_that_matter(tmp_path):
    """The curated tab's Link column shows a LABEL; the real URL is the cell's hyperlink
    target. Reading the display text would ship "click here" to the user as the source."""
    from app.rag.ingest.sskb import parse_curated

    path = tmp_path / "content.xlsx"
    path.write_bytes(_curated_xlsx_bytes())
    records = parse_curated(str(path))

    assert len(records) == 2, "a row with no Name is not a record"
    by_title = {r["title"]: r for r in records}
    candor = by_title["Radical Candor"]
    assert candor["source_link"] == "https://www.youtube.com/watch?v=candor", (
        "the URL comes from the cell HYPERLINK, not the 'click here' display text"
    )
    assert candor["content_format"] == "video", "format is inferred from the resolved URL"
    assert candor["author"] == "Kim Scott" and candor["source"] == "curated"
    assert candor["item_type"] == "curated_content" and candor["kb"] == "sskb"
    assert candor["text"] == "Radical Candor. Care personally, challenge directly"
    assert candor["synopsis"] == "Care personally, challenge directly"
    assert candor["skill"] == "feedback" and candor["topic"] == "manager"
    assert candor["cluster"] == "senior" and "level" not in candor, (
        "`level` is reserved for USER seniority — the sheet's levels are content taxonomy, "
        "so they go to topic/cluster and `level` is left unset (the store defaults it to '')"
    )
    assert by_title["Deep Work"]["content_format"] == "pdf", "a literal URL is used as-is"

    # A workbook without the tab is a mis-upload, not a crash.
    from openpyxl import Workbook

    wb = Workbook()
    wb.active.title = "wrong tab"
    other = tmp_path / "other.xlsx"
    wb.save(other)
    assert parse_curated(str(other)) == []

    # A library with no Link column (and a blank header cell) still yields records — they
    # simply carry no source_link, which is what `source_required` is checked against.
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "content library"
    ws2.append(["Name", None, "Synopsis"])
    ws2.append(["Untitled Talk", "ignored", "no link anywhere"])
    linkless = tmp_path / "linkless.xlsx"
    wb2.save(linkless)
    (record,) = parse_curated(str(linkless))
    assert record["source_link"] == "" and record["content_format"] == ""
    assert record["title"] == "Untitled Talk" and record["synopsis"] == "no link anywhere"


def test_sskb_ingestion_is_directory_first_and_incremental(lance, fake_embedder, fake_s3):
    """The SUBFOLDER is the authority for what a document is — filename sniffing used to
    misclassify docs and leak the wrong source_link. And ingestion must be incremental:
    unchanged docs (same ETag) are skipped, so a boot doesn't re-embed the whole corpus."""
    from app.rag.ingest.sskb import ingest_sskb

    s3 = fake_s3({
        "sskb/sskb_concept/concepts.pdf": _pdf_bytes("Psychological safety"),
        "sskb/sskb_microlearning/bites.pdf": _pdf_bytes("A two minute bite"),
        "sskb/sskb_competency/master.pdf": _pdf_bytes("Competency clusters"),
        "sskb/sskb_curated/library.xlsx": _curated_xlsx_bytes(),
        "sskb/sskb_unknown_dir/mystery.pdf": _pdf_bytes("who am i"),
        "sskb/loose_file_at_root.pdf": _pdf_bytes("no subdir"),
        "sskb/sskb_concept/notes.png": b"an image, not a document",
    })

    first = ingest_sskb()
    assert first["sskb_concept"] == 1 and first["sskb_microlearning"] == 1
    assert first["sskb_competency"] == 1
    assert first["sskb_curated"] == 2, "the xlsx yields one record per content-library row"
    assert first["skipped_type"] == 3, "png + unknown subdir + root-level file"
    assert first["skipped_existing"] == 0 and first["pruned"] == 0
    assert lance.count("sskb") == 5

    # The directory decided `source`, and `source` is what Extract1/6/7/8 filter on.
    concept = lance.search("sskb", _fake_vector("Psychological safety"),
                           filters={"source": "concept"}, top_k=5)
    assert len(concept) == 1
    assert concept[0]["item_type"] == "concept"
    assert concept[0]["doc_key"] == "sskb/sskb_concept/concepts.pdf"
    assert concept[0]["s3_etag"] == s3.etag("sskb/sskb_concept/concepts.pdf")
    assert concept[0]["source_link"].endswith("sskb/sskb_concept/concepts.pdf")
    assert concept[0]["title"] == "concepts", "title is the filename stem"
    assert {h["source"] for h in lance.search("sskb", _fake_vector("bite"),
                                              filters={"source": "micro_learning"})} == {"micro_learning"}

    # Re-running with nothing changed must embed NOTHING (this is the whole point of the
    # ETag check — a full re-embed on every boot is a real, recurring bill).
    second = ingest_sskb()
    assert second["skipped_existing"] == 4 and sum(second[d] for d in
                                                   ("sskb_concept", "sskb_microlearning",
                                                    "sskb_competency", "sskb_curated")) == 0
    assert lance.count("sskb") == 5

    # A CHANGED doc is re-embedded, and its stale chunks are dropped first.
    s3.objects["sskb/sskb_concept/concepts.pdf"] = _pdf_bytes("Psychological safety rewritten")
    third = ingest_sskb()
    assert third["sskb_concept"] == 1 and lance.count("sskb") == 5, "replaced, not duplicated"
    assert lance.search("sskb", _fake_vector("Psychological safety rewritten"),
                        filters={"source": "concept"})[0]["text"].endswith("rewritten")

    # A doc DELETED from S3 has its chunks pruned — otherwise retrieval keeps citing a
    # document the client has already taken down.
    del s3.objects["sskb/sskb_competency/master.pdf"]
    fourth = ingest_sskb()
    assert fourth["pruned"] == 1 and lance.count("sskb") == 4
    assert lance.search("sskb", _fake_vector("Competency clusters"),
                        filters={"source": "competency"}) == []


def test_cskb_ingestion_reads_the_org_and_doc_type_off_the_path(lance, fake_embedder, fake_s3):
    """cskb/<orgId>/<group>/<file>: the org and the doc_type both come from the KEY. Get
    this wrong and one client's framework answers another client's coaching prompt."""
    from app.rag.ingest.cskb import classify_doc_type, ingest_cskb

    fake_s3({
        "cskb/acme/cskb_org_framework/leadership.pdf": _pdf_bytes("Acme leadership framework"),
        "cskb/acme/cskb_values/values.docx": _docx_bytes(["Integrity", "Courage"]),
        "cskb/acme/cskb_learning_aid/tool.pptx": _pptx_bytes({1: "Acme delegation tool"}),
        "cskb/acme/cskb_competency/comp.pdf": _pdf_bytes("Acme competencies"),
        "cskb/acme/random_folder/notes.pdf": _pdf_bytes("Acme uncategorised"),
        "cskb/rival/cskb_org_framework/leadership.pdf": _pdf_bytes("Rival leadership framework"),
        "cskb/acme/cskb_values/logo.png": b"not a document",
    })

    result = ingest_cskb()
    assert result["files"] == 6 and result["skipped_type"] == 1
    assert result["org_id"] == "ALL" and result["by_file"]["leadership.pdf"] == 1
    assert lance.count("cskb") == 6

    query = _fake_vector("leadership framework")
    acme = lance.search("cskb", query, filters={"org_id": "acme", "doc_type": "frameworks"}, top_k=10)
    assert len(acme) == 1 and "Acme" in acme[0]["text"]
    assert acme[0]["doc_group"] == "cskb_org_framework" and acme[0]["source"] == "client"

    rival = lance.search("cskb", query, filters={"org_id": "rival"}, top_k=10)
    assert [h["org_id"] for h in rival] == ["rival"], "org scoping is absolute"

    # Every typed subfolder maps to the doc_type its extraction filters on.
    types = {h["doc_group"]: h["doc_type"] for h in lance.search("cskb", query,
                                                                 filters={"org_id": "acme"}, top_k=10)}
    assert types == {
        "cskb_org_framework": "frameworks", "cskb_values": "values",
        "cskb_learning_aid": "learning_aids", "cskb_competency": "competencies",
        "random_folder": "general",
    }
    assert classify_doc_type("CSKB_Values") == "values", "case-insensitive folder match"
    assert classify_doc_type("anything else") == "general", "unrecognised ⇒ general, never guessed"

    # `general` is filtered out of every extraction — an uncategorised doc can never be
    # served as a framework/value/competency/learning aid.
    for doc_type in ("frameworks", "values", "competencies", "learning_aids"):
        hits = lance.search("cskb", _fake_vector("Acme uncategorised"),
                            filters={"org_id": "acme", "doc_type": doc_type}, top_k=10)
        assert all("uncategorised" not in h["text"] for h in hits)

    # A single-org re-ingest must not prune the OTHER orgs' documents.
    again = ingest_cskb(org_id="acme")
    assert again["pruned"] == 0 and again["skipped_existing"] == 5
    assert lance.count("cskb") == 6


def test_a_document_with_no_extractable_text_is_skipped_not_embedded(lance, fake_embedder):
    """An image-only PDF extracts to "". Embedding "" would store a meaningless vector that
    can still win a top-k — a passage of nothing, cited as evidence."""
    from app.rag.ingest import embed_and_upsert

    assert embed_and_upsert("sskb", [{"id": "empty", "text": "   ", "kb": "sskb"}]) == 0
    assert embed_and_upsert("sskb", []) == 0
    assert lance.count("sskb") == 0
    assert embed_and_upsert("sskb", [{"id": "real", "text": "content", "kb": "sskb"}]) == 1


def test_one_unreadable_document_does_not_abort_the_whole_ingestion(lance, fake_embedder, fake_s3):
    """Ingestion walks the entire bucket. If a single object blows up (a permissions blip, a
    truncated upload), the run must log it and carry on — the alternative is one bad file
    keeping the ENTIRE knowledge base out of the index."""
    from app.rag.ingest.cskb import ingest_cskb
    from app.rag.ingest.sskb import ingest_sskb

    s3 = fake_s3({
        "sskb/sskb_concept/good.pdf": _pdf_bytes("Psychological safety"),
        "sskb/sskb_concept/boom.pdf": _pdf_bytes("this one fails to download"),
        "cskb/acme/cskb_values/good.pdf": _pdf_bytes("Integrity"),
        "cskb/acme/cskb_values/boom.pdf": _pdf_bytes("this one fails too"),
    })
    original = s3.download_file

    def _explode(bucket, key, dest):
        if "boom" in key:
            raise OSError("connection reset while downloading")
        original(bucket, key, dest)

    s3.download_file = _explode

    assert ingest_sskb()["sskb_concept"] == 1, "the good doc is still ingested"
    assert ingest_cskb()["chunks"] == 1
    assert lance.count("sskb") == 1 and lance.count("cskb") == 1
    assert lance.search("sskb", _fake_vector("Psychological safety"))[0]["text"].startswith(
        "Psychological safety")


def test_an_unconfigured_bucket_must_never_be_read_as_everything_was_deleted(
    lance, fake_embedder, fake_s3, monkeypatch
):
    """The prune step deletes any indexed doc that wasn't listed this run. With no bucket
    configured, the listing is EMPTY — so an unguarded prune would wipe a perfectly good
    index on the first boot of a misconfigured instance. The guard is the only thing
    standing between a missing env var and total data loss."""
    from app.rag import ingest
    from app.rag.ingest.cskb import ingest_cskb
    from app.rag.ingest.sskb import ingest_sskb

    fake_s3({})
    lance.upsert("sskb", [_rec("a", "concept", kb="sskb", doc_key="sskb/a.pdf", s3_etag="E")])
    lance.upsert("cskb", [_rec("b", "values", kb="cskb", doc_key="cskb/acme/v.pdf", s3_etag="E")])
    monkeypatch.setattr(ingest.config, "RAG_S3_BUCKET", "")

    assert ingest_sskb()["pruned"] == 0
    assert ingest_cskb()["pruned"] == 0
    assert lance.count("sskb") == 1 and lance.count("cskb") == 1, "the index MUST survive"


def test_a_changed_client_doc_is_re_embedded_and_a_deleted_one_is_pruned(
    lance, fake_embedder, fake_s3
):
    """A client uploads a new version of their values doc, or takes one down. Both must be
    reflected: retrieval that keeps citing a withdrawn document is a compliance problem."""
    from app.rag.ingest.cskb import ingest_cskb

    s3 = fake_s3({
        "cskb/acme/cskb_values/values.pdf": _pdf_bytes("Integrity and courage"),
        "cskb/acme/cskb_org_framework/f.pdf": _pdf_bytes("Acme framework"),
    })
    assert ingest_cskb()["files"] == 2 and lance.count("cskb") == 2

    s3.objects["cskb/acme/cskb_values/values.pdf"] = _pdf_bytes("Integrity, courage and candor")
    changed = ingest_cskb()
    assert changed["files"] == 1 and changed["skipped_existing"] == 1
    assert lance.count("cskb") == 2, "the stale chunk is dropped before the new one lands"
    hits = lance.search("cskb", _fake_vector("Integrity, courage and candor"),
                        filters={"org_id": "acme", "doc_type": "values"})
    assert len(hits) == 1 and "candor" in hits[0]["text"]

    del s3.objects["cskb/acme/cskb_values/values.pdf"]
    assert ingest_cskb()["pruned"] == 1
    assert lance.count("cskb") == 1
    assert lance.search("cskb", _fake_vector("Integrity"),
                        filters={"org_id": "acme", "doc_type": "values"}) == []


def test_a_key_that_is_not_under_an_org_folder_yields_no_records():
    """`cskb/` itself, or a stray object at the prefix root, has no org — and a record with
    no org_id would be visible to EVERY client."""
    from app.rag.ingest.cskb import _records_for_key
    from app.rag.ingest.sskb import _records_for_key as sskb_records

    assert _records_for_key("cskb/") == []
    assert sskb_records("sskb/unknown_folder/x.pdf") == [], "an unknown subfolder is skipped"


def test_the_local_dev_ingest_path_loads_the_curated_library(lance, fake_embedder, tmp_path):
    """`ingest_sskb_local` is the no-S3 developer path. The curated library is a real local
    file read, so it must work with nothing but a path."""
    from app.rag.ingest.sskb import ingest_sskb_local

    curated = tmp_path / "content.xlsx"
    curated.write_bytes(_curated_xlsx_bytes())

    written = ingest_sskb_local(curated_path=str(curated))
    assert written == {"curated": 2, "micro_learning": 0, "concepts": 0, "competency": 0}
    assert lance.count("sskb") == 2
    assert ingest_sskb_local() == {"curated": 0, "micro_learning": 0, "concepts": 0,
                                   "competency": 0}, "no paths → nothing ingested, no error"


def test_the_vector_store_never_ingests_its_own_index_files(fake_s3):
    """The LanceDB index lives in the SAME bucket as the source docs. Listing must skip
    everything under the embeddings prefix, or ingestion feeds the index back into itself."""
    from app import config
    from app.rag.ingest import iter_s3_keys

    fake_s3({
        f"{config.RAG_LANCEDB_PREFIX}/sskb.lance/data.manifest": b"binary index data",
        "sskb/sskb_concept/real.pdf": _pdf_bytes("real doc"),
        "sskb/sskb_concept/": b"",  # a folder marker
    })
    assert list(iter_s3_keys("")) == ["sskb/sskb_concept/real.pdf"]


def test_ingestion_without_a_bucket_lists_nothing_instead_of_erroring(monkeypatch):
    """No bucket configured = a local/offline deployment. It must yield nothing (and the
    prune step is guarded on the same flag, so an unset bucket can never be mistaken for
    'every document was deleted' and wipe the index)."""
    from app.rag import ingest

    monkeypatch.setattr(ingest.config, "RAG_S3_BUCKET", "")
    assert list(ingest.iter_s3_objects("sskb/")) == []


def test_public_url_percent_encodes_the_key(monkeypatch):
    """source_link is rendered to the user. An unencoded space produces a dead link."""
    from app.rag import ingest

    monkeypatch.setattr(ingest.config, "RAG_S3_BUCKET", "b")
    monkeypatch.setattr(ingest.config, "AWS_REGION", "us-east-1")
    assert ingest.public_url("cskb/acme/my doc.pdf") == (
        "https://b.s3.us-east-1.amazonaws.com/cskb/acme/my%20doc.pdf"
    )

    # The S3 client is built with the SAME region the links are minted with — a mismatch
    # yields signed requests the bucket rejects.
    monkeypatch.setitem(sys.modules, "boto3", _FakeBoto3())
    assert ingest.s3_client() is not None


# ════════════════════════════════════════════════════════════════════════════
#  startup — boot-time ingestion must never block the service
# ════════════════════════════════════════════════════════════════════════════


def test_startup_ingest_can_be_switched_off(monkeypatch):
    from app.rag import startup

    monkeypatch.setattr(startup.config, "RAG_INGEST_ON_STARTUP", False)
    assert startup.run_startup() == {"ingested": False}


def test_startup_skips_cleanly_when_the_environment_cannot_ingest(monkeypatch):
    """A lean install (no lancedb/boto3) or a creds-less box must log a one-line hint and
    boot. A traceback at startup on a machine that was never meant to ingest reads like an
    outage and gets 'fixed' by disabling RAG entirely."""
    from app.rag import startup

    monkeypatch.setattr(startup.config, "RAG_INGEST_ON_STARTUP", True)
    monkeypatch.setattr(startup, "_REQUIRED_INGEST_MODULES", ("lancedb", "no_such_module_xyz"))
    result = startup.run_startup()
    assert result == {"ingested": False, "reason": "missing_deps", "missing": ["no_such_module_xyz"]}

    # Deps present, but no AWS credentials resolve → skip, don't raise NoCredentialsError.
    monkeypatch.setattr(startup, "_REQUIRED_INGEST_MODULES", ("lancedb", "boto3"))
    monkeypatch.setitem(sys.modules, "boto3", _FakeBoto3(creds=False))
    monkeypatch.setattr(startup.config, "RAG_S3_BUCKET", "some-bucket")
    assert startup.run_startup() == {"ingested": False, "reason": "no_aws_credentials"}


def test_the_credential_and_error_probes_never_raise_at_boot(monkeypatch):
    """Both preflight probes run before anything else at startup. Each has to answer a
    question, not raise one — a traceback here reads as "the service is broken" when the
    truth is "this box cannot ingest, and that's fine"."""
    from app.rag import startup

    class _BrokenBoto3(_FakeBoto3):
        def Session(self):  # noqa: N802
            raise RuntimeError("botocore is half-installed")

    monkeypatch.setitem(sys.modules, "boto3", _BrokenBoto3())
    assert startup._has_aws_credentials() is False

    monkeypatch.setitem(sys.modules, "boto3", _FakeBoto3(creds=True))
    assert startup._has_aws_credentials() is True

    # No botocore at all → nothing can be classified as an S3 access error, so every
    # failure is treated as a real bug and surfaces. Silence is the wrong default.
    monkeypatch.setitem(sys.modules, "botocore.exceptions", None)
    assert startup._is_s3_access_error(RuntimeError("anything")) is False


def test_startup_ingests_both_knowledge_bases_and_reports_row_counts(
    lance, fake_embedder, fake_s3, monkeypatch
):
    """The happy path, end to end: S3 → chunks → embeddings → LanceDB, with the summary the
    boot log (and /rag health) is read from. A summary that reports rows the store doesn't
    have is how a broken ingest looks healthy."""
    from app.rag import startup

    monkeypatch.setattr(startup.config, "RAG_INGEST_ON_STARTUP", True)
    monkeypatch.setitem(sys.modules, "boto3", _FakeBoto3(creds=True))
    fake_s3({
        "sskb/sskb_concept/c.pdf": _pdf_bytes("Psychological safety"),
        "cskb/acme/cskb_values/v.docx": _docx_bytes(["Integrity"]),
    })

    summary = startup.run_startup()

    assert summary["ingested"] is True and summary["reindex"] is False
    assert summary["sskb_rows"] == 1 == lance.count("sskb")
    assert summary["cskb_rows"] == 1 == lance.count("cskb")
    assert summary["sskb"]["sskb_concept"] == 1 and summary["cskb"]["files"] == 1
    assert "elapsed_s" in summary


def test_reindex_drops_the_old_index_before_rebuilding(lance, fake_embedder, fake_s3, monkeypatch):
    """RAG_REINDEX exists because a doc removed from the corpus, or an embedder change,
    leaves rows that can still be retrieved. The drop must actually happen — otherwise
    'reindex' quietly means 'append'."""
    from app.rag import startup

    monkeypatch.setattr(startup.config, "RAG_INGEST_ON_STARTUP", False)  # reindex overrides it
    monkeypatch.setitem(sys.modules, "boto3", _FakeBoto3(creds=True))
    lance.upsert("sskb", [_rec("stale", "a row from a previous embedding model", kb="sskb")])
    fake_s3({"sskb/sskb_concept/c.pdf": _pdf_bytes("Psychological safety")})

    summary = startup.run_startup(reindex=True)

    assert summary["ingested"] is True and summary["reindex"] is True
    assert lance.count("sskb") == 1, "the stale row is gone, only the re-ingested doc remains"
    assert lance.search("sskb", _fake_vector("a row from a previous embedding model"),
                        top_k=5)[0]["id"] != "stale"


def test_an_s3_outage_degrades_to_a_skip_but_a_real_bug_still_raises(
    lance, fake_embedder, monkeypatch
):
    """Expired token / access denied / no bucket = "can't ingest HERE", so boot continues
    with an empty KB. Anything else is a genuine bug and MUST surface — swallowing every
    exception is exactly how a broken ingest hides."""
    from botocore.exceptions import ClientError

    from app.rag import ingest, startup

    monkeypatch.setattr(startup.config, "RAG_INGEST_ON_STARTUP", True)
    monkeypatch.setitem(sys.modules, "boto3", _FakeBoto3(creds=True))

    def _expired():
        raise ClientError({"Error": {"Code": "ExpiredToken"}}, "ListObjectsV2")

    monkeypatch.setattr(ingest, "s3_client", _expired)
    result = startup.run_startup()
    assert result["ingested"] is False and result["reason"] == "s3_access_error"
    assert "ExpiredToken" in result["error"]

    def _bug():
        raise ValueError("a genuine programming error")

    monkeypatch.setattr(ingest, "s3_client", _bug)
    with pytest.raises(ValueError, match="genuine programming error"):
        startup.run_startup()


# ════════════════════════════════════════════════════════════════════════════
#  pgvector_store — the S3-free backend, against a REAL Postgres
# ════════════════════════════════════════════════════════════════════════════


def test_the_pgvector_backend_is_opt_in(monkeypatch):
    """An existing client must stay on LanceDB until they explicitly switch. A backend that
    switches itself on would search an empty Postgres and return nothing, forever."""
    from app.rag import pgvector_store as pgv

    monkeypatch.delenv("CEREBROZEN_RAG_BACKEND", raising=False)
    assert pgv.enabled() is False
    monkeypatch.setenv("CEREBROZEN_RAG_BACKEND", "lancedb")
    assert pgv.enabled() is False
    monkeypatch.setenv("CEREBROZEN_RAG_BACKEND", "  PGVECTOR  ")
    assert pgv.enabled() is True, "the switch is case/whitespace tolerant"


@requires_pg
def test_pgvector_upsert_search_count_and_delete_against_real_postgres(pgvec, fake_embedder):
    """The full surface, on a live pgvector index. `_score` here is a SIMILARITY (1 = same
    vector) — the inverse of LanceDB's distance."""
    rows = [
        {"id": "a", "doc_key": "sskb/x.pdf", "text": "psychological safety in teams",
         "embedding": _fake_vector("psychological safety in teams"),
         "kb": "sskb", "source": "concept", "title": "Safety"},
        {"id": "b", "doc_key": "sskb/y.pdf", "text": "quarterly revenue forecasting",
         "embedding": _fake_vector("quarterly revenue forecasting"),
         "kb": "sskb", "source": "concept", "title": "Revenue"},
    ]
    assert pgvec.upsert("sskb", rows) == 2
    assert pgvec.count("sskb") == 2

    hits = pgvec.search("sskb", _fake_vector("psychological safety in teams"), top_k=2)
    assert [h["text"] for h in hits] == ["psychological safety in teams",
                                         "quarterly revenue forecasting"]
    # `_score` is a DISTANCE — lower is closer, so an exact match scores 0.0.
    #
    # This test originally asserted the opposite (a similarity, 1.0 for an exact match),
    # which was the behaviour the suite ALSO pinned as a bug three hundred lines below: the
    # two backends disagreed about which direction "better" points, and
    # `extractors._extract_learning_aid` picks with `min(_score)` — so on pgvector it chose
    # the LEAST relevant aid in the knowledge base. Distance is the contract LanceDB set and
    # every consumer was written against, so pgvector now matches it, and this assertion is
    # corrected rather than the code being bent to fit it.
    assert hits[0]["_score"] == pytest.approx(0.0, abs=1e-6), "an exact vector match is distance 0"
    assert hits[0]["_score"] < hits[1]["_score"], "_score is a DISTANCE — lower is closer"
    assert hits[0]["title"] == "Safety" and hits[0]["source"] == "concept"

    assert len(pgvec.search("sskb", _fake_vector("anything"), top_k=1)) == 1

    # Re-ingesting the same id UPDATES it (ON CONFLICT) — the corpus cannot double.
    rows[0]["text"] = "psychological safety, rewritten"
    assert pgvec.upsert("sskb", rows) == 2 and pgvec.count("sskb") == 2
    assert pgvec.search("sskb", _fake_vector("psychological safety, rewritten"),
                        top_k=1)[0]["text"] == "psychological safety, rewritten"

    pgvec.delete_by_doc_key("sskb", "sskb/x.pdf")
    assert pgvec.count("sskb") == 1
    pgvec.delete_by_doc_key("sskb", "")  # empty key → no-op, never a table wipe
    assert pgvec.count("sskb") == 1

    assert pgvec.upsert("sskb", []) == 0, "nothing to write is not an error"
    assert pgvec.count("cskb") == 0, "a table that was never created counts as empty"


@requires_pg
def test_pgvector_prefilters_by_metadata_so_an_org_can_never_see_another(pgvec, fake_embedder):
    """The WHERE runs BEFORE the ranking (`meta @> …` then ORDER BY distance). A post-filter
    would let a rival's document take the top-k slot and then be dropped — silently
    returning nothing instead of the client's own best match."""
    same = "leadership presence"
    pgvec.upsert("cskb", [
        {"id": f"{org}-{dt}", "doc_key": f"cskb/{org}/d.pdf", "text": f"{same} {org} {dt}",
         "embedding": _fake_vector(same), "kb": "cskb", "org_id": org, "doc_type": dt}
        for org in ("acme", "rival") for dt in ("frameworks", "values")
    ])
    query = _fake_vector(same)

    acme = pgvec.search("cskb", query, filters={"org_id": "acme"}, top_k=10)
    assert len(acme) == 2 and {h["org_id"] for h in acme} == {"acme"}

    scoped = pgvec.search("cskb", query, filters={"org_id": "acme", "doc_type": "frameworks"},
                          top_k=10)
    assert len(scoped) == 1 and scoped[0]["doc_type"] == "frameworks"

    assert pgvec.search("cskb", query, filters={"org_id": "nobody"}, top_k=10) == []
    assert len(pgvec.search("cskb", query, top_k=10)) == 4, "no filter → the whole table"


@requires_pg
def test_pgvector_builds_an_HNSW_COSINE_index(pgvec, fake_embedder, pgdb):
    """HNSW over cosine, created with the table. IVFFlat's lists must be trained AFTER the
    data is loaded — indexing an empty table with IVFFlat silently yields terrible recall.
    And the index must be COSINE: the query orders by `<=>`, so an l2 index is never used."""
    pgvec.upsert("sskb", [{"id": "a", "text": "x", "embedding": _fake_vector("x")}])

    indexes = {r[0]: r[1] for r in pgdb.sql(
        "SELECT indexname, indexdef FROM pg_indexes WHERE tablename = 'rag_sskb'")}
    hnsw = indexes["rag_sskb_emb_hnsw"]
    assert "USING hnsw" in hnsw and "vector_cosine_ops" in hnsw
    assert "rag_sskb_meta_gin" in indexes and "rag_sskb_doc_key" in indexes


@requires_pg
def test_a_dimension_mismatch_raises_an_ACTIONABLE_error(pgvec, pgdb):
    """Embeddings are NOT portable: 1536-dim (OpenAI) vectors cannot search a 768-dim
    (nomic) index. Postgres rejects it with "expected 8 dimensions, not 1536", which tells
    an operator nothing. The guard must name the cause AND the only fix — re-ingest.

    Silently returning garbage here would be worse than the error: the coach would cite
    passages that have nothing to do with the query."""
    pgvec.upsert("sskb", [{"id": "a", "text": "x", "embedding": [0.1] * _DIM}])

    pgvec._ready.clear()  # a fresh process: the table exists, the embedder changed
    with pytest.raises(RuntimeError) as exc:
        pgvec.upsert("sskb", [{"id": "b", "text": "y", "embedding": [0.1] * 8}])
    message = str(exc.value)
    assert "re-index" in message or "re-ingest" in message, "the guard must say what to DO"
    assert "DROP TABLE rag_sskb" in message and str(_DIM) in message and "8-dim" in message

    pgvec._ready.clear()
    with pytest.raises(RuntimeError, match="re-ingest"):
        pgvec.search("sskb", [0.1] * 8)

    assert pgdb.sql("SELECT count(*) FROM rag_sskb")[0][0] == 1, "the index is left intact"


@requires_pg
def test_store_delegates_every_operation_to_pgvector_when_it_is_enabled(
    pgvec, pgdb, lance, fake_embedder, monkeypatch
):
    """`store.py` is the only import the extractors know. With the backend switched on,
    every one of its functions must reach Postgres — one that still writes to LanceDB
    would split the corpus across two stores, and the half nobody searches just vanishes."""
    from app.rag import store

    monkeypatch.setenv("CEREBROZEN_RAG_BACKEND", "pgvector")
    written = store.upsert("sskb", [
        {"id": "a", "doc_key": "sskb/a.pdf", "text": "delegation and trust",
         "embedding": _fake_vector("delegation and trust"), "kb": "sskb", "source": "concept"},
    ])
    assert written == 1

    assert store.count("sskb") == 1
    assert pgdb.sql("SELECT count(*) FROM rag_sskb")[0][0] == 1, "the row is in POSTGRES"
    assert lance._connect().table_names() == [], "and nothing was written to LanceDB"

    hits = store.search("sskb", _fake_vector("delegation and trust"),
                        filters={"source": "concept"}, top_k=3)
    assert [h["text"] for h in hits] == ["delegation and trust"]
    # {doc_key: s3_etag} — NOT {doc_key: doc_key}. This row was upserted without an
    # `s3_etag` in its metadata, so the ETag is empty, and empty is exactly right: it means
    # "unknown", so ingestion re-indexes rather than wrongly skipping. Returning the key as
    # its own ETag made `indexed.get(key) == etag` impossible to satisfy, so the entire
    # corpus was deleted and re-embedded on every boot.
    assert store.indexed_docs("sskb") == {"sskb/a.pdf": ""}

    store.delete_by_doc_key("sskb", "sskb/a.pdf")
    assert store.count("sskb") == 0


@requires_pg
def test_pgvector_is_inert_when_postgres_is_not_configured(pgvec, monkeypatch):
    """The backend flag can be on while POSTGRES_URL is empty (a half-finished config). It
    must return neutral values rather than raise — the graph keeps coaching, unaided."""
    from app.stores import pg

    monkeypatch.setenv("POSTGRES_URL", "")
    monkeypatch.setattr(pg, "_pool", None)

    assert pgvec.upsert("sskb", [{"id": "a", "text": "t", "embedding": [0.1] * 4}]) == 0
    assert pgvec.search("sskb", [0.1] * 4) == []
    assert pgvec.count("sskb") == 0 and pgvec.indexed_docs("sskb") == {}
    assert pgvec.delete_by_doc_key("sskb", "k") is None
    assert pgvec._ensure("sskb", 4) is None, "table creation is skipped, not attempted"


@requires_pg
def test_pgvector_reports_an_empty_index_before_anything_is_ingested(pgvec):
    """Before the first ingest the tables do not exist yet. /rag health and the incremental
    ingest both query them anyway — and both must read that as "empty", not as an error."""
    assert pgvec.count("sskb") == 0
    assert pgvec.indexed_docs("sskb") == {}, "no table → nothing indexed → ingest everything"
    assert pgvec.search("sskb", [0.1] * _DIM) == [], "searching creates the table and finds nothing"
    assert pgvec.count("sskb") == 0


def test_pgvector_maps_the_kb_name_to_its_own_table():
    """Two tables, because SSKB (global) and CSKB (per-org) must stay structurally
    isolated: an SSKB query physically cannot return a client's documents."""
    from app.rag.pgvector_store import _table

    assert _table("sskb") == "rag_sskb" and _table(" SSKB ") == "rag_sskb"
    assert _table("cskb") == "rag_cskb" and _table("") == "rag_sskb"


# ════════════════════════════════════════════════════════════════════════════
#  stores/pg.py — the Mongo-compatible shim, against a REAL Postgres
# ════════════════════════════════════════════════════════════════════════════


@requires_pg
def test_the_mongo_update_operators_land_correctly_in_jsonb(pgdb):
    """This is the shim's whole job. Every operator is asserted by reading the row STRAIGHT
    OUT OF POSTGRES — not through the shim that wrote it, which would hide an error that
    round-trips consistently."""
    coll = pgdb.collection("agentic")
    table = coll.name

    coll.update_one(
        {"user_id": "u1"},
        {"$setOnInsert": {"created_at": "T0"},
         "$set": {"user_id": "u1", "intake_vars.userRoleContext": "EM"},
         "$inc": {"turns": 1},
         "$push": {"messages": {"$each": [{"n": 1}, {"n": 2}]}},
         "$addToSet": {"done": {"$each": ["s1", "s2"]}}},
        upsert=True,
    )
    coll.update_one(
        {"user_id": "u1"},
        {"$setOnInsert": {"created_at": "T-LATER"},
         "$set": {"intake_vars.level": "senior"},
         "$inc": {"turns": 2},
         "$push": {"messages": {"$each": [{"n": 3}, {"n": 4}], "$slice": -3}},
         "$addToSet": {"done": {"$each": ["s2", "s3"]}}},
    )

    (doc,) = pgdb.rows(table)
    assert doc["created_at"] == "T0", "$setOnInsert must NOT fire on an update"
    assert doc["intake_vars"] == {"userRoleContext": "EM", "level": "senior"}, (
        "dotted paths must build a nested subdocument, not a literal 'a.b' key"
    )
    assert doc["turns"] == 3, "$inc accumulates"
    assert doc["messages"] == [{"n": 2}, {"n": 3}, {"n": 4}], "$slice -3 keeps the LAST three"
    assert doc["done"] == ["s1", "s2", "s3"], "$addToSet must not duplicate (idempotent re-runs)"

    # The row is keyed by the filter's identity field, so one user is one row.
    keys = pgdb.sql(f'SELECT _id FROM "{table}"')
    assert keys == [("u1",)]


@requires_pg
def test_push_slice_positive_keeps_the_head_and_plain_push_appends(pgdb):
    from app.stores.pg import _apply_update

    doc = _apply_update({}, {"$push": {"q": {"$each": [1, 2, 3, 4], "$slice": 2}}}, inserted=True)
    assert doc["q"] == [1, 2], "a POSITIVE $slice keeps the FIRST n"
    doc = _apply_update(doc, {"$push": {"q": 9}}, inserted=False)
    assert doc["q"] == [1, 2, 9], "a bare $push appends a single element"
    doc = _apply_update(doc, {"$addToSet": {"tags": "x"}}, inserted=False)
    assert doc["tags"] == ["x"], "$addToSet without $each adds one element"
    # A scalar sitting where a subdocument is needed must not corrupt the document.
    doc = _apply_update({"a": 1}, {"$set": {"a.b": 2}}, inserted=False)
    assert doc["a"] == 1, "the write is dropped rather than clobbering the scalar"


@requires_pg
def test_an_upsert_creates_the_document_a_plain_update_does_not(pgdb):
    """`upsert=False` on a missing document must be a NO-OP. Creating the row instead would
    resurrect sessions that were deliberately deleted."""
    coll = pgdb.collection("conv")

    result = coll.update_one({"session_id": "ghost"}, {"$set": {"x": 1}})
    assert result.matched_count == 0 and pgdb.rows(coll.name) == []

    created = coll.update_one({"session_id": "s1"}, {"$set": {"x": 1}}, upsert=True)
    assert created.matched_count == 0 and created.modified_count == 1
    assert pgdb.rows(coll.name) == [{"session_id": "s1", "x": 1}], (
        "the filter's scalar fields seed the new document (Mongo's upsert semantics)"
    )

    updated = coll.update_one({"session_id": "s1"}, {"$set": {"x": 2}}, upsert=True)
    assert updated.matched_count == 1, "an existing row reports matched=1"
    assert pgdb.rows(coll.name)[0]["x"] == 2

    # A filter with no identity field can't be keyed → a silent no-op (not a crash).
    assert coll.update_one({"status": "open"}, {"$set": {"x": 3}}, upsert=True).modified_count == 0
    assert len(pgdb.rows(coll.name)) == 1
    # $ne conditions in the filter are not part of the identity.
    assert coll.update_one({"session_id": {"$ne": "s1"}}, {"$set": {"x": 4}}).matched_count == 0


@requires_pg
def test_concurrent_turns_on_one_session_do_not_lose_writes(pgdb):
    """update_one is a read-modify-write. Without the `SELECT … FOR UPDATE` row lock, two
    turns landing together each read the same transcript, each append their own message, and
    the second write CLOBBERS the first — a user message that vanishes from history.

    Eight threads, one document: every append must survive."""
    coll = pgdb.collection("conv")
    coll.update_one({"session_id": "s1"}, {"$set": {"messages": []}}, upsert=True)

    start = threading.Barrier(8)

    def _append(n: int):
        start.wait()
        coll.update_one({"session_id": "s1"},
                        {"$push": {"messages": {"$each": [{"n": n}]}}, "$inc": {"turns": 1}})

    threads = [threading.Thread(target=_append, args=(i,)) for i in range(8)]
    for t in threads:
        t.start()
    for t in threads:
        t.join()

    (doc,) = pgdb.rows(coll.name)
    assert sorted(m["n"] for m in doc["messages"]) == list(range(8)), "a lost update"
    assert doc["turns"] == 8


@requires_pg
def test_find_returns_a_CURSOR_because_the_stores_chain_sort_and_limit(pgdb):
    """`coll.find(...).sort("updated_at", -1).limit(n)` is what the history list does. When
    find() returned a LIST this raised TypeError, the store swallowed it, and every repeat
    user's history came back EMPTY — they looked like a brand-new user."""
    coll = pgdb.collection("conv")
    for sid, updated in (("s1", "2026-01-01"), ("s2", "2026-03-01"), ("s3", "2026-02-01")):
        coll.update_one({"session_id": sid},
                        {"$set": {"session_id": sid, "user_id": "u1", "updated_at": updated}},
                        upsert=True)

    cursor = coll.find({"user_id": "u1"})
    assert hasattr(cursor, "sort") and hasattr(cursor, "limit")
    assert len(cursor) == 3

    newest = coll.find({"user_id": "u1"}).sort("updated_at", -1).limit(2)
    assert [d["session_id"] for d in newest] == ["s2", "s3"]
    oldest = coll.find({"user_id": "u1"}).sort("updated_at", 1)
    assert [d["session_id"] for d in oldest] == ["s1", "s3", "s2"]

    # A document missing the sort key must not raise (it sorts as "").
    coll.update_one({"session_id": "s4"}, {"$set": {"session_id": "s4", "user_id": "u1"}},
                    upsert=True)
    assert coll.find({"user_id": "u1"}).sort("updated_at", 1)._docs[0]["session_id"] == "s4"

    assert list(coll.find({"user_id": "nobody"})) == []
    assert len(coll.find()) == 4, "no filter → every document"


@requires_pg
def test_count_documents_honours_limit_because_the_stores_pass_it(pgdb):
    """`count_documents(query, limit=1)` is how "has this user been here before?" is asked.
    pymongo's `limit=` kwarg used to raise TypeError here, the store swallowed the error and
    returned False — so every REPEAT user was greeted as a brand-new one."""
    coll = pgdb.collection("conv")
    for sid in ("s1", "s2", "s3"):
        coll.update_one({"session_id": sid},
                        {"$set": {"session_id": sid, "user_id": "u1", "ended": True}},
                        upsert=True)

    assert coll.count_documents({"user_id": "u1"}) == 3
    assert coll.count_documents({"user_id": "u1"}, limit=1) == 1, "limit must CAP the count"
    assert coll.count_documents({"user_id": "u1", "session_id": {"$ne": "s1"}}, limit=1) == 1
    assert coll.count_documents({"user_id": "u1", "ended": True}, limit=10) == 3
    assert coll.count_documents({"user_id": "nobody"}, limit=1) == 0
    assert coll.count_documents({}) == 3


@requires_pg
def test_the_query_operators_the_stores_use_match_real_rows(pgdb):
    """$ne excludes the CURRENT session from "prior sessions"; $exists/$in/$nin gate the
    profile reads. An operator that silently matches everything turns "prior sessions" into
    "all sessions", and the coach greets a first-time user with a recap."""
    coll = pgdb.collection("conv")
    coll.update_one({"session_id": "s1"}, {"$set": {
        "session_id": "s1", "user_id": "u1", "ended": True, "org": {"id": "acme"}}}, upsert=True)
    coll.update_one({"session_id": "s2"}, {"$set": {
        "session_id": "s2", "user_id": "u1", "org": {"id": "rival"}}}, upsert=True)

    ids = lambda flt: sorted(d["session_id"] for d in coll.find(flt))  # noqa: E731
    assert ids({"user_id": "u1", "session_id": {"$ne": "s1"}}) == ["s2"]
    assert ids({"ended": {"$exists": True}}) == ["s1"]
    assert ids({"ended": {"$exists": False}}) == ["s2"], "a missing field does not 'exist'"
    assert ids({"session_id": {"$in": ["s1", "s9"]}}) == ["s1"]
    assert ids({"session_id": {"$nin": ["s1"]}}) == ["s2"]
    assert ids({"org.id": "acme"}) == ["s1"], "dotted paths match nested fields"
    assert ids({"org.id": "nope"}) == []
    assert ids({}) == ["s1", "s2"]

    # find_one applies the WHOLE filter, not just the key it looked the row up by.
    assert coll.find_one({"session_id": "s1", "ended": True})["user_id"] == "u1"
    assert coll.find_one({"session_id": "s1", "ended": False}) is None
    assert coll.find_one({"org.id": "rival"})["session_id"] == "s2", "keyless find_one scans"
    assert coll.find_one({"session_id": "missing"}) is None


@requires_pg
def test_the_slice_projection_the_transcript_reads_depend_on(pgdb):
    """`{"messages": {"$slice": 4}}` builds the fallback title from the FIRST 4 messages;
    `{"$slice": -10}` reads the last 10 for context. Getting the sign backwards feeds the
    coach the wrong end of the conversation."""
    coll = pgdb.collection("conv")
    messages = [{"n": i} for i in range(12)]
    coll.update_one({"session_id": "s1"}, {"$set": {
        "session_id": "s1", "user_id": "u1", "title": "T", "messages": messages}}, upsert=True)

    head = coll.find_one({"session_id": "s1"}, {"title": 1, "messages": {"$slice": 4}})
    assert [m["n"] for m in head["messages"]] == [0, 1, 2, 3]
    assert head["title"] == "T" and "user_id" not in head, "an inclusion projection excludes"

    tail = coll.find_one({"session_id": "s1"}, {"messages": {"$slice": -10}})
    assert [m["n"] for m in tail["messages"]] == list(range(2, 12))

    # A projected field the document doesn't carry is simply absent (not None).
    assert "missing" not in coll.find_one({"session_id": "s1"}, {"title": 1, "missing": 1})
    # An exclusion-only projection returns the whole document.
    whole = coll.find_one({"session_id": "s1"}, {"_id": 0})
    assert whole["user_id"] == "u1" and len(whole["messages"]) == 12
    assert coll.find_one({"session_id": "s1"}, None)["user_id"] == "u1"

    # The projection also applies through find().
    projected = list(coll.find({"user_id": "u1"}, {"messages": {"$slice": 2}, "session_id": 1}))
    assert projected == [{"messages": [{"n": 0}, {"n": 1}], "session_id": "s1"}]


@requires_pg
def test_the_one_aggregation_pipeline_the_stores_run(pgdb):
    """conversation.py's summary pipeline: $match, then $project with $size/$ifNull (total
    message count) and $slice (the last 10, to find the last bot message). It is emulated
    directly — so it has to be pinned directly."""
    coll = pgdb.collection("conv")
    coll.update_one({"session_id": "s1"}, {"$set": {
        "session_id": "s1", "user_id": "u1", "title": "T",
        "messages": [{"role": "user" if i % 2 else "bot", "text": str(i)} for i in range(14)],
    }}, upsert=True)
    coll.update_one({"session_id": "s2"}, {"$set": {"session_id": "s2", "user_id": "u2"}},
                    upsert=True)

    docs = coll.aggregate([
        {"$match": {"session_id": "s1"}},
        {"$project": {"_id": 0, "title": 1, "session_id": 1,
                      "total": {"$size": {"$ifNull": ["$messages", []]}},
                      "tail": {"$slice": ["$messages", -10]}}},
    ])
    assert len(docs) == 1
    assert docs[0]["total"] == 14 and docs[0]["title"] == "T"
    assert [m["text"] for m in docs[0]["tail"]] == [str(i) for i in range(4, 14)]
    assert "_id" not in docs[0]

    # $ifNull is what stops a session with NO messages from erroring out mid-pipeline.
    empty = coll.aggregate([
        {"$match": {"session_id": "s2"}},
        {"$project": {"total": {"$size": {"$ifNull": ["$messages", []]}},
                      "head": {"$slice": ["$messages", 3]}}},
    ])
    assert empty == [{"total": 0, "head": []}]
    assert coll.aggregate([{"$match": {"session_id": "nope"}}]) == []


@requires_pg
def test_delete_one_and_delete_many(pgdb):
    """Deletes are keyed off the document's identity field. A keyless filter must delete
    NOTHING — a delete_one({}) that fell through to a table scan would wipe the collection."""
    coll = pgdb.collection("conv")
    for sid in ("s1", "s2", "s3"):
        coll.update_one({"session_id": sid},
                        {"$set": {"session_id": sid, "user_id": "u1" if sid != "s3" else "u2"}},
                        upsert=True)

    assert coll.delete_one({"session_id": "s1"}).deleted_count == 1
    assert coll.delete_one({"session_id": "s1"}).deleted_count == 0, "deleting twice is not an error"
    assert coll.delete_one({}).deleted_count == 0, "a keyless delete_one must NOT wipe the table"
    assert len(pgdb.rows(coll.name)) == 2

    # delete_many on a USER-keyed collection (agentic/dynamic_vars are keyed by user_id).
    agentic = pgdb.collection("agentic")
    for uid in ("u1", "u2"):
        agentic.update_one({"user_id": uid}, {"$set": {"user_id": uid, "tier": "free"}},
                           upsert=True)
    assert agentic.delete_many({"user_id": "u1"}).deleted_count == 1
    assert [d["user_id"] for d in pgdb.rows(agentic.name)] == ["u2"]
    assert agentic.delete_many({"user_id": "nobody"}).deleted_count == 0
    assert agentic.delete_many({"tier": "free"}).deleted_count == 1, "a non-key filter still scans"
    assert pgdb.rows(agentic.name) == []


@requires_pg
def test_the_collection_name_cannot_carry_sql_into_the_ddl(pgdb):
    """Every statement interpolates the table name into the SQL string (a table name cannot
    be a bind parameter). The sanitiser is therefore the ONLY thing standing between a
    config-supplied collection name and DDL injection."""
    from app.stores.pg import PgCollection

    assert PgCollection('conv"; DROP TABLE users; --').name == "convDROPTABLEusers"
    assert PgCollection("agentic_docs-2026").name == "agentic_docs2026"


@requires_pg
def test_a_collections_table_is_created_once_not_on_every_call(pgdb):
    """`_ensure` runs CREATE TABLE/INDEX. Without the `_ensured` guard, EVERY find_one on
    every request would pay two DDL round-trips against the same table."""
    from app.stores import pg

    coll = pgdb.collection("agentic")
    assert coll.name in pg._ensured

    again = pg.PgCollection(coll.name)  # a second handle: the DDL must be skipped
    again.update_one({"user_id": "u1"}, {"$set": {"user_id": "u1"}}, upsert=True)
    assert again.find_one({"user_id": "u1"}) == {"user_id": "u1"}


@requires_pg
def test_index_management_is_a_no_op_on_postgres(pgdb):
    """The Mongo-era index code still calls these on every boot; the GIN index is created
    with the table instead. They must be inert, not raise."""
    coll = pgdb.collection("conv")
    assert coll.create_index("user_id") == ""
    assert coll.index_information() == {}
    assert coll.drop_index("user_id") is None


@requires_pg
def test_the_client_shim_gives_every_store_the_same_seam(pgdb):
    """`client[db][collection]` is how mongo.read_user_context (the profile READ path)
    reaches its data, while the other stores call `collection()`. Both must land on the same
    table, or a user's profile is written to one place and read from another."""
    from app.stores import pg

    coll = pgdb.collection("users")
    coll.update_one({"user_id": "u1"}, {"$set": {"username": "Ada"}}, upsert=True)

    client = pg.client()
    assert client is not None
    via_client = client["any_db_name_at_all"][coll.name]
    assert via_client.find_one({"user_id": "u1"})["username"] == "Ada", (
        "the Mongo database name is ignored — Postgres has one database, and the "
        "collections are tables inside it"
    )
    assert pg.collection(coll.name) is via_client, "collections are cached, not rebuilt"


def test_without_postgres_every_store_call_is_a_quiet_no_op(monkeypatch):
    """No POSTGRES_URL = the app is on Mongo. `collection()`/`client()` must return None so
    the stores fall through to Mongo — and a PgCollection built before the pool died must
    degrade to neutral values rather than raise on a None pool."""
    from app.stores import pg

    monkeypatch.setenv("POSTGRES_URL", "")
    monkeypatch.setattr(pg, "_pool", None)
    monkeypatch.setattr(pg, "_ensured", set())
    monkeypatch.setattr(pg, "_collections", {})

    assert pg.get_pool() is None
    assert pg.collection("conv") is None and pg.client() is None

    orphan = pg.PgCollection("conv")  # constructed with no pool: _ensure() must not raise
    assert orphan.find_one({"user_id": "u1"}) is None
    assert list(orphan.find({})) == [] and orphan.count_documents({}) == 0
    assert orphan.update_one({"user_id": "u1"}, {"$set": {"x": 1}}, upsert=True).modified_count == 0
    assert orphan.delete_one({"user_id": "u1"}).deleted_count == 0
    assert orphan.delete_many({}).deleted_count == 0
    assert orphan.aggregate([{"$match": {}}]) == []


def test_a_pool_that_cannot_be_built_is_remembered_not_retried_forever(monkeypatch):
    """If the pool can't be constructed (here: a typo'd CEREBROZEN_PG_POOL_MAX), the app must
    fall back to Mongo — and must NOT re-attempt on every single store call. The failure is
    cached in a sentinel; a retry storm on every find_one would be worse than the outage."""
    from app.stores import pg

    monkeypatch.setenv("POSTGRES_URL", PG_URL)
    monkeypatch.setenv("CEREBROZEN_PG_POOL_MAX", "ten")  # not an int
    monkeypatch.setattr(pg, "_pool", None)

    assert pg.get_pool() is None
    assert pg._pool is False, "the failure is cached as a sentinel, not re-attempted"
    assert pg.get_pool() is None, "and a second call short-circuits on it"
    assert pg.collection("conv") is None and pg.client() is None, "→ the stores use Mongo"


@requires_pg
def test_get_pool_opens_a_real_pool_from_the_env(monkeypatch):
    """The env var is the switch that moves the whole app from Mongo to Postgres."""
    from app.stores import pg

    monkeypatch.setenv("POSTGRES_URL", PG_URL)
    monkeypatch.setattr(pg, "_pool", None)
    try:
        pool = pg.get_pool()
        assert pool is not None and pg.get_pool() is pool, "the pool is opened once and cached"
        with pool.connection() as conn:
            assert conn.execute("SELECT 1").fetchone() == (1,)
        assert pg.postgres_url() == PG_URL
    finally:
        if pg._pool:
            pg._pool.close()


def test_the_stores_only_use_operators_the_shim_actually_emulates():
    """The shim emulates a FIXED operator set. An unsupported operator is not rejected — it
    is IGNORED, and `_matches` then returns True for every document (e.g. {"n": {"$gt": 5}}
    matches n=1). So the day someone adds a $gt/$or/$elemMatch query to a store, the
    Postgres deployment silently starts matching the wrong rows.

    This test is the tripwire: it fails the moment a store uses an operator pg.py cannot
    honour."""
    from app.stores.pg import _matches

    assert _matches({"n": 1}, {"n": {"$gt": 5}}) is True, (
        "documents the hole this test guards: an unknown operator matches EVERYTHING"
    )

    supported = {"$set", "$setOnInsert", "$inc", "$push", "$each", "$slice", "$addToSet",
                 "$ne", "$exists", "$in", "$nin", "$match", "$project", "$size", "$ifNull"}
    store_dir = Path(__file__).resolve().parent.parent / "app" / "stores"
    used: set[str] = set()
    for path in store_dir.glob("*.py"):
        if path.name in ("pg.py", "__init__.py"):
            continue
        # Operators appear as KEYS ("$set": …); "$messages" and friends are field
        # references (values inside an aggregation expression), not operators.
        used |= set(re.findall(r'"(\$[a-zA-Z]+)"\s*:', path.read_text()))

    unsupported = used - supported
    assert not unsupported, (
        f"{sorted(unsupported)} is used by a store but NOT emulated by app/stores/pg.py — "
        "on Postgres it will be silently ignored and match the wrong documents"
    )


# ════════════════════════════════════════════════════════════════════════════
#  BUGS FOUND — each of these fails today. They are pinned, not worked around.
# ════════════════════════════════════════════════════════════════════════════


@requires_pg
def test_delete_session_actually_deletes_the_session_on_postgres(pgdb):
    """app/stores/conversation.py:530 sends `delete_one({"session_id": s, "user_id": u})`.

    The conversation row's _id IS the session_id (it was written by a filter that only
    carried session_id). But `_key` scans ("_id", "user_id", "session_id") IN THAT ORDER,
    so it picks the USER id and deletes `WHERE _id = <user_id>` — which matches nothing.
    "Delete my session" returns False and the transcript stays in the database forever.

    The second half is worse and is masked by the first: delete_one deletes purely BY KEY
    and never applies `_matches`, so the user_id in that filter is not an ownership check
    at all. Fixing the key order alone would turn this into a cross-user delete."""
    coll = pgdb.collection("conv")
    coll.update_one(  # exactly how conversation.py writes a session
        {"session_id": "s1"},
        {"$set": {"session_id": "s1", "user_id": "owner", "messages": [{"n": 1}]}},
        upsert=True,
    )

    result = coll.delete_one({"session_id": "s1", "user_id": "owner"})
    assert result.deleted_count == 1, "the owner must be able to delete their own session"
    assert pgdb.rows(coll.name) == []


@requires_pg
def test_delete_one_honours_the_whole_filter_so_it_cannot_delete_another_users_session(pgdb):
    """delete_one resolves ONE identity field out of the filter and runs
    `DELETE … WHERE _id = <that>`. Every other condition is dropped on the floor.

    Mongo would delete nothing here (the precondition does not match). The shim deletes the
    row. Today the damage is capped by the `_key` ordering bug above (the wrong field is
    picked, so nothing matches at all) — but that means the ownership scoping
    `conversation.delete_session` documents is not enforced ANYWHERE, and fixing `_key`
    alone converts this into a cross-user delete."""
    coll = pgdb.collection("agentic")  # keyed by user_id, so the key DOES resolve
    coll.update_one({"user_id": "u1"}, {"$set": {"user_id": "u1", "tier": "paid"}}, upsert=True)

    result = coll.delete_one({"user_id": "u1", "tier": "free"})  # precondition is FALSE
    assert result.deleted_count == 0, "a filter condition that does not match must not delete"
    assert len(pgdb.rows(coll.name)) == 1


@requires_pg
def test_delete_many_reports_what_the_database_actually_removed(pgdb):
    """delete_many re-derives the key from each matched DOCUMENT — and `_key` prefers
    user_id, while a conversation row's _id is its session_id. So the DELETE matches no row,
    yet `n` is incremented anyway and the caller is told the documents are gone.

    A deletion routine that lies about its result is the worst kind: a GDPR/erase path built
    on this would report success and leave every transcript in place."""
    coll = pgdb.collection("conv")
    for sid in ("s1", "s2"):
        coll.update_one({"session_id": sid},
                        {"$set": {"session_id": sid, "user_id": "u1"}}, upsert=True)

    result = coll.delete_many({"user_id": "u1"})
    assert (result.deleted_count, pgdb.rows(coll.name)) == (2, []), (
        "delete_many must remove what it says it removed"
    )


@requires_pg
def test_the_ingest_pipeline_can_write_to_the_pgvector_backend(
    pgvec, fake_embedder, monkeypatch
):
    """`embed_and_upsert` sets `rec["vector"]` and hands the records to `store.upsert`,
    which forwards them to `pgvector_store.upsert`, which reads `r["embedding"]` → KeyError.

    ingest_sskb/ingest_cskb catch that per-document, log it, and report 0 chunks. So on the
    pgvector backend the corpus stays EMPTY, every extraction resolves to null, and the only
    symptom is a coach that has quietly stopped citing anything."""
    from app.rag import store
    from app.rag.ingest import embed_and_upsert

    monkeypatch.setenv("CEREBROZEN_RAG_BACKEND", "pgvector")
    written = embed_and_upsert("sskb", [
        {"id": "a", "doc_key": "sskb/a.pdf", "text": "psychological safety", "kb": "sskb",
         "source": "concept"},
    ])
    assert written == 1, "the ingest pipeline must be able to fill the pgvector index"
    assert store.count("sskb") == 1


def test_the_local_dev_ingest_path_never_touches_s3(lance, fake_embedder, fake_s3, tmp_path):
    """`ingest_sskb_local(micro_path=…)` is documented as "Dev path: ingest from local files
    (no S3)". It hands the local path to `chunk_doc()`, which treats every path as an S3 KEY
    and downloads it — so on a laptop with no AWS credentials the only usable local path is
    `curated_path`, and the three document KBs cannot be loaded offline at all.

    (The S3 client is faked here purely so the test cannot make a network call; that it is
    reached at all is the finding.)"""
    from app.rag.ingest.sskb import ingest_sskb_local

    fake_s3({})  # an empty bucket: ANY download attempt is the bug
    doc = tmp_path / "kb.pdf"
    doc.write_bytes(_pdf_bytes("A two minute bite"))

    written = {}
    for field, kwarg in (("micro_learning", "micro_path"), ("concepts", "concepts_path"),
                         ("competency", "competency_path")):
        try:
            written[field] = ingest_sskb_local(**{kwarg: str(doc)})[field]
        except Exception as exc:  # noqa: BLE001 — the S3 round-trip failing IS the finding
            written[field] = f"{type(exc).__name__}: {exc}"

    assert written == {"micro_learning": 1, "concepts": 1, "competency": 1}, (
        "the local dev path must be able to read local files"
    )


@requires_pg
def test_both_backends_agree_that_score_is_a_distance(pgvec, lance, fake_embedder,
                                                              monkeypatch):
    """`app/rag/store.py` documents `_score` as "LanceDB distance, lower = closer", and
    `extractors._extract_learning_aid` relies on it:

        idx = min(range(len(pool)), key=lambda i: pool[i].get("_score") …)

    `pgvector_store.search` returns `1 - (embedding <=> query)` — a similarity, where HIGHER
    is closer. Same function, same field, inverted meaning: the min() fallback then picks
    the worst match in the pool and shows it to the user as their learning aid."""
    from app.rag import store

    near, far = "delegation and trust", "quarterly revenue spreadsheets"

    # Postgres-first: with POSTGRES_URL set, pgvector is the default backend,
    # so the LanceDB half of this comparison must opt out explicitly.
    monkeypatch.setenv("CEREBROZEN_RAG_BACKEND", "lancedb")
    lance.upsert("sskb", [_rec("near", near, kb="sskb"), _rec("far", far, kb="sskb")])
    lance_hits = {h["id"]: h["_score"] for h in store.search("sskb", _fake_vector(near), top_k=2)}
    assert lance_hits["near"] < lance_hits["far"], "LanceDB: lower is closer"

    monkeypatch.setenv("CEREBROZEN_RAG_BACKEND", "pgvector")
    pgvec.upsert("sskb", [
        {"id": "near", "text": near, "embedding": _fake_vector(near)},
        {"id": "far", "text": far, "embedding": _fake_vector(far)},
    ])
    pg_hits = {h["text"]: h["_score"] for h in store.search("sskb", _fake_vector(near), top_k=2)}
    assert pg_hits[near] < pg_hits[far], (
        "both backends must agree on what _score MEANS — every consumer of store.search() "
        "ranks on it"
    )


@requires_pg
def test_pgvector_indexed_docs_returns_the_etag_so_ingestion_can_skip_unchanged_docs(pgvec, fake_embedder):
    """`indexed_docs()` exists for exactly one purpose: `ingest_*` compares its values to
    the S3 ETag to decide whether a document changed. pgvector returns the doc_key as the
    value, so the comparison always fails — the entire corpus is re-embedded on every
    restart (a recurring embedding bill), and each doc's chunks are deleted first, so a
    concurrent read during boot sees an empty knowledge base.

    The ETag IS stored — it rides in the `meta` JSONB (`meta->>'s3_etag'`)."""
    pgvec.upsert("sskb", [{"id": "a", "doc_key": "sskb/x.pdf", "text": "t",
                           "embedding": _fake_vector("t"), "s3_etag": "ETAG-1"}])

    assert pgvec.indexed_docs("sskb") == {"sskb/x.pdf": "ETAG-1"}


@requires_pg
def test_the_pgvector_backend_bootstraps_its_own_extension_on_a_clean_database(monkeypatch):
    """The pgvector Docker image ships the extension but does NOT enable it in your
    database; `CREATE EXTENSION vector` is a per-database step. `_ensure` creates the table,
    the indexes and guards the dimension — but never creates the extension, and there is no
    migration/bootstrap script in the repo that does. A correct deployment therefore fails
    at the first upsert, and (because store.search's caller swallows it) looks exactly like
    an empty knowledge base."""
    import psycopg
    from psycopg_pool import ConnectionPool

    from app.rag import pgvector_store as pgv
    from app.stores import pg

    admin = PG_URL.rsplit("/", 1)[0] + "/postgres"
    with psycopg.connect(admin, autocommit=True) as conn:
        conn.execute("DROP DATABASE IF EXISTS cerebrozen_rag_fresh WITH (FORCE)")
        conn.execute("CREATE DATABASE cerebrozen_rag_fresh")
    fresh_url = PG_URL.rsplit("/", 1)[0] + "/cerebrozen_rag_fresh"

    # min_size must be given explicitly: psycopg_pool defaults it to 4, so a bare
    # `max_size=2` is rejected with "max_size must be greater or equal than min_size".
    pool = ConnectionPool(conninfo=fresh_url, kwargs={"autocommit": True}, open=True,
                          min_size=1, max_size=2)
    monkeypatch.setattr(pg, "_pool", pool)
    monkeypatch.setattr(pgv, "_ready", set())
    try:
        assert pgv.upsert("sskb", [{"id": "a", "text": "t", "embedding": [0.1] * 4}]) == 1, (
            "the backend must bootstrap its own extension on a clean database"
        )
    finally:
        pool.close()
        with psycopg.connect(admin, autocommit=True) as conn:
            conn.execute("DROP DATABASE IF EXISTS cerebrozen_rag_fresh WITH (FORCE)")
